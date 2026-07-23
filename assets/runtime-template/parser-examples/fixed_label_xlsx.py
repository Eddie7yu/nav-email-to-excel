"""完全虚构的固定标签/值 XLSX + 配套 PDF 本地解析器示例。

复制到运行目录 app/parsers/ 后，必须先把下列常量改成已由本地证据核实的值，
再用脱敏夹具和真实只读候选验证。不要直接启用这个示例。
"""

from __future__ import annotations

import datetime as dt
import io
import math
import re
from email.message import Message
from pathlib import Path
from typing import Any

import openpyxl

from nav_parse import NavRow, ParseError


EXPECTED_PRODUCT_CODE = "SAMPLE01"
EXPECTED_PRODUCT_NAME = "虚构示例产品"
EXPECTED_MANAGER = "虚构示例管理人"
EXPECTED_CUSTODIAN = "虚构示例托管人"
EXPECTED_SHEET = "净值通知"
MAX_XLSX_BYTES = 5 * 1024 * 1024
LABELS = (
    ("产品代码", "code"),
    ("产品名称", "product_name"),
    ("净值日期", "date"),
    ("单位净值", "unit"),
    ("累计净值", "cumulative"),
    ("管理人", "manager"),
    ("托管人", "custodian"),
)


def _payload(part: Message, kind: str) -> bytes:
    payload = part.get_payload(decode=True)
    if not isinstance(payload, bytes) or not payload:
        raise ParseError(f"{kind} attachment is empty")
    return payload


def _subject_date(subject: str) -> dt.date:
    matches = {
        dt.date(int(year), int(month), int(day))
        for year, month, day in re.findall(
            r"(?<!\d)(20\d{2})[-年/](\d{1,2})[-月/](\d{1,2})(?:日)?(?!\d)",
            subject,
        )
    }
    if len(matches) != 1:
        raise ParseError("subject must contain exactly one NAV date")
    return next(iter(matches))


def _date(value: Any) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    match = re.fullmatch(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", text)
    if not match:
        raise ParseError("worksheet NAV date is invalid")
    return dt.date(*(int(group) for group in match.groups()))


def _number(value: Any, label: str) -> float:
    if isinstance(value, bool):
        raise ParseError(f"{label} is invalid")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ParseError(f"{label} is invalid") from exc
    if not math.isfinite(number) or not 0 < number < 1_000_000:
        raise ParseError(f"{label} is outside the accepted range")
    return number


def parse_message(message: Message) -> list[NavRow]:
    subject = str(message.get("Subject") or "")
    if not re.search(
        rf"(?<![A-Za-z0-9]){re.escape(EXPECTED_PRODUCT_CODE)}(?![A-Za-z0-9])",
        subject,
        re.IGNORECASE,
    ):
        raise ParseError("subject product code does not match")
    subject_date = _subject_date(subject)

    attachments = list(message.iter_attachments())
    xlsx_parts = [
        part
        for part in attachments
        if Path(str(part.get_filename() or "")).suffix.casefold() == ".xlsx"
    ]
    pdf_parts = [
        part
        for part in attachments
        if Path(str(part.get_filename() or "")).suffix.casefold() == ".pdf"
    ]
    if len(attachments) != 2 or len(xlsx_parts) != 1 or len(pdf_parts) != 1:
        raise ParseError("expected exactly one XLSX and one PDF attachment")

    pdf_payload = _payload(pdf_parts[0], "PDF")
    if not pdf_payload.startswith(b"%PDF-"):
        raise ParseError("PDF companion attachment is invalid")
    xlsx_payload = _payload(xlsx_parts[0], "XLSX")
    if len(xlsx_payload) > MAX_XLSX_BYTES:
        raise ParseError("XLSX attachment exceeds the parser limit")

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(xlsx_payload), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ParseError("XLSX attachment cannot be opened") from exc
    try:
        if workbook.sheetnames != [EXPECTED_SHEET]:
            raise ParseError("XLSX worksheet identity does not match")
        sheet = workbook[EXPECTED_SHEET]
        values: dict[str, Any] = {}
        for row, (expected_label, field) in enumerate(LABELS, 1):
            label = str(sheet.cell(row, 1).value or "").strip()
            if label != expected_label:
                raise ParseError("XLSX fixed label layout does not match")
            values[field] = sheet.cell(row, 2).value
    finally:
        workbook.close()

    if str(values["code"] or "").strip().upper() != EXPECTED_PRODUCT_CODE:
        raise ParseError("worksheet product code does not match")
    if str(values["product_name"] or "").strip() != EXPECTED_PRODUCT_NAME:
        raise ParseError("worksheet product name does not match")
    if str(values["manager"] or "").strip() != EXPECTED_MANAGER:
        raise ParseError("worksheet manager identity does not match")
    if str(values["custodian"] or "").strip() != EXPECTED_CUSTODIAN:
        raise ParseError("worksheet custodian identity does not match")

    nav_date = _date(values["date"])
    if nav_date != subject_date:
        raise ParseError("subject and worksheet NAV dates disagree")
    unit = _number(values["unit"], "unit NAV")
    cumulative = _number(values["cumulative"], "cumulative NAV")
    return [
        NavRow(
            date=nav_date,
            unit=unit,
            cumulative=cumulative,
            code=EXPECTED_PRODUCT_CODE,
            source="local:fixed-label-xlsx",
        )
    ]
