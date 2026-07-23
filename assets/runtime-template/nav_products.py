from __future__ import annotations

import datetime as dt
import json
import re
import shutil
from copy import deepcopy
from pathlib import Path
from typing import Any

import openpyxl

from nav_automation import approve, status as approval_status
from nav_config import (
    STATE_ROOT,
    active_routes,
    benchmark_requires_review,
    benchmark_review_issue,
    normalize_code,
    validate_config,
    write_json_atomic,
)
from nav_product_workbook import (
    clone_product_sheet,
    prepare_clone_spec,
    restore_after_config_failure,
)
from nav_template import add_template_product
from nav_workbook import (
    _data_frequency,
    _summary_reserved_row,
    discover_layout,
    existing_rows,
)


PROPOSALS = STATE_ROOT / "route-proposals.json"
PARTIAL_PROPOSALS = STATE_ROOT / "route-proposals.partial.json"


class ProductError(RuntimeError):
    pass


def _coverage_matrix(
    config: dict[str, Any],
    sheets: list[str],
    route_sheets: set[str],
    source_sheets: set[str],
) -> dict[str, Any]:
    reviews = config.get("sheet_reviews") or {}
    active_sheet_names = {
        str(route.get("sheet") or "")
        for route in active_routes(config)
        if isinstance(route, dict)
    }
    review_route_issues = {
        str(route.get("sheet") or ""): str(benchmark_review_issue(route))
        for route in active_routes(config)
        if isinstance(route, dict) and benchmark_requires_review(route)
    }
    review_route_sheets = set(review_route_issues)
    paused_sheet_names = route_sheets - active_sheet_names
    statuses = (
        "active",
        "active_review_required",
        "paused",
        "reference",
        "excluded",
        "no_mail_evidence",
        "local_parser_required",
        "business_review",
        "unclassified",
    )
    counts = {status: 0 for status in statuses}
    items: list[dict[str, Any]] = []
    for sheet in sheets:
        reason = None
        if sheet in review_route_sheets:
            status = "active_review_required"
            reason = review_route_issues[sheet]
        elif sheet in active_sheet_names:
            status = "active"
        elif sheet in paused_sheet_names:
            status = "paused"
        elif sheet in source_sheets:
            status = "reference"
        elif isinstance(reviews.get(sheet), dict):
            status = str(reviews[sheet].get("status") or "unclassified")
            reason = reviews[sheet].get("reason")
        else:
            status = "unclassified"
        counts[status] += 1
        item = {"sheet": sheet, "status": status}
        if reason:
            item["reason"] = reason
        items.append(item)
    action_required_statuses = {
        "active_review_required",
        "local_parser_required",
        "business_review",
        "unclassified",
    }
    return {
        "total_workbook_sheets": len(sheets),
        "classified_sheets": len(sheets) - counts["unclassified"],
        "all_sheets_classified": counts["unclassified"] == 0,
        "action_required_sheets": sum(
            counts[status] for status in action_required_statuses
        ),
        "counts": counts,
        "sheets": items,
        "reviews_for_missing_sheets": sorted(
            str(sheet) for sheet in reviews if sheet not in sheets
        ),
    }


def _workbook_sheets(config: dict[str, Any]) -> list[str]:
    path = Path(config["workbook_path"]).expanduser().resolve()
    if not path.is_file():
        return []
    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=path.suffix.lower() == ".xlsm",
    )
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _read_proposals() -> dict[str, Any]:
    if PARTIAL_PROPOSALS.is_file():
        raise ProductError(
            "候选扫描尚未完成；请先运行 propose --resume，不能使用旧候选接管"
        )
    try:
        report = json.loads(PROPOSALS.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ProductError("缺少产品候选报告；请先运行 products sync") from exc
    except json.JSONDecodeError as exc:
        raise ProductError("产品候选报告损坏；请重新运行 products sync") from exc
    if not isinstance(report, dict) or not isinstance(report.get("candidates"), list):
        raise ProductError("产品候选报告格式无效；请重新运行 products sync")
    return report


def _proposal_items(report: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for proposal_index, candidate in enumerate(report.get("candidates") or [], 1):
        if not isinstance(candidate, dict):
            continue
        sender = str(candidate.get("sender") or "").strip().lower()
        codes = [
            normalize_code(code)
            for code in candidate.get("detected_codes") or []
            if normalize_code(code)
        ]
        for code in codes or [None]:
            items.append(
                {
                    "proposal_index": proposal_index,
                    "sender": sender,
                    "code": code,
                    "first_date": candidate.get("first_date"),
                    "latest_date": candidate.get("latest_date"),
                    "message_count": candidate.get("message_count", 0),
                    "subject_examples": candidate.get("subject_examples") or [],
                }
            )
    return items


def _route_matches_item(route: dict[str, Any], item: dict[str, Any]) -> bool:
    if str(route.get("sender") or "").strip().lower() != item["sender"]:
        return False
    expected = normalize_code(route.get("code"))
    if expected:
        accepted = {expected}
        accepted.update(
            normalized
            for alias in route.get("code_aliases") or []
            if (normalized := normalize_code(alias))
        )
        return item["code"] in accepted
    return item["code"] is None


def _possible_share_alias(route: dict[str, Any], item: dict[str, Any]) -> bool:
    if str(route.get("sender") or "").strip().lower() != item["sender"]:
        return False
    expected = normalize_code(route.get("code"))
    candidate = normalize_code(item.get("code"))
    if not expected or not candidate or expected == candidate:
        return False
    shorter, longer = sorted((expected, candidate), key=len)
    suffix = longer[len(shorter) :] if longer.startswith(shorter) else ""
    return bool(suffix and len(suffix) <= 12 and re.fullmatch(r"[A-Z0-9_-]+", suffix))


def _analysis(config: dict[str, Any], report: dict[str, Any] | None) -> dict[str, Any]:
    routes = [route for route in config.get("routes") or [] if isinstance(route, dict)]
    sheets = _workbook_sheets(config)
    route_sheets = {str(route.get("sheet") or "") for route in routes}
    source_sheets = {
        str(route["benchmark"]["source_sheet"])
        for route in routes
        if isinstance(route.get("benchmark"), dict)
        and route["benchmark"].get("source_sheet")
    }
    items = _proposal_items(report or {"candidates": []})
    matched: list[dict[str, Any]] = []
    new_candidates: list[dict[str, Any]] = []
    alias_pending: list[dict[str, Any]] = []
    for item in items:
        matches = [route for route in routes if _route_matches_item(route, item)]
        if matches:
            for route in matches:
                matched.append(
                    {
                        **item,
                        "sheet": route["sheet"],
                        "paused": bool(route.get("paused", False)),
                    }
                )
        else:
            possible_aliases = [
                route for route in routes if _possible_share_alias(route, item)
            ]
            if len(possible_aliases) == 1:
                alias_pending.append(
                    {
                        **item,
                        "possible_sheet": possible_aliases[0]["sheet"],
                        "reason": "share-class-suffix-difference",
                        "requires_confirmation": True,
                    }
                )
            else:
                new_candidates.append(item)
    configured_not_seen = [
        {
            "sheet": route.get("sheet"),
            "sender": route.get("sender"),
            "code": normalize_code(route.get("code")),
            "paused": bool(route.get("paused", False)),
        }
        for route in routes
        if not any(_route_matches_item(route, item) for item in items)
    ]
    route_status = [
        {
            "sheet": route.get("sheet"),
            "sender": route.get("sender"),
            "code": normalize_code(route.get("code")),
            "product_name": route.get("product_name"),
            "paused": bool(route.get("paused", False)),
            "pause_reason": route.get("pause_reason"),
            "review_required": benchmark_requires_review(route),
            "review_issue": benchmark_review_issue(route),
            "sheet_present": route.get("sheet") in sheets,
        }
        for route in routes
    ]
    coverage = _coverage_matrix(config, sheets, route_sheets, source_sheets)
    return {
        "passed": True,
        "changed": False,
        "configured": len(routes),
        "active": len(active_routes(config)),
        "paused": len(routes) - len(active_routes(config)),
        "review_required": sum(
            1
            for route in active_routes(config)
            if benchmark_requires_review(route)
        ),
        "routes": route_status,
        "matched_candidates": matched,
        "new_candidates": new_candidates,
        "alias_pending": alias_pending,
        "configured_not_seen_in_lookback": configured_not_seen,
        "workbook_missing_sheets": [
            item["sheet"] for item in route_status if not item["sheet_present"]
        ],
        "unmanaged_workbook_sheets": [
            name for name in sheets if name not in route_sheets | source_sheets
        ],
        "workbook_coverage": coverage,
        "automatic_updates": approval_status(config),
    }


def status(config: dict[str, Any]) -> dict[str, Any]:
    report = None
    if PROPOSALS.is_file():
        try:
            report = _read_proposals()
        except ProductError:
            report = None
    return _analysis(config, report)


def sync(
    config: dict[str, Any],
    refresh: bool = True,
    *,
    sender: str | None = None,
    subject_contains: str | None = None,
) -> dict[str, Any]:
    if refresh:
        from nav_service import propose_routes

        report = propose_routes(
            config, sender=sender, subject_contains=subject_contains
        )
    else:
        report = _read_proposals()
    result = _analysis(config, report)
    result["scan_passed"] = bool(report.get("passed"))
    result["warnings"] = report.get("warnings") or []
    result["errors"] = report.get("errors") or []
    result["passed"] = bool(report.get("passed"))
    return result


def _candidate(proposal_index: int) -> dict[str, Any]:
    report = _read_proposals()
    candidates = report.get("candidates") or []
    if not 1 <= proposal_index <= len(candidates):
        raise ProductError(f"proposal index must be between 1 and {len(candidates)}")
    candidate = candidates[proposal_index - 1]
    if not isinstance(candidate, dict):
        raise ProductError("Selected proposal is invalid")
    return candidate


def _candidate_subject(
    candidate: dict[str, Any], requested: str | None
) -> str | None:
    selection = candidate.get("selection")
    if selection is None:
        selection = _read_proposals().get("selection")
    if selection is not None and not isinstance(selection, dict):
        raise ProductError("产品候选的邮箱选择范围无效；请重新运行 products sync")
    inherited = None
    if isinstance(selection, dict):
        selected_sender = str(selection.get("sender") or "").strip().casefold()
        candidate_sender = str(candidate.get("sender") or "").strip().casefold()
        if selected_sender and selected_sender != candidate_sender:
            raise ProductError("产品候选与邮箱选择范围的发件人不一致；请重新扫描")
        value = selection.get("subject_contains")
        inherited = str(value).strip() if value is not None else None
        inherited = inherited or None
    explicit = str(requested).strip() if requested is not None else None
    explicit = explicit or None
    if inherited and explicit and inherited.casefold() != explicit.casefold():
        raise ProductError(
            "不能把选择式邮箱发现的主题范围替换为其他范围；请重新按目标范围运行 propose"
        )
    return inherited or explicit


def _candidate_subject_product_code(candidate: dict[str, Any]) -> str | None:
    selection = candidate.get("selection")
    if selection is None:
        selection = _read_proposals().get("selection")
    if not isinstance(selection, dict):
        return None
    binding = normalize_code(selection.get("subject_product_code"))
    if not binding:
        return None
    subject = str(selection.get("subject_contains") or "").strip()
    if normalize_code(subject) != binding:
        raise ProductError("主题产品代码绑定与候选主题范围不一致；请重新扫描")
    return binding


def _benchmark(
    source_sheet: str | None,
    source_type: str,
    source_date: str,
    source_value: str,
    display_name: str | None,
) -> dict[str, Any] | None:
    if not source_sheet:
        return None
    result = {
        "source_sheet": source_sheet.strip(),
        "source_type": source_type,
        "source_date": source_date.strip().upper(),
        "source_value": source_value.strip().upper(),
    }
    if display_name:
        result["display_name"] = display_name.strip()
    return result


def _detected_codes(candidate: dict[str, Any]) -> set[str]:
    return {
        normalized
        for value in candidate.get("detected_codes") or []
        if (normalized := normalize_code(value))
    }


def _select_code(
    candidate: dict[str, Any],
    requested: str | None,
    workbook_codes: set[str] | None = None,
) -> str | None:
    detected = _detected_codes(candidate)
    selected = normalize_code(requested)
    if selected:
        if detected and selected not in detected:
            raise ProductError("指定的产品代码未出现在该邮箱候选中")
        return selected
    workbook_codes = {value for value in (workbook_codes or set()) if value}
    overlap = detected & workbook_codes
    if len(overlap) == 1:
        return next(iter(overlap))
    if len(overlap) > 1:
        raise ProductError("邮箱候选与工作表匹配到多个产品代码，无法唯一接管")
    if len(detected) == 1:
        return next(iter(detected))
    if len(detected) > 1:
        raise ProductError(
            "该邮箱候选包含多个产品代码，AI 必须先按工作表证据唯一确定代码"
        )
    return None


def _candidate_observations(
    candidate: dict[str, Any], code: str | None
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for item in candidate.get("observations") or []:
        if not isinstance(item, dict):
            continue
        item_code = normalize_code(item.get("code"))
        if code and item_code != code:
            continue
        try:
            dt.date.fromisoformat(str(item["date"]))
            float(item["unit"])
        except (KeyError, TypeError, ValueError):
            continue
        result.append(item)
    return sorted(result, key=lambda item: str(item["date"]))


def _sheet_values(sheet, layout, field: str) -> list[str]:
    column = layout.columns.get(field)
    if not column:
        return []
    end = layout.summary_row - 1 if layout.mode == "summary" else layout.last_data_row
    result: list[str] = []
    for row in range(layout.data_start, max(layout.data_start, end) + 1):
        value = str(sheet.cell(row, column).value or "").strip()
        if value:
            result.append(value)
    return result


def _infer_existing_sheet(
    config: dict[str, Any],
    candidate: dict[str, Any],
    *,
    sheet_name: str,
    requested_code: str | None,
    requested_product_name: str | None,
) -> dict[str, Any]:
    master = Path(config["workbook_path"]).expanduser().resolve()
    if str(config.get("workbook_mode", "existing")) != "existing":
        raise ProductError("products adopt 只用于用户已有工作簿")
    if any(
        str(route.get("sheet")) == sheet_name
        for route in config.get("routes") or []
        if isinstance(route, dict)
    ):
        raise ProductError(f"该 Sheet 已经受管：{sheet_name}")
    workbook = openpyxl.load_workbook(
        master,
        data_only=False,
        keep_vba=master.suffix.lower() == ".xlsm",
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ProductError(f"工作簿中找不到目标 Sheet：{sheet_name}")
        sheet = workbook[sheet_name]
        detected_codes = _detected_codes(candidate)
        if requested_code:
            provisional_code = _select_code(candidate, requested_code)
        elif len(detected_codes) == 1:
            provisional_code = next(iter(detected_codes))
        else:
            # Keep layout discovery independent from a still-ambiguous mailbox
            # candidate.  Once the code column is known, the worksheet can provide
            # the unique overlap that resolves a multi-product sender safely.
            provisional_code = None
        provisional_name = (requested_product_name or sheet_name).strip()
        layouts: list[tuple[str, Any]] = []
        errors: list[str] = []
        for mode in ("summary", "append"):
            route = {
                "sheet_mode": mode,
                "code": provisional_code,
                "product_name": provisional_name,
                "data_frequency": "auto",
            }
            try:
                layouts.append(
                    (
                        mode,
                        discover_layout(
                            sheet,
                            (config.get("column_overrides") or {}).get(sheet_name),
                            route,
                        ),
                    )
                )
            except RuntimeError as exc:
                errors.append(f"{mode}: {exc}")
        if not layouts:
            raise ProductError(
                f"无法从产品名/代码、日期和表头识别 {sheet_name}：{'；'.join(errors)}"
            )
        mode, layout = layouts[0]
        workbook_codes = {
            normalized
            for value in _sheet_values(sheet, layout, "code")
            if (normalized := normalize_code(value))
        }
        selected_code = _select_code(candidate, requested_code, workbook_codes)
        names = _sheet_values(sheet, layout, "name")
        product_name = (
            requested_product_name.strip()
            if requested_product_name and requested_product_name.strip()
            else names[-1]
            if names
            else sheet_name
        )
        route = {
            "sheet_mode": mode,
            "code": selected_code,
            "product_name": product_name,
            "data_frequency": "auto",
        }
        layout = discover_layout(
            sheet,
            (config.get("column_overrides") or {}).get(sheet_name),
            route,
        )
        current = existing_rows(sheet, layout)
        reserved = _summary_reserved_row(sheet, layout, route, current)
        adoption_series_start = (
            None
            if not current or reserved is not None
            else (max(current) + dt.timedelta(days=1)).isoformat()
        )
        frequency, frequency_source, _weekday = _data_frequency(
            sheet, layout, route, current, reserved
        )
        if frequency not in {"daily", "weekly"}:
            raise ProductError(f"无法从 {sheet_name} 的日期规律或表头确定日频/周频")
        benchmark_columns = {
            semantic: layout.columns[semantic]
            for semantic in ("benchmark_level", "benchmark_return", "excess")
            if layout.columns.get(semantic)
        }
        column_roles = {
            semantic: {
                "column": openpyxl.utils.get_column_letter(column),
                "header": str(
                    sheet.cell(layout.header_row, column).value or ""
                ).strip(),
            }
            for semantic, column in sorted(
                layout.columns.items(), key=lambda item: item[1]
            )
        }
        benchmark_review_only = bool(
            benchmark_columns
            and any(
                sheet.cell(row, column).value not in (None, "")
                for column in benchmark_columns.values()
                for row in range(layout.data_start, layout.summary_row + 1)
            )
        )
        observations = _candidate_observations(candidate, selected_code)
        if layout.columns.get("cumulative"):
            if observations and all(
                item.get("cumulative") is not None for item in observations
            ):
                cumulative_policy = "require"
                return_basis = "cumulative"
            else:
                comparable = [
                    values
                    for values in current.values()
                    if values.get("unit") is not None
                    and values.get("cumulative") is not None
                ]
                if comparable and all(
                    abs(float(values["unit"]) - float(values["cumulative"])) <= 1e-9
                    for values in comparable
                ):
                    cumulative_policy = "unit"
                    return_basis = "unit"
                else:
                    raise ProductError(
                        "表内要求累计净值，但邮箱候选没有提供，且历史不能证明应以单位净值代替；这是需要用户确认的业务口径"
                    )
        else:
            cumulative_policy = "unit"
            return_basis = "unit"
        return {
            "sheet_mode": mode,
            "code": selected_code,
            "product_name": product_name,
            "frequency": frequency,
            "frequency_source": frequency_source,
            "series_start": adoption_series_start,
            "series_start_reason": (
                "existing-sheet-tail-plus-one-day"
                if adoption_series_start
                else "cold-start-or-reserved-row"
            ),
            "cumulative_policy": cumulative_policy,
            "return_basis": return_basis,
            "benchmark_review_only": benchmark_review_only,
            "benchmark_columns_present": sorted(benchmark_columns),
            "column_roles": column_roles,
        }
    finally:
        workbook.close()


def add(
    config: dict[str, Any],
    config_path: Path,
    *,
    proposal_index: int,
    sheet: str,
    frequency: str,
    code: str | None = None,
    product_name: str | None = None,
    subject_contains: str | None = None,
    sheet_mode: str | None = None,
    parser: str = "auto",
    cumulative_policy: str = "require",
    cumulative_offset: float | None = None,
    return_basis: str = "cumulative",
    series_start: str | None = None,
    baseline_overlap: str | None = None,
    max_staleness_days: int = 14,
    benchmark_source_sheet: str | None = None,
    benchmark_source_type: str = "level",
    benchmark_source_date: str = "A",
    benchmark_source_value: str = "B",
    benchmark_display_name: str | None = None,
    benchmark_review_only: bool = False,
) -> dict[str, Any]:
    candidate = _candidate(proposal_index)
    resolved_subject = _candidate_subject(candidate, subject_contains)
    subject_product_code = _candidate_subject_product_code(candidate)
    detected = {
        normalize_code(value)
        for value in candidate.get("detected_codes") or []
        if normalize_code(value)
    }
    selected_code = normalize_code(code)
    if selected_code and detected and selected_code not in detected:
        raise ProductError("Selected code was not found in this proposal")
    if selected_code is None:
        if len(detected) == 1:
            selected_code = next(iter(detected))
        elif len(detected) > 1:
            raise ProductError(
                "This proposal contains multiple product codes; use --code"
            )
    workbook_mode = str(config.get("workbook_mode", "existing"))
    if workbook_mode == "bundled-template":
        resolved_mode = "template"
        data_frequency = frequency
    else:
        if sheet_mode not in {"summary", "append"}:
            raise ProductError(
                "An existing workbook requires --sheet-mode summary or append"
            )
        resolved_mode = str(sheet_mode)
        data_frequency = "auto"
        workbook_sheets = _workbook_sheets(config)
        if sheet not in workbook_sheets:
            raise ProductError(
                "Existing workbook does not contain the target sheet; create or choose the product sheet before adding the route"
            )
        if benchmark_source_sheet and benchmark_source_sheet not in workbook_sheets:
            raise ProductError(
                "Existing workbook does not contain the selected benchmark source sheet"
            )
    route: dict[str, Any] = {
        "sender": str(candidate.get("sender") or "").strip().lower(),
        "subject_contains": resolved_subject,
        "sheet": sheet.strip(),
        "sheet_mode": resolved_mode,
        "code": selected_code,
        "product_name": (product_name or sheet).strip(),
        "parser": parser,
        "paused": False,
        "allow_sender_only": selected_code is None,
        "cumulative_policy": cumulative_policy,
        "return_basis": return_basis,
        "return_frequency": frequency,
        "data_frequency": data_frequency,
        "max_staleness_days": max_staleness_days,
        "benchmark": _benchmark(
            benchmark_source_sheet,
            benchmark_source_type,
            benchmark_source_date,
            benchmark_source_value,
            benchmark_display_name,
        ),
    }
    if benchmark_review_only:
        route["benchmark_review_only"] = True
    if subject_product_code:
        if selected_code != subject_product_code:
            raise ProductError("主题绑定代码与目标产品代码不一致；拒绝接管")
        route["subject_product_code"] = subject_product_code
    if cumulative_policy == "offset":
        if cumulative_offset is None:
            raise ProductError("--cumulative-offset is required for offset policy")
        route["cumulative_offset"] = cumulative_offset
    if series_start:
        route["series_start"] = series_start
    if baseline_overlap:
        route["baseline_overlap"] = baseline_overlap

    updated = deepcopy(config)
    updated.setdefault("sheet_reviews", {}).pop(route["sheet"], None)
    updated.setdefault("routes", []).append(route)
    validate_config(updated)
    template_result = None
    try:
        if workbook_mode == "bundled-template":
            template_result = add_template_product(updated, route)
        write_json_atomic(config_path, updated)
    except Exception:
        if template_result and template_result.get("backup"):
            backup = Path(str(template_result["backup"]))
            workbook = Path(updated["workbook_path"])
            if backup.is_file():
                shutil.copy2(backup, workbook)
        raise
    return {
        "passed": True,
        "changed": True,
        "action": "added",
        "sheet": route["sheet"],
        "code": route["code"],
        "route": route,
        "template": template_result,
        "requires_preview_approval": True,
        "automatic_updates": approval_status(updated),
    }


def adopt(
    config: dict[str, Any],
    config_path: Path,
    *,
    proposal_index: int,
    sheet: str,
    code: str | None = None,
    product_name: str | None = None,
    subject_contains: str | None = None,
    history_scope: str = "tail",
    inspect_only: bool = False,
) -> dict[str, Any]:
    """Connect a user-prepared sheet without asking them for internal modes."""

    if history_scope not in {"tail", "mail-history"}:
        raise ProductError(
            "history_scope must be tail or mail-history"
        )
    candidate = _candidate(proposal_index)
    profile = _infer_existing_sheet(
        config,
        candidate,
        sheet_name=sheet.strip(),
        requested_code=code,
        requested_product_name=product_name,
    )
    profile["history_scope"] = history_scope
    profile["history_scope_default"] = "tail"
    if history_scope == "mail-history":
        profile["series_start"] = None
        profile["series_start_reason"] = "supervised-mail-history"
        profile["baseline_overlap"] = None
    else:
        profile["baseline_overlap"] = (
            "last_existing_point" if profile["series_start"] else None
        )
    if inspect_only:
        _candidate_subject(candidate, subject_contains)
        return {
            "passed": True,
            "changed": False,
            "action": "inspected-existing-sheet",
            "sheet": sheet,
            "review_required": bool(profile["benchmark_review_only"]),
            "ready_for_direct_adoption": not bool(
                profile["benchmark_review_only"]
            ),
            "inference": profile,
            "master_unchanged": True,
        }
    result = add(
        config,
        config_path,
        proposal_index=proposal_index,
        sheet=sheet,
        frequency=profile["frequency"],
        code=profile["code"],
        product_name=profile["product_name"],
        subject_contains=subject_contains,
        sheet_mode=profile["sheet_mode"],
        cumulative_policy=profile["cumulative_policy"],
        return_basis=profile["return_basis"],
        series_start=profile["series_start"],
        baseline_overlap=profile["baseline_overlap"],
        benchmark_review_only=profile["benchmark_review_only"],
    )
    result["action"] = "adopted-existing-sheet"
    result["inference"] = profile
    return result


def add_code_alias(
    config: dict[str, Any], config_path: Path, *, sheet: str, code: str
) -> dict[str, Any]:
    """Persist one evidence-backed exact mailbox code alias for a managed route."""

    normalized = normalize_code(code)
    if not normalized:
        raise ProductError("别名代码不能为空")
    updated = deepcopy(config)
    route = next(
        (
            item
            for item in updated.get("routes") or []
            if str(item.get("sheet") or "") == sheet
        ),
        None,
    )
    if route is None:
        raise ProductError(f"找不到受管 Sheet：{sheet}")
    primary = normalize_code(route.get("code"))
    if not primary:
        raise ProductError("没有主产品代码的路由不能添加代码别名")
    if normalized == primary:
        return {
            "passed": True,
            "changed": False,
            "sheet": sheet,
            "code_alias": normalized,
            "automatic_updates": approval_status(config),
        }
    report = _read_proposals()
    observed = {
        normalize_code(value)
        for candidate in report.get("candidates") or []
        if str(candidate.get("sender") or "").strip().casefold()
        == str(route.get("sender") or "").strip().casefold()
        for value in candidate.get("detected_codes") or []
        if normalize_code(value)
    }
    if normalized not in observed:
        raise ProductError("该别名未出现在当前发件人的候选报告中；请先重新扫描并核实")
    aliases = {
        alias
        for value in route.get("code_aliases") or []
        if (alias := normalize_code(value))
    }
    aliases.add(normalized)
    route["code_aliases"] = sorted(aliases)
    validate_config(updated)
    write_json_atomic(config_path, updated)
    return {
        "passed": True,
        "changed": True,
        "sheet": sheet,
        "code_alias": normalized,
        "requires_preview_approval": True,
        "automatic_updates": approval_status(updated),
    }


def clone(
    config: dict[str, Any],
    config_path: Path,
    *,
    proposal_index: int,
    sheet: str,
    copy_from: str,
    code: str | None = None,
    product_name: str | None = None,
    subject_contains: str | None = None,
    inherit_benchmark: bool = False,
) -> dict[str, Any]:
    """Clone a managed summary sheet, remove its business data, and add a route."""

    target_sheet = sheet.strip()
    if (
        not target_sheet
        or len(target_sheet) > 31
        or re.search(r"[\\/*?:\[\]]", target_sheet)
    ):
        raise ProductError("目标 Sheet 名称为空、超过 31 字符或包含 Excel 禁用字符")
    candidate = _candidate(proposal_index)
    resolved_subject = _candidate_subject(candidate, subject_contains)
    subject_product_code = _candidate_subject_product_code(candidate)
    selected_code = _select_code(candidate, code)
    observations = _candidate_observations(candidate, selected_code)
    if not observations:
        raise ProductError("该邮箱候选没有可用于新产品冷启动的真实净值日期")
    first_date = dt.date.fromisoformat(str(observations[0]["date"]))
    source_route = _route(config, copy_from.strip())
    spec = prepare_clone_spec(config, source_route, target_sheet)
    source_benchmark = source_route.get("benchmark")
    if source_benchmark and not inherit_benchmark:
        raise ProductError(
            "参考 Sheet 含基准指数；“格式相同”不能证明新产品基准相同。AI 只有在可靠资料明确证明后才能增加 --inherit-benchmark"
        )
    if spec.columns.get("cumulative"):
        if all(item.get("cumulative") is not None for item in observations):
            cumulative_policy = "require"
            return_basis = "cumulative"
        elif str(source_route.get("cumulative_policy")) == "unit":
            cumulative_policy = "unit"
            return_basis = "unit"
        else:
            raise ProductError(
                "参考格式包含累计净值，但新产品邮件没有累计净值，且不能从参考产品推导新产品口径"
            )
    else:
        cumulative_policy = "unit"
        return_basis = "unit"
    frequency = str(source_route.get("return_frequency", ""))
    if frequency not in {"daily", "weekly"}:
        raise ProductError("参考 Sheet 的日频/周频配置无效，无法安全复制")
    route: dict[str, Any] = {
        "sender": str(candidate.get("sender") or "").strip().lower(),
        "subject_contains": resolved_subject,
        "sheet": target_sheet,
        "sheet_mode": "summary",
        "code": selected_code,
        "product_name": (product_name or target_sheet).strip(),
        "parser": "auto",
        "paused": False,
        "allow_sender_only": selected_code is None,
        "cumulative_policy": cumulative_policy,
        "return_basis": return_basis,
        "return_frequency": frequency,
        "data_frequency": "auto",
        "max_staleness_days": int(source_route.get("max_staleness_days", 14)),
        "benchmark": deepcopy(source_benchmark) if inherit_benchmark else None,
    }
    if subject_product_code:
        if selected_code != subject_product_code:
            raise ProductError("主题绑定代码与目标产品代码不一致；拒绝新建")
        route["subject_product_code"] = subject_product_code
    updated = deepcopy(config)
    updated.setdefault("sheet_reviews", {}).pop(target_sheet, None)
    updated.setdefault("routes", []).append(route)
    source_overrides = (config.get("column_overrides") or {}).get(copy_from.strip())
    if source_overrides:
        updated.setdefault("column_overrides", {})[target_sheet] = deepcopy(
            source_overrides
        )
    validate_config(updated)
    workbook_result = clone_product_sheet(config, source_route, route, first_date)
    try:
        write_json_atomic(config_path, updated)
    except Exception:
        backup = Path(str(workbook_result["backup"]))
        restore_after_config_failure(
            Path(config["workbook_path"]).expanduser().resolve(), backup
        )
        backup.unlink(missing_ok=True)
        raise
    return {
        "passed": True,
        "changed": True,
        "action": "cloned-and-adopted-sheet",
        "sheet": target_sheet,
        "copy_from": copy_from.strip(),
        "code": selected_code,
        "route": route,
        "workbook": workbook_result,
        "requires_preview_approval": True,
        "automatic_updates": approval_status(updated),
    }


def _route(updated: dict[str, Any], sheet: str) -> dict[str, Any]:
    matches = [
        route
        for route in updated.get("routes") or []
        if isinstance(route, dict) and str(route.get("sheet")) == sheet
    ]
    if len(matches) != 1:
        raise ProductError(
            f"Expected exactly one configured product sheet named: {sheet}"
        )
    return matches[0]


def pause(
    config: dict[str, Any], config_path: Path, *, sheet: str, reason: str
) -> dict[str, Any]:
    updated = deepcopy(config)
    route = _route(updated, sheet)
    if route.get("paused", False):
        return {
            "passed": True,
            "changed": False,
            "action": "already-paused",
            "sheet": sheet,
            "automatic_updates": approval_status(config),
        }
    was_approved = bool(approval_status(config)["approved"])
    route["paused"] = True
    route["pause_reason"] = reason.strip()
    validate_config(updated)
    write_json_atomic(config_path, updated)
    all_paused = not active_routes(updated)
    approval_preserved = bool(was_approved and not all_paused)
    if approval_preserved:
        approve(updated)
    return {
        "passed": True,
        "changed": True,
        "action": "paused",
        "sheet": sheet,
        "all_products_paused": all_paused,
        "approval_preserved": approval_preserved,
        "requires_schedule_removal": all_paused,
        "automatic_updates": approval_status(updated),
    }


def resume(config: dict[str, Any], config_path: Path, *, sheet: str) -> dict[str, Any]:
    updated = deepcopy(config)
    route = _route(updated, sheet)
    if not route.get("paused", False):
        return {
            "passed": True,
            "changed": False,
            "action": "already-active",
            "sheet": sheet,
            "automatic_updates": approval_status(config),
        }
    if sheet not in _workbook_sheets(updated):
        raise ProductError("Cannot resume a product whose worksheet is missing")
    route["paused"] = False
    route.pop("pause_reason", None)
    validate_config(updated)
    write_json_atomic(config_path, updated)
    return {
        "passed": True,
        "changed": True,
        "action": "resumed",
        "sheet": sheet,
        "requires_preview_approval": True,
        "automatic_updates": approval_status(updated),
    }
