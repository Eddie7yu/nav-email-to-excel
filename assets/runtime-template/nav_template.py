from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import uuid
from copy import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.properties import CalcProperties

from nav_config import ROOT, STATE_ROOT, active_routes


ASSET_NAME = "nav-standard-cn.xlsx"
TEMPLATE_SHEETS = {
    ("weekly", False): "模板-周度-无指数",
    ("weekly", True): "模板-周度-有指数",
    ("daily", False): "模板-日度-无指数",
    ("daily", True): "模板-日度-有指数",
}
INDEX_TEMPLATE = "模板-指数数据"
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


class TemplateError(RuntimeError):
    pass


def _column(value: Any) -> int:
    if isinstance(value, int) and value > 0:
        return value
    text = str(value or "").strip().upper()
    if text.isdigit() and int(text) > 0:
        return int(text)
    if re.fullmatch(r"[A-Z]{1,3}", text):
        return column_index_from_string(text)
    raise TemplateError(f"Invalid benchmark source column: {value!r}")


def _valid_sheet_name(name: str) -> bool:
    return bool(name) and len(name) <= 31 and not INVALID_SHEET_CHARS.search(name)


def _return_columns(frequency: str, has_benchmark: bool) -> list[int]:
    if frequency == "weekly":
        return [6, 8, 9] if has_benchmark else [6]
    return [6, 7, 8, 9] if has_benchmark else [6, 7]


def _add_return_rules(sheet, columns: list[int], threshold: float) -> None:
    for column in columns:
        target = f"{get_column_letter(column)}3:{get_column_letter(column)}10000"
        sheet.conditional_formatting.add(
            target,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(threshold)],
                font=Font(color="FFFF0000"),
            ),
        )
        sheet.conditional_formatting.add(
            target,
            CellIsRule(
                operator="lessThanOrEqual",
                formula=[str(-threshold)],
                font=Font(color="FF00B050"),
            ),
        )


def _copy_template_sheet(source, workbook, title: str, index: int | None = None):
    sheet = workbook.create_sheet(title=title, index=index)
    for row in source.iter_rows():
        for source_cell in row:
            target = sheet.cell(source_cell.row, source_cell.column)
            target.value = source_cell.value
            if source_cell.has_style:
                target._style = copy(source_cell._style)
            target.number_format = source_cell.number_format
            target.alignment = copy(source_cell.alignment)
            target.protection = copy(source_cell.protection)
    for merged in source.merged_cells.ranges:
        sheet.merge_cells(str(merged))
    for key, dimension in source.column_dimensions.items():
        target = sheet.column_dimensions[key]
        target.width = dimension.width
        target.hidden = dimension.hidden
        target.bestFit = dimension.bestFit
        target.outlineLevel = dimension.outlineLevel
    for key, dimension in source.row_dimensions.items():
        target = sheet.row_dimensions[key]
        target.height = dimension.height
        target.hidden = dimension.hidden
        target.outlineLevel = dimension.outlineLevel
    sheet.freeze_panes = source.freeze_panes
    sheet.sheet_format = copy(source.sheet_format)
    sheet.sheet_properties = copy(source.sheet_properties)
    sheet.page_margins = copy(source.page_margins)
    sheet.page_setup = copy(source.page_setup)
    sheet.print_options = copy(source.print_options)
    sheet.sheet_view.showGridLines = source.sheet_view.showGridLines
    sheet.auto_filter.ref = source.auto_filter.ref
    return sheet


def _configure_product_sheet(sheet, route: dict[str, Any], threshold: float) -> None:
    frequency = str(route["data_frequency"])
    has_benchmark = bool(route.get("benchmark"))
    sheet["A1"] = None
    display = str(
        (route.get("benchmark") or {}).get("display_name") or "基准指数"
    ).strip()
    if frequency == "weekly" and has_benchmark:
        sheet["G2"] = display
        sheet["H2"] = "指数收益(周度)"
    elif frequency == "daily" and has_benchmark:
        sheet["H2"] = (
            f"{display}收益(日度)" if display != "基准指数" else "指数收益(日度)"
        )
    _add_return_rules(sheet, _return_columns(frequency, has_benchmark), threshold)


def _sanitize_metadata(workbook) -> None:
    workbook.properties.creator = ""
    workbook.properties.lastModifiedBy = ""
    workbook.properties.title = ""
    workbook.properties.subject = ""
    workbook.properties.description = ""
    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def _set_index_headers(sheet, benchmark: dict[str, Any]) -> None:
    date_column = _column(benchmark["source_date"])
    value_column = _column(benchmark["source_value"])
    if date_column == value_column:
        raise TemplateError("benchmark source_date and source_value must differ")
    source_column = next(
        column
        for column in range(1, max(date_column, value_column, 3) + 2)
        if column not in {date_column, value_column}
    )
    originals = [sheet.cell(1, column) for column in range(1, 4)]
    templates = {
        "date": originals[0],
        "value": originals[1],
        "source": originals[2],
    }
    for row in range(1, max(sheet.max_row, 2) + 1):
        for column in range(1, max(sheet.max_column, source_column) + 1):
            sheet.cell(row, column).value = None
    display = str(benchmark.get("display_name") or "基准指数").strip()
    value_header = (
        f"{display}点位"
        if str(benchmark.get("source_type", "level")) == "level"
        else f"{display}收益"
    )
    assignments = (
        (date_column, "日期", templates["date"]),
        (value_column, value_header, templates["value"]),
        (source_column, "来源", templates["source"]),
    )
    for column, label, template in assignments:
        cell = sheet.cell(1, column)
        cell.value = label
        cell._style = copy(template._style)
        cell.alignment = copy(template.alignment)
        cell.number_format = template.number_format
        body = sheet.cell(2, column)
        body._style = copy(sheet.cell(2, min(template.column, 3))._style)
        body.alignment = copy(sheet.cell(2, min(template.column, 3)).alignment)
        sheet.column_dimensions[get_column_letter(column)].width = (
            42 if label == "来源" else 18 if column == value_column else 14
        )
    sheet.cell(2, date_column).number_format = "yyyy-mm-dd"
    sheet.cell(2, value_column).number_format = (
        "0.00%"
        if str(benchmark.get("source_type", "level")) == "aligned_return"
        else "0.0000"
    )


def _source_definitions(routes: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    definitions: dict[str, dict[str, Any]] = {}
    for route in routes:
        benchmark = route.get("benchmark")
        if not benchmark:
            continue
        name = str(benchmark["source_sheet"]).strip()
        normalized = {
            "source_sheet": name,
            "source_type": str(benchmark.get("source_type", "level")),
            "source_date": str(benchmark["source_date"]).strip().upper(),
            "source_value": str(benchmark["source_value"]).strip().upper(),
            "display_name": str(benchmark.get("display_name") or "").strip(),
        }
        previous = definitions.get(name)
        if previous is not None and previous != normalized:
            raise TemplateError(
                f"Shared benchmark sheet {name!r} has conflicting definitions"
            )
        definitions[name] = normalized
    return definitions


def init_template(config: dict[str, Any]) -> dict[str, Any]:
    if str(config.get("workbook_mode", "existing")) != "bundled-template":
        raise TemplateError(
            "workbook init-template is only available after bootstrap --new-workbook"
        )
    target = Path(config["workbook_path"]).expanduser().resolve()
    if target.exists():
        raise TemplateError(
            f"Target workbook already exists; refusing to overwrite it: {target}"
        )
    if target.suffix.lower() != ".xlsx":
        raise TemplateError("Template workbook target must end in .xlsx")
    if not target.parent.is_dir():
        raise TemplateError(f"Target workbook parent does not exist: {target.parent}")
    routes = active_routes(config)
    if not routes:
        raise TemplateError("No active routes are configured")
    product_names = [str(route.get("sheet") or "").strip() for route in routes]
    if any(not _valid_sheet_name(name) for name in product_names):
        raise TemplateError(
            "Every product sheet name must be 1-31 characters and contain no \\ / * ? : [ ]"
        )
    if len(product_names) != len(set(product_names)):
        raise TemplateError("Product sheet names must be unique")
    for route in routes:
        if str(route.get("sheet_mode", "")) != "template":
            raise TemplateError(
                f"{route['sheet']}: sheet_mode must be template for a generated workbook"
            )
        if str(route.get("data_frequency", "auto")) not in {"daily", "weekly"}:
            raise TemplateError(
                f"{route['sheet']}: data_frequency must be daily or weekly"
            )
    sources = _source_definitions(routes)
    if any(not _valid_sheet_name(name) for name in sources):
        raise TemplateError(
            "Every benchmark source sheet name must be 1-31 characters and contain no invalid characters"
        )
    if set(product_names) & set(sources):
        raise TemplateError(
            "Product sheets and benchmark source sheets cannot share a name"
        )

    asset = STATE_ROOT / "assets" / ASSET_NAME
    if not asset.is_file():
        raise TemplateError(f"Bundled workbook template is missing: {asset}")
    workbook = openpyxl.load_workbook(asset, data_only=False, read_only=False)
    created_products: list[str] = []
    created_sources: list[str] = []
    temporary = target.parent / f".{target.name}.template-{uuid.uuid4().hex[:8]}.tmp"
    try:
        threshold = float((config.get("style") or {}).get("zero_threshold", 0.00005))
        for route in routes:
            frequency = str(route["data_frequency"])
            has_benchmark = bool(route.get("benchmark"))
            source_title = TEMPLATE_SHEETS[(frequency, has_benchmark)]
            sheet = workbook.copy_worksheet(workbook[source_title])
            sheet.title = str(route["sheet"])
            _configure_product_sheet(sheet, route, threshold)
            created_products.append(sheet.title)
        for source_name, benchmark in sources.items():
            sheet = workbook.copy_worksheet(workbook[INDEX_TEMPLATE])
            sheet.title = source_name
            _set_index_headers(sheet, benchmark)
            created_sources.append(sheet.title)
        for title in list(TEMPLATE_SHEETS.values()) + [INDEX_TEMPLATE]:
            del workbook[title]
        _sanitize_metadata(workbook)
        workbook.save(temporary)
        try:
            os.link(temporary, target)
        except FileExistsError as exc:
            raise TemplateError(
                f"Target workbook appeared during initialization; refusing to overwrite it: {target}"
            ) from exc
        except OSError as exc:
            raise TemplateError(
                "Could not create the workbook with atomic no-overwrite protection; "
                "use a new path on a local NTFS/APFS/ext4 volume"
            ) from exc
        temporary.unlink()
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()
    return {
        "created": True,
        "workbook_path": str(target),
        "product_sheets": created_products,
        "benchmark_source_sheets": created_sources,
        "style_mode": "cn-red-up-green-down",
        "overwrote_existing": False,
    }


def add_template_product(
    config: dict[str, Any], route: dict[str, Any]
) -> dict[str, Any]:
    if str(config.get("workbook_mode", "existing")) != "bundled-template":
        raise TemplateError(
            "Only a workbook created from the bundled template can add a template product page"
        )
    target = Path(config["workbook_path"]).expanduser().resolve()
    if not target.is_file() or target.suffix.lower() != ".xlsx":
        raise TemplateError("The initialized bundled-template workbook is missing")
    title = str(route.get("sheet") or "").strip()
    frequency = str(route.get("data_frequency", "auto"))
    if not _valid_sheet_name(title):
        raise TemplateError(
            "Product sheet name must be 1-31 characters and contain no \\ / * ? : [ ]"
        )
    if str(route.get("sheet_mode", "")) != "template":
        raise TemplateError("A bundled-template product must use sheet_mode: template")
    if frequency not in {"daily", "weekly"}:
        raise TemplateError("A bundled-template product must be daily or weekly")

    routes = active_routes(config)
    sources = _source_definitions(routes)
    asset_path = STATE_ROOT / "assets" / ASSET_NAME
    if not asset_path.is_file():
        raise TemplateError(f"Bundled workbook template is missing: {asset_path}")

    workbook = openpyxl.load_workbook(target, data_only=False, read_only=False)
    asset = openpyxl.load_workbook(asset_path, data_only=False, read_only=False)
    temporary = (
        target.parent
        / f".{target.stem}.product-{uuid.uuid4().hex[:8]}.tmp{target.suffix.lower()}"
    )
    backup_directory = ROOT / "backups"
    backup_directory.mkdir(parents=True, exist_ok=True)
    backup = backup_directory / (
        f"before-product-add-{dt.datetime.now():%Y%m%d-%H%M%S}-{uuid.uuid4().hex[:6]}"
        f"{target.suffix.lower()}"
    )
    created_sources: list[str] = []
    try:
        if title in workbook.sheetnames:
            raise TemplateError(f"Workbook already contains sheet: {title}")
        conflicting_sources = set(sources) & {title}
        if conflicting_sources:
            raise TemplateError(
                "Product sheet and benchmark source sheet names conflict"
            )
        for source_name in sources:
            if not _valid_sheet_name(source_name):
                raise TemplateError(
                    f"Invalid benchmark source sheet name: {source_name}"
                )

        source_positions = [
            workbook.sheetnames.index(name)
            for name in sources
            if name in workbook.sheetnames
        ]
        insert_at = (
            min(source_positions) if source_positions else len(workbook.sheetnames)
        )
        source_title = TEMPLATE_SHEETS[(frequency, bool(route.get("benchmark")))]
        sheet = _copy_template_sheet(asset[source_title], workbook, title, insert_at)
        threshold = float((config.get("style") or {}).get("zero_threshold", 0.00005))
        _configure_product_sheet(sheet, route, threshold)

        for source_name, definition in sources.items():
            if source_name in workbook.sheetnames:
                continue
            source = _copy_template_sheet(
                asset[INDEX_TEMPLATE], workbook, source_name, len(workbook.sheetnames)
            )
            _set_index_headers(source, definition)
            created_sources.append(source_name)

        _sanitize_metadata(workbook)
        shutil.copy2(target, backup)
        workbook.save(temporary)
        check = openpyxl.load_workbook(temporary, read_only=True, data_only=False)
        try:
            if title not in check.sheetnames:
                raise TemplateError("Saved workbook is missing the new product sheet")
        finally:
            check.close()
        os.replace(temporary, target)
    except Exception:
        temporary.unlink(missing_ok=True)
        if backup.is_file() and not target.is_file():
            shutil.copy2(backup, target)
        raise
    finally:
        asset.close()
        workbook.close()
    return {
        "created": True,
        "workbook_path": str(target),
        "product_sheet": title,
        "benchmark_source_sheets_created": created_sources,
        "backup": str(backup),
        "overwrote_existing_sheet": False,
    }
