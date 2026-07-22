from __future__ import annotations

import datetime as dt
import gc
import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from nav_commit import ensure_process_exit, spreadsheet_app
from nav_config import ROOT
from nav_workbook import discover_layout, file_sha256


class ProductWorkbookError(RuntimeError):
    pass


TOTAL_WORDS = {"累计", "合计", "total", "cumulative"}


def _norm(value: Any) -> str:
    return re.sub(r"[\s_:/：()（）\[\]-]", "", str(value or "")).lower()


def _has_values(sheet, start_row: int, end_row: int | None = None) -> bool:
    end = sheet.max_row if end_row is None else end_row
    for row in range(start_row, end + 1):
        for column in range(1, sheet.max_column + 1):
            if sheet.cell(row, column).value not in {None, ""}:
                return True
    return False


@dataclass(frozen=True)
class CloneSpec:
    source_sheet: str
    target_sheet: str
    header_row: int
    data_row: int
    source_summary_row: int
    target_summary_row: int
    max_column: int
    columns: dict[str, int]
    summary_labels: dict[int, str]


def prepare_clone_spec(
    config: dict[str, Any], source_route: dict[str, Any], target_sheet: str
) -> CloneSpec:
    if str(config.get("workbook_mode", "existing")) != "existing":
        raise ProductWorkbookError("照现有 Sheet 新建只用于用户已有工作簿")
    master = Path(config["workbook_path"]).expanduser().resolve()
    if not master.is_file():
        raise ProductWorkbookError(f"工作簿不存在：{master}")
    keep_vba = master.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(master, data_only=False, keep_vba=keep_vba)
    try:
        source_sheet = str(source_route["sheet"])
        if source_sheet not in workbook.sheetnames:
            raise ProductWorkbookError(f"参考 Sheet 不存在：{source_sheet}")
        if target_sheet in workbook.sheetnames:
            raise ProductWorkbookError(f"目标 Sheet 已存在：{target_sheet}")
        if str(source_route.get("sheet_mode", "summary")) != "summary":
            raise ProductWorkbookError(
                "当前只支持复制带累计/合计行的 summary Sheet；纯追加表请先在 Excel/WPS 中复制好 Sheet，再使用 products adopt 接管"
            )
        layout = discover_layout(
            workbook[source_sheet],
            (config.get("column_overrides") or {}).get(source_sheet),
            source_route,
        )
        labels = {
            column: str(workbook[source_sheet].cell(layout.summary_row, column).value)
            for column in range(1, workbook[source_sheet].max_column + 1)
            if _norm(workbook[source_sheet].cell(layout.summary_row, column).value)
            in TOTAL_WORDS
        }
        if not labels:
            raise ProductWorkbookError(
                f"{source_sheet} 的累计/合计行无法安全识别，拒绝复制"
            )
        return CloneSpec(
            source_sheet=source_sheet,
            target_sheet=target_sheet,
            header_row=layout.header_row,
            data_row=layout.data_start,
            source_summary_row=layout.summary_row,
            target_summary_row=layout.data_start + 1,
            max_column=max(workbook[source_sheet].max_column, *layout.columns.values()),
            columns=dict(layout.columns),
            summary_labels=labels,
        )
    finally:
        workbook.close()


def _apply_clone_with_com(
    temp: Path,
    spec: CloneSpec,
    *,
    code: str | None,
    product_name: str,
    first_date: dt.date,
) -> str:
    app, progid, process_id = spreadsheet_app()
    book = None
    source = None
    target = None
    success = False
    try:
        book = app.Workbooks.Open(
            str(temp.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        source = book.Worksheets(spec.source_sheet)
        sheet_count = int(book.Worksheets.Count)
        source_index = int(source.Index)
        source.Copy(None, source)
        if int(book.Worksheets.Count) != sheet_count + 1:
            raise ProductWorkbookError("Excel/WPS 没有创建参考 Sheet 的副本")
        target = book.Worksheets(source_index + 1)
        target.Name = spec.target_sheet

        if spec.source_summary_row > spec.data_row + 1:
            target.Rows(
                f"{spec.data_row + 1}:{spec.source_summary_row - 1}"
            ).Delete()
        if spec.header_row > 1:
            target.Range(
                target.Cells(1, 1),
                target.Cells(spec.header_row - 1, spec.max_column),
            ).ClearContents()
        target.Range(
            target.Cells(spec.data_row, 1),
            target.Cells(spec.data_row, spec.max_column),
        ).ClearContents()
        target.Range(
            target.Cells(spec.target_summary_row, 1),
            target.Cells(spec.target_summary_row, spec.max_column),
        ).ClearContents()

        if code and spec.columns.get("code"):
            target.Cells(spec.data_row, spec.columns["code"]).Value2 = code
        if spec.columns.get("name"):
            target.Cells(spec.data_row, spec.columns["name"]).Value2 = product_name
        date_base = dt.date(1904, 1, 1) if bool(book.Date1904) else dt.date(1899, 12, 30)
        target.Cells(spec.data_row, spec.columns["date"]).Value2 = (
            first_date - date_base
        ).days
        for column, label in spec.summary_labels.items():
            target.Cells(spec.target_summary_row, column).Value2 = label

        book.Save()
        success = True
        return progid
    finally:
        if book is not None:
            try:
                book.Close(SaveChanges=success)
            except Exception:
                pass
        try:
            app.Quit()
        except Exception:
            pass
        source = None
        target = None
        book = None
        app = None
        gc.collect()
        ensure_process_exit(process_id)


def _same_cell(left: Any, right: Any) -> bool:
    if isinstance(left, dt.datetime):
        left = left.date()
    if isinstance(right, dt.datetime):
        right = right.date()
    return left == right


def _verify_other_sheets_unchanged(before, after, target_sheet: str) -> None:
    for name in before.sheetnames:
        left = before[name]
        right = after[name]
        if left.max_row != right.max_row or left.max_column != right.max_column:
            raise ProductWorkbookError(f"复制过程意外改变了原有 Sheet 结构：{name}")
        for row in range(1, left.max_row + 1):
            for column in range(1, left.max_column + 1):
                if not _same_cell(
                    left.cell(row, column).value, right.cell(row, column).value
                ):
                    raise ProductWorkbookError(
                        f"复制过程意外改变了 {name}!{left.cell(row, column).coordinate}"
                    )
    if target_sheet not in after.sheetnames:
        raise ProductWorkbookError("复制结果缺少目标 Sheet")


def _style_signature(cell) -> tuple[Any, ...]:
    color = cell.fill.fgColor
    rgb = color.rgb
    if isinstance(rgb, str) and len(rgb) == 8:
        rgb = rgb[-6:]
    return (
        str(cell.number_format).replace("\\-", "-").replace("\\/", "/"),
        cell.font.name,
        cell.font.sz,
        cell.font.bold,
        cell.fill.fill_type,
        color.type,
        rgb,
        color.indexed,
        color.theme,
        color.tint,
        cell.border.left.style,
        cell.border.right.style,
        cell.border.top.style,
        cell.border.bottom.style,
        cell.alignment.horizontal,
        cell.alignment.wrap_text,
    )


def verify_clone(
    original: Path,
    result: Path,
    spec: CloneSpec,
    *,
    code: str | None,
    product_name: str,
    first_date: dt.date,
) -> None:
    keep_vba = original.suffix.lower() == ".xlsm"
    before = openpyxl.load_workbook(original, data_only=False, keep_vba=keep_vba)
    after = openpyxl.load_workbook(result, data_only=False, keep_vba=keep_vba)
    try:
        expected_names = list(before.sheetnames)
        source_index = expected_names.index(spec.source_sheet)
        expected_names.insert(source_index + 1, spec.target_sheet)
        if after.sheetnames != expected_names:
            raise ProductWorkbookError(
                f"复制结果的 Sheet 顺序不正确：expected={expected_names!r}, actual={after.sheetnames!r}"
            )
        _verify_other_sheets_unchanged(before, after, spec.target_sheet)
        source = before[spec.source_sheet]
        target = after[spec.target_sheet]
        for column in range(1, spec.max_column + 1):
            if target.cell(spec.header_row, column).value != source.cell(
                spec.header_row, column
            ).value:
                raise ProductWorkbookError("复制结果没有保留参考 Sheet 的表头")
            target_header_style = _style_signature(
                target.cell(spec.header_row, column)
            )
            source_header_style = _style_signature(
                source.cell(spec.header_row, column)
            )
            if target_header_style != source_header_style:
                raise ProductWorkbookError(
                    "复制结果没有保留参考 Sheet 的表头样式："
                    f"column={column}, source={source_header_style!r}, target={target_header_style!r}"
                )
            target_data_style = _style_signature(target.cell(spec.data_row, column))
            source_data_style = _style_signature(source.cell(spec.data_row, column))
            if target_data_style != source_data_style:
                raise ProductWorkbookError(
                    "复制结果没有保留参考 Sheet 的数据行样式："
                    f"column={column}, source={source_data_style!r}, target={target_data_style!r}"
                )
        if spec.header_row > 1 and _has_values(target, 1, spec.header_row - 1):
            raise ProductWorkbookError("复制结果残留了参考产品说明")
        allowed = {
            spec.columns["date"],
            *(
                [spec.columns["code"]]
                if code and spec.columns.get("code")
                else []
            ),
            *([spec.columns["name"]] if spec.columns.get("name") else []),
        }
        for column in range(1, spec.max_column + 1):
            value = target.cell(spec.data_row, column).value
            if column not in allowed and value not in {None, ""}:
                raise ProductWorkbookError("复制结果残留了参考产品历史数据或公式")
        if code and spec.columns.get("code") and str(
            target.cell(spec.data_row, spec.columns["code"]).value
        ) != str(code):
            raise ProductWorkbookError("复制结果的产品代码不正确")
        if spec.columns.get("name") and target.cell(
            spec.data_row, spec.columns["name"]
        ).value != product_name:
            raise ProductWorkbookError("复制结果的产品名称不正确")
        date_value = target.cell(spec.data_row, spec.columns["date"]).value
        if isinstance(date_value, dt.datetime):
            date_value = date_value.date()
        if date_value != first_date:
            raise ProductWorkbookError(
                f"复制结果的冷启动日期不正确：expected={first_date!r}, actual={date_value!r}"
            )
        for column in range(1, spec.max_column + 1):
            value = target.cell(spec.target_summary_row, column).value
            expected = spec.summary_labels.get(column)
            if value not in {None, ""} and value != expected:
                raise ProductWorkbookError("复制结果的累计行残留了参考产品内容")
        if _has_values(target, spec.target_summary_row + 1):
            raise ProductWorkbookError("复制结果在累计行下方残留了参考产品内容")
    finally:
        before.close()
        after.close()


def clone_product_sheet(
    config: dict[str, Any],
    source_route: dict[str, Any],
    target_route: dict[str, Any],
    first_date: dt.date,
) -> dict[str, Any]:
    master = Path(config["workbook_path"]).expanduser().resolve()
    spec = prepare_clone_spec(config, source_route, str(target_route["sheet"]))
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_dir = ROOT / "backups"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"{master.stem}-before-add-product-{timestamp}{master.suffix}"
    temp = master.with_name(
        f".{master.stem}.nav-add-product-{os.getpid()}{master.suffix}"
    )
    if temp.exists():
        raise ProductWorkbookError(f"临时文件已存在：{temp.name}")
    before_hash = file_sha256(master)
    try:
        shutil.copy2(master, backup)
        shutil.copy2(master, temp)
        application = _apply_clone_with_com(
            temp,
            spec,
            code=target_route.get("code"),
            product_name=str(target_route.get("product_name") or target_route["sheet"]),
            first_date=first_date,
        )
        verify_clone(
            master,
            temp,
            spec,
            code=target_route.get("code"),
            product_name=str(target_route.get("product_name") or target_route["sheet"]),
            first_date=first_date,
        )
        if file_sha256(master) != before_hash:
            raise ProductWorkbookError("复制期间正式工作簿发生变化，拒绝替换")
        try:
            os.replace(temp, master)
        except OSError as exc:
            if isinstance(exc, PermissionError) or getattr(exc, "winerror", None) in {
                5,
                32,
                33,
            }:
                raise ProductWorkbookError(
                    "无法新增 Sheet；请关闭正在打开该工作簿的 Excel/WPS 窗口后重试"
                ) from exc
            raise
    except Exception:
        temp.unlink(missing_ok=True)
        backup.unlink(missing_ok=True)
        if not master.is_file() or file_sha256(master) != before_hash:
            raise ProductWorkbookError("新增 Sheet 失败且正式工作簿发生了意外变化")
        raise
    keep = int((config.get("retention") or {}).get("backup_count", 10))
    old_backups = sorted(
        backup_dir.glob(f"{master.stem}-before-add-product-*{master.suffix}"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for old in old_backups[keep:]:
        try:
            old.unlink()
        except OSError:
            pass
    return {
        "application": application,
        "backup": str(backup),
        "source_sheet": spec.source_sheet,
        "target_sheet": spec.target_sheet,
        "header_row": spec.header_row,
        "reserved_row": spec.data_row,
        "summary_row": spec.target_summary_row,
    }


def restore_after_config_failure(master: Path, backup: Path) -> None:
    try:
        shutil.copy2(backup, master)
    except OSError as exc:
        raise ProductWorkbookError(
            "路由保存失败，且无法从新增产品前备份恢复工作簿"
        ) from exc
