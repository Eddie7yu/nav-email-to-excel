from __future__ import annotations

import csv
import datetime as dt
import html
import importlib.util
import io
import re
import sys
import zipfile
from dataclasses import dataclass
from email.message import Message
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import xlrd
from pypdf import PdfReader

from nav_config import STATE_ROOT, normalize_code


DATE_WORDS = (
    "净值日期",
    "估值日期",
    "估值基准日",
    "业务日期",
    "navdate",
    "date",
    "日期",
)
CODE_WORDS = ("产品代码", "基金代码", "productcode", "fundcode", "代码")
UNIT_WORDS = ("单位净值", "份额净值", "最新净值", "unitnav", "nav")
CUM_WORDS = ("累计单位净值", "累计净值", "cumulativenav", "accumulatednav")
EPOCH = dt.date(1899, 12, 30)
MAX_TEXT_CHARS = 2_000_000
MAX_SHEETS = 20
MAX_ROWS = 20_000
MAX_COLUMNS = 200
MAX_CELLS = 2_000_000
MAX_PDF_PAGES = 100
MAX_ZIP_MEMBERS = 2_000
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024


class ParseError(ValueError):
    pass


@dataclass(frozen=True)
class NavRow:
    date: dt.date
    unit: float
    cumulative: float | None = None
    code: str | None = None
    source: str = "unknown"

    def key(self) -> tuple[str, str | None, float, float | None]:
        return (
            self.date.isoformat(),
            self.code,
            round(self.unit, 12),
            None if self.cumulative is None else round(self.cumulative, 12),
        )


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace(",", "").replace("，", "")
        if "%" in text:
            return None
        match = re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)", text)
        if not match:
            return None
        number = float(text)
    if not (0 < number < 1000000):
        return None
    return number


def parse_date(value: Any, pivot: int = 80) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and 20000 < float(value) < 100000:
        return EPOCH + dt.timedelta(days=int(value))
    text = str(value or "").strip()
    if not text:
        return None
    normalized = re.sub(r"[年/.]", "-", text).replace("月", "-").replace("日", "")
    match = re.search(r"(?<!\d)(20\d{2})-?(\d{1,2})-?(\d{1,2})(?!\d)", normalized)
    if match:
        try:
            return dt.date(
                int(match.group(1)), int(match.group(2)), int(match.group(3))
            )
        except ValueError:
            return None
    match = re.search(r"(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)", re.sub(r"\D", "", text))
    if match:
        year = int(match.group(1))
        year += 1900 if year >= pivot else 2000
        try:
            return dt.date(year, int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def normalized_header(value: Any) -> str:
    return re.sub(r"[\s_:/：()（）\[\]-]", "", str(value or "")).lower()


def classify_header(value: Any) -> str | None:
    text = normalized_header(value)
    if not text:
        return None
    if any(normalized_header(word) in text for word in CUM_WORDS):
        return "cumulative"
    if any(normalized_header(word) in text for word in DATE_WORDS):
        return "date"
    if any(normalized_header(word) in text for word in CODE_WORDS):
        return "code"
    if any(normalized_header(word) in text for word in UNIT_WORDS):
        return "unit"
    return None


def rows_from_matrix(matrix: list[list[Any]], source: str) -> list[NavRow]:
    if len(matrix) > MAX_ROWS or any(len(row) > MAX_COLUMNS for row in matrix):
        raise ParseError(f"Table limits exceeded in {source}")
    if sum(len(row) for row in matrix) > MAX_CELLS:
        raise ParseError(f"Cell limit exceeded in {source}")
    results: list[NavRow] = []
    for header_index, row in enumerate(matrix[:50]):
        candidates: dict[str, list[int]] = {}
        for column, value in enumerate(row):
            field = classify_header(value)
            if field:
                candidates.setdefault(field, []).append(column)
        if not {"date", "unit"} <= candidates.keys():
            continue
        if len(candidates["date"]) != 1:
            raise ParseError(f"Ambiguous date columns in {source}")
        mapping = {field: columns[0] for field, columns in candidates.items()}
        for data_row in matrix[header_index + 1 :]:
            if mapping["date"] >= len(data_row) or mapping["unit"] >= len(data_row):
                continue
            date = parse_date(data_row[mapping["date"]])
            unit = parse_number(data_row[mapping["unit"]])
            if not date or unit is None:
                continue
            cumulative = None
            if "cumulative" in mapping and mapping["cumulative"] < len(data_row):
                cumulative = parse_number(data_row[mapping["cumulative"]])
            code = None
            if "code" in mapping and mapping["code"] < len(data_row):
                code = normalize_code(data_row[mapping["code"]])
            results.append(
                NavRow(
                    date=date,
                    unit=unit,
                    cumulative=cumulative,
                    code=code,
                    source=source,
                )
            )
        if results:
            break
    return results


def rows_from_text(text: str, source: str, subject: str = "") -> list[NavRow]:
    if len(text) > MAX_TEXT_CHARS:
        raise ParseError(f"Text limit exceeded in {source}")
    clean = html.unescape(text).replace("\u00a0", " ")
    matrices: list[list[list[str]]] = []
    for delimiter in ("|", "\t", ",", ";"):
        lines = [line for line in clean.splitlines() if delimiter in line]
        if lines:
            matrices.append(
                [[part.strip() for part in line.split(delimiter)] for line in lines]
            )
    spaced = [
        re.split(r"\s{2,}", line.strip())
        for line in clean.splitlines()
        if re.search(r"\s{2,}", line.strip())
    ]
    if spaced:
        matrices.append(spaced)
    results: list[NavRow] = []
    for matrix in matrices:
        results.extend(rows_from_matrix(matrix, source))

    if not results:
        values: dict[str, str] = {}
        for line in clean.splitlines():
            match = re.match(r"\s*([^:：]{1,30})\s*[:：]\s*(.+?)\s*$", line)
            if not match:
                continue
            field = classify_header(match.group(1))
            if field:
                if field in values:
                    raise ParseError(f"Repeated labelled {field} field in {source}")
                values[field] = match.group(2)
        date = parse_date(values.get("date") or subject)
        unit = parse_number(values.get("unit"))
        if date and unit is not None:
            results.append(
                NavRow(
                    date=date,
                    unit=unit,
                    cumulative=parse_number(values.get("cumulative")),
                    code=normalize_code(values.get("code")),
                    source=source,
                )
            )
    return deduplicate(results)


def rows_from_xlsx(payload: bytes, source: str) -> list[NavRow]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            if (
                len(members) > MAX_ZIP_MEMBERS
                or sum(item.file_size for item in members) > MAX_UNCOMPRESSED_BYTES
            ):
                raise ParseError(f"Compressed workbook limits exceeded in {source}")
            for item in members:
                if (
                    item.file_size > 10 * 1024 * 1024
                    and item.file_size > max(item.compress_size, 1) * 200
                ):
                    raise ParseError(
                        f"Suspicious workbook compression ratio in {source}"
                    )
    except zipfile.BadZipFile as exc:
        raise ParseError(f"Invalid workbook archive in {source}") from exc
    workbook = openpyxl.load_workbook(
        io.BytesIO(payload), data_only=True, read_only=True
    )
    rows: list[NavRow] = []
    try:
        if len(workbook.worksheets) > MAX_SHEETS:
            raise ParseError(f"Worksheet limit exceeded in {source}")
        for sheet in workbook.worksheets:
            if sheet.max_row > MAX_ROWS or sheet.max_column > MAX_COLUMNS:
                raise ParseError(f"Worksheet dimensions exceeded in {source}")
            matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
            rows.extend(rows_from_matrix(matrix, f"{source}:{sheet.title}"))
    finally:
        workbook.close()
    return deduplicate(rows)


def rows_from_xls(payload: bytes, source: str) -> list[NavRow]:
    workbook = xlrd.open_workbook(file_contents=payload)
    rows: list[NavRow] = []
    if workbook.nsheets > MAX_SHEETS:
        raise ParseError(f"Worksheet limit exceeded in {source}")
    for sheet in workbook.sheets():
        if sheet.nrows > MAX_ROWS or sheet.ncols > MAX_COLUMNS:
            raise ParseError(f"Worksheet dimensions exceeded in {source}")
        matrix = [sheet.row_values(index) for index in range(sheet.nrows)]
        rows.extend(rows_from_matrix(matrix, f"{source}:{sheet.name}"))
    return deduplicate(rows)


def rows_from_csv(payload: bytes, source: str) -> list[NavRow]:
    text = None
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            text = payload.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ParseError(f"Unsupported CSV encoding in {source}")
    try:
        dialect = csv.Sniffer().sniff(text[:4096])
        matrix = list(csv.reader(io.StringIO(text), dialect))
    except csv.Error:
        matrix = list(csv.reader(io.StringIO(text)))
    return rows_from_matrix(matrix, source)


def rows_from_pdf(payload: bytes, source: str, subject: str) -> list[NavRow]:
    reader = PdfReader(io.BytesIO(payload))
    if len(reader.pages) > MAX_PDF_PAGES:
        raise ParseError(f"PDF page limit exceeded in {source}")
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if len(text) > MAX_TEXT_CHARS:
        raise ParseError(f"PDF text limit exceeded in {source}")
    return rows_from_text(text, source, subject)


class _HTMLText(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.casefold() in {"br", "div", "li", "p", "td", "th", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        self.parts.append(data)


class _HTMLTables(HTMLParser):
    def __init__(self, source: str) -> None:
        super().__init__(convert_charrefs=True)
        self.source = source
        self.tables: list[list[list[tuple[str, int, int]]]] = []
        self.stack: list[dict[str, Any]] = []

    def _span(self, attrs: list[tuple[str, str | None]], name: str, limit: int) -> int:
        raw = dict(attrs).get(name, "1")
        try:
            value = int(str(raw or "1"))
        except ValueError as exc:
            raise ParseError(f"Invalid HTML {name} in {self.source}") from exc
        if not 1 <= value <= limit:
            raise ParseError(f"Invalid HTML {name} in {self.source}")
        return value

    def _finish_cell(self, context: dict[str, Any]) -> None:
        cell = context.get("cell")
        if cell is None:
            return
        text = " ".join(part.strip() for part in cell["parts"] if str(part).strip())
        context["row"].append((text, cell["rowspan"], cell["colspan"]))
        context["cell"] = None

    def _finish_row(self, context: dict[str, Any]) -> None:
        self._finish_cell(context)
        row = context.get("row")
        if row:
            context["rows"].append(row)
        context["row"] = None

    def _finish_table(self) -> None:
        if not self.stack:
            return
        context = self.stack.pop()
        self._finish_row(context)
        if context["rows"]:
            self.tables.append(context["rows"])
            if len(self.tables) > MAX_SHEETS:
                raise ParseError(f"HTML table limit exceeded in {self.source}")

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.casefold()
        if tag == "table":
            self.stack.append({"rows": [], "row": None, "cell": None})
            if len(self.stack) > MAX_SHEETS:
                raise ParseError(f"Nested HTML table limit exceeded in {self.source}")
            return
        if not self.stack:
            return
        context = self.stack[-1]
        if tag == "tr":
            self._finish_row(context)
            context["row"] = []
        elif tag in {"td", "th"}:
            if context["row"] is None:
                context["row"] = []
            self._finish_cell(context)
            context["cell"] = {
                "parts": [],
                "rowspan": self._span(attrs, "rowspan", MAX_ROWS),
                "colspan": self._span(attrs, "colspan", MAX_COLUMNS),
            }
        elif tag == "br":
            for active in self.stack:
                if active.get("cell") is not None:
                    active["cell"]["parts"].append(" ")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.casefold()
        if tag == "table":
            self._finish_table()
        elif self.stack and tag in {"td", "th"}:
            self._finish_cell(self.stack[-1])
        elif self.stack and tag == "tr":
            self._finish_row(self.stack[-1])

    def handle_data(self, data: str) -> None:
        for context in self.stack:
            if context.get("cell") is not None:
                context["cell"]["parts"].append(data)

    def finish(self) -> None:
        while self.stack:
            self._finish_table()


def _expand_html_table(
    rows: list[list[tuple[str, int, int]]], source: str
) -> list[list[str]]:
    occupied: dict[tuple[int, int], str] = {}
    for row_index, row in enumerate(rows):
        column = 0
        for value, rowspan, colspan in row:
            while (row_index, column) in occupied:
                column += 1
            if column + colspan > MAX_COLUMNS or row_index + rowspan > MAX_ROWS:
                raise ParseError(f"HTML table dimensions exceeded in {source}")
            for target_row in range(row_index, row_index + rowspan):
                for target_column in range(column, column + colspan):
                    key = (target_row, target_column)
                    if key in occupied:
                        raise ParseError(f"Overlapping HTML spans in {source}")
                    occupied[key] = (
                        value if colspan == 1 and target_column == column else ""
                    )
            column += colspan
    if not occupied:
        return []
    row_count = max(row for row, _ in occupied) + 1
    column_count = max(column for _, column in occupied) + 1
    if row_count * column_count > MAX_CELLS:
        raise ParseError(f"Cell limit exceeded in {source}")
    return [
        [occupied.get((row, column), "") for column in range(column_count)]
        for row in range(row_count)
    ]


def rows_from_html(text: str, source: str, subject: str = "") -> list[NavRow]:
    if len(text) > MAX_TEXT_CHARS:
        raise ParseError(f"Text limit exceeded in {source}")
    extractor = _HTMLTables(source)
    try:
        extractor.feed(text)
        extractor.close()
        extractor.finish()
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(f"Invalid HTML body in {source}") from exc
    results: list[NavRow] = []
    total_cells = 0
    for index, raw_table in enumerate(extractor.tables, start=1):
        table_source = f"{source}:table:{index}"
        matrix = _expand_html_table(raw_table, table_source)
        total_cells += sum(len(row) for row in matrix)
        if total_cells > MAX_CELLS:
            raise ParseError(f"Cell limit exceeded in {source}")
        results.extend(rows_from_matrix(matrix, table_source))

    flattened = _HTMLText()
    flattened.feed(text)
    flattened.close()
    results.extend(rows_from_text("".join(flattened.parts), source, subject))
    return deduplicate(results)


def _message_body_parts(message: Message) -> tuple[list[str], list[str]]:
    plain: list[str] = []
    rich: list[str] = []
    parts: Iterable[Message] = message.walk() if message.is_multipart() else [message]
    for part in parts:
        if part.get_content_disposition() == "attachment":
            continue
        kind = part.get_content_type()
        if kind not in {"text/plain", "text/html"}:
            continue
        payload = part.get_payload(decode=True) or b""
        charset = part.get_content_charset() or "utf-8"
        try:
            text = payload.decode(charset, errors="replace")
        except LookupError:
            text = payload.decode("utf-8", errors="replace")
        if kind == "text/plain":
            plain.append(text)
        else:
            rich.append(text)
    return plain, rich


def message_text(message: Message) -> str:
    plain, rich = _message_body_parts(message)
    if plain:
        return "\n".join(plain)
    flattened: list[str] = []
    for text in rich:
        parser = _HTMLText()
        parser.feed(text)
        parser.close()
        flattened.append("".join(parser.parts))
    return "\n".join(flattened)


def _rows_from_message_auto(message: Message) -> list[NavRow]:
    subject = str(message.get("Subject") or "")
    rows: list[NavRow] = []
    plain, rich = _message_body_parts(message)
    for text in plain:
        rows.extend(rows_from_text(text, "body:text", subject))
    for text in rich:
        rows.extend(rows_from_html(text, "body:html", subject))
    for part in message.walk():
        filename = part.get_filename()
        if not filename:
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        suffix = Path(filename).suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                rows.extend(rows_from_xlsx(payload, f"attachment:{suffix}"))
            elif suffix == ".xls":
                rows.extend(rows_from_xls(payload, "attachment:.xls"))
            elif suffix in {".csv", ".txt"}:
                rows.extend(rows_from_csv(payload, f"attachment:{suffix}"))
            elif suffix == ".pdf":
                rows.extend(rows_from_pdf(payload, "attachment:.pdf", subject))
        except Exception as exc:
            raise ParseError(
                f"Could not parse a {suffix or 'named'} attachment"
            ) from exc
    return deduplicate(rows)


def _rows_from_local_parser(message: Message, parser_name: str) -> list[NavRow]:
    name = parser_name.removeprefix("local:")
    if not re.fullmatch(r"[a-z][a-z0-9_-]{0,63}", name):
        raise ParseError("Local parser name is invalid")
    directory = (STATE_ROOT / "parsers").resolve()
    path = (directory / f"{name}.py").resolve()
    if path.parent != directory or not path.is_file():
        raise ParseError(f"Local parser is missing: {name}")
    module_name = f"nav_local_parser_{name.replace('-', '_')}"
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ParseError(f"Local parser could not be loaded: {name}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    try:
        spec.loader.exec_module(module)
        parse_message = getattr(module, "parse_message", None)
        if not callable(parse_message):
            raise ParseError(f"Local parser {name} must define parse_message(message)")
        result = parse_message(message)
        rows = list(result)
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(f"Local parser failed: {name}") from exc
    finally:
        sys.modules.pop(module_name, None)
    if any(not isinstance(row, NavRow) for row in rows):
        raise ParseError(f"Local parser {name} returned an invalid row")
    return deduplicate(rows)


def rows_from_message(message: Message, parser_name: str = "auto") -> list[NavRow]:
    if parser_name == "auto":
        return _rows_from_message_auto(message)
    if parser_name.startswith("local:"):
        return _rows_from_local_parser(message, parser_name)
    raise ParseError(f"Unsupported parser: {parser_name}")


def deduplicate(rows: Iterable[NavRow]) -> list[NavRow]:
    unique: dict[tuple[str, str | None, float, float | None], NavRow] = {}
    for row in rows:
        unique[row.key()] = row
    return sorted(
        unique.values(), key=lambda item: (item.date, item.code or "", item.unit)
    )


def choose_route_rows(
    rows: Iterable[NavRow], route: dict[str, Any], single_sender_route: bool
) -> list[NavRow]:
    expected = normalize_code(route.get("code"))
    selected: list[NavRow] = []
    for row in rows:
        if expected and row.code and row.code != expected:
            continue
        if (
            expected
            and not row.code
            and not (single_sender_route and route.get("allow_sender_only"))
        ):
            continue
        selected.append(
            NavRow(
                date=row.date,
                unit=row.unit,
                cumulative=row.cumulative,
                code=expected or row.code,
                source=row.source,
            )
        )

    by_date: dict[dt.date, list[NavRow]] = {}
    for row in selected:
        by_date.setdefault(row.date, []).append(row)
    resolved: list[NavRow] = []
    for date, candidates in sorted(by_date.items()):
        signatures = {
            (
                round(item.unit, 10),
                None if item.cumulative is None else round(item.cumulative, 10),
            )
            for item in candidates
        }
        compatible = len({value[0] for value in signatures}) == 1
        cumulative_values = {value[1] for value in signatures if value[1] is not None}
        if not compatible or len(cumulative_values) > 1:
            raise ParseError(f"Conflicting NAV candidates for {date.isoformat()}")
        richer = sorted(
            candidates, key=lambda item: item.cumulative is not None, reverse=True
        )[0]
        resolved.append(richer)
    return resolved
