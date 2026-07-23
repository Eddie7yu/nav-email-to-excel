from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
import shutil
import stat
import statistics
import uuid
from collections import Counter
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    quote_sheetname,
    range_boundaries,
)
from openpyxl.worksheet.formula import ArrayFormula

from nav_config import (
    ROOT,
    STATE_ROOT,
    active_routes,
    benchmark_requires_review,
    benchmark_review_issue,
    normalize_code,
    write_json_atomic,
)
from nav_parse import NavRow, parse_date, parse_number


HEADER_WORDS = {
    "date": ("净值日期", "估值日期", "业务日期", "nav date", "date", "日期"),
    "code": ("产品代码", "基金代码", "product code", "fund code", "代码"),
    "name": ("产品名称", "基金名称", "product name", "fund name", "名称"),
    "unit": ("单位净值", "份额净值", "unit nav"),
    "cumulative": ("累计单位净值", "累计净值", "cumulative nav"),
    "return": ("产品收益", "基金收益", "return"),
    "daily_return": ("日收益", "日度收益", "收益（日度）", "daily return"),
    "weekly_return": ("周收益", "周度收益", "收益（周度）", "weekly return"),
    "benchmark_level": ("指数点位", "基准点位", "benchmark level", "index level"),
    "benchmark_return": ("指数收益", "基准收益", "benchmark return", "index return"),
    "excess": ("超额", "excess", "alpha"),
}
APPEND_HEADER_LABELS = {field: words[0] for field, words in HEADER_WORDS.items()}
TOTAL_WORDS = {"累计", "合计", "total", "cumulative"}
CONCURRENCY_REPORT = STATE_ROOT / "concurrency-report.json"
MANIFEST_HEAD_ROWS = 10
MANIFEST_TAIL_ROWS = 30


class WorkbookError(RuntimeError):
    pass


def make_file_writable(path: Path) -> None:
    try:
        path.chmod(path.stat().st_mode | stat.S_IWRITE)
    except FileNotFoundError:
        pass


def make_review_file_read_only(path: Path) -> None:
    path.chmod(path.stat().st_mode & ~0o222)


def review_file_is_read_only(path: Path) -> bool:
    return not bool(path.stat().st_mode & stat.S_IWRITE)


def _cell_values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, ArrayFormula) or isinstance(right, ArrayFormula):
        return (
            type(left) is type(right)
            and left.ref == right.ref
            and left.text == right.text
        )
    return left == right


@dataclass
class Layout:
    sheet: str
    header_row: int
    data_start: int
    summary_row: int
    last_data_row: int
    columns: dict[str, int]
    mode: str = "summary"
    headers_to_write: dict[int, str] | None = None


def _norm(value: Any) -> str:
    return re.sub(r"[\s_:/：()（）\[\]-]", "", str(value or "")).lower()


def _field(value: Any, route: dict[str, Any] | None = None) -> str | None:
    text = _norm(value)
    if not text:
        return None
    # Match date before unit NAV: a header such as "NAV Date" contains both
    # concepts, and treating it as NAV would make otherwise valid layouts fail.
    for field in ("cumulative", "date", "code", "name", "excess"):
        if any(_norm(word) in text for word in HEADER_WORDS[field]):
            return field
    benchmark = (route or {}).get("benchmark")
    display_name = (
        _norm(benchmark.get("display_name"))
        if isinstance(benchmark, dict)
        else ""
    )
    identifies_benchmark = (
        "指数" in text
        or "基准" in text
        or bool(display_name and display_name in text)
    )
    if identifies_benchmark:
        if any(
            marker in text
            for marker in ("日收益", "周收益", "收益率", "收益", "return")
        ):
            return "benchmark_return"
        if any(
            marker in text
            for marker in ("点位", "收盘", "level", "close")
        ) or (display_name and text == display_name):
            return "benchmark_level"
    for field in (
        "benchmark_return",
        "benchmark_level",
        "daily_return",
        "weekly_return",
        "return",
        "unit",
    ):
        if any(_norm(word) in text for word in HEADER_WORDS[field]):
            return field
    return None


def _return_header_base(value: Any) -> str:
    text = _norm(value)
    for marker in (
        "dailyreturn",
        "weeklyreturn",
        "benchmarkreturn",
        "indexreturn",
        "日收益率",
        "周收益率",
        "日收益",
        "周收益",
        "收益率",
        "收益",
        "return",
    ):
        text = text.replace(_norm(marker), "")
    return text


def _infer_named_benchmark_columns(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    candidate: dict[str, int],
    route: dict[str, Any],
) -> dict[str, int]:
    """Recognize a concrete benchmark name paired with its return header."""

    if {"benchmark_level", "benchmark_return"} <= candidate.keys():
        return candidate
    values = {
        column: sheet.cell(row, column).value
        for column in range(1, min(sheet.max_column, 80) + 1)
    }
    used = set(candidate.values())
    benchmark = route.get("benchmark")
    display_name = (
        _norm(benchmark.get("display_name"))
        if isinstance(benchmark, dict)
        else ""
    )
    return_columns = (
        [candidate["benchmark_return"]]
        if candidate.get("benchmark_return")
        else list(values)
    )
    for return_column in return_columns:
        value = values[return_column]
        if (
            return_column in used
            and candidate.get("benchmark_return") != return_column
        ):
            continue
        normalized = _norm(value)
        base = _return_header_base(value)
        if not base or base == normalized:
            continue
        explicit_benchmark = (
            "指数" in normalized
            or "基准" in normalized
            or bool(display_name and display_name in normalized)
        )
        level_column = next(
            (
                column
                for column, level_value in values.items()
                if column not in used
                and column != return_column
                and _norm(level_value) == base
            ),
            None,
        )
        paired_with_excess = bool(candidate.get("excess") and level_column)
        if not explicit_benchmark and not paired_with_excess:
            continue
        candidate.setdefault("benchmark_return", return_column)
        if level_column:
            candidate.setdefault("benchmark_level", level_column)
        break
    return candidate


def _column(value: Any) -> int:
    if isinstance(value, int) and value > 0:
        return value
    text = str(value or "").strip().upper()
    if text.isdigit() and int(text) > 0:
        return int(text)
    if re.fullmatch(r"[A-Z]{1,3}", text):
        return column_index_from_string(text)
    raise WorkbookError(f"Invalid column override: {value!r}")


def _has_values(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int = 1,
    end_row: int | None = None,
) -> bool:
    end = sheet.max_row if end_row is None else end_row
    for row in range(start_row, end + 1):
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row, column).value
            if value is not None and value != "":
                return True
    return False


def _append_columns(route: dict[str, Any], override: dict[str, Any]) -> dict[str, int]:
    explicit = {
        field: _column(override[field]) for field in HEADER_WORDS if field in override
    }
    if explicit:
        return explicit
    fields = ["date"]
    if normalize_code(route.get("code")):
        fields.append("code")
    if str(route.get("product_name") or "").strip():
        fields.append("name")
    fields.extend(("unit", "cumulative"))
    if route.get("benchmark"):
        fields.append("return")
        if str(route["benchmark"].get("source_type", "level")) == "level":
            fields.append("benchmark_level")
        fields.extend(("benchmark_return", "excess"))
    return {field: index for index, field in enumerate(fields, 1)}


def _require_append_identity(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    columns: dict[str, int],
    route: dict[str, Any],
) -> None:
    code_identity = bool(columns.get("code") and normalize_code(route.get("code")))
    name_identity = bool(
        columns.get("name") and str(route.get("product_name") or "").strip()
    )
    if not code_identity and not name_identity:
        raise WorkbookError(
            f"{sheet.title}: append mode needs a writable product code or product name column"
        )


def discover_layout(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    override: dict[str, Any] | None = None,
    route: dict[str, Any] | None = None,
) -> Layout:
    override = override or {}
    route = route or {}
    mode = str(route.get("sheet_mode", "summary"))
    if mode not in {"summary", "append", "template"}:
        raise WorkbookError(f"{sheet.title}: unsupported sheet mode {mode}")
    header_row = int(override.get("header_row") or 0)
    columns: dict[str, int] = {}
    sheet_is_blank = not _has_values(sheet)
    if mode == "append" and sheet_is_blank:
        header_row = header_row or 1
        columns = _append_columns(route, override)
        if not {"date", "unit"} <= columns.keys():
            raise WorkbookError(
                f"{sheet.title}: append mode requires explicit date and unit NAV columns"
            )
        _require_append_identity(sheet, columns, route)
        headers = {
            column: APPEND_HEADER_LABELS[field] for field, column in columns.items()
        }
        return Layout(
            sheet.title,
            header_row,
            header_row + 1,
            header_row + 1,
            header_row,
            columns,
            mode,
            headers,
        )
    if header_row:
        for field in HEADER_WORDS:
            if field in override:
                columns[field] = _column(override[field])
    else:
        best_score = -1
        for row in range(1, min(sheet.max_row, 30) + 1):
            candidate: dict[str, int] = {}
            for column in range(1, min(sheet.max_column, 80) + 1):
                field = _field(sheet.cell(row, column).value, route)
                if field and field not in candidate:
                    candidate[field] = column
            candidate = _infer_named_benchmark_columns(
                sheet, row, candidate, route
            )
            score = len(candidate) + (3 if {"date", "unit"} <= candidate.keys() else 0)
            if score > best_score:
                best_score = score
                header_row = row
                columns = candidate
        for field in HEADER_WORDS:
            if field in override:
                columns[field] = _column(override[field])

    if not {"date", "unit"} <= columns.keys():
        raise WorkbookError(
            f"{sheet.title}: date and unit NAV columns are not unambiguous"
        )
    data_start = header_row + 1
    if mode == "template":
        frequency = str(route.get("data_frequency", "auto"))
        if (
            frequency == "weekly"
            and route.get("benchmark")
            and str(route["benchmark"].get("source_type", "level")) == "level"
            and "benchmark_level" not in columns
            and columns.get("benchmark_return")
        ):
            columns["benchmark_level"] = columns["benchmark_return"] - 1
        required_returns = (
            {"daily_return", "weekly_return"}
            if frequency == "daily"
            else {"weekly_return"}
        )
        if frequency not in {"daily", "weekly"}:
            raise WorkbookError(
                f"{sheet.title}: template mode requires explicit daily or weekly data_frequency"
            )
        if not required_returns <= columns.keys():
            raise WorkbookError(
                f"{sheet.title}: template return columns do not match {frequency} frequency"
            )
        if not ({"code", "name"} & columns.keys()):
            raise WorkbookError(
                f"{sheet.title}: template mode requires a product code or product name column"
            )
        if (
            route.get("benchmark")
            and not {
                "benchmark_return",
                "excess",
            }
            <= columns.keys()
        ):
            raise WorkbookError(
                f"{sheet.title}: benchmark template is missing benchmark return or excess columns"
            )
        if (
            not route.get("benchmark")
            and {
                "benchmark_return",
                "benchmark_level",
                "excess",
            }
            & columns.keys()
        ):
            raise WorkbookError(
                f"{sheet.title}: non-benchmark template contains benchmark columns"
            )
    headers_to_write: dict[int, str] = {}
    if mode == "append":
        _require_append_identity(sheet, columns, route)
        headers_to_write = {
            column: APPEND_HEADER_LABELS[field]
            for field, column in columns.items()
            if sheet.cell(header_row, column).value in {None, ""}
        }
    dated_rows: list[int] = []
    for row in range(data_start, sheet.max_row + 1):
        if parse_date(sheet.cell(row, columns["date"]).value):
            dated_rows.append(row)
    if not dated_rows:
        if mode == "summary":
            raise WorkbookError(f"{sheet.title}: no dated NAV rows found")
        if mode == "template":
            summary_row = data_start + 1
            if header_row != 2 or sheet.max_row < summary_row:
                raise WorkbookError(
                    f"{sheet.title}: generated template structure is incomplete"
                )
            if _has_values(sheet, data_start, data_start):
                raise WorkbookError(
                    f"{sheet.title}: template cold-start row must be empty"
                )
            values = {
                _norm(sheet.cell(summary_row, column).value)
                for column in range(1, min(sheet.max_column, 6) + 1)
            }
            if not values & TOTAL_WORDS or _has_values(sheet, summary_row + 1):
                raise WorkbookError(
                    f"{sheet.title}: generated template must end with its cumulative row"
                )
            return Layout(
                sheet.title,
                header_row,
                data_start,
                summary_row,
                header_row,
                columns,
                mode,
                {},
            )
        if _has_values(sheet, data_start):
            raise WorkbookError(
                f"{sheet.title}: append mode found content below the header but no dated NAV rows"
            )
        return Layout(
            sheet.title,
            header_row,
            data_start,
            data_start,
            header_row,
            columns,
            mode,
            headers_to_write,
        )
    last_data_row = max(dated_rows)
    if mode == "append":
        if _has_values(sheet, last_data_row + 1):
            raise WorkbookError(
                f"{sheet.title}: append mode requires no footer content below the last dated row"
            )
        return Layout(
            sheet.title,
            header_row,
            data_start,
            last_data_row + 1,
            last_data_row,
            columns,
            mode,
            headers_to_write,
        )
    summary_row = 0
    for row in range(last_data_row + 1, sheet.max_row + 2):
        values = {
            _norm(sheet.cell(row, column).value)
            for column in range(1, min(sheet.max_column, 6) + 1)
        }
        if values & TOTAL_WORDS:
            summary_row = row
            break
    if not summary_row:
        raise WorkbookError(f"{sheet.title}: summary row was not found after the data")
    if summary_row != last_data_row + 1:
        raise WorkbookError(
            f"{sheet.title}: summary row must immediately follow the last dated row"
        )
    return Layout(
        sheet.title,
        header_row,
        data_start,
        summary_row,
        last_data_row,
        columns,
        mode,
        headers_to_write,
    )


def effective_cumulative(row: NavRow, route: dict[str, Any]) -> float:
    policy = str(route.get("cumulative_policy", "require"))
    if row.cumulative is not None:
        return row.cumulative
    if policy == "unit":
        return row.unit
    if policy == "offset":
        return row.unit + float(route["cumulative_offset"])
    raise WorkbookError(
        f"{route['sheet']}: cumulative NAV is required for {row.date.isoformat()}"
    )


def existing_rows(
    sheet: openpyxl.worksheet.worksheet.Worksheet, layout: Layout
) -> dict[dt.date, dict[str, Any]]:
    output: dict[dt.date, dict[str, Any]] = {}
    ordered_dates: list[dt.date] = []
    for row in range(layout.data_start, layout.summary_row):
        date = parse_date(sheet.cell(row, layout.columns["date"]).value)
        if not date:
            continue
        if date in output:
            raise WorkbookError(f"{sheet.title}: duplicate date {date.isoformat()}")
        ordered_dates.append(date)
        output[date] = {
            "row": row,
            "unit": parse_number(sheet.cell(row, layout.columns["unit"]).value),
            "cumulative": parse_number(
                sheet.cell(row, layout.columns.get("cumulative", 0)).value
            )
            if layout.columns.get("cumulative")
            else None,
            "code": normalize_code(sheet.cell(row, layout.columns.get("code", 0)).value)
            if layout.columns.get("code")
            else None,
            "name": str(
                sheet.cell(row, layout.columns.get("name", 0)).value or ""
            ).strip()
            if layout.columns.get("name")
            else None,
        }
    if ordered_dates != sorted(ordered_dates):
        raise WorkbookError(f"{sheet.title}: dated NAV rows are not in ascending order")
    return output


def _history_value_id(*values: Any) -> str:
    import hashlib

    encoded = "\x1f".join("" if value is None else str(value) for value in values)
    return hashlib.sha256(encoded.encode("utf-8", errors="replace")).hexdigest()[:12]


def _history_integrity(
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
    tolerance: float,
) -> dict[str, Any]:
    expected_code = normalize_code(route.get("code"))
    code_rows = [
        (date, values)
        for date, values in sorted(existing.items())
        if values.get("code")
    ]
    observed_codes = {str(values["code"]) for _, values in code_rows}
    code_constant = len(observed_codes) <= 1 and (
        not expected_code or not observed_codes or observed_codes == {expected_code}
    )
    unexpected_code_rows = [
        {
            "row": int(values["row"]),
            "date": date.isoformat(),
            "code_id": _history_value_id(values.get("code")),
        }
        for date, values in code_rows
        if (expected_code and values.get("code") != expected_code)
        or (not expected_code and len(observed_codes) > 1)
    ]

    cumulative_anomalies: list[dict[str, Any]] = []
    complete: list[tuple[dt.date, dict[str, Any], float]] = []
    if layout.columns.get("cumulative"):
        for date, values in sorted(existing.items()):
            unit = values.get("unit")
            cumulative = values.get("cumulative")
            if unit is None:
                cumulative_anomalies.append(
                    {
                        "row": int(values["row"]),
                        "date": date.isoformat(),
                        "issue": "missing-unit-nav",
                    }
                )
                continue
            if cumulative is None:
                cumulative_anomalies.append(
                    {
                        "row": int(values["row"]),
                        "date": date.isoformat(),
                        "issue": "missing-cumulative-nav",
                    }
                )
                continue
            spread = float(cumulative) - float(unit)
            complete.append((date, values, spread))
        policy = str(route.get("cumulative_policy", "require"))
        spread_tolerance = max(tolerance * 10, 1e-8)
        if policy in {"unit", "offset"}:
            expected_spread = (
                0.0 if policy == "unit" else float(route["cumulative_offset"])
            )
            for date, values, spread in complete:
                if abs(spread - expected_spread) > spread_tolerance:
                    cumulative_anomalies.append(
                        {
                            "row": int(values["row"]),
                            "date": date.isoformat(),
                            "issue": "configured-cumulative-policy-break",
                            "spread_id": _history_value_id(f"{spread:.10f}"),
                        }
                    )
        elif len(complete) >= 3:
            for previous, current, following in zip(
                complete, complete[1:], complete[2:]
            ):
                if (
                    abs(previous[2] - following[2]) <= spread_tolerance
                    and abs(current[2] - previous[2]) > spread_tolerance
                ):
                    cumulative_anomalies.append(
                        {
                            "row": int(current[1]["row"]),
                            "date": current[0].isoformat(),
                            "issue": "isolated-unit-cumulative-spread-break",
                            "spread_id": _history_value_id(f"{current[2]:.10f}"),
                            "neighbor_spread_id": _history_value_id(
                                f"{previous[2]:.10f}"
                            ),
                        }
                    )
    else:
        cumulative_anomalies.extend(
            {
                "row": int(values["row"]),
                "date": date.isoformat(),
                "issue": "missing-unit-nav",
            }
            for date, values in sorted(existing.items())
            if values.get("unit") is None
        )
    return {
        "passed": code_constant and not cumulative_anomalies,
        "repair_required": not code_constant or bool(cumulative_anomalies),
        "code_column": {
            "passed": code_constant,
            "nonblank_code_count": len(code_rows),
            "distinct_code_count": len(observed_codes),
            "unexpected_rows": unexpected_code_rows,
        },
        "cumulative_sequence": {
            "passed": not cumulative_anomalies,
            "anomalies": cumulative_anomalies,
        },
    }


def _identity_evidence_rows(
    existing: dict[dt.date, dict[str, Any]], route: dict[str, Any]
) -> dict[dt.date, dict[str, Any]]:
    """Use managed history for identity, or only the preserved tail at the boundary."""

    if not existing:
        return {}
    start = parse_date(route.get("series_start")) or dt.date.min
    managed = {date: values for date, values in existing.items() if date >= start}
    if managed:
        return managed
    tail = max(existing)
    return {tail: existing[tail]}


def _summary_reserved_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
) -> tuple[dt.date, int] | None:
    """Identify a strictly bounded summary-mode cold-start placeholder.

    A complete real data row is never treated as a placeholder. The accepted shape
    is one dated row immediately above the summary row, with a matching configured
    identity, no return/formula content, and at least one missing NAV field. A user
    may prefill the unit or cumulative NAV; validation must prove that partial value
    against the email for the same date before the row can be replaced.
    """

    if layout.mode != "summary" or len(existing) != 1:
        return None
    date, values = next(iter(existing.items()))
    row = int(values["row"])
    if row != layout.data_start or row != layout.summary_row - 1:
        return None

    expected_code = normalize_code(route.get("code"))
    expected_name = str(route.get("product_name") or "").strip()
    identity_columns: set[int] = set()
    if expected_code and layout.columns.get("code"):
        code_column = layout.columns["code"]
        code_value = normalize_code(sheet.cell(row, code_column).value)
        if code_value:
            if code_value != expected_code:
                return None
            identity_columns.add(code_column)
    if expected_name and layout.columns.get("name"):
        name_column = layout.columns["name"]
        name_value = str(sheet.cell(row, name_column).value or "").strip()
        if name_value:
            if name_value != expected_name:
                return None
            identity_columns.add(name_column)
    if not identity_columns:
        return None

    nav_columns = {
        column
        for column in (
            layout.columns.get("unit"),
            layout.columns.get("cumulative"),
        )
        if column
    }
    if nav_columns and all(
        sheet.cell(row, column).value not in {None, ""} for column in nav_columns
    ):
        return None
    allowed_columns = {layout.columns["date"], *identity_columns, *nav_columns}
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(row, column).value
        if column not in allowed_columns and value is not None and value != "":
            return None
    return date, row


def _summary_reviewed_onboarding(
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
    selected_dates: set[dt.date],
    candidates: list[NavRow],
    conflicts: int,
) -> str | None:
    """Allow a review-gated preview when identity is unique but history is sparse."""

    expected_code = normalize_code(route.get("code"))
    expected_name = _norm(route.get("product_name"))
    identity_existing = _identity_evidence_rows(existing, route)
    observed_codes = {
        values["code"]
        for values in identity_existing.values()
        if values.get("code")
    }
    observed_names = {
        _norm(values["name"])
        for values in identity_existing.values()
        if values.get("name")
    }
    candidate_codes = {candidate.code for candidate in candidates if candidate.code}
    code_identity = bool(
        expected_code
        and layout.columns.get("code")
        and observed_codes == {expected_code}
    )
    name_identity = bool(
        expected_name
        and (
            _norm(layout.sheet) == expected_name
            or (layout.columns.get("name") and observed_names == {expected_name})
        )
    )
    complete_history = all(
        values.get("unit") is not None
        and (
            not layout.columns.get("cumulative") or values.get("cumulative") is not None
        )
        for values in identity_existing.values()
    )
    if not (
        layout.mode == "summary"
        and conflicts == 0
        and existing
        and complete_history
        and selected_dates
        and candidates
        and len(candidate_codes) <= 1
        and (code_identity or name_identity)
        and all(
            candidate.date in existing or candidate.date > max(existing)
            for candidate in candidates
        )
    ):
        return None
    if (
        code_identity
        and expected_code
        and any(candidate.code not in {None, expected_code} for candidate in candidates)
    ):
        return None
    identity = (
        "product code and product name"
        if code_identity and name_identity
        else "product code"
        if code_identity
        else "product name"
    )
    return identity


def _withheld_current_week_dates(
    data_frequency: str,
    existing: dict[dt.date, dict[str, Any]],
    selected_candidates: list[NavRow],
    candidates: list[NavRow],
) -> list[dt.date]:
    """Return real tail dates intentionally withheld until the ISO week completes."""

    if data_frequency != "weekly" or not existing:
        return []
    today = dt.date.today()
    current_iso = today.isocalendar()
    current_week = (current_iso.year, current_iso.week)
    latest_existing = max(existing)
    selected_dates = {candidate.date for candidate in selected_candidates}
    return sorted(
        {
            candidate.date
            for candidate in candidates
            if candidate.date not in existing
            and candidate.date not in selected_dates
            and latest_existing < candidate.date <= today
            and (
                candidate.date.isocalendar().year,
                candidate.date.isocalendar().week,
            )
            == current_week
        }
    )


def _weekly_pending_baseline_identity(
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
    selected_candidates: list[NavRow],
    candidates: list[NavRow],
    conflicts: int,
    data_frequency: str,
) -> tuple[str, list[dt.date]] | None:
    """Allow a zero-add baseline when only the unfinished current week is pending."""

    withheld_dates = _withheld_current_week_dates(
        data_frequency, existing, selected_candidates, candidates
    )
    if not withheld_dates or conflicts or layout.mode not in {"summary", "append"}:
        return None
    tail_age_days = (dt.date.today() - max(existing)).days
    if tail_age_days > int(route.get("max_staleness_days", 14)):
        return None
    if any(candidate.date not in existing for candidate in selected_candidates):
        return None
    allowed_dates = set(existing) | set(withheld_dates)
    if not candidates or any(
        candidate.date not in allowed_dates for candidate in candidates
    ):
        return None
    identity_existing = _identity_evidence_rows(existing, route)
    complete_history = all(
        values.get("unit") is not None
        and (
            not layout.columns.get("cumulative")
            or values.get("cumulative") is not None
        )
        for values in identity_existing.values()
    )
    if not complete_history:
        return None

    expected_code = normalize_code(route.get("code"))
    expected_name = _norm(route.get("product_name"))
    observed_codes = {
        values["code"]
        for values in identity_existing.values()
        if values.get("code")
    }
    observed_names = {
        _norm(values["name"])
        for values in identity_existing.values()
        if values.get("name")
    }
    candidate_codes = {candidate.code for candidate in candidates if candidate.code}
    code_identity = bool(
        expected_code
        and layout.columns.get("code")
        and observed_codes == {expected_code}
        and all(candidate.code == expected_code for candidate in candidates)
    )
    name_identity = bool(
        expected_name
        and len(candidate_codes) <= 1
        and (
            _norm(layout.sheet) == expected_name
            or (layout.columns.get("name") and observed_names == {expected_name})
        )
    )
    if not (code_identity or name_identity):
        return None
    identity = (
        "product code and product name"
        if code_identity and name_identity
        else "product code"
        if code_identity
        else "product name"
    )
    return identity, withheld_dates


def _boundary_anchor_identity(
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
    pre_managed_candidates: list[NavRow],
    tolerance: float,
) -> tuple[str, dt.date] | None:
    """Verify the preserved workbook tail immediately before series_start."""

    if route.get("baseline_overlap") != "last_existing_point":
        return None
    start = parse_date(route.get("series_start"))
    if not start or start == dt.date.min or not existing:
        return None
    anchor_date = start - dt.timedelta(days=1)
    if max(existing) != anchor_date or anchor_date not in existing:
        return None
    candidates = [
        candidate
        for candidate in pre_managed_candidates
        if candidate.date == anchor_date
    ]
    signatures = {
        (
            round(candidate.unit, 10),
            None
            if candidate.cumulative is None
            else round(candidate.cumulative, 10),
            candidate.code,
        )
        for candidate in candidates
    }
    if len(signatures) != 1:
        return None
    candidate = candidates[0]
    observed = existing[anchor_date]
    if observed.get("unit") is None or abs(
        float(observed["unit"]) - candidate.unit
    ) > tolerance:
        return None
    if layout.columns.get("cumulative"):
        if observed.get("cumulative") is None or abs(
            float(observed["cumulative"])
            - effective_cumulative(candidate, route)
        ) > tolerance:
            return None
    expected_code = normalize_code(route.get("code"))
    expected_name = _norm(route.get("product_name"))
    code_identity = bool(
        expected_code
        and layout.columns.get("code")
        and observed.get("code") == expected_code
        and candidate.code == expected_code
    )
    name_identity = bool(
        expected_name
        and (
            _norm(layout.sheet) == expected_name
            or (
                layout.columns.get("name")
                and _norm(observed.get("name")) == expected_name
            )
        )
    )
    if not (code_identity or name_identity):
        return None
    identity = (
        "product code and product name"
        if code_identity and name_identity
        else "product code"
        if code_identity
        else "product name"
    )
    return identity, anchor_date


def _header_data_frequency(
    sheet: openpyxl.worksheet.worksheet.Worksheet, layout: Layout
) -> str | None:
    if layout.columns.get("daily_return"):
        return "daily"
    if layout.columns.get("weekly_return"):
        return "weekly"
    return_column = layout.columns.get("return")
    if not return_column:
        return None
    header = _norm(sheet.cell(layout.header_row, return_column).value)
    if "周" in header or "week" in header:
        return "weekly"
    if "日收益" in header or "日度" in header or "daily" in header:
        return "daily"
    return None


def _weekly_like_gaps(gaps: list[int]) -> bool:
    if not gaps:
        return False
    weekly = sum(1 for gap in gaps if gap >= 5 and min(gap % 7, (7 - gap % 7) % 7) <= 1)
    return weekly / len(gaps) >= 0.7


def _data_frequency(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
    reserved: tuple[dt.date, int] | None,
) -> tuple[str, str, int | None]:
    configured = str(route.get("data_frequency", "auto"))
    history_dates = sorted(
        date for date in existing if reserved is None or date != reserved[0]
    )
    inferred: str | None = None
    source = "email-source"
    preferred_weekday: int | None = None
    if len(history_dates) >= 2:
        gaps = [
            (current - previous).days
            for previous, current in zip(history_dates, history_dates[1:])
        ]
        median_gap = float(statistics.median(gaps))
        if median_gap <= 4:
            inferred = "daily"
        elif median_gap <= 10 or _weekly_like_gaps(gaps):
            inferred = "weekly"
        else:
            raise WorkbookError(
                f"{sheet.title}: existing dates do not prove a daily or weekly data frequency"
            )
        source = "workbook-history"
        if inferred == "weekly":
            counts = Counter(date.weekday() for date in history_dates)
            last_seen = {
                weekday: max(
                    index
                    for index, date in enumerate(history_dates)
                    if date.weekday() == weekday
                )
                for weekday in counts
            }
            preferred_weekday = max(
                counts, key=lambda weekday: (counts[weekday], last_seen[weekday])
            )
    else:
        inferred = _header_data_frequency(sheet, layout)
        if inferred:
            source = "workbook-header"

    if configured != "auto":
        if inferred and configured != inferred:
            raise WorkbookError(
                f"{sheet.title}: configured data_frequency {configured} conflicts with "
                f"the {inferred} workbook template"
            )
        return configured, "route-config", preferred_weekday
    return inferred or "daily", source, preferred_weekday


def _select_data_rows(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
    existing: dict[dt.date, dict[str, Any]],
    reserved: tuple[dt.date, int] | None,
    candidates: list[NavRow],
) -> tuple[list[NavRow], str, str]:
    frequency, source, preferred_weekday = _data_frequency(
        sheet, layout, route, existing, reserved
    )
    if frequency == "daily":
        return candidates, frequency, source

    groups: dict[tuple[int, int], list[NavRow]] = {}
    for candidate in candidates:
        iso = candidate.date.isocalendar()
        groups.setdefault((iso.year, iso.week), []).append(candidate)
    current_iso = dt.date.today().isocalendar()
    current_week = (current_iso.year, current_iso.week)
    existing_dates = set(existing)
    selected: list[NavRow] = []
    for week in sorted(groups):
        rows = sorted(groups[week], key=lambda item: item.date)
        exact = [row for row in rows if row.date in existing_dates]
        if exact:
            selected.extend(exact)
            continue
        if week >= current_week:
            continue
        if preferred_weekday is None:
            selected.append(rows[-1])
            continue
        on_or_before = [row for row in rows if row.date.weekday() <= preferred_weekday]
        selected.append(on_or_before[-1] if on_or_before else rows[0])
    return selected, frequency, source


def validate_history(
    config: dict[str, Any], route_rows: dict[str, list[NavRow]]
) -> dict[str, Any]:
    path = Path(config["workbook_path"])
    workbook = openpyxl.load_workbook(
        path, data_only=True, read_only=False, keep_vba=path.suffix.lower() == ".xlsm"
    )
    minimum = int((config.get("validation") or {}).get("minimum_history_dates", 2))
    tolerance = float((config.get("validation") or {}).get("tolerance", 1e-6))
    max_future_days = int((config.get("validation") or {}).get("max_future_days", 0))
    max_period_change = float(
        (config.get("validation") or {}).get("max_period_change", 0.5)
    )
    reports: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []
    try:
        for route in active_routes(config):
            sheet_name = str(route["sheet"])
            if benchmark_requires_review(route):
                issue = benchmark_review_issue(route)
                warnings.append(
                    (
                        f"{sheet_name}: 基准数据使用许可尚未确认；"
                        if issue == "benchmark-license-unresolved"
                        else f"{sheet_name}: 基准/超额来源尚未确认；"
                    )
                    + "本次只生成审查预览，AI 解决对应审查项并重新预览前"
                    "不能提交或启用自动更新"
                )
            if sheet_name not in workbook.sheetnames:
                errors.append(f"Missing managed sheet: {sheet_name}")
                continue
            layout = discover_layout(
                workbook[sheet_name],
                (config.get("column_overrides") or {}).get(sheet_name),
                route,
            )
            existing = existing_rows(workbook[sheet_name], layout)
            start = parse_date(route.get("series_start")) or dt.date.min
            reserved = _summary_reserved_row(
                workbook[sheet_name], layout, route, existing
            )
            integrity_existing = {
                date: values
                for date, values in existing.items()
                if date >= start and (reserved is None or date != reserved[0])
            }
            pre_managed_existing = {
                date: values for date, values in existing.items() if date < start
            }
            history_integrity = _history_integrity(
                layout, route, integrity_existing, tolerance
            )
            pre_managed_integrity = _history_integrity(
                layout, route, pre_managed_existing, tolerance
            )
            history_integrity["scope"] = {
                "series_start": start.isoformat(),
                "pre_managed_rows": len(pre_managed_existing),
                "managed_rows": len(integrity_existing),
            }
            history_integrity["pre_managed_diagnostics"] = {
                "repair_required": pre_managed_integrity["repair_required"],
                "code_column": pre_managed_integrity["code_column"],
                "cumulative_sequence": pre_managed_integrity[
                    "cumulative_sequence"
                ],
            }
            if pre_managed_integrity["repair_required"]:
                warnings.append(
                    f"{sheet_name}: preserved pre-managed history before series_start "
                    "contains integrity diagnostics; it is not modified and does not "
                    "block managed-tail validation"
                )
            if not history_integrity["code_column"]["passed"]:
                errors.append(
                    f"{sheet_name}: managed historical product code column is not constant; "
                    "inspect history_integrity before any supervised repair"
                )
            if not history_integrity["cumulative_sequence"]["passed"]:
                errors.append(
                    f"{sheet_name}: managed unit/cumulative NAV history has a sequence break; "
                    "inspect history_integrity before any supervised repair"
                )
            matches = 0
            conflicts = 0
            pre_managed_candidates = sorted(
                (
                    candidate
                    for candidate in route_rows.get(sheet_name, [])
                    if candidate.date < start
                ),
                key=lambda item: item.date,
            )
            candidates = sorted(
                (
                    candidate
                    for candidate in route_rows.get(sheet_name, [])
                    if candidate.date >= start
                ),
                key=lambda item: item.date,
            )
            if reserved:
                reserved_date = reserved[0]
                observed = existing[reserved_date]
                matching = [
                    candidate
                    for candidate in candidates
                    if candidate.date == reserved_date
                ]
                if observed.get("unit") is not None:
                    if not matching or all(
                        abs(float(observed["unit"]) - candidate.unit) > tolerance
                        for candidate in matching
                    ):
                        errors.append(
                            f"{sheet_name}: partial onboarding unit NAV cannot be verified on {reserved_date.isoformat()}"
                        )
                if observed.get("cumulative") is not None:
                    if not matching or all(
                        abs(
                            float(observed["cumulative"])
                            - effective_cumulative(candidate, route)
                        )
                        > tolerance
                        for candidate in matching
                    ):
                        errors.append(
                            f"{sheet_name}: partial onboarding cumulative NAV cannot be verified on {reserved_date.isoformat()}"
                        )
            selected_candidates, data_frequency, frequency_source = _select_data_rows(
                workbook[sheet_name],
                layout,
                route,
                existing,
                reserved,
                candidates,
            )
            selected_dates = {candidate.date for candidate in selected_candidates}
            reserved_date = reserved[0] if reserved else None
            candidate_dates = selected_dates
            reserved_source_date = (
                min(candidate_dates) if reserved and candidate_dates else None
            )
            known_units = {
                date: values["unit"]
                for date, values in existing.items()
                if date >= start and values["unit"] is not None
            }
            for candidate in candidates:
                expected_code = normalize_code(route.get("code"))
                if expected_code and candidate.code != expected_code:
                    errors.append(
                        f"{sheet_name}: candidate code mismatch on {candidate.date.isoformat()}"
                    )
                if candidate.date > dt.date.today() + dt.timedelta(
                    days=max_future_days
                ):
                    errors.append(
                        f"{sheet_name}: NAV date {candidate.date.isoformat()} exceeds max_future_days"
                    )
                if (
                    reserved_source_date is not None
                    and candidate.date == reserved_source_date
                ):
                    if layout.columns.get("cumulative"):
                        effective_cumulative(candidate, route)
                    known_units[candidate.date] = candidate.unit
                    continue
                if candidate.date not in existing:
                    prior = [
                        (date, unit)
                        for date, unit in known_units.items()
                        if date < candidate.date
                    ]
                    if prior:
                        previous_unit = max(prior)[1]
                        if (
                            abs(candidate.unit / float(previous_unit) - 1)
                            > max_period_change
                        ):
                            errors.append(
                                f"{sheet_name}: NAV change exceeds max_period_change on {candidate.date.isoformat()}"
                            )
                    known_units[candidate.date] = candidate.unit
                if (
                    candidate.date < start
                    or candidate.date not in existing
                    or candidate.date not in selected_dates
                ):
                    continue
                observed = existing[candidate.date]
                unit_ok = (
                    observed["unit"] is not None
                    and abs(float(observed["unit"]) - candidate.unit) <= tolerance
                )
                cumulative_ok = True
                if layout.columns.get("cumulative"):
                    expected_cumulative = effective_cumulative(candidate, route)
                    cumulative_ok = (
                        observed["cumulative"] is not None
                        and abs(float(observed["cumulative"]) - expected_cumulative)
                        <= tolerance
                    )
                expected_code = normalize_code(route.get("code"))
                code_verified = bool(
                    expected_code
                    and layout.columns.get("code")
                    and observed["code"] == expected_code
                )
                expected_name = _norm(route.get("product_name"))
                name_verified = bool(
                    expected_name
                    and (
                        _norm(layout.sheet) == expected_name
                        or (
                            layout.columns.get("name")
                            and _norm(observed.get("name")) == expected_name
                        )
                    )
                )
                code_conflict = bool(
                    expected_code
                    and layout.columns.get("code")
                    and observed["code"]
                    and observed["code"] != expected_code
                )
                name_conflict = bool(
                    expected_name
                    and layout.columns.get("name")
                    and observed.get("name")
                    and _norm(observed["name"]) != expected_name
                )
                identity_ok = (
                    code_verified
                    or name_verified
                    or not (code_conflict or name_conflict)
                )
                if unit_ok and cumulative_ok and identity_ok:
                    matches += 1
                else:
                    conflicts += 1
                    errors.append(
                        f"{sheet_name}: historical value conflict on {candidate.date.isoformat()}"
                    )
            summary_cold_start = reserved_date is not None
            template_cold_start = layout.mode == "template" and matches < minimum
            weekly_pending_baseline = (
                reserved_date is None
                and matches < minimum
                and _weekly_pending_baseline_identity(
                    layout,
                    route,
                    existing,
                    selected_candidates,
                    candidates,
                    conflicts,
                    data_frequency,
                )
            )
            boundary_anchor = _boundary_anchor_identity(
                layout,
                route,
                existing,
                pre_managed_candidates,
                tolerance,
            )
            boundary_zero_baseline = bool(boundary_anchor and not candidates)
            if (
                route.get("baseline_overlap") == "last_existing_point"
                and not boundary_anchor
            ):
                errors.append(
                    f"{sheet_name}: baseline_overlap last_existing_point could not "
                    "verify the workbook tail immediately before series_start"
                )
            summary_review_identity = (
                reserved_date is None
                and matches < minimum
                and not boundary_zero_baseline
                and not weekly_pending_baseline
                and _summary_reviewed_onboarding(
                    layout,
                    route,
                    existing,
                    selected_dates,
                    candidates,
                    conflicts,
                )
            )
            if summary_cold_start:
                if not candidate_dates:
                    errors.append(
                        f"{sheet_name}: summary-mode cold start needs at least one real email NAV date"
                    )
                else:
                    warnings.append(
                        f"{sheet_name}: detected a reserved summary row cold start with {len(candidate_dates)} email date(s); replace the empty placeholder with the earliest email NAV, keep the summary row, and review the first preview"
                    )
            elif template_cold_start:
                if candidate_dates and not conflicts:
                    warnings.append(
                        f"{sheet_name}: bundled template cold start has {len(candidate_dates)} real email date(s); review the first preview, and strict validation resumes after {minimum} verified dates"
                    )
                else:
                    errors.append(
                        f"{sheet_name}: bundled template cold start needs at least one real email NAV date"
                    )
            elif weekly_pending_baseline:
                identity, withheld_dates = weekly_pending_baseline
                warnings.append(
                    f"{sheet_name}: a unique {identity} match and the workbook tail were verified; "
                    f"{len(withheld_dates)} real email date(s) belong to the unfinished current week, "
                    "so generate a zero-add baseline and do not write that week yet"
                )
            elif boundary_zero_baseline and boundary_anchor:
                identity, anchor_date = boundary_anchor
                warnings.append(
                    f"{sheet_name}: the preserved workbook tail on "
                    f"{anchor_date.isoformat()} and a unique {identity} match were "
                    "verified at the adoption boundary; generate a zero-add baseline "
                    "without reopening earlier history"
                )
            elif summary_review_identity:
                warnings.append(
                    f"{sheet_name}: only {matches} historical date(s) were verified, but a unique {summary_review_identity} match allows a review-gated preview; verify the product, dates and NAV values before approval"
                )
            elif matches < minimum:
                message = f"{sheet_name}: only {matches} verified historical dates; {minimum} required"
                if layout.mode == "append":
                    warnings.append(
                        f"{message}; append-mode cold start requires review"
                    )
                else:
                    errors.append(message)
            reports.append(
                {
                    "sheet": sheet_name,
                    "sheet_mode": layout.mode,
                    "cold_start": summary_cold_start
                    or template_cold_start
                    or bool(summary_review_identity)
                    or (
                        layout.mode == "append"
                        and matches < minimum
                        and not boundary_zero_baseline
                        and not weekly_pending_baseline
                    ),
                    "cold_start_kind": (
                        "summary-reserved-row"
                        if summary_cold_start
                        else "bundled-template"
                        if template_cold_start
                        else "summary-reviewed-preview"
                        if summary_review_identity
                        else "append"
                        if (
                            layout.mode == "append"
                            and matches < minimum
                            and not boundary_zero_baseline
                            and not weekly_pending_baseline
                        )
                        else None
                    ),
                    "matched_history_dates": matches,
                    "conflicts": conflicts,
                    "data_frequency": data_frequency,
                    "data_frequency_source": frequency_source,
                    "pending_current_week_baseline": bool(weekly_pending_baseline),
                    "boundary_anchor_verified": bool(boundary_anchor),
                    "boundary_anchor_date": (
                        boundary_anchor[1].isoformat() if boundary_anchor else None
                    ),
                    "withheld_current_week_dates": [
                        date.isoformat()
                        for date in (
                            weekly_pending_baseline[1]
                            if weekly_pending_baseline
                            else []
                        )
                    ],
                    "history_integrity": history_integrity,
                }
            )
    finally:
        workbook.close()
    report = {
        "passed": not errors,
        "routes": reports,
        "warnings": warnings,
        "errors": errors,
        "history_repairs_required": sum(
            1
            for route_report in reports
            if route_report["history_integrity"]["repair_required"]
        ),
    }
    write_json_atomic(STATE_ROOT / "validation-report.json", report)
    return report


def _copy_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    source: int,
    target: int,
    max_column: int,
) -> None:
    sheet.row_dimensions[target].height = sheet.row_dimensions[source].height
    for column in range(1, max_column + 1):
        original = sheet.cell(source, column)
        cell = sheet.cell(target, column)
        if original.has_style:
            cell._style = copy(original._style)
        cell.number_format = original.number_format
        cell.alignment = copy(original.alignment)
        cell.protection = copy(original.protection)
        if isinstance(original.value, ArrayFormula):
            raise WorkbookError(
                f"{sheet.title}: array formula at {original.coordinate} cannot be "
                "copied safely during automatic row insertion"
            )
        if isinstance(original.value, str) and original.value.startswith("="):
            try:
                cell.value = Translator(
                    original.value, origin=original.coordinate
                ).translate_formula(cell.coordinate)
            except Exception as exc:
                raise WorkbookError(
                    f"{sheet.title}: could not translate formula at {original.coordinate}"
                ) from exc
        else:
            cell.value = original.value


def _date_rows(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    summary_row: int,
    start: dt.date,
) -> list[tuple[dt.date, int]]:
    rows: list[tuple[dt.date, int]] = []
    for row in range(layout.data_start, summary_row):
        date = parse_date(sheet.cell(row, layout.columns["date"]).value)
        if date and date >= start:
            rows.append((date, row))
    return sorted(rows)


def _formula_series_start(route: dict[str, Any]) -> dt.date:
    start = parse_date(route.get("series_start")) or dt.date.min
    if route.get("baseline_overlap") == "last_existing_point":
        # This mode marks an adoption/write boundary, not an investment-series
        # reset.  Existing return and summary formulas must continue to use the
        # workbook's full preserved history.
        return dt.date.min
    return start


def _primary_return_column(layout: Layout, route: dict[str, Any]) -> int | None:
    frequency = str(route.get("return_frequency", "weekly"))
    semantic = "daily_return" if frequency == "daily" else "weekly_return"
    return layout.columns.get(semantic) or layout.columns.get("return")


def _period_rows(
    rows: list[tuple[dt.date, int]], frequency: str
) -> tuple[list[int], dict[tuple[int, int], list[tuple[dt.date, int]]]]:
    if frequency == "daily":
        return [row for _, row in rows], {}
    groups: dict[tuple[int, int], list[tuple[dt.date, int]]] = {}
    for date, row in rows:
        iso = date.isocalendar()
        groups.setdefault((iso.year, iso.week), []).append((date, row))
    current_iso = dt.date.today().isocalendar()
    completed = [
        key for key in sorted(groups) if key < (current_iso.year, current_iso.week)
    ]
    return [max(groups[key])[1] for key in completed], groups


def _set_return_formulas(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
    summary_row: int,
    new_rows: set[int],
    changed: set[tuple[int, int]],
    write_summary: bool = True,
) -> tuple[list[int], set[int]]:
    primary_column = _primary_return_column(layout, route)
    semantic_columns = {
        "daily": layout.columns.get("daily_return"),
        "weekly": layout.columns.get("weekly_return"),
    }
    if not any(semantic_columns.values()) and layout.columns.get("return"):
        semantic_columns[str(route.get("return_frequency", "weekly"))] = layout.columns[
            "return"
        ]
    if not primary_column:
        return [], set()
    basis_name = (
        "cumulative"
        if route.get("return_basis", "cumulative") == "cumulative"
        else "unit"
    )
    basis_column = layout.columns.get(basis_name)
    if not basis_column:
        raise WorkbookError(
            f"{sheet.title}: return basis column {basis_name} is missing"
        )
    start = _formula_series_start(route)
    rows = _date_rows(sheet, layout, summary_row, start)
    if not rows:
        return [], set()
    letter = get_column_letter(basis_column)
    results: dict[str, tuple[list[int], set[int]]] = {}
    for frequency, return_column in semantic_columns.items():
        if not return_column:
            continue
        period_rows, groups = _period_rows(rows, frequency)
        affected_rows: set[int] = set()
        if frequency == "daily":
            for index, (_, row) in enumerate(rows):
                if row not in new_rows:
                    continue
                value = (
                    None
                    if index == 0
                    else f"={letter}{row}/{letter}{rows[index - 1][1]}-1"
                )
                if sheet.cell(row, return_column).value != value:
                    sheet.cell(row, return_column).value = value
                    changed.add((row, return_column))
                affected_rows.add(row)
        else:
            new_week_keys = {
                (date.isocalendar().year, date.isocalendar().week)
                for date, row in rows
                if row in new_rows
            }
            for key in new_week_keys:
                week_rows = [row for _, row in groups[key]]
                affected_rows.update(week_rows)
                target = (
                    max(groups[key])[1]
                    if target_key_exists(key, period_rows, groups)
                    else None
                )
                previous = (
                    period_rows[period_rows.index(target) - 1]
                    if target in period_rows and period_rows.index(target) > 0
                    else None
                )
                for row in week_rows:
                    value = (
                        f"={letter}{row}/{letter}{previous}-1"
                        if row == target and previous is not None
                        else None
                    )
                    if sheet.cell(row, return_column).value != value:
                        sheet.cell(row, return_column).value = value
                        changed.add((row, return_column))
        results[frequency] = (period_rows, affected_rows)
    frequency = str(route.get("return_frequency", "weekly"))
    period_rows, affected_rows = results.get(frequency, ([], set()))
    if write_summary:
        summary_value = None
        if period_rows:
            summary_value = f"={letter}{period_rows[-1]}/{letter}{period_rows[0]}-1"
        sheet.cell(summary_row, primary_column).value = summary_value
        changed.add((summary_row, primary_column))
    return period_rows, affected_rows


def target_key_exists(
    key: tuple[int, int],
    period_rows: list[int],
    groups: dict[tuple[int, int], list[tuple[dt.date, int]]],
) -> bool:
    return bool(groups.get(key)) and max(groups[key])[1] in period_rows


def _benchmark_source(
    workbook: openpyxl.Workbook, benchmark: dict[str, Any]
) -> tuple[dict[dt.date, tuple[int, Any]], int]:
    source_name = str(benchmark["source_sheet"])
    if source_name not in workbook.sheetnames:
        raise WorkbookError(f"Missing benchmark source sheet: {source_name}")
    sheet = workbook[source_name]
    date_column = _column(benchmark["source_date"])
    value_column = _column(benchmark["source_value"])
    values: dict[dt.date, tuple[int, Any]] = {}
    for row in range(1, sheet.max_row + 1):
        date = parse_date(sheet.cell(row, date_column).value)
        value = sheet.cell(row, value_column).value
        if date and value is not None:
            if date in values:
                raise WorkbookError(
                    f"{source_name}: duplicate benchmark date {date.isoformat()}"
                )
            values[date] = (row, value)
    return values, value_column


def _product_formula(column: int, rows: list[int]) -> str:
    expressions = [f"1+{get_column_letter(column)}{row}" for row in rows]
    while len(expressions) > 200:
        expressions = [
            f"PRODUCT({','.join(expressions[index : index + 200])})"
            for index in range(0, len(expressions), 200)
        ]
    return f"=PRODUCT({','.join(expressions)})-1"


def _set_benchmark_formulas(
    workbook: openpyxl.Workbook,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
    summary_row: int,
    period_rows: list[int],
    affected_rows: set[int],
    changed: set[tuple[int, int]],
    write_summary: bool = True,
    review_rows: set[int] | None = None,
) -> None:
    if benchmark_requires_review(route):
        review_columns = {
            column
            for column in (
                layout.columns.get("benchmark_level"),
                layout.columns.get("benchmark_return"),
                layout.columns.get("excess"),
            )
            if column
        }
        rows_to_clear = set(review_rows if review_rows is not None else affected_rows)
        if write_summary:
            rows_to_clear.add(summary_row)
        for row in rows_to_clear:
            for column in review_columns:
                if sheet.cell(row, column).value is not None:
                    sheet.cell(row, column).value = None
                    changed.add((row, column))
        return
    benchmark = route.get("benchmark")
    if not benchmark:
        return
    source_values, source_column = _benchmark_source(workbook, benchmark)
    source_sheet = quote_sheetname(str(benchmark["source_sheet"]))
    source_letter = get_column_letter(source_column)
    target_return = layout.columns.get("benchmark_return")
    excess = layout.columns.get("excess")
    product_return = _primary_return_column(layout, route)
    if not target_return or not excess or not product_return:
        raise WorkbookError(
            f"{sheet.title}: benchmark return, product return, and excess columns are required"
        )
    start = _formula_series_start(route)
    rows = _date_rows(sheet, layout, summary_row, start)
    by_row = {row: date for date, row in rows}
    target_rows = [
        row for row in period_rows if sheet.cell(row, product_return).value is not None
    ]
    affected_targets = [row for row in target_rows if row in affected_rows]
    source_type = str(benchmark.get("source_type", "level"))
    if source_type not in {"aligned_return", "level"}:
        raise WorkbookError("benchmark.source_type must be aligned_return or level")
    level_column = layout.columns.get("benchmark_level")
    for row in affected_rows:
        if row in affected_targets:
            continue
        for column in (
            target_return,
            excess,
            level_column if source_type == "level" else None,
        ):
            if column and sheet.cell(row, column).value is not None:
                sheet.cell(row, column).value = None
                changed.add((row, column))
    if source_type == "aligned_return":
        for row in affected_targets:
            date = by_row[row]
            source = source_values.get(date)
            if not source:
                raise WorkbookError(
                    f"{sheet.title}: benchmark is missing for {date.isoformat()}"
                )
            source_row = source[0]
            sheet.cell(
                row, target_return
            ).value = f"={source_sheet}!{source_letter}{source_row}"
            changed.add((row, target_return))
            sheet.cell(
                row, excess
            ).value = f"={get_column_letter(product_return)}{row}-{get_column_letter(target_return)}{row}"
            changed.add((row, excess))
    else:
        for row in affected_targets:
            index = period_rows.index(row)
            if index == 0:
                raise WorkbookError(
                    f"{sheet.title}: benchmark level has no prior period anchor"
                )
            previous_level_row = period_rows[index - 1]
            for anchor_row in (previous_level_row, row):
                date = by_row[anchor_row]
                source = source_values.get(date)
                if not source:
                    raise WorkbookError(
                        f"{sheet.title}: benchmark is missing for {date.isoformat()}"
                    )
                if level_column and (
                    anchor_row == row
                    or sheet.cell(anchor_row, level_column).value is None
                ):
                    sheet.cell(
                        anchor_row, level_column
                    ).value = f"={source_sheet}!{source_letter}{source[0]}"
                    changed.add((anchor_row, level_column))
            if level_column:
                level_letter = get_column_letter(level_column)
                sheet.cell(
                    row, target_return
                ).value = f"={level_letter}{row}/{level_letter}{previous_level_row}-1"
            else:
                current_source = source_values[by_row[row]][0]
                previous_source = source_values[by_row[previous_level_row]][0]
                sheet.cell(row, target_return).value = (
                    f"={source_sheet}!{source_letter}{current_source}/"
                    f"{source_sheet}!{source_letter}{previous_source}-1"
                )
            changed.add((row, target_return))
            sheet.cell(
                row, excess
            ).value = f"={get_column_letter(product_return)}{row}-{get_column_letter(target_return)}{row}"
            changed.add((row, excess))

    if not write_summary:
        return
    if source_type == "level" and period_rows:
        first_date, last_date = by_row[period_rows[0]], by_row[period_rows[-1]]
        first_source, last_source = (
            source_values.get(first_date),
            source_values.get(last_date),
        )
        if not first_source or not last_source:
            raise WorkbookError(f"{sheet.title}: benchmark summary anchors are missing")
        summary_value = f"={source_sheet}!{source_letter}{last_source[0]}/{source_sheet}!{source_letter}{first_source[0]}-1"
    elif period_rows:
        summary_value = _product_formula(target_return, period_rows)
    else:
        summary_value = None
    sheet.cell(summary_row, target_return).value = summary_value
    sheet.cell(summary_row, excess).value = (
        f"={get_column_letter(product_return)}{summary_row}-{get_column_letter(target_return)}{summary_row}"
        if summary_value is not None
        else None
    )
    changed.update({(summary_row, target_return), (summary_row, excess)})


def _ensure_summary_formula_safety(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
) -> None:
    managed = {
        column
        for column in (
            layout.columns.get("return"),
            layout.columns.get("daily_return"),
            layout.columns.get("weekly_return"),
        )
        if column
    }
    if route.get("benchmark") or benchmark_requires_review(route):
        managed.update(
            column
            for column in (
                layout.columns.get("benchmark_level"),
                layout.columns.get("benchmark_return"),
                layout.columns.get("excess"),
            )
            if column
        )
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(layout.summary_row, column).value
        if isinstance(value, ArrayFormula):
            if _review_summary_array_column(sheet, layout, route, column):
                continue
            coordinate = sheet.cell(layout.summary_row, column).coordinate
            raise WorkbookError(
                f"{sheet.title}: array formula at {coordinate} cannot be moved "
                "safely during automatic row insertion"
            )
        if isinstance(value, str) and value.startswith("=") and column not in managed:
            coordinate = sheet.cell(layout.summary_row, column).coordinate
            raise WorkbookError(
                f"{sheet.title}: summary formula at {coordinate} is not managed; "
                "automatic insertion cannot prove that its range will expand safely"
            )


def _review_summary_array_column(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
    column: int,
) -> bool:
    if not benchmark_requires_review(route) or column not in {
        layout.columns.get("benchmark_level"),
        layout.columns.get("benchmark_return"),
        layout.columns.get("excess"),
    }:
        return False
    cell = sheet.cell(layout.summary_row, column)
    value = cell.value
    if not isinstance(value, ArrayFormula):
        return False
    try:
        min_column, min_row, max_column, max_row = range_boundaries(value.ref)
    except (TypeError, ValueError):
        return False
    return (
        min_column == max_column == cell.column
        and min_row == max_row == cell.row
    )


def _clear_review_summary_arrays(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    layout: Layout,
    route: dict[str, Any],
) -> list[int]:
    columns = [
        column
        for column in range(1, sheet.max_column + 1)
        if _review_summary_array_column(sheet, layout, route, column)
    ]
    for column in columns:
        sheet.cell(layout.summary_row, column).value = None
    return columns


def _ensure_array_formula_insertion_safety(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    insert_before: int,
) -> None:
    template_row = insert_before - 1
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, ArrayFormula):
                continue
            try:
                min_column, min_row, max_column, max_row = range_boundaries(value.ref)
            except (TypeError, ValueError) as exc:
                raise WorkbookError(
                    f"{sheet.title}: array formula at {cell.coordinate} has an invalid range"
                ) from exc
            if not (
                min_column <= cell.column <= max_column
                and min_row <= cell.row <= max_row
            ):
                raise WorkbookError(
                    f"{sheet.title}: array formula anchor {cell.coordinate} is outside "
                    f"its declared range {value.ref}"
                )
            if cell.row == template_row or max_row >= insert_before:
                raise WorkbookError(
                    f"{sheet.title}: array formula at {cell.coordinate} with range "
                    f"{value.ref} cannot be copied or moved safely during automatic "
                    "row insertion"
                )


def _write_nav_values(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    layout: Layout,
    route: dict[str, Any],
    nav: NavRow,
    changed: set[tuple[int, int]] | None = None,
) -> None:
    values: list[tuple[int, Any]] = [
        (layout.columns["date"], nav.date),
        (layout.columns["unit"], nav.unit),
    ]
    if layout.columns.get("cumulative"):
        values.append((layout.columns["cumulative"], effective_cumulative(nav, route)))
    code_value = normalize_code(route.get("code")) or nav.code
    if layout.columns.get("code") and code_value:
        values.append((layout.columns["code"], code_value))
    if layout.columns.get("name") and route.get("product_name"):
        values.append((layout.columns["name"], str(route["product_name"]).strip()))
    for column, value in values:
        cell = sheet.cell(row, column)
        same = (
            parse_date(cell.value) == value
            if column == layout.columns["date"]
            else cell.value == value
        )
        if not same:
            cell.value = value
            if changed is not None:
                changed.add((row, column))


def build_preview(
    config: dict[str, Any],
    route_rows: dict[str, list[NavRow]],
    warnings: list[str] | None = None,
    *,
    diagnostic_only: bool = False,
) -> dict[str, Any]:
    (STATE_ROOT / "plan.json").unlink(missing_ok=True)
    CONCURRENCY_REPORT.unlink(missing_ok=True)
    master = Path(config["workbook_path"])
    master_sha256 = file_sha256(master)
    master_manifest = workbook_manifest(master)
    if file_sha256(master) != master_sha256:
        write_concurrency_report(
            master,
            phase="preview-baseline",
            plan_id=None,
            expected_sha256=master_sha256,
            expected_manifest=master_manifest,
        )
        raise WorkbookError(
            "正式工作簿在建立预览基线时被外部进程修改；"
            "未生成计划，请查看 concurrency-report.json 后重新预览"
        )
    preview_dir = ROOT / "previews"
    preview_dir.mkdir(exist_ok=True)
    keep = int((config.get("retention") or {}).get("preview_count", 10))
    old_previews = sorted(
        (
            path
            for path in preview_dir.glob("preview-*")
            if path.suffix.lower() in {".xlsx", ".xlsm", ".txt"}
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for old in old_previews[max(keep - 1, 0) :]:
        make_file_writable(old)
        old.unlink()
    preview = (
        preview_dir
        / f"preview-只读审查-{dt.datetime.now():%Y%m%d-%H%M%S}-{uuid.uuid4().hex[:8]}{master.suffix.lower()}"
    )
    shutil.copy2(master, preview)
    make_file_writable(preview)
    keep_vba = master.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(preview, data_only=False, keep_vba=keep_vba)
    plan_sheets: list[dict[str, Any]] = []
    baseline_routes: list[dict[str, Any]] = []
    blocking_reviews: list[dict[str, str]] = []
    if diagnostic_only:
        blocking_reviews.append(
            {
                "sheet": "、".join(
                    str(route["sheet"]) for route in active_routes(config)
                ),
                "issue": "scoped-diagnostic-preview",
            }
        )
    try:
        for route in active_routes(config):
            sheet_name = str(route["sheet"])
            if benchmark_requires_review(route):
                blocking_reviews.append(
                    {
                        "sheet": sheet_name,
                        "issue": str(benchmark_review_issue(route)),
                    }
                )
            sheet = workbook[sheet_name]
            layout = discover_layout(
                sheet, (config.get("column_overrides") or {}).get(sheet_name), route
            )
            current = existing_rows(sheet, layout)
            start = parse_date(route.get("series_start")) or dt.date.min
            observed_candidates = sorted(
                route_rows.get(sheet_name, []), key=lambda row: row.date
            )
            candidates = [
                row for row in observed_candidates if row.date >= start
            ]
            reserved = _summary_reserved_row(sheet, layout, route, current)
            candidates, data_frequency, frequency_source = _select_data_rows(
                sheet, layout, route, current, reserved, candidates
            )
            withheld_current_week_dates = _withheld_current_week_dates(
                data_frequency, current, candidates, observed_candidates
            )
            reserved_date = reserved[0] if reserved else None
            reserved_row = reserved[1] if reserved else None
            reserved_nav: NavRow | None = None
            if reserved_date is not None:
                distinct_dates = {row.date for row in candidates}
                if not distinct_dates:
                    raise WorkbookError(
                        f"{sheet_name}: summary-mode cold start needs at least one real email NAV date"
                    )
                reserved_nav = candidates[0]
            elif layout.mode == "template" and not current:
                if not candidates:
                    raise WorkbookError(
                        f"{sheet_name}: bundled template cold start needs at least one real email NAV date"
                    )
                reserved_nav = candidates[0]
                reserved_row = layout.data_start
            additions = (
                candidates[1:]
                if reserved_nav is not None
                else [row for row in candidates if row.date not in current]
            )
            baseline_routes.append(
                {
                    "sheet": sheet_name,
                    "sheet_mode": layout.mode,
                    "data_frequency": data_frequency,
                    "data_frequency_source": frequency_source,
                    "matched_history_dates": len(
                        set(current) & {row.date for row in candidates}
                    ),
                    "workbook_latest_date": (
                        max(current).isoformat() if current else None
                    ),
                    "email_latest_date": (
                        max(row.date for row in observed_candidates).isoformat()
                        if observed_candidates
                        else None
                    ),
                    "withheld_current_week_dates": [
                        date.isoformat() for date in withheld_current_week_dates
                    ],
                    "verification_anchor_date": (
                        max(
                            (
                                row.date
                                for row in observed_candidates
                                if row.date < start
                            ),
                            default=None,
                        ).isoformat()
                        if any(row.date < start for row in observed_candidates)
                        else None
                    ),
                    "pending_dates": len(additions)
                    + (1 if reserved_nav is not None else 0),
                }
            )
            if not additions and reserved_nav is None:
                continue
            if layout.mode in {"summary", "template"}:
                _ensure_summary_formula_safety(sheet, layout, route)
            latest_existing = max(current) if current else None
            gaps = (
                [row.date for row in additions if row.date <= latest_existing]
                if latest_existing
                else []
            )
            if gaps:
                dates = ", ".join(date.isoformat() for date in gaps)
                raise WorkbookError(
                    f"{sheet_name}: internal historical gaps require supervised repair: {dates}"
                )
            old_summary = layout.summary_row
            count = len(additions)
            max_column = max([sheet.max_column, *layout.columns.values()])
            changed: set[tuple[int, int]] = set()
            review_array_columns = _clear_review_summary_arrays(
                sheet, layout, route
            )
            filled_existing_rows: list[int] = []
            for column, label in (layout.headers_to_write or {}).items():
                cell = sheet.cell(layout.header_row, column)
                cell.value = label
                font = copy(cell.font)
                font.bold = True
                cell.font = font
                changed.add((layout.header_row, column))
            if reserved_nav is not None and reserved_row is not None:
                _write_nav_values(
                    sheet,
                    reserved_row,
                    layout,
                    route,
                    reserved_nav,
                    changed,
                )
                filled_existing_rows.append(reserved_row)
            if count:
                _ensure_array_formula_insertion_safety(sheet, old_summary)
                sheet.insert_rows(old_summary, count)
            for offset, nav in enumerate(additions):
                target = old_summary + offset
                if current or layout.mode in {"summary", "template"}:
                    template = old_summary - 1 if offset == 0 else target - 1
                    _copy_row(sheet, template, target, max_column)
                _write_nav_values(sheet, target, layout, route, nav)
                if layout.mode == "append" and not current:
                    sheet.cell(
                        target, layout.columns["date"]
                    ).number_format = "yyyy-mm-dd"
                    sheet.cell(
                        target, layout.columns["unit"]
                    ).number_format = "0.000000"
                if layout.columns.get("cumulative"):
                    if layout.mode == "append" and not current:
                        sheet.cell(
                            target, layout.columns["cumulative"]
                        ).number_format = "0.000000"
            new_summary = old_summary + count
            changed.update(
                (new_summary, column) for column in review_array_columns
            )
            new_rows = set(range(old_summary, new_summary))
            managed_rows = new_rows | set(filled_existing_rows)
            period_rows, affected_rows = _set_return_formulas(
                sheet,
                layout,
                route,
                new_summary,
                managed_rows,
                changed,
                write_summary=layout.mode in {"summary", "template"},
            )
            _set_benchmark_formulas(
                workbook,
                sheet,
                layout,
                route,
                new_summary,
                period_rows,
                affected_rows,
                changed,
                write_summary=layout.mode in {"summary", "template"},
                review_rows=managed_rows,
            )
            plan_sheets.append(
                {
                    "sheet": sheet_name,
                    "sheet_mode": layout.mode,
                    "data_frequency": data_frequency,
                    "data_frequency_source": frequency_source,
                    "header_row": layout.header_row,
                    "insert_before": old_summary,
                    "insert_count": count,
                    "populated_count": count + len(filled_existing_rows),
                    "new_rows": list(range(old_summary, new_summary)),
                    "filled_existing_rows": filled_existing_rows,
                    "summary_row": new_summary,
                    "copy_template_rows": bool(current)
                    or layout.mode in {"summary", "template"},
                    "format_rows": sorted(
                        {
                            *([layout.header_row] if layout.headers_to_write else []),
                            *(
                                range(old_summary, new_summary)
                                if layout.mode == "append" and not current
                                else []
                            ),
                        }
                    ),
                    "changed_cells": sorted(
                        [{"row": row, "column": column} for row, column in changed],
                        key=lambda item: (item["row"], item["column"]),
                    ),
                    "new_dates": [
                        row.date.isoformat()
                        for row in (
                            ([reserved_nav] if reserved_nav is not None else [])
                            + additions
                        )
                    ],
                    "review_array_formulas_cleared": [
                        f"{get_column_letter(column)}{new_summary}"
                        for column in review_array_columns
                    ],
                    "return_columns": [
                        column
                        for column in (
                            layout.columns.get("return"),
                            layout.columns.get("daily_return"),
                            layout.columns.get("weekly_return"),
                            layout.columns.get("benchmark_return"),
                            layout.columns.get("excess"),
                        )
                        if column
                    ],
                }
            )
        workbook.save(preview)
    except Exception:
        workbook.close()
        preview.unlink(missing_ok=True)
        (STATE_ROOT / "plan.json").unlink(missing_ok=True)
        raise
    else:
        workbook.close()
    if file_sha256(master) != master_sha256:
        preview.unlink(missing_ok=True)
        write_concurrency_report(
            master,
            phase="preview-build",
            plan_id=None,
            expected_sha256=master_sha256,
            expected_manifest=master_manifest,
        )
        raise WorkbookError(
            "正式工作簿在生成预览期间被外部进程修改；"
            "旧基线已失效，未生成计划，请查看 concurrency-report.json"
        )
    plan = {
        "schema_version": 1,
        "plan_id": str(uuid.uuid4()),
        "created": dt.datetime.now().isoformat(timespec="seconds"),
        "config_sha256": payload_sha256(config),
        "master_path": str(master.resolve()),
        "master_sha256": master_sha256,
        "master_manifest": master_manifest,
        "preview_path": str(preview.resolve()) if plan_sheets else None,
        "preview_sha256": file_sha256(preview) if plan_sheets else None,
        "preview_display_name": preview.name if plan_sheets else None,
        "preview_read_only": bool(plan_sheets),
        "approval_kind": ("workbook-preview" if plan_sheets else "validated-no-change"),
        "review_path": None,
        "review_sha256": None,
        "warnings": list(warnings or []),
        "committable": not blocking_reviews,
        "diagnostic_only": diagnostic_only,
        "blocking_reviews": blocking_reviews,
        "sheets": plan_sheets,
    }
    if not plan_sheets:
        preview.unlink(missing_ok=True)
        review = preview.with_suffix(".txt")
        lines = [
            "零新增基线验收报告",
            "",
            f"生成时间：{plan['created']}",
            "结论：邮箱历史与现有工作簿已通过校验，当前没有待写入的新日期。",
            "正式工作簿：未修改；批准本报告不会创建备份或写入工作簿。",
            (
                "批准用途：当前仍有待确认项，本报告不能批准为自动更新权限。"
                if blocking_reviews
                else "批准用途：记录当前写表配置的自动更新权限，供后续计划任务使用。"
            ),
            "",
            f"受管产品页数量：{len(baseline_routes)}",
        ]
        for item in baseline_routes:
            lines.extend(
                [
                    "",
                    f"产品页：{item['sheet']}",
                    f"工作表模式：{item['sheet_mode']}",
                    f"数据频率：{item['data_frequency']}（{item['data_frequency_source']}）",
                    f"已核验历史日期数：{item['matched_history_dates']}",
                    f"工作簿最新日期：{item['workbook_latest_date'] or '无'}",
                    f"邮箱最新日期：{item['email_latest_date'] or '无'}",
                    f"待写入日期数：{item['pending_dates']}",
                    "未完成自然周暂缓日期数："
                    f"{len(item['withheld_current_week_dates'])}",
                    f"接管边界核验日期：{item['verification_anchor_date'] or '无'}",
                ]
            )
        if warnings:
            lines.extend(["", "警告：", *[f"- {item}" for item in warnings]])
        if blocking_reviews:
            lines.extend(
                [
                    "",
                    "待 AI 解决后重新预览：",
                    *[
                        f"- {item['sheet']}：基准/超额来源尚未确认"
                        if item["issue"] == "benchmark-source-unresolved"
                        else (
                            f"- {item['sheet']}：基准数据使用许可尚未确认"
                            if item["issue"] == "benchmark-license-unresolved"
                            else f"- {item['sheet']}：这是局部诊断预览，必须完成一次完整预览"
                        )
                        for item in blocking_reviews
                    ],
                ]
            )
        review.write_text("\n".join(lines) + "\n", encoding="utf-8")
        plan["review_path"] = str(review.resolve())
        plan["review_sha256"] = file_sha256(review)
        plan["preview_display_name"] = review.name
        plan["preview_read_only"] = True
        make_review_file_read_only(review)
        try:
            write_json_atomic(STATE_ROOT / "plan.json", plan)
        except Exception:
            make_file_writable(review)
            review.unlink(missing_ok=True)
            raise
        return plan
    try:
        validate_preview(config, plan)
    except Exception:
        preview.unlink(missing_ok=True)
        (STATE_ROOT / "plan.json").unlink(missing_ok=True)
        raise
    make_review_file_read_only(preview)
    try:
        write_json_atomic(STATE_ROOT / "plan.json", plan)
    except Exception:
        make_file_writable(preview)
        preview.unlink(missing_ok=True)
        raise
    return plan


def validate_preview(config: dict[str, Any], plan: dict[str, Any]) -> None:
    master = Path(plan["master_path"])
    preview = Path(plan["preview_path"])
    keep_vba = master.suffix.lower() == ".xlsm"
    original = openpyxl.load_workbook(
        master, data_only=False, read_only=False, keep_vba=keep_vba
    )
    candidate = openpyxl.load_workbook(
        preview, data_only=False, read_only=False, keep_vba=keep_vba
    )
    try:
        if original.sheetnames != candidate.sheetnames:
            raise WorkbookError("Preview changed the workbook sheet topology")
        plans = {item["sheet"]: item for item in plan["sheets"]}
        routes = {str(route["sheet"]): route for route in active_routes(config)}
        for name in original.sheetnames:
            left, right = original[name], candidate[name]
            sheet_plan = plans.get(name)
            if sheet_plan is None:
                if left.max_row != right.max_row or left.max_column != right.max_column:
                    raise WorkbookError(f"{name}: preview changed an unmanaged sheet")
                for row in range(1, left.max_row + 1):
                    for column in range(1, left.max_column + 1):
                        if not _cell_values_equal(
                            left.cell(row, column).value,
                            right.cell(row, column).value,
                        ):
                            raise WorkbookError(
                                f"{name}: preview changed an unmanaged cell"
                            )
                continue
            insert_before = int(sheet_plan["insert_before"])
            new_summary = int(sheet_plan["summary_row"])
            allowed = {
                (int(item["row"]), int(item["column"]))
                for item in sheet_plan.get("changed_cells") or []
            }
            max_column = max(left.max_column, right.max_column)
            for row in range(1, insert_before):
                for column in range(1, max_column + 1):
                    if (row, column) not in allowed and not _cell_values_equal(
                        left.cell(row, column).value,
                        right.cell(row, column).value,
                    ):
                        raise WorkbookError(
                            f"{name}: preview changed an unapproved historical cell"
                        )
            for column in range(1, max_column + 1):
                if (new_summary, column) not in allowed and not _cell_values_equal(
                    left.cell(insert_before, column).value,
                    right.cell(new_summary, column).value,
                ):
                    raise WorkbookError(
                        f"{name}: preview changed an unapproved summary cell at "
                        f"{get_column_letter(column)}{new_summary}"
                    )
        for sheet_plan in plan["sheets"]:
            sheet = candidate[sheet_plan["sheet"]]
            layout = discover_layout(
                sheet,
                (config.get("column_overrides") or {}).get(sheet.title),
                routes[sheet.title],
            )
            dates = []
            for row in range(layout.data_start, layout.summary_row):
                date = parse_date(sheet.cell(row, layout.columns["date"]).value)
                if date:
                    dates.append(date)
            if len(dates) != len(set(dates)):
                raise WorkbookError(f"{sheet.title}: preview contains duplicate dates")
            if not set(sheet_plan["new_dates"]) <= {date.isoformat() for date in dates}:
                raise WorkbookError(f"{sheet.title}: preview is missing proposed dates")
            for row in range(1, sheet.max_row + 1):
                for cell in sheet[row]:
                    if (
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                        and "#REF!" in cell.value.upper()
                    ):
                        raise WorkbookError(
                            f"{sheet.title}: preview contains a broken formula at {cell.coordinate}"
                        )
    finally:
        original.close()
        candidate.close()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _manifest_value(value: Any) -> tuple[str, str]:
    if isinstance(value, ArrayFormula):
        return "array-formula", f"{value.ref}\x1f{value.text}"
    if isinstance(value, str) and value.startswith("="):
        return "formula", value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return "date-time", value.isoformat()
    if isinstance(value, bytes):
        return "bytes", hashlib.sha256(value).hexdigest()
    return type(value).__name__, repr(value)


def _manifest_cell_hash(value: Any) -> tuple[str, bool]:
    kind, normalized = _manifest_value(value)
    encoded = f"{kind}\x1f{normalized}".encode("utf-8", errors="backslashreplace")
    return hashlib.sha256(encoded).hexdigest(), kind in {
        "formula",
        "array-formula",
    }


def workbook_manifest(path: Path) -> dict[str, Any]:
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(
        path,
        data_only=False,
        read_only=True,
        keep_vba=keep_vba,
    )
    sheets: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            values_digest = hashlib.sha256()
            formulas_digest = hashlib.sha256()
            nonempty_cells = 0
            formula_cells = 0
            sampled_cells: dict[str, dict[str, Any]] = {}
            tail_start = max(1, sheet.max_row - MANIFEST_TAIL_ROWS + 1)
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value in (None, ""):
                        continue
                    cell_hash, is_formula = _manifest_cell_hash(value)
                    token = (
                        f"{cell.coordinate}\x1f{cell_hash}\x1e"
                    ).encode("ascii")
                    values_digest.update(token)
                    nonempty_cells += 1
                    if is_formula:
                        formulas_digest.update(token)
                        formula_cells += 1
                    if cell.row <= MANIFEST_HEAD_ROWS or cell.row >= tail_start:
                        sampled_cells[cell.coordinate] = {
                            "hash": cell_hash,
                            "formula": is_formula,
                        }
            sheets.append(
                {
                    "sheet": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "nonempty_cells": nonempty_cells,
                    "formula_cells": formula_cells,
                    "values_sha256": values_digest.hexdigest(),
                    "formulas_sha256": formulas_digest.hexdigest(),
                    "sampled_cells": sampled_cells,
                }
            )
    finally:
        workbook.close()
    return {
        "schema_version": 1,
        "sampled_regions": {
            "head_rows": MANIFEST_HEAD_ROWS,
            "tail_rows": MANIFEST_TAIL_ROWS,
        },
        "sheet_order": [item["sheet"] for item in sheets],
        "sheets": sheets,
    }


def workbook_manifest_diff(
    expected: dict[str, Any], current: dict[str, Any]
) -> dict[str, Any]:
    expected_sheets = {
        str(item.get("sheet") or ""): item
        for item in expected.get("sheets") or []
        if isinstance(item, dict)
    }
    current_sheets = {
        str(item.get("sheet") or ""): item
        for item in current.get("sheets") or []
        if isinstance(item, dict)
    }
    changes: list[dict[str, Any]] = []
    for name in sorted(set(expected_sheets) | set(current_sheets)):
        before = expected_sheets.get(name)
        after = current_sheets.get(name)
        if before is None or after is None:
            changes.append(
                {
                    "sheet": name,
                    "status": "added" if before is None else "removed",
                }
            )
            continue
        fields = (
            "max_row",
            "max_column",
            "nonempty_cells",
            "formula_cells",
            "values_sha256",
            "formulas_sha256",
        )
        if all(before.get(field) == after.get(field) for field in fields):
            continue
        before_cells = before.get("sampled_cells") or {}
        after_cells = after.get("sampled_cells") or {}
        changed_coordinates = sorted(
            coordinate
            for coordinate in set(before_cells) | set(after_cells)
            if before_cells.get(coordinate) != after_cells.get(coordinate)
        )
        changes.append(
            {
                "sheet": name,
                "status": "changed",
                "rows_before": before.get("max_row"),
                "rows_after": after.get("max_row"),
                "row_delta": int(after.get("max_row") or 0)
                - int(before.get("max_row") or 0),
                "columns_before": before.get("max_column"),
                "columns_after": after.get("max_column"),
                "column_delta": int(after.get("max_column") or 0)
                - int(before.get("max_column") or 0),
                "nonempty_cells_before": before.get("nonempty_cells"),
                "nonempty_cells_after": after.get("nonempty_cells"),
                "formula_cells_before": before.get("formula_cells"),
                "formula_cells_after": after.get("formula_cells"),
                "values_changed": (
                    before.get("values_sha256") != after.get("values_sha256")
                ),
                "formulas_changed": (
                    before.get("formulas_sha256")
                    != after.get("formulas_sha256")
                ),
                "sampled_changed_cells": len(changed_coordinates),
                "sample_coordinates": changed_coordinates[:20],
                "sample_truncated": len(changed_coordinates) > 20,
            }
        )
    topology_changed = expected.get("sheet_order") != current.get("sheet_order")
    return {
        "sheet_topology_changed": topology_changed,
        "changed_sheet_count": len(changes),
        "sheets": changes,
    }


def write_concurrency_report(
    master: Path,
    *,
    phase: str,
    plan_id: str | None,
    expected_sha256: str | None,
    expected_manifest: dict[str, Any] | None,
) -> dict[str, Any]:
    current_exists = master.is_file()
    try:
        current_sha256 = file_sha256(master) if current_exists else None
    except OSError:
        current_sha256 = None
    manifest_available = False
    current_manifest: dict[str, Any] = {}
    if current_exists:
        try:
            current_manifest = workbook_manifest(master)
            manifest_available = True
        except Exception:
            # A workbook can be briefly unreadable while another application is
            # replacing it. The concurrency block must still produce a safe,
            # value-free report instead of masking the original condition.
            current_manifest = {}
    differences = (
        workbook_manifest_diff(expected_manifest or {}, current_manifest)
        if manifest_available
        else {
            "sheet_topology_changed": None,
            "changed_sheet_count": None,
            "sheets": [],
        }
    )
    report = {
        "schema_version": 1,
        "blocked": True,
        "kind": "master-workbook-concurrency",
        "detected_at": dt.datetime.now().isoformat(timespec="seconds"),
        "phase": phase,
        "plan_id": plan_id,
        "master_file": master.name,
        "master_exists": current_exists,
        "expected_sha256": expected_sha256,
        "current_sha256": current_sha256,
        "binary_hash_changed": expected_sha256 != current_sha256,
        "current_manifest_available": manifest_available,
        "binary_or_metadata_only": bool(
            current_exists
            and manifest_available
            and expected_sha256 != current_sha256
            and not differences["sheet_topology_changed"]
            and not differences["changed_sheet_count"]
        ),
        **differences,
        "action": (
            "旧预览计划已失效；不得回滚、覆盖或自动接受当前正式表。"
            "先核实外部保存来源，再以当前正式表重新生成完整预览。"
        ),
    }
    write_json_atomic(CONCURRENCY_REPORT, report)
    return report


def payload_sha256(payload: Any) -> str:
    encoded = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()
