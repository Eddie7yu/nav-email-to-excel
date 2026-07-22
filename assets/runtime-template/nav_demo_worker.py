from __future__ import annotations

import argparse
import datetime as dt
import json
import os
from email.message import EmailMessage
from email.parser import BytesParser
from email.policy import default
from pathlib import Path
from typing import Any

import openpyxl

import nav_mail
from nav_commit import commit
from nav_config import load_config, write_json_atomic
from nav_service import discover, doctor, preview, validate


ROOT = Path(__file__).resolve().parent
STATE_PATH = ROOT / "demo-state.json"
FIXTURE_VALUE = "-".join(("fixture", "only"))


def require(condition: bool, message: str) -> None:
    if not condition:
        raise RuntimeError(message)


def date_series() -> tuple[dt.date, dt.date, dt.date]:
    today = dt.date.today()
    current_monday = today - dt.timedelta(days=today.weekday())
    latest = current_monday - dt.timedelta(days=3)
    return latest - dt.timedelta(days=14), latest - dt.timedelta(days=7), latest


def create_workbook(path: Path, dates: tuple[dt.date, dt.date, dt.date]) -> None:
    first, second, latest = dates
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Demo Fund"
    sheet.append(
        [
            "NAV Date",
            "Fund Return",
            "Product Name",
            "Unit NAV",
            "Product Code",
            "Cumulative NAV",
            "Audit Formula",
            "Benchmark Return",
            "Excess",
        ]
    )
    sheet.append([first, None, "Example Fund", 1.0, "DEMO01", 1.0, "=D2*2"])
    sheet.append(
        [
            second,
            "=F3/F2-1",
            "Example Fund",
            1.01,
            "DEMO01",
            1.01,
            "=D3*2",
            0.005,
            "=B3-H3",
        ]
    )
    sheet.append(["TOTAL", "=F3/F2-1", None, None, None, None, None, 0.005, "=B4-H4"])
    benchmark = book.create_sheet("Demo Benchmark")
    benchmark.append(["Date", "Benchmark Return"])
    benchmark.append([first, 0.0])
    benchmark.append([second, 0.005])
    benchmark.append([latest, -0.002])
    book.save(path)


def create_config(
    workbook: Path, dates: tuple[dt.date, dt.date, dt.date]
) -> dict[str, Any]:
    first, _, _ = dates
    return {
        "schema_version": 1,
        "runtime_id": "00000000-0000-4000-8000-000000000002",
        "workbook_path": str(workbook.resolve()),
        "imap": {
            "host": "imap.example.invalid",
            "port": 993,
            "user": "user@example.invalid",
            "mailbox": "INBOX",
            "lookback_days": 30,
            "max_messages": 20,
            "max_message_bytes": 1048576,
            "max_total_bytes": 4194304,
        },
        "routes": [
            {
                "sender": "sender@example.invalid",
                "subject_contains": "NAV",
                "sheet": "Demo Fund",
                "code": "DEMO01",
                "parser": "auto",
                "allow_sender_only": False,
                "cumulative_policy": "require",
                "cumulative_offset": None,
                "return_basis": "cumulative",
                "return_frequency": "weekly",
                "series_start": first.isoformat(),
                "max_staleness_days": 14,
                "benchmark": {
                    "source_sheet": "Demo Benchmark",
                    "source_type": "aligned_return",
                    "source_date": "A",
                    "source_value": "B",
                },
            }
        ],
        "column_overrides": {},
        "style": {"mode": "infer", "zero_threshold": 0.00005},
        "schedule": [],
        "validation": {
            "minimum_history_dates": 2,
            "tolerance": 0.000001,
            "max_future_days": 0,
            "max_period_change": 0.5,
        },
        "retention": {"backup_count": 3, "preview_count": 3, "log_days": 7},
    }


def create_message(date: str, unit: str, subject: str = "NAV notice") -> bytes:
    message = EmailMessage()
    message["From"] = "Example Sender <sender@example.invalid>"
    message["To"] = "user@example.invalid"
    message["Subject"] = subject
    message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        f"DEMO01 | {date} | {unit} | {unit}"
    )
    return message.as_bytes()


def messages_for(dates: tuple[dt.date, dt.date, dt.date]) -> dict[bytes, bytes]:
    first, second, latest = dates
    return {
        b"1": create_message(first.isoformat(), "1.0000"),
        b"2": create_message(second.isoformat(), "1.0100"),
        b"3": create_message(latest.isoformat(), "1.0200"),
        b"4": create_message(latest.isoformat(), "9.9999", "unrelated notice"),
    }


class FakeIMAP:
    messages: dict[bytes, bytes] = {}

    def __init__(self, *_args: Any, **_kwargs: Any):
        pass

    def login(self, user: str, password: str) -> tuple[str, list[bytes]]:
        require(user == "user@example.invalid", "演练使用了意外的邮箱账号")
        require(password == FIXTURE_VALUE, "演练使用了意外的邮箱密钥")
        return "OK", [b""]

    def select(self, mailbox: str, readonly: bool = False) -> tuple[str, list[bytes]]:
        require(mailbox == "INBOX" and readonly, "演练邮箱未以只读方式打开")
        return "OK", [b""]

    def uid(self, command: str, *args: Any) -> tuple[str, list[Any]]:
        if command == "search":
            query = str(args[-1])
            require("SINCE" in query and "FROM" in query, "演练 IMAP 搜索条件不安全")
            return "OK", [b"1 2 3 4"]
        uid = args[0]
        request = args[1]
        if request == "(UID RFC822.SIZE)":
            return "OK", [
                b"1 (UID "
                + item_uid
                + b" RFC822.SIZE "
                + str(len(self.messages[item_uid])).encode("ascii")
                + b")"
                for item_uid in uid.split(b",")
            ]
        if request == "(UID BODY.PEEK[HEADER.FIELDS (FROM SUBJECT)])":
            parts: list[Any] = []
            for item_uid in uid.split(b","):
                message = BytesParser(policy=default).parsebytes(
                    self.messages[item_uid]
                )
                header = EmailMessage()
                header["From"] = str(message.get("From") or "")
                header["Subject"] = str(message.get("Subject") or "")
                parts.append(
                    (
                        b"1 (UID " + item_uid + b" BODY[HEADER] {1})",
                        header.as_bytes(),
                    )
                )
            return "OK", parts
        payload = self.messages[uid]
        if request == "(BODY.PEEK[])":
            return "OK", [(b"1 (BODY[])", payload)]
        raise RuntimeError(f"意外的演练 IMAP 请求：{request}")

    def close(self) -> tuple[str, list[bytes]]:
        return "OK", [b""]

    def logout(self) -> tuple[str, list[bytes]]:
        return "BYE", [b""]


def install_fake_mailbox(dates: tuple[dt.date, dt.date, dt.date]) -> None:
    os.environ["NAV_EMAIL_PASSWORD"] = FIXTURE_VALUE
    FakeIMAP.messages = messages_for(dates)
    nav_mail.imaplib.IMAP4_SSL = FakeIMAP


def prepare() -> dict[str, Any]:
    dates = date_series()
    workbook = ROOT / "离线演练工作簿.xlsx"
    create_workbook(workbook, dates)
    config = create_config(workbook, dates)
    write_json_atomic(ROOT / "config.json", config)
    install_fake_mailbox(dates)

    readiness = doctor(config)
    require(readiness["preview_ready"], "演练环境未达到预览条件")
    rows, route_report = discover(config)
    require(route_report["passed"], "演练邮件发现失败")
    require(
        route_report["routes"][0]["messages_scanned"] == 3
        and route_report["routes"][0]["messages_filtered_by_subject"] == 0,
        "演练邮件头主题预筛选失败",
    )
    validation = validate(config, rows)
    require(validation["passed"], "演练历史核验失败")
    plan = preview(config, rows)
    require(len(plan["sheets"]) == 1, "演练预览没有生成唯一工作表方案")
    require(
        plan["sheets"][0]["new_dates"] == [dates[-1].isoformat()], "演练补录日期错误"
    )

    state = {
        "dates": [value.isoformat() for value in dates],
        "preview_path": plan["preview_path"],
        "committed": False,
    }
    write_json_atomic(STATE_PATH, state)
    return {
        "passed": True,
        "demo": True,
        "real_data_used": False,
        "formal_workbook_unchanged": True,
        "stages": [
            {"name": "环境检查", "passed": True},
            {"name": "虚构邮件发现与主题过滤", "passed": True},
            {"name": "两个历史日期核验", "passed": True},
            {"name": "生成一条虚构净值预览", "passed": True},
        ],
        "preview_path": plan["preview_path"],
        "commit_available": bool(readiness["commit_ready"]),
        "spreadsheet": next(
            (
                item["detail"]
                for item in readiness["checks"]
                if item["name"] == "spreadsheet-com" and item["passed"]
            ),
            None,
        ),
    }


def commit_demo() -> dict[str, Any]:
    state = json.loads(STATE_PATH.read_text(encoding="utf-8"))
    require(not state.get("committed"), "该演练已经完成过正式写入")
    dates = tuple(dt.date.fromisoformat(value) for value in state["dates"])
    require(len(dates) == 3, "演练日期状态无效")
    config = load_config(ROOT / "config.json")
    result = commit(config)
    require(result["changed"] and result["rows"] == 1, "演练 COM 没有写入一行")

    workbook = openpyxl.load_workbook(config["workbook_path"], data_only=True)
    try:
        sheet = workbook["Demo Fund"]
        require(
            abs(sheet["B4"].value - (1.02 / 1.01 - 1)) < 1e-10,
            "演练产品收益计算错误",
        )
        require(abs(sheet["H4"].value - (-0.002)) < 1e-12, "演练基准收益错误")
        require(
            abs(sheet["I4"].value - ((1.02 / 1.01 - 1) + 0.002)) < 1e-10,
            "演练超额收益计算错误",
        )
    finally:
        workbook.close()

    install_fake_mailbox(dates)
    rows, route_report = discover(config)
    require(route_report["passed"], "演练写入后的邮件发现失败")
    second = preview(config, rows)
    require(
        not second["sheets"] and second["preview_path"] is None, "演练重复运行不幂等"
    )
    require(not (ROOT / "plan.json").exists(), "无新增数据仍留下了提交计划")
    state["committed"] = True
    write_json_atomic(STATE_PATH, state)
    return {
        "passed": True,
        "demo": True,
        "real_data_used": False,
        "application": result["application"],
        "backup_created": True,
        "rows_written": result["rows"],
        "stages": [
            {"name": "Excel/WPS COM 写入虚构副本", "passed": True},
            {"name": "产品、基准和超额收益复核", "passed": True},
            {"name": "第二次运行无重复写入", "passed": True},
            {"name": "无新增数据不保留提交计划", "passed": True},
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=("prepare", "commit"))
    args = parser.parse_args()
    try:
        payload = prepare() if args.action == "prepare" else commit_demo()
    except Exception as exc:
        payload = {
            "passed": False,
            "demo": True,
            "real_data_used": False,
            "error": str(exc),
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 2
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
