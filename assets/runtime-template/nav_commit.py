from __future__ import annotations

import collections
import ctypes
import datetime as dt
import gc
import json
import os
import shutil
import sys
import uuid
from ctypes import wintypes
from pathlib import Path
from typing import Any

import openpyxl

from nav_config import ROOT, STATE_ROOT, active_routes
from nav_workbook import file_sha256, payload_sha256, validate_preview


class CommitError(RuntimeError):
    pass


SPREADSHEET_PROGIDS = (
    ("Excel.Application", "Microsoft Excel"),
    ("ket.Application", "WPS 表格"),
)


def detected_spreadsheet_apps() -> list[str]:
    if sys.platform != "win32":
        return []
    import winreg

    detected: set[str] = set()
    for progid, name in SPREADSHEET_PROGIDS:
        try:
            with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, f"{progid}\\CLSID"):
                detected.add(name)
        except OSError:
            pass
    app_paths = (("excel.exe", "Microsoft Excel"), ("et.exe", "WPS 表格"))
    views = (0, winreg.KEY_WOW64_32KEY, winreg.KEY_WOW64_64KEY)
    for executable, name in app_paths:
        if shutil.which(executable):
            detected.add(name)
        key_path = (
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths" + "\\" + executable
        )
        for root in (winreg.HKEY_CURRENT_USER, winreg.HKEY_LOCAL_MACHINE):
            for view in views:
                try:
                    with winreg.OpenKey(root, key_path, 0, winreg.KEY_READ | view):
                        detected.add(name)
                except OSError:
                    pass
    return sorted(detected)


def _activation_error(exc: Exception) -> str:
    code = getattr(exc, "hresult", None)
    if isinstance(code, int):
        return f"{type(exc).__name__}(0x{code & 0xFFFFFFFF:08X})"
    return type(exc).__name__


def _process_id(app) -> int:
    process_id = ctypes.c_ulong()
    ctypes.windll.user32.GetWindowThreadProcessId(
        int(app.Hwnd), ctypes.byref(process_id)
    )
    if not process_id.value:
        raise CommitError("Could not identify the dedicated spreadsheet process")
    return int(process_id.value)


def ensure_process_exit(process_id: int, timeout_ms: int = 5000) -> None:
    access = 0x00100000 | 0x00001000 | 0x0001
    kernel32 = ctypes.windll.kernel32
    kernel32.OpenProcess.argtypes = (wintypes.DWORD, wintypes.BOOL, wintypes.DWORD)
    kernel32.OpenProcess.restype = wintypes.HANDLE
    kernel32.WaitForSingleObject.argtypes = (wintypes.HANDLE, wintypes.DWORD)
    kernel32.WaitForSingleObject.restype = wintypes.DWORD
    kernel32.TerminateProcess.argtypes = (wintypes.HANDLE, wintypes.UINT)
    kernel32.TerminateProcess.restype = wintypes.BOOL
    kernel32.CloseHandle.argtypes = (wintypes.HANDLE,)
    kernel32.CloseHandle.restype = wintypes.BOOL
    handle = kernel32.OpenProcess(access, False, process_id)
    if not handle:
        return
    try:
        if kernel32.WaitForSingleObject(handle, timeout_ms) == 0x00000102:
            if not kernel32.TerminateProcess(handle, 1):
                raise CommitError("The dedicated spreadsheet process did not exit")
            kernel32.WaitForSingleObject(handle, 2000)
    finally:
        kernel32.CloseHandle(handle)


def spreadsheet_app():
    if sys.platform != "win32":
        raise CommitError("Formal commit requires Windows Excel or WPS Spreadsheet COM")
    from win32com.client import DispatchEx

    errors: list[str] = []
    for progid, _name in SPREADSHEET_PROGIDS:
        app = None
        try:
            app = DispatchEx(progid)
            app.Visible = False
            app.DisplayAlerts = False
            app.EnableEvents = False
            app.AskToUpdateLinks = False
            app.AutomationSecurity = 3
            if bool(app.EnableEvents) or int(app.AutomationSecurity) != 3:
                raise CommitError(
                    "Spreadsheet macro/event safety settings were not accepted"
                )
            return app, progid, _process_id(app)
        except Exception as exc:
            errors.append(f"{progid}={_activation_error(exc)}")
            try:
                if app is not None:
                    app.Quit()
            except Exception:
                pass
    detected = detected_spreadsheet_apps()
    detected_text = "、".join(detected) if detected else "未确认"
    raise CommitError(
        "无法启动 Excel/WPS COM 自动化接口；"
        f"已安装软件探测：{detected_text}；启动结果：{'; '.join(errors)}。"
        "这不等于软件没有安装：请先确认使用原生 Windows Python，"
        "再手动启动一次 Excel/WPS、完成首次设置并关闭，然后重新运行 doctor；"
        "仍失败时再修复或启用 Office/WPS 的 COM 自动化组件"
    )


def _copy_value(source, target) -> None:
    try:
        has_formula = bool(source.HasFormula)
    except Exception:
        has_formula = isinstance(source.Formula, str) and source.Formula.startswith("=")
    if has_formula:
        target.Formula = source.Formula
    else:
        target.Value2 = source.Value2


def _profile(sheet, rows: range, columns: list[int]) -> dict[str, Any]:
    positive: collections.Counter[int] = collections.Counter()
    negative: collections.Counter[int] = collections.Counter()
    zero: collections.Counter[int] = collections.Counter()
    fonts: collections.Counter[tuple[str, float, bool, str]] = collections.Counter()
    for row in rows:
        for column in columns:
            cell = sheet.Cells(row, column)
            value = cell.Value2
            if not isinstance(value, (int, float)):
                continue
            color = int(cell.Font.Color)
            fonts[
                (
                    str(cell.Font.Name),
                    float(cell.Font.Size),
                    bool(cell.Font.Bold),
                    str(cell.NumberFormat),
                )
            ] += 1
            if value > 0:
                positive[color] += 1
            elif value < 0:
                negative[color] += 1
            else:
                zero[color] += 1
    font = fonts.most_common(1)[0][0] if fonts else None
    return {
        "positive": positive.most_common(1)[0][0] if positive else None,
        "negative": negative.most_common(1)[0][0] if negative else None,
        "zero": zero.most_common(1)[0][0] if zero else None,
        "font": font,
    }


def _apply_return_style(
    sheet, rows: set[int], columns: list[int], profile: dict[str, Any], threshold: float
) -> None:
    for row in rows:
        for column in columns:
            cell = sheet.Cells(row, column)
            value = cell.Value2
            if not isinstance(value, (int, float)):
                continue
            if profile["font"]:
                name, size, bold, number_format = profile["font"]
                cell.Font.Name = name
                cell.Font.Size = size
                cell.Font.Bold = bold
                cell.NumberFormat = number_format
            color = profile["zero"]
            if value >= threshold and profile["positive"] is not None:
                color = profile["positive"]
            elif value <= -threshold and profile["negative"] is not None:
                color = profile["negative"]
            if color is not None:
                cell.Font.Color = color


def _apply_plan_with_com(
    temp: Path, preview: Path, plan: dict[str, Any], config: dict[str, Any]
) -> str:
    app, progid, process_id = spreadsheet_app()
    target_book = None
    preview_book = None
    target = None
    source = None
    success = False
    try:
        target_book = app.Workbooks.Open(
            str(temp.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        preview_book = app.Workbooks.Open(
            str(preview.resolve()),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        profiles: dict[str, dict[str, Any]] = {}
        for sheet_plan in plan["sheets"]:
            name = sheet_plan["sheet"]
            target = target_book.Worksheets(name)
            source = preview_book.Worksheets(name)
            insert_before = int(sheet_plan["insert_before"])
            count = int(sheet_plan["insert_count"])
            return_columns = [
                int(value) for value in sheet_plan.get("return_columns") or []
            ]
            if return_columns:
                profiles[name] = _profile(
                    target, range(1, insert_before), return_columns
                )
            if count:
                target.Rows(f"{insert_before}:{insert_before + count - 1}").Insert(
                    -4121
                )
                if bool(sheet_plan.get("copy_template_rows", True)):
                    for row in range(insert_before, insert_before + count):
                        target.Rows(insert_before - 1).Copy(
                            Destination=target.Rows(row)
                        )
            max_column = int(source.UsedRange.Columns.Count)
            for row in sheet_plan.get("format_rows") or []:
                row = int(row)
                source_range = source.Range(
                    source.Cells(row, 1), source.Cells(row, max_column)
                )
                target_range = target.Range(
                    target.Cells(row, 1), target.Cells(row, max_column)
                )
                source_range.Copy()
                target_range.PasteSpecial(Paste=-4122)
                target.Rows(row).RowHeight = source.Rows(row).RowHeight
                app.CutCopyMode = False
            for row in sheet_plan["new_rows"]:
                for column in range(1, max_column + 1):
                    _copy_value(
                        source.Cells(int(row), column), target.Cells(int(row), column)
                    )
            summary_row = int(sheet_plan["summary_row"])
            for column in range(1, max_column + 1):
                _copy_value(
                    source.Cells(summary_row, column), target.Cells(summary_row, column)
                )
            new_rows = {int(row) for row in sheet_plan["new_rows"]}
            for item in sheet_plan.get("changed_cells") or []:
                row, column = int(item["row"]), int(item["column"])
                if row not in new_rows and row != summary_row:
                    _copy_value(source.Cells(row, column), target.Cells(row, column))
        app.CalculateFull()
        threshold = float((config.get("style") or {}).get("zero_threshold", 0.00005))
        if str((config.get("style") or {}).get("mode", "infer")) == "infer":
            for sheet_plan in plan["sheets"]:
                name = sheet_plan["sheet"]
                target = target_book.Worksheets(name)
                styled_rows = {int(row) for row in sheet_plan["new_rows"]}
                styled_rows.add(int(sheet_plan["summary_row"]))
                styled_rows.update(
                    int(item["row"]) for item in sheet_plan.get("changed_cells") or []
                )
                _apply_return_style(
                    target,
                    styled_rows,
                    [int(value) for value in sheet_plan.get("return_columns") or []],
                    profiles.get(
                        name,
                        {
                            "positive": None,
                            "negative": None,
                            "zero": None,
                            "font": None,
                        },
                    ),
                    threshold,
                )
        target_book.Save()
        success = True
        return progid
    finally:
        if preview_book is not None:
            try:
                preview_book.Close(SaveChanges=False)
            except Exception:
                pass
        if target_book is not None:
            try:
                target_book.Close(SaveChanges=success)
            except Exception:
                pass
        try:
            app.Quit()
        except Exception:
            pass
        target = None
        source = None
        target_book = None
        preview_book = None
        app = None
        gc.collect()
        ensure_process_exit(process_id)


def _same_cell(left: Any, right: Any) -> bool:
    if isinstance(left, dt.datetime):
        left = left.date()
    if isinstance(right, dt.datetime):
        right = right.date()
    if isinstance(left, dt.date) and isinstance(right, dt.date):
        return left == right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= 1e-9
    return left == right


def _same_number_format(left: str, right: str) -> bool:
    def normalized(value: str) -> str:
        return str(value).replace("\\-", "-").replace("\\/", "/")

    return normalized(left) == normalized(right)


def _verify_temp(temp: Path, preview: Path, plan: dict[str, Any]) -> None:
    keep_vba = temp.suffix.lower() == ".xlsm"
    target = openpyxl.load_workbook(temp, data_only=False, keep_vba=keep_vba)
    expected = openpyxl.load_workbook(preview, data_only=False, keep_vba=keep_vba)
    calculated = openpyxl.load_workbook(temp, data_only=True, keep_vba=keep_vba)
    try:
        if target.sheetnames != expected.sheetnames:
            raise CommitError("COM result changed workbook sheet topology")
        for sheet_plan in plan["sheets"]:
            name = sheet_plan["sheet"]
            rows = set(int(row) for row in sheet_plan["new_rows"])
            rows.add(int(sheet_plan["summary_row"]))
            rows.update(
                int(item["row"]) for item in sheet_plan.get("changed_cells") or []
            )
            for row in rows:
                for column in range(1, expected[name].max_column + 1):
                    if not _same_cell(
                        target[name].cell(row, column).value,
                        expected[name].cell(row, column).value,
                    ):
                        raise CommitError(
                            f"COM verification failed at {name}!{target[name].cell(row, column).coordinate}"
                        )
                    formula = expected[name].cell(row, column).value
                    value = calculated[name].cell(row, column).value
                    if isinstance(formula, str) and formula.startswith("="):
                        if value is None or (
                            isinstance(value, str) and value.startswith("#")
                        ):
                            raise CommitError(
                                f"COM calculation failed at {name}!{target[name].cell(row, column).coordinate}"
                            )
            for row in (int(value) for value in sheet_plan.get("format_rows") or []):
                for column in range(1, expected[name].max_column + 1):
                    target_cell = target[name].cell(row, column)
                    expected_cell = expected[name].cell(row, column)
                    if (
                        not _same_number_format(
                            target_cell.number_format, expected_cell.number_format
                        )
                        or target_cell.font.bold != expected_cell.font.bold
                    ):
                        raise CommitError(
                            f"COM format verification failed at {name}!{target_cell.coordinate}"
                        )
    finally:
        target.close()
        expected.close()
        calculated.close()


def commit(config: dict[str, Any]) -> dict[str, Any]:
    plan_path = STATE_ROOT / "plan.json"
    try:
        plan = json.loads(plan_path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError) as exc:
        raise CommitError("No valid plan.json. Run preview again.") from exc
    if plan.get("schema_version") != 1 or plan.get("config_sha256") != payload_sha256(
        config
    ):
        raise CommitError(
            "The runtime configuration changed after preview; regenerate the preview"
        )
    try:
        uuid.UUID(str(plan.get("plan_id") or ""))
        created = dt.datetime.fromisoformat(str(plan["created"]))
    except (ValueError, KeyError) as exc:
        raise CommitError("The preview plan metadata is invalid") from exc
    if dt.datetime.now() - created > dt.timedelta(
        hours=24
    ) or created > dt.datetime.now() + dt.timedelta(minutes=5):
        raise CommitError(
            "The preview plan is stale or future-dated; regenerate the preview"
        )
    master = Path(str(plan.get("master_path") or "")).resolve()
    configured_master = Path(config["workbook_path"]).resolve()
    if master != configured_master:
        raise CommitError("The plan does not belong to the configured master workbook")
    if file_sha256(master) != plan["master_sha256"]:
        raise CommitError(
            "The master workbook changed after preview; regenerate the preview"
        )
    if not plan.get("sheets"):
        return {"changed": False, "message": "No rows require commit"}
    preview = Path(str(plan.get("preview_path") or "")).resolve()
    if (
        preview.parent != (ROOT / "previews").resolve()
        or preview.suffix.lower() != master.suffix.lower()
    ):
        raise CommitError(
            "The plan preview path is outside the runtime preview directory"
        )
    if not preview.is_file() or file_sha256(preview) != plan.get("preview_sha256"):
        raise CommitError(
            "The reviewed preview is missing or changed; regenerate the preview"
        )
    allowed_sheets = {str(route["sheet"]) for route in active_routes(config)}
    allowed_modes = {
        str(route["sheet"]): str(route.get("sheet_mode", "summary"))
        for route in active_routes(config)
    }
    planned_sheets = {str(item.get("sheet") or "") for item in plan["sheets"]}
    if not planned_sheets or not planned_sheets <= allowed_sheets:
        raise CommitError(
            "The plan contains a sheet not authorized by the current configuration"
        )
    for item in plan["sheets"]:
        sheet_name = str(item.get("sheet") or "")
        if str(item.get("sheet_mode", "summary")) != allowed_modes.get(sheet_name):
            raise CommitError("The plan sheet mode does not match the configuration")
        try:
            header_row = int(item["header_row"])
            insert_before = int(item["insert_before"])
            insert_count = int(item["insert_count"])
            populated_count = int(item.get("populated_count", insert_count))
            summary_row = int(item["summary_row"])
            new_rows = [int(row) for row in item["new_rows"]]
            filled_existing_rows = [
                int(row) for row in item.get("filled_existing_rows") or []
            ]
            format_rows = [int(row) for row in item.get("format_rows") or []]
            changed_rows = {
                int(cell["row"]) for cell in item.get("changed_cells") or []
            }
        except (KeyError, TypeError, ValueError) as exc:
            raise CommitError("The plan row structure is invalid") from exc
        if (
            header_row < 1
            or insert_before <= header_row
            or insert_count < 1
            or new_rows != list(range(insert_before, insert_before + insert_count))
        ):
            raise CommitError("The plan insertion range is invalid")
        if (
            summary_row != insert_before + insert_count
            or populated_count != insert_count + len(filled_existing_rows)
            or len(item.get("new_dates") or []) != populated_count
        ):
            raise CommitError("The plan summary row or new-date count is invalid")
        if filled_existing_rows and (
            str(item.get("sheet_mode", "summary")) != "summary"
            or filled_existing_rows != [insert_before - 1]
            or filled_existing_rows[0] <= header_row
        ):
            raise CommitError("The plan reserved-row range is invalid")
        if not set(filled_existing_rows) <= changed_rows:
            raise CommitError("The plan does not populate its reserved row")
        if not set(format_rows) <= {*new_rows, header_row}:
            raise CommitError("The plan format range is invalid")
        if not isinstance(item.get("copy_template_rows", True), bool):
            raise CommitError("The plan template-copy flag is invalid")
    validate_preview(config, plan)
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_dir = ROOT / "backups"
    backup_dir.mkdir(exist_ok=True)
    keep = int((config.get("retention") or {}).get("backup_count", 10))
    backup = backup_dir / f"{master.stem}-before-{timestamp}{master.suffix}"
    temp = master.with_name(f".{master.stem}.nav-write-{os.getpid()}{master.suffix}")
    if temp.exists():
        raise CommitError(f"Temporary target already exists: {temp.name}")
    before = file_sha256(master)
    try:
        shutil.copy2(master, backup)
        shutil.copy2(master, temp)
        application = _apply_plan_with_com(temp, preview, plan, config)
        _verify_temp(temp, preview, plan)
        if file_sha256(preview) != plan["preview_sha256"]:
            raise CommitError("The reviewed preview changed during commit")
        if file_sha256(master) != before:
            raise CommitError("The master changed during commit; refusing replacement")
        try:
            os.replace(temp, master)
        except OSError as exc:
            if isinstance(exc, PermissionError) or getattr(exc, "winerror", None) in {
                5,
                32,
                33,
            }:
                raise CommitError(
                    "无法替换正式工作簿；请关闭正在打开该文件的 Excel/WPS 窗口后重试"
                ) from exc
            raise
    except Exception:
        if temp.exists():
            try:
                temp.unlink()
            except OSError:
                pass
        if backup.exists():
            try:
                backup.unlink()
            except OSError:
                pass
        if not master.is_file() or file_sha256(master) != before:
            raise CommitError("Commit failed and the master hash changed unexpectedly")
        raise
    old_backups = sorted(
        backup_dir.glob(f"{master.stem}-before-*{master.suffix}"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    cleanup_failed = False
    for old in old_backups[keep:]:
        try:
            old.unlink()
        except OSError:
            cleanup_failed = True
    return {
        "changed": True,
        "application": application,
        "backup": str(backup),
        "master_sha256": file_sha256(master),
        "sheets": len(plan["sheets"]),
        "rows": sum(
            int(sheet.get("populated_count", sheet["insert_count"]))
            for sheet in plan["sheets"]
        ),
        "backup_cleanup_failed": cleanup_failed,
    }
