# -*- coding: utf-8 -*-
"""生成一个【脱敏样例】净值表 示例净值表.xlsx，结构与 registry.example.json 对齐，
用于离线查看表布局 / 试跑写入逻辑。纯造数据，无任何真实信息。

用法：
    python examples/make_sample_workbook.py            # 生成到 src/ 的上一级目录
    python examples/make_sample_workbook.py 输出路径.xlsx
"""
import os, sys, datetime
import openpyxl
from openpyxl.styles import Font

EPOCH = datetime.date(1899, 12, 30)
HEADER = ["产品代码", "产品名称", "单位净值", "累计单位净值", "净值日期", "收益（周度）"]

# (sheet, 代码, 名称, 是否带指数列)
SHEETS = [
    ("基金01", "DEMO04", "示例指数增强1号", False),
    ("基金11", "DEMO05", "示例中性500指增", True),
    ("基金16", "DEMO06", "示例日频CTA", False),
    ("基金20", "DEMO07", "示例附件净值产品", False),
    ("基金27", "DEMOA", "示例PDF周报产品A类", False),
]


def build(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, code, fund, has_idx in SHEETS:
        ws = wb.create_sheet(name)
        ws.cell(1, 1).value = f"{fund}  净值跟踪（脱敏样例）"
        cols = HEADER + (["中证500", "指数收益(周度)"] if has_idx else [])
        for c, h in enumerate(cols, 1):
            cell = ws.cell(2, c); cell.value = h; cell.font = Font(name="宋体", size=10, bold=True)
        nav = 1.0000
        d = datetime.date(2026, 5, 15)
        for i in range(4):
            r = 3 + i
            ws.cell(r, 1).value = code
            ws.cell(r, 2).value = fund
            ws.cell(r, 3).value = round(nav, 4)
            ws.cell(r, 4).value = round(nav, 4)
            ws.cell(r, 5).value = (d - EPOCH).days
            if i > 0:
                ws.cell(r, 6).value = f"=C{r}/C{r-1}-1"
            if has_idx:
                ws.cell(r, 7).value = 6000 + i * 30
                if i > 0:
                    ws.cell(r, 8).value = f"=G{r}/G{r-1}-1"
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).font = Font(name="等线", size=10)
            nav *= 1.012
            d += datetime.timedelta(days=7)
        sr = 3 + 4
        ws.cell(sr, 1).value = "累计"
        ws.cell(sr, 6).value = f"=C{sr-1}/C3-1"
        if has_idx:
            ws.cell(sr, 8).value = f"=G{sr-1}/G3-1"
        for c in range(1, len(cols) + 1):
            ws.cell(sr, c).font = Font(name="宋体", size=10, bold=True)
    wb.save(path)
    print("已生成样例表 ->", path)


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    default = os.path.join(here, "..", "示例净值表.xlsx")
    build(sys.argv[1] if len(sys.argv) > 1 else os.path.abspath(default))
