# -*- coding: utf-8 -*-
"""ONE-TIME 补丁脚本（跑一次即可，跑完可删）：
  1) 改表名：基金21→基金21(友)；基金28/基金08/基金09/基金18/基金12(友）追加(赎)。
  2) 基金27：只清排版(补累计净值列/周收益公式、清累计行下残留空格)，净值不动(由write.py写)。
  3) 累计行收益率配色：整表统一盈红亏绿（只动累计行）。
  4) 基金26：2000指数列右边加 1000指数列(引用基金10)。
  5) 指数列空缺回填（基金15/基金26 最新行 引用源表）。
  6) 字体统一：以基金15为准，表头宋体/10粗、数据等线/10、行高统一；保留所有字体颜色。

这是【一次性】把现有净值表升级到新结构用的脚本，跑成功一次后即可删除本文件。
用法（公司 nav_tool 目录，建议先 pushd 映射网络盘）：
    python init_master_once.py            # 预览：写到 *_补丁预览.xlsx，正式表不动
    python init_master_once.py --commit   # 备份正式表后原地修改
    python init_master_once.py <路径> [--commit]   # 指定文件（测试用）
"""
import os, sys, re, shutil, datetime
from copy import copy
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula
import navlib as L

RED = "FFFF0000"
GREEN = "FF008000"
EPOCH = datetime.date(1899, 12, 30)

# 表名变更（旧名必须与 Excel 标签完全一致，含全角括号）
RENAME = {
    "基金21": "基金21(友)",
    "基金28": "基金28(赎)", "基金08": "基金08(赎)", "基金09": "基金09(赎)",
    "基金18": "基金18(赎)", "基金12(友）": "基金12(友）(赎)",
}
HEADER_FONT, DATA_FONT, FONT_SIZE, ROW_H = "宋体", "等线", 10.0, 14.25


def rename_tabs(wb):
    done = []
    for old, new in RENAME.items():
        if old in wb.sheetnames and new not in wb.sheetnames:
            wb[old].title = new
            done.append(f"{old} → {new}")
        elif new in wb.sheetnames:
            done.append(f"{new}（已是新名,跳过）")
        else:
            done.append(f"!! 找不到表 {old}")
    return done


def fix_yiyuan(wb):
    """基金27【只清排版, 不碰净值】：净值由 write.py 按邮件/PDF 真实值写入；本函数仅
    (a) 补累计净值列(=单位净值, 该产品无分红); (b) 清累计行以下的残留(如 B10 空格);
    (c) 数据行间补 F 周收益公式。绝不写死任何净值数字。"""
    if "基金27" not in wb.sheetnames:
        return "无基金27表"
    ws = wb["基金27"]
    accr = next((r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value).strip() == "累计"), None)
    if not accr:
        return "!! 基金27找不到累计行, 跳过"
    first = 4
    # 清掉代码/名称单元格首尾的换行/空白(Excel 里显示成乱码方块)
    for r in range(first, accr):
        for c in (1, 2):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v != v.strip():
                ws.cell(r, c).value = v.strip()
    notes, prev_data_row = [], None
    for r in range(first, accr):
        c_unit = ws.cell(r, 3).value
        if not isinstance(c_unit, (int, float)):
            continue
        # 累计净值列(D)补齐 = 单位净值(C)
        if ws.cell(r, 4).value in (None, ""):
            ws.cell(r, 4).value = ws.cell(r, 3).value
            notes.append(f"补 D{r}")
        # 周收益(F)：与上一数据行比, 缺则补公式
        if prev_data_row is not None and ws.cell(r, 6).value in (None, ""):
            ws.cell(r, 6).value = f"=C{r}/C{prev_data_row}-1"
            notes.append(f"补 F{r}=C{r}/C{prev_data_row}-1")
        prev_data_row = r
    # 清掉累计行以下的残留(如 B10 的空格)，否则 write 会判"累计行后有内容"而跳过
    cleared = 0
    for r in range(accr + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value not in (None, ""):
                ws.cell(r, c).value = None
                cleared += 1
    return f"累计行r{accr}; {', '.join(notes) or '无需补列'}; 清累计行下残留 {cleared} 格 (未改动任何净值)"


def normalize_fonts(wb):
    """字体/行高统一(参考基金15)：表头行宋体/10粗，其余等线/10非粗，行高统一；保留所有颜色。"""
    n = 0
    for sh in wb.sheetnames:
        ws = wb[sh]
        hrow = next((r for r in range(1, min(ws.max_row, 8) + 1)
                     if str(ws.cell(r, 1).value).strip() == "产品代码"), None)
        try:
            ws.sheet_format.defaultRowHeight = ROW_H
        except Exception:
            pass
        for r in list(ws.row_dimensions):       # 清掉参差的显式行高 -> 用默认
            if ws.row_dimensions[r].height not in (None, ROW_H):
                ws.row_dimensions[r].height = None
        for r in range(1, ws.max_row + 1):
            is_h = (r == hrow)
            name, bold = (HEADER_FONT, True) if is_h else (DATA_FONT, False)
            for c in range(1, ws.max_column + 1):
                f = ws.cell(r, c).font
                if f.name == name and f.size == FONT_SIZE and bool(f.bold) == bold:
                    continue
                ws.cell(r, c).font = Font(name=name, size=FONT_SIZE, bold=bold,
                                          italic=f.italic, vertAlign=f.vertAlign,
                                          underline=f.underline, strike=f.strike, color=f.color)
                n += 1
    return n

def font_kind(cell):
    """'red' / 'green' / None  —— 判断该格字体大类（按红/绿通道）。"""
    f = cell.font
    if not f or not f.color or f.color.type != "rgb":
        return None
    rgb = f.color.rgb
    if not rgb or rgb in ("00000000", "FF000000"):
        return None
    try:
        r = int(rgb[2:4], 16); g = int(rgb[4:6], 16)
    except (ValueError, TypeError):
        return None
    if r >= 0x80 and g < 0x80:
        return "red"
    if g >= 0x80 and r < 0x80:
        return "green"
    return None

def set_color(cell, rgb):
    f = cell.font
    cell.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                     vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                     color=rgb)

def cellnum(ws, ref):
    m = re.match(r"\$?([A-Z]+)\$?(\d+)$", ref)
    if not m:
        return None
    v = ws.cell(int(m.group(2)), column_index_from_string(m.group(1))).value
    return v if isinstance(v, (int, float)) else None

def eval_accum(ws, val):
    """计算累计收益格的值（不依赖 Excel 缓存）。支持 A/B-1、PRODUCT(1+col s:e)-1、F-H。"""
    text = val.text if isinstance(val, ArrayFormula) else str(val)
    t = text[1:] if text.startswith("=") else text
    t = t.replace(" ", "")
    m = re.fullmatch(r"([A-Z]+\d+)/([A-Z]+\d+)-1", t)
    if m:
        a, b = cellnum(ws, m.group(1)), cellnum(ws, m.group(2))
        return (a / b - 1) if (a is not None and b not in (None, 0)) else None
    m = re.fullmatch(r"PRODUCT\(1\+([A-Z]+)(\d+):([A-Z]+)(\d+)\)-1", t)
    if m:
        col = column_index_from_string(m.group(1))
        prod, any_ = 1.0, False
        for r in range(int(m.group(2)), int(m.group(4)) + 1):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                prod *= (1 + v); any_ = True
        return (prod - 1) if any_ else None
    m = re.fullmatch(r"([A-Z]+\d+)-([A-Z]+\d+)", t)   # 超额 F-H：两边再各自求值
    if m:
        def side(ref):
            mm = re.match(r"([A-Z]+)(\d+)", ref)
            c = ws.cell(int(mm.group(2)), column_index_from_string(mm.group(1)))
            return eval_accum(ws, c.value) if isinstance(c.value, str) and c.value.startswith("=") else cellnum(ws, ref)
        a, b = side(m.group(1)), side(m.group(2))
        return (a - b) if (a is not None and b is not None) else None
    return None

def fix_colors(wb):
    """整表统一：累计行所有可计算的收益率格，盈利红、亏损绿。无法计算(跨表引用)的跳过。
    只动累计行；其它行/其它格一律不碰。"""
    fixed = []
    for sh in wb.sheetnames:
        ws = wb[sh]
        accr = next((r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "累计"), None)
        if not accr:
            continue
        for c in range(6, ws.max_column + 1):
            cell = ws.cell(accr, c)
            if cell.value in (None, ""):
                continue
            val = eval_accum(ws, cell.value)
            if val is None:               # 跨表引用等无法计算 -> 保持原样
                continue
            cur = font_kind(cell)
            want = "green" if val < 0 else "red"
            if cur != want:
                set_color(cell, GREEN if val < 0 else RED)
                fixed.append(f"{sh}!{get_column_letter(c)}{accr} ({val:+.4f}) {cur or '黑'}→{want}")
    return fixed

def add_1000_to_aifanzhe(wb):
    if "基金26" not in wb.sheetnames:
        return "未找到基金26表"
    if "基金10" not in wb.sheetnames:
        return "未找到基金10表（1000 来源），跳过"
    ws = wb["基金26"]; zh = wb["基金10"]
    if ws["H3"].value not in (None, ""):
        return f"H 列非空(H3={ws['H3'].value!r})，疑似已加过，跳过"
    # 基金10 日期(E)->行 映射
    z_date2row = {}
    for r in range(3, zh.max_row + 1):
        e = zh.cell(r, 5).value
        if isinstance(e, (int, float)):
            z_date2row[int(e)] = r
    accr = next((r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "累计"), None)
    first = 4
    # 复制 G 列样式到 H，H3 写表头，数据行按日期引用基金10!H，累计行写数组公式
    for r in range(3, accr + 1):
        ws.cell(r, 8)._style = copy(ws.cell(r, 7)._style)
    ws.cell(3, 8).value = "1000指数"
    refs = []
    for r in range(first, accr):
        e = ws.cell(r, 5).value
        gref = ws.cell(r, 7).value      # 2000 列该行是否已填(=基金03!H..)
        if isinstance(e, (int, float)) and int(e) in z_date2row and gref not in (None, ""):
            ws.cell(r, 8).value = f"=基金10!H{z_date2row[int(e)]}"
            refs.append(f"H{r}=基金10!H{z_date2row[int(e)]}")
        else:
            ws.cell(r, 8).value = None  # 与 2000 列留空行保持一致
    ws.cell(accr, 8).value = ArrayFormula(ref=f"H{accr}", text=f"=PRODUCT(1+H{first}:H{accr-1})-1")
    return "OK: 表头H3=1000指数; " + ", ".join(refs) + f"; 累计H{accr}=PRODUCT(1+H{first}:H{accr-1})-1"

def main():
    args = [a for a in sys.argv[1:] if a != "--commit"]
    commit = "--commit" in sys.argv
    path = args[0] if args else L.CFG["master_path"]
    wb = openpyxl.load_workbook(path, data_only=False)

    ren = rename_tabs(wb)                 # 1. 先改表名(后续按新名)
    yiyuan = fix_yiyuan(wb)               # 2. 基金27 6-12
    fixed = fix_colors(wb)                # 3. 累计配色
    aiz = add_1000_to_aifanzhe(wb)        # 4. 基金26 1000 列
    from write import fill_crossref_index
    idx_filled, idx_blank = fill_crossref_index(wb, L.load_registry())  # 5. 指数回填
    nfont = normalize_fonts(wb)           # 6. 字体统一(最后,保留颜色)

    print("=== 1. 表名变更 ===")
    for f in ren: print("  " + f)
    print("=== 2. 基金27 ===\n  " + yiyuan)
    print("=== 3. 累计行配色(整表统一 盈红亏绿) ===")
    for f in fixed: print("  " + f)
    print(f"  共上色 {len(fixed)} 格")
    print("=== 4. 基金26 1000 列 ===\n  " + aiz)
    print("=== 5. 指数列空缺回填 ===")
    for f in idx_filled: print("  " + f)
    print(f"  共回填 {len(idx_filled)} 格" + (f"；仍空 {len(idx_blank)}(源表无该日期,如成立首行)" if idx_blank else ""))
    print(f"=== 6. 字体/行高统一: 调整 {nfont} 个单元格字体 ===")

    if commit:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = os.path.splitext(os.path.basename(path))[0]
        bakdir = os.path.join(os.path.dirname(os.path.abspath(path)), "backups")
        try:
            bakdir2 = os.path.join(L.HERE, "backups"); os.makedirs(bakdir2, exist_ok=True); bakdir = bakdir2
        except Exception:
            os.makedirs(bakdir, exist_ok=True)
        bak = os.path.join(bakdir, f"{stem}_补丁前_{ts}.xlsx")
        shutil.copy2(path, bak)
        wb.save(path)
        print(f"\n已备份 -> {bak}\n已写入 -> {path}")
    else:
        d = os.path.dirname(os.path.abspath(path))
        stem = os.path.splitext(os.path.basename(path))[0]
        out = os.path.join(d, stem + "_补丁预览.xlsx")
        wb.save(out)
        print(f"\n[预览] 已写 -> {out}（正式表未动；确认后加 --commit）")

if __name__ == "__main__":
    main()
