# -*- coding: utf-8 -*-
"""Phase-2 products that deliver NAV via .xls/.xlsx attachments (or labeled body
tables). Parses the attachment into rows {code,date,unit,cum} and routes to the
correct sheet. Routing is RESTRICTED to phase-2 sheets so these senders cannot
override phase-1 products (custodian A also mails 基金01/基金08 估算净值).

Sender -> fmt (示例，真实地址请填进 config.json 的 senders):
  custodian_a   nav@custodian-a.example.com   (净值表 / 虚拟计提净值表)
  custodian_b   nav@custodian-b.example.com
  custodian_c   nav@custodian-c.example.com   (labeled layout)
  custodian_d   nav@custodian-d.example.com   (虚拟业绩报酬, code=TA891A)
  custodian_e   nav@custodian-e.example.com   (TA虚拟净值)
"""
import io, re, datetime
import navlib as L

FMTS = {"cms", "dwzq", "csc", "htsc_leap", "gtht_ta", "yiyuan"}
CODE_ALIAS = {"TA891A": "STA891"}   # 托管代码 -> 协会备案代码(表里用的)

# header synonyms (matched by substring)
H_DATE = ["净值日期", "业务日期", "估值日期", "日期"]
H_CODE = ["产品代码", "基金代码", "资产代码", "备案编码"]
H_UNIT = ["资产份额净值", "基金份额净值", "单位净值"]
H_CUM  = ["资产份额累计净值", "基金份额累计净值", "累计单位净值", "累计净值"]

def _norm_code(c):
    c = L.base_code(str(c)) if c else c
    return CODE_ALIAS.get(c, c)

def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and v > 40000:  # excel serial
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    return L.parse_date(str(v))

def _match(cell, syns):
    s = str(cell or "")
    for i, kw in enumerate(syns):
        if kw in s:
            return i
    return None

def _rows_from_attachment(msg):
    """Return (rows, source_name). rows = list[list[cell]] of the first sheet of
    the first readable .xls/.xlsx attachment."""
    for p in msg.walk():
        fn = p.get_filename()
        if not fn:
            continue
        fn = L.dh(fn); low = fn.lower()
        payload = p.get_payload(decode=True)
        if not payload:
            continue
        try:
            if low.endswith(".xlsx"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
                ws = wb[wb.sheetnames[0]]
                return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                        for r in range(1, ws.max_row + 1)], fn
            elif low.endswith(".xls"):
                import xlrd
                book = xlrd.open_workbook(file_contents=payload)
                sh = book.sheet_by_index(0)
                return [sh.row_values(r) for r in range(sh.nrows)], fn
        except Exception:
            continue
    return None, None

def parse_pdf_nav(msg):
    """基金27周报 PDF：抽取 单位净值 + 净值日期(+代码)。返回 [{code,date,unit,cum}] 或 []。
    无分红产品(单位净值-1=成立至今收益率) -> 累计净值=单位净值。"""
    for p in msg.walk():
        fn = p.get_filename()
        if not fn or not L.dh(fn).lower().endswith(".pdf"):
            continue
        payload = p.get_payload(decode=True)
        if not payload:
            continue
        try:
            from pypdf import PdfReader
            txt = "\n".join(pg.extract_text() or "" for pg in PdfReader(io.BytesIO(payload)).pages)
        except Exception:
            return []
        # 持仓情况行："... 最新净值 更新日期 ..." -> 净值 紧邻 日期(YYYY-MM-DD)
        m = re.search(r"(\d+\.\d+)\s+(\d{4}-\d{2}-\d{2})", txt)
        unit = d = None
        if m:
            unit = L.num(m.group(1)); d = L.parse_date(m.group(2))
        # 兜底/交叉校验：基金表现表的"单位净值"首个数字
        m2 = re.search(r"单位净值[\s\S]{0,40}?(\d+\.\d+)\s+-?\d", txt)
        if m2 and unit is None:
            unit = L.num(m2.group(1))
        if d is None:                      # 日期再兜底：任意 YYYY 年 MM 月 DD 日 / YYYY-MM-DD
            d = L.parse_date(txt)
        mc = re.search(r"基金编号\s*([A-Z0-9]{5,8})", txt)
        code = _norm_code(mc.group(1)) if mc else None
        if unit is None or d is None:
            return []
        return [dict(code=code, date=d, unit=unit, cum=unit)]
    return []

def parse_rows(msg):
    """Parse a phase-2 email's attachment into [{code,date,unit,cum}].
    Handles tabular / labeled .xls(x) layouts, and 基金27 PDF 周报."""
    pdf = parse_pdf_nav(msg)
    if pdf:
        return pdf
    rows, _ = _rows_from_attachment(msg)
    if not rows:
        return []
    # --- find a tabular header row: has a unit synonym AND a date synonym ---
    header_idx = None
    for i, row in enumerate(rows[:8]):
        cells = [str(c or "") for c in row]
        has_unit = any(_match(c, H_UNIT) is not None for c in cells)
        has_date = any(_match(c, H_DATE) is not None for c in cells)
        if has_unit and has_date:
            header_idx = i
            break
    out = []
    if header_idx is not None:
        hdr = [str(c or "") for c in rows[header_idx]]
        def find_col(syns):
            best = (None, 99)
            for ci, h in enumerate(hdr):
                m = _match(h, syns)
                if m is not None and m < best[1]:
                    best = (ci, m)
            return best[0]
        c_date, c_code = find_col(H_DATE), find_col(H_CODE)
        c_unit, c_cum = find_col(H_UNIT), find_col(H_CUM)
        for row in rows[header_idx + 1:]:
            if c_unit is None or c_unit >= len(row):
                continue
            unit = L.num(row[c_unit])
            d = _to_date(row[c_date]) if (c_date is not None and c_date < len(row)) else None
            if unit is None or d is None:
                continue
            code = _norm_code(row[c_code]) if (c_code is not None and c_code < len(row)) else None
            cum = L.num(row[c_cum]) if (c_cum is not None and c_cum < len(row)) else None
            out.append(dict(code=code, date=d, unit=unit, cum=cum))
        return out
    # --- labeled layout (custodian C): key in col A, value in col B ---
    kv = {}
    for row in rows:
        if len(row) >= 2 and row[0]:
            kv[str(row[0]).replace("：", "").replace(":", "").strip()] = row[1]
    if kv:
        def get(syns):
            for k, v in kv.items():
                if any(s.replace("：", "") in k for s in syns):
                    return v
            return None
        unit = L.num(get(H_UNIT)); d = _to_date(get(H_DATE))
        if unit is not None and d is not None:
            out.append(dict(code=_norm_code(get(H_CODE)), date=d, unit=unit, cum=L.num(get(H_CUM))))
    return out

# ----------------------------------------------------------------- routing
def build_routers():
    """code -> sheet and (name_key, sheet) list, restricted to phase-2 sheets."""
    reg = L.load_registry()
    p2 = L.CFG.get("scope_sheets_p2", [])
    code2sheet, name_keys = {}, []
    for sheet in p2:
        for c in reg[sheet]["codes"]:
            code2sheet[_norm_code(c)] = sheet
        # routing keyword = the sheet name (short, distinctive)
        name_keys.append((sheet.strip(), sheet))
    # longest name first so e.g. avoid partial overlaps
    name_keys.sort(key=lambda x: -len(x[0]))
    return code2sheet, name_keys

def route(subject, code2sheet, name_keys):
    """Return (sheet, code, date) for a phase-2 email, or (None,None,date)."""
    subj = subject or ""
    # 1) any code token in subject that maps to a phase-2 sheet
    for tok in re.findall(r"[A-Z0-9]{5,8}", subj):
        s = code2sheet.get(_norm_code(tok))
        if s:
            return s, _norm_code(tok), _subj_date(subj)
    # 1b) If the subject carries an explicit product code (5-8 chars, letters+digits)
    #     that maps to none of our products, this is a DIFFERENT product from the same
    #     manager (e.g. 基金21 also mails 示例策略甲-DEMO01 / 示例策略乙-DEMO02). Do NOT fall
    #     back to manager-name routing, or we'd write the wrong product's NAV into the
    #     sheet. (The真身 e.g. 中性一号-DEMO03 已在第1步命中。)
    codeish = [t for t in re.findall(r"[A-Z0-9]{5,8}", subj)
               if re.search(r"[A-Z]", t) and re.search(r"\d", t)]
    if codeish and not any(code2sheet.get(_norm_code(t)) for t in codeish):
        return None, None, _subj_date(subj)
    # 2) product name keyword (only when no explicit foreign product code present)
    for key, sheet in name_keys:
        if key and key in subj:
            # find that sheet's code
            code = next((c for c, sh in code2sheet.items() if sh == sheet), None)
            return sheet, code, _subj_date(subj)
    return None, None, _subj_date(subj)

def _subj_date(subj):
    """Latest date appearing in the subject (handles ranges like 20210422_20260611)."""
    dates = []
    for m in re.finditer(r"(\d{4})\D?(\d{2})\D?(\d{2})", subj):
        try:
            dates.append(datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
        except ValueError:
            pass
    return max(dates) if dates else None
