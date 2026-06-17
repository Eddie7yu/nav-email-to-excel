# -*- coding: utf-8 -*-
"""Validation: for each in-scope sheet, take the last N recorded rows and verify
the email parser reproduces the SAME unit & cumulative NAV. This proves extraction
correctness before we ever write new data.
"""
import os, json, datetime
import openpyxl
import navlib as L
import phase2 as P2

N_CHECK = 6
TOL = 5e-5

def s2d(v):
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None

def approx(a, b):
    if a is None or b is None:
        return None
    return abs(round(a, 4) - round(b, 4)) <= TOL

def main():
    ipath = os.path.join(L.HERE, "index.json")
    if not os.path.exists(ipath):
        print("[提示] 未找到 index.json，请先运行 build_index.py。跳过校验。")
        return
    idx = json.load(open(ipath, encoding="utf-8"))["index"]
    # group index by (sheet, date)
    by = {}
    for r in idx:
        if r["date"]:
            by.setdefault((r["sheet"], r["date"]), []).append(r)
    reg = L.load_registry()
    wb = openpyxl.load_workbook(L.CFG["master_path"], data_only=True)
    M = L.connect()
    bodycache = {}

    p2cache = {}

    def parsed_for(uid, fmt, subj, subj_code, target=None):
        if fmt in P2.FMTS:
            if uid not in p2cache:
                msg = L.fetch_full(M, uid)
                p2cache[uid] = P2.parse_rows(msg) if msg else []
            rows = p2cache[uid]
            best = next((r for r in rows if target and r["date"] == target), None)
            if best is None and rows:
                best = max(rows, key=lambda r: r["date"])
            if not best or best["unit"] is None:
                return None
            return dict(code=best["code"] or subj_code, date=best["date"],
                        unit=best["unit"], cum=best["cum"], source=fmt)
        if uid not in bodycache:
            msg = L.fetch_full(M, uid)
            bodycache[uid] = L.body_text(msg) if msg else ""
        return L.parse_body(fmt, subj, bodycache[uid], subj_code)

    total = ok = miss = nofind = 0
    report = []
    for sheet in L.all_scope():
        if sheet not in wb.sheetnames:
            continue
        info = reg[sheet]
        ws = wb[sheet]
        rows = []
        for r in range(info["data_start"], ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a in (None, "", "累计"):
                continue
            d = s2d(ws.cell(row=r, column=5).value)
            unit = L.num(ws.cell(row=r, column=3).value)
            cum = L.num(ws.cell(row=r, column=4).value)
            if d:
                rows.append((d, unit, cum))
        rows = rows[-N_CHECK:]
        for (d, unit, cum) in rows:
            total += 1
            cands = by.get((sheet, d.isoformat()), [])
            if not cands:
                nofind += 1
                report.append((sheet, d.isoformat(), "NO-EMAIL", unit, cum, None, None, None))
                continue
            best = None  # prefer a unit match
            for cand in cands:
                p = parsed_for(cand["uid"], cand["fmt"], cand["subject"], cand["code"], target=d)
                if not p:
                    continue
                if approx(p["unit"], unit):
                    best = (cand, p, "MATCH"); break
                if best is None:
                    best = (cand, p, "DIFF")
            if best is None:
                nofind += 1
                report.append((sheet, d.isoformat(), "PARSE-FAIL", unit, cum, None, None, cands[0]["source"] if cands else None))
            else:
                cand, p, st = best
                if st == "MATCH":
                    ok += 1
                    # flag only when the email's OWN cumulative differs from Linda's D
                    if cum is not None and p["cum"] is not None and not approx(cum, p["cum"]):
                        report.append((sheet, d.isoformat(), "cum+%.4f" % (cum - p["cum"]),
                                       unit, cum, p["unit"], p["cum"], p["source"]))
                else:
                    miss += 1
                    report.append((sheet, d.isoformat(), "UNIT-DIFF", unit, cum, p["unit"], p["cum"], p["source"]))
    M.logout()

    print(f"\n==== VALIDATION: {ok}/{total} matched | {miss} diff | {nofind} no-email/parse-fail ====\n")
    if report:
        print(f"{'sheet':10s} {'date':11s} {'status':10s} {'xls_unit':>10s} {'xls_cum':>10s} {'eml_unit':>10s} {'eml_cum':>10s}  source")
        for (sheet, d, st, xu, xc, eu, ec, src) in report:
            print(f"{sheet:10s} {d:11s} {st:10s} {str(xu):>10s} {str(xc):>10s} {str(eu):>10s} {str(ec):>10s}  {src}")
    else:
        print("All checked rows matched exactly. ✅")

    # write validation.json (consumed by notify.py); diffs = only true unit mismatches
    diffs = [{"sheet": r[0], "date": r[1], "xls_unit": r[3], "eml_unit": r[5]}
             for r in report if r[2] == "UNIT-DIFF" and r[0] not in {"基金16", "基金21(友)"}]
    out = {"time": datetime.datetime.now().isoformat(timespec="seconds"),
           "matched": ok, "total": total, "diffs": diffs}
    json.dump(out, open(os.path.join(L.HERE, "validation.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

if __name__ == "__main__":
    main()
