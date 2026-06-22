# -*- coding: utf-8 -*-
"""(Re)generate registry.json: per-sheet structure (codes, header row, data range,
last NAV date, and whether the weekly-return formula uses C or D). Run this if the
master's sheet layout changes (e.g. a product added/removed)."""
import os, json, datetime, re
import openpyxl
import navlib as L

EPOCH = datetime.date(1899, 12, 30)

def serial_to_date(v):
    if isinstance(v, (int, float)):
        return (EPOCH + datetime.timedelta(days=int(v))).isoformat()
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    return str(v)

def to_serial(v):
    """Make the E-column value JSON-safe. Some workbooks store the NAV date as a
    real date object (not an Excel serial int) -> json.dump would crash. Normalize
    any date/datetime to an Excel serial int; leave numbers as-is."""
    if isinstance(v, datetime.datetime):
        v = v.date()
    if isinstance(v, datetime.date):
        return (v - EPOCH).days
    return v

def clean(v):
    return v.strip() if isinstance(v, str) else v

def main():
    wb = openpyxl.load_workbook(L.CFG["master_path"], data_only=False)
    reg = {}
    for ws in wb.worksheets:
        header_row = None
        for r in range(1, min(ws.max_row, 5) + 1):
            if any("产品代码" in str(ws.cell(r, c).value or "") for c in range(1, 7)):
                header_row = r; break
        data_start = (header_row + 1) if header_row else 1
        codes, names, last_dr, last_date = set(), set(), None, None
        for r in range(data_start, ws.max_row + 1):
            a = clean(ws.cell(r, 1).value)
            if a in (None, "", "累计"):
                continue
            codes.add(str(a))
            b = clean(ws.cell(r, 2).value)
            if b: names.add(str(b))
            last_dr = r
            ev = ws.cell(r, 5).value
            if ev not in (None, ""): last_date = ev
        fbase = None
        if last_dr:
            fv = ws.cell(last_dr, 6).value
            if isinstance(fv, str) and fv.startswith("="):
                if re.search(r"\bC\d", fv): fbase = "C"
                elif re.search(r"\bD\d", fv): fbase = "D"
        reg[ws.title] = dict(header_row=header_row, data_start=data_start, last_data_row=last_dr,
                             max_col=ws.max_column, codes=sorted(codes), names=sorted(names),
                             return_base=fbase, last_date_serial=to_serial(last_date),
                             last_date=serial_to_date(last_date) if last_date is not None else None)
    out = os.path.join(L.HERE, "registry.json")
    json.dump(reg, open(out, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"wrote {out} ({len(reg)} sheets)")

if __name__ == "__main__":
    main()
