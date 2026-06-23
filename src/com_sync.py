# -*- coding: utf-8 -*-
"""COM 同步器 —— 把"自动更新预览副本"里【比正式表多出来的新行】移植进正式表。

这是整套自动更新【唯一】驱动真实 Excel(win32com) 写正式表的地方:
  write.py / obs_daily.py 只负责把新行算进一份一次性"预览副本"(它们的复杂逻辑一行不动);
  本模块再 diff 预览 vs 正式表, 仅把新增的【数据行 + 累计行】用 COM 搬进正式表。

为什么这样: openpyxl 往返会毁手工格式 + 复活主题色(到 Linda 电脑变色)。COM 只碰
"copy 上一行→插入→逐格覆盖值/公式", 格式天然继承; 颜色按【清晰一致规则】重算成写死 RGB:
  百分比(0.00%)的收益格 -> 红涨(>0,FF0000) / 绿跌(<0,008000) / 0 黑;
  指数列(表头含 指数/中证/沪深/北证/红利) -> 黑(指数不红绿);
  其它(净值/日期/代码/名称) -> 不动(继承上一行的黑)。
日期一律转 Excel 序列号写入(直接传 datetime 会被 pywin32 按时区倒退 1 天)。

打开正式表用纯晚绑定(dynamic.Dispatch): 起独立 Excel 实例, 不打扰用户已开的 Excel,
也绕开新版 Python 上会损坏的 gen_py(makepy) 缓存。

用法:
  python com_sync.py            # 干跑: 只报告"会搬哪些新行/补哪些空格", 不动正式表
  python com_sync.py --commit   # 备份正式表后, 把预览的新行/空格 COM 写入正式表
  可选 --preview PATH 指定预览副本(默认 <主表同目录>/各私募净值_自动更新预览.xlsx)
"""
import os, sys, re, shutil, datetime
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
import navlib as L

EPOCH = datetime.date(1899, 12, 30)
RED, GREEN, BLACK = 255, 32768, 0          # COM(BGR)整数: 红FF0000 / 绿008000 / 黑
xlDown, xlLeft = -4121, -4131
xlCalcManual, xlCalcAuto = -4135, -4105
IDX_RE = re.compile(r"指数|中证|沪深|北证|红利")          # 指数列表头特征 -> 该列黑(不红绿)

def s2d(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, (int, float)) and v > 40000:
        return EPOCH + datetime.timedelta(days=int(v))
    return None

def cellpack(cell):
    """读一格内容 -> ('ARR'|'F'|'V', payload), 供之后照原样写回。"""
    v = cell.value
    if isinstance(v, ArrayFormula): return ("ARR", v.text)
    if isinstance(v, str) and v.startswith("="): return ("F", v)
    return ("V", v)

def find_accum(ws):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == "累计":
            return r
    return None

def data_rows(ws, accr):
    """{date -> 行号}, 只数有产品代码且E列是日期的数据行(不含累计/空行)。"""
    out = {}
    for r in range(2, (accr or ws.max_row + 1)):
        a = ws.cell(r, 1).value
        if a in (None, "") or str(a).strip() == "累计":
            continue
        d = s2d(ws.cell(r, 5).value)
        if d:
            out[d] = r
    return out

def index_cols(ws, maxc, data_start):
    cols = set()
    for c in range(6, maxc + 1):
        for r in range(1, (data_start or 3) + 1):
            if IDX_RE.search(str(ws.cell(r, c).value or "")):
                cols.add(c); break
    return cols

def _blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")

def plan_sheet(name, msheet, psheet):
    """对一张表算移植方案; 返回 dict 或 None(无变化) 或 ('SKIP', 原因)。
    含: ①新增数据行 ②存量行里【正式表空着、预览填上了】的格(如刚结束那周的周收益; 绝不覆盖已有值) ③累计行。"""
    macc, pacc = find_accum(msheet), find_accum(psheet)
    if not macc or not pacc:
        return ("SKIP", "没找到累计行")
    md, pd = data_rows(msheet, macc), data_rows(psheet, pacc)
    if not md or not pd:
        return ("SKIP", "无数据行")
    last_md = max(md); last_row = md[last_md]
    if macc != last_row + 1:
        return ("SKIP", f"累计行不紧跟数据(累计{macc},末行{last_row})")
    for d, r in md.items():                              # 结构对齐: 已有日期必须在预览同一行号
        if pd.get(d) != r:
            return ("SKIP", f"结构不一致(日期{d} 正式表第{r}行/预览第{pd.get(d)}行)")
    maxc = max(msheet.max_column, psheet.max_column)

    def packcell(sheet, r, c):
        kind, val = cellpack(sheet.cell(r, c))
        if c == 5:                                       # 日期列 -> 序列号(防 pywin32 时区倒退1天)
            dd = s2d(val); val, kind = ((dd - EPOCH).days if dd else val), "V"
        return kind, val

    # ① 新增行(预览有、正式表无的日期; 必须紧跟末行连续排到累计行前)
    new_dates = sorted(d for d in pd if d not in md)
    expect = last_row + 1
    rows = []
    for d in new_dates:
        pr = pd[d]
        if pr != expect:
            return ("SKIP", f"新行不连续(日期{d} 预览第{pr}行, 期望{expect})")
        rows.append((d, [packcell(psheet, pr, c) for c in range(1, maxc + 1)]))
        expect += 1

    # ② 存量行补空: 仅当正式表该格为空、预览该格有值(典型=刚结束那周的周收益)。决不动已有值, 保护手填。
    fills = []
    for d, r in md.items():
        for c in range(1, maxc + 1):
            if _blank(msheet.cell(r, c).value) and not _blank(psheet.cell(r, c).value):
                kind, val = packcell(psheet, r, c)
                fills.append((r, c, kind, val))

    if not rows and not fills:
        return None
    acc = [cellpack(psheet.cell(pacc, c)) for c in range(1, maxc + 1)]
    ds = next((r for r in range(1, msheet.max_row + 1) if s2d(msheet.cell(r, 5).value)), 2)
    return dict(name=name, last_row=last_row, macc=macc, maxc=maxc, N=len(rows),
                rows=rows, fills=fills, acc=acc, idx=index_cols(msheet, maxc, ds),
                new_dates=[d.isoformat() for d in new_dates])

def build_plans(master, preview):
    mwb = openpyxl.load_workbook(master, data_only=False)
    pwb = openpyxl.load_workbook(preview, data_only=False)
    plans, skips = {}, []
    for name in pwb.sheetnames:
        if name not in mwb.sheetnames:
            continue
        p = plan_sheet(name, mwb[name], pwb[name])
        if p is None:
            continue
        if isinstance(p, tuple):
            skips.append((name, p[1]))
        else:
            plans[name] = p
    mwb.close(); pwb.close()
    return plans, skips

def _set(cell, kind, val):
    if kind == "ARR":
        cell.FormulaArray = val
    elif kind == "F":
        cell.Formula = val
    elif val is None:
        cell.ClearContents()
    else:
        cell.Value = val

def _color_one(ws, r, c, idx):
    """按规则给单格上色: 第6列起, 指数列黑; 百分比收益格红涨(>0)绿跌(<0)0黑; 其余不动。"""
    if c < 6:
        return                                           # 前5列(代码/名称/净值/累计/日期)保持黑, 不动
    cell = ws.Cells(r, c)
    if c in idx:                                         # 指数列 -> 黑(不随涨跌)
        cell.Font.Color = BLACK
        return
    if "%" in str(cell.NumberFormat):                    # 百分比=收益 -> 红绿
        v = cell.Value
        if isinstance(v, (int, float)):
            cell.Font.Color = RED if v > 0 else (GREEN if v < 0 else BLACK)

def _color_returns(ws, r, maxc, idx):
    """对一整行(新数据行或累计行)按规则上色。"""
    for c in range(6, maxc + 1):
        _color_one(ws, r, c, idx)

def _excel():
    """纯晚绑定 Excel 实例: 独立进程(不打扰用户已开的 Excel), 绕开 gen_py(makepy)缓存。"""
    from win32com.client import dynamic
    xl = dynamic.Dispatch("Excel.Application")
    xl.Visible = False
    try:
        xl.DisplayAlerts = False
    except Exception:
        pass
    return xl

def apply_com(plans, master):
    xl = _excel()
    try:
        wb = xl.Workbooks.Open(os.path.abspath(master))
    except Exception as e:
        try: xl.Quit()
        except Exception: pass
        raise RuntimeError(f"打不开正式表(可能正被人打开/占用), 本轮不写、下次自动补: {e}")
    xl.Calculation = xlCalcManual
    try:
        # 1) 插新行 + 移植值/公式; 补存量空格; 重写累计行
        for p in plans.values():
            ws = wb.Worksheets(p["name"]); N, lr = p["N"], p["last_row"]
            if N:
                ws.Rows(lr).Copy()
                ws.Rows(f"{lr+1}:{lr+N}").Insert(xlDown)
                xl.CutCopyMode = False
                for i, (d, cells) in enumerate(p["rows"]):
                    rr = lr + 1 + i
                    for c, (kind, val) in enumerate(cells, 1):
                        _set(ws.Cells(rr, c), kind, val)
            for (r, c, kind, val) in p["fills"]:
                _set(ws.Cells(r, c), kind, val)
            ar = lr + 1 + N
            for c, (kind, val) in enumerate(p["acc"], 1):
                _set(ws.Cells(ar, c), kind, val)
        xl.CalculateFull()
        # 2) 上色(写死 RGB): 新行整行 + 被补空格的存量行 + 累计行 —— 都按"百分比红绿/指数黑/其余不动"
        for p in plans.values():
            ws = wb.Worksheets(p["name"]); N, lr, maxc, idx = p["N"], p["last_row"], p["maxc"], p["idx"]
            for i in range(N):                           # 新行: 左对齐+字号10 + 整行按规则上色
                r = lr + 1 + i
                ws.Rows(r).HorizontalAlignment = xlLeft
                ws.Rows(r).Font.Size = 10
                _color_returns(ws, r, maxc, idx)
            for (r, c, kind, val) in p["fills"]:         # 补的存量空格: 只给被补的那格上色(不动该行其它手工色)
                _color_one(ws, r, c, idx)
            _color_returns(ws, lr + 1 + N, maxc, idx)    # 累计行
        wb.Save()
    finally:
        xl.Calculation = xlCalcAuto
        wb.Close(SaveChanges=True); xl.Quit()

def main():
    commit = "--commit" in sys.argv
    preview = sys.argv[sys.argv.index("--preview") + 1] if "--preview" in sys.argv else None
    master = L.CFG["master_path"]
    if not preview:
        d = os.path.dirname(master)
        stem = os.path.splitext(os.path.basename(master))[0]
        preview = os.path.join(d, stem + "_自动更新预览.xlsx")
    if not os.path.exists(preview):
        print(f"!! 预览副本不存在: {preview}\n   (应由 write.py/obs_daily.py --book 先生成)"); return 2

    plans, skips = build_plans(master, preview)
    print(f"主表 : {master}")
    print(f"预览 : {preview}")
    print(f"模式 : {'写入(--commit, 先备份)' if commit else '干跑(只报告, 不动正式表)'}\n")
    if skips:
        print("跳过的表(结构不符, 本轮不动, 不影响其它表):")
        for n, why in skips:
            print(f"   {n:16s} {why}")
        print()
    if not plans:
        print("没有要移植的新行/空格(正式表已是最新)。")
        return 0
    total = fills_total = 0
    print("将变更的表:")
    for n, p in plans.items():
        total += p["N"]; fills_total += len(p["fills"])
        nd = ("新增 " + ", ".join(p["new_dates"])) if p["new_dates"] else "无新增"
        fl = f" | 补空格{len(p['fills'])}处" if p["fills"] else ""
        print(f"   {n:16s} {nd}{fl}")
    print(f"\n合计 新增 {total} 行 + 补空格 {fills_total} 处 / {len(plans)} 张表。")
    if not commit:
        print("\n(干跑结束, 正式表未改动。确认无误后加 --commit 写入。)")
        return 0
    bk = os.path.join(os.path.dirname(master), "backups")
    os.makedirs(bk, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = os.path.splitext(os.path.basename(master))[0]
    bak = os.path.join(bk, f"{stem}_synced_{ts}.xlsx")
    shutil.copy2(master, bak)
    print(f"\n已备份正式表 -> {bak}")
    apply_com(plans, master)
    print(f"已用 COM 写入正式表 {total} 行 + 补 {fills_total} 处(格式保留, RGB红绿, 无主题色)。")
    return 0

if __name__ == "__main__":
    sys.exit(main())
