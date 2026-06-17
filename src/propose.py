# -*- coding: utf-8 -*-
"""Proposal engine. compute_proposals() returns the rows that SHOULD be added to
each in-scope sheet (completed weeks only), plus current-week previews and
anomalies. It writes nothing. Re-used by write.py.
"""
import os, json, csv, datetime
import openpyxl
import navlib as L
import phase2 as P2

EXCLUDE = set()   # 基金21 2026-06-15 启用: 以某托管发来的净值为准(此前对不上是因她上周临时填了周四值)
# 基金16: 来源=某券商 glrfw 每日值, 记录每周五(同 基金15); 此前 06-05 的差异是她临时手改, 非系统性
RET_FLAG = 0.20               # weekly return magnitude that triggers a review flag
DORMANT_DAYS = 45             # if newest email older than this, product likely redeemed -> don't auto-write
# Normally we never write the unfinished current week. NAV_TEST_ALLWEEKS=1 lifts
# that guard (testing only) so current-week rows become writable.
CURRENT_WK = (9999, 99) if os.environ.get("NAV_TEST_ALLWEEKS") else datetime.date.today().isocalendar()[:2]

def s2d(v):
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    if isinstance(v, datetime.datetime):
        return v.date()
    return v

def last_recorded(ws, info):
    last_d = last_u = last_cum = None
    for r in range(info["data_start"], ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a in (None, "", "累计"):
            continue
        d = s2d(ws.cell(row=r, column=5).value)
        if d:
            last_d = d
            last_u = L.num(ws.cell(row=r, column=3).value)
            last_cum = L.num(ws.cell(row=r, column=4).value)
    return last_d, last_u, last_cum

def compute_proposals():
    idx = json.load(open(os.path.join(L.HERE, "index.json"), encoding="utf-8"))["index"]
    by_sheet = {}
    for r in idx:
        if r["date"]:
            by_sheet.setdefault(r["sheet"], []).append(r)
    reg = L.load_registry()
    wb = openpyxl.load_workbook(L.CFG["master_path"], data_only=True)
    M = L.connect()
    cache = {}
    p2cache = {}

    def parsed(cand):
        fmt = cand["fmt"]
        if fmt in P2.FMTS:
            if cand["uid"] not in p2cache:
                msg = L.fetch_full(M, cand["uid"])
                p2cache[cand["uid"]] = P2.parse_rows(msg) if msg else []
            rows = p2cache[cand["uid"]]
            td = datetime.date.fromisoformat(cand["date"]) if cand["date"] else None
            best = next((r for r in rows if td and r["date"] == td), None)
            if best is None and rows:
                best = max(rows, key=lambda r: r["date"])
            if not best or best["unit"] is None:
                return None
            return dict(code=best["code"] or cand["code"], date=best["date"],
                        unit=best["unit"], cum=best["cum"], source=fmt)
        body = cache.get(cand["uid"])
        if not body:
            msg = L.fetch_full(M, cand["uid"])
            body = L.body_text(msg) if msg else ""
            if body:
                cache[cand["uid"]] = body
        return L.parse_body(fmt, cand["subject"], body, cand["code"])

    proposals, anomalies, preview, summary = [], [], [], []
    dormant, excluded, held = [], [], []
    today = datetime.date.today()
    for sheet in L.all_scope():
        if sheet not in wb.sheetnames:        # 表名不一致(尚未改名/初始化)时跳过, 不崩整轮
            summary.append((sheet, None, 0, "表不存在(尚未改名/初始化?)，跳过"))
            continue
        info = reg[sheet]
        ws = wb[sheet]
        last_d, last_u, last_cum = last_recorded(ws, info)
        offset = (last_cum - last_u) if (last_cum is not None and last_u is not None) else 0.0
        if sheet in EXCLUDE:
            summary.append((sheet, last_d, 0, "排除：来源不符，人工处理"))
            excluded.append(sheet)
            continue
        # dormant guard: if the newest email for this product is long stale, the
        # product was likely redeemed/closed -> flag, don't silently backfill.
        sheet_dates = [datetime.date.fromisoformat(c["date"]) for c in by_sheet.get(sheet, []) if c["date"]]
        newest = max(sheet_dates) if sheet_dates else None
        if newest and (datetime.date.today() - newest).days > DORMANT_DAYS:
            summary.append((sheet, last_d, 0, f"疑似已赎回/停更(最新邮件{newest})，未自动写入，请人工确认"))
            dormant.append((sheet, newest.isoformat()))
            continue
        last_wk = last_d.isocalendar()[:2] if last_d else (0, 0)
        weeks = {}
        for cand in by_sheet.get(sheet, []):
            cd = datetime.date.fromisoformat(cand["date"])
            if cd <= (last_d or datetime.date.min):
                continue
            wk = cd.isocalendar()[:2]
            if wk <= last_wk:
                continue
            if wk not in weeks or cd > datetime.date.fromisoformat(weeks[wk]["date"]):
                weeks[wk] = cand
        prev_u = last_u
        added = []
        for wk in sorted(weeks):
            cand = weeks[wk]
            p = parsed(cand)
            d = datetime.date.fromisoformat(cand["date"])
            if not p or p["unit"] is None:
                anomalies.append(dict(sheet=sheet, date=d, flag="解析失败", source=cand["fmt"], subject=cand["subject"]))
                continue
            # safety: the code parsed from the body must match the expected product
            if p.get("code") and L.base_code(p["code"]) != cand["code"]:
                anomalies.append(dict(sheet=sheet, date=d, flag="代码不一致(取数校验失败)",
                                      source=cand["fmt"], subject=cand["subject"]))
                continue
            # safety: the parsed NAV date must match the email's subject date
            if p.get("date") and p["date"] != d:
                anomalies.append(dict(sheet=sheet, date=d, flag="日期不一致(取数校验失败)",
                                      source=cand["fmt"], subject=cand["subject"]))
                continue
            unit = p["unit"]
            if p["source"] == "citics_virtual":
                cum = round(unit + offset, 4)
            elif p["cum"] is not None:
                cum = p["cum"]
            else:
                cum = round(unit + offset, 4)
            ret = (unit / prev_u - 1) if prev_u else None
            flag = ""
            if ret is not None and abs(ret) > RET_FLAG:
                flag = "周收益异常 %.1f%%" % (ret * 100)
            if unit <= 0:
                flag = "净值<=0"
            # cadence guard: Linda records each week's last trading day (≈周五).
            # If a just-finished week's latest available date is before that
            # week's Friday and is still fresh (<=6 天), the Friday NAV is most
            # likely still in transit -> HOLD (don't write, don't advance the
            # week pointer) so the real Friday row can land on the next run.
            # After ~一周 with no Friday data we accept it (handles 基金24 这类不规则发
            # 布 / 节假日短周). 当周(未截止)仍走 preview.
            pending_friday = (not (wk >= CURRENT_WK)) and (d.weekday() < 4) and ((today - d).days <= 6)
            row = dict(sheet=sheet, code=cand["code"], date=d, unit=unit, cum=cum,
                       ret=ret, source=p["source"], subject=cand["subject"],
                       current=(wk >= CURRENT_WK), flag=flag, hold=pending_friday)
            if row["current"]:
                preview.append(row)
            elif pending_friday:
                held.append(row)
            elif flag:
                anomalies.append(dict(sheet=sheet, date=d, flag=flag, unit=unit, cum=cum,
                                      source=p["source"], subject=cand["subject"]))
            else:
                proposals.append(row)
            added.append(row)
            prev_u = unit
        summary.append((sheet, last_d, len(added),
                        ", ".join(f"{a['date']}={a['unit']}" + ("~" if a['current'] else ("待五" if a.get('hold') else ("!" if a['flag'] else ""))) for a in added) or "(无新数据)"))
    M.logout()
    return dict(proposals=proposals, anomalies=anomalies, preview=preview,
                summary=summary, dormant=dormant, excluded=excluded, held=held)

def main():
    R = compute_proposals()
    print(f"{'sheet':10s} {'last rec':11s} {'+rows':>5s}  detail   (~=本周未截止预览, !=异常)")
    for sheet, last_d, n, det in R["summary"]:
        print(f"{sheet:10s} {str(last_d):11s} {n:>5d}  {det}")
    pj = os.path.join(L.HERE, "proposed_updates.csv")
    with open(pj, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sheet", "code", "date", "unit_nav", "cum_nav", "weekly_ret", "source", "subject"])
        for r in R["proposals"]:
            w.writerow([r["sheet"], r["code"], r["date"], r["unit"], r["cum"],
                        ("%.4f" % r["ret"] if r["ret"] is not None else ""), r["source"], r["subject"]])
    print(f"\n已截止周·可写入: {len(R['proposals'])} 行 -> {pj}")
    print(f"异常待核: {len(R['anomalies'])} 行 | 本周未截止预览: {len(R['preview'])} 行 | 待周五暂缓: {len(R.get('held', []))} 行")
    if R["preview"]:
        print("\n本周未截止预览（暂不写入）：")
        for r in R["preview"]:
            print(f"   {r['sheet']:10s} {r['date']} 单位={r['unit']} 累计={r['cum']} 周收益={r['ret']:.4f}")
    if R.get("held"):
        print("\n待本周五净值·暂缓（最新仅到周内非周五，下次自动补；不会写错日期）：")
        for r in R["held"]:
            print(f"   {r['sheet']:10s} {r['date']}(周{r['date'].isoweekday()}) 单位={r['unit']} 累计={r['cum']}")

if __name__ == "__main__":
    main()
