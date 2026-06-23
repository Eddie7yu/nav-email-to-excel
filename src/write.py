# -*- coding: utf-8 -*-
"""Writer. Appends proposed weekly NAV rows into each sheet, before the 累计
summary row, copying cell styles from the row above and rebuilding formulas.

  python write.py            -> writes a PREVIEW COPY (master untouched)
  python write.py --commit   -> backs up master, then writes master in place

Index-fund rows are written only when the benchmark close is available in
index_cache.json; otherwise they are DEFERRED (kept consistent, retried later).
"""
import os, sys, re, json, csv, shutil, datetime
from copy import copy
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula
import navlib as L
from propose import compute_proposals

EPOCH = datetime.date(1899, 12, 30)
IDX_RE = re.compile(r"中证\s*(500|1000|2000)|北证")
# 盈利=深红(标准色 C00000); 亏损=绿 着色6 深色50%(主题 Accent6, Darker 50%)。
# 绿用主题引用, 与 Excel 调色板里"着色6 深色50%"那一格完全一致(随主题), 不写死 RGB。
RED_COL = Color(rgb="FFC00000")
GREEN_COL = Color(theme=9, tint=-0.4999847412109375)
BLACK = Color(theme=1, tint=0.0)                      # 黑(默认文字色); 指数列保持黑白不上红绿
# 指数列表头特征: 含 指数/中证/沪深/北证 -> 这类列(及累计)不上红绿, 用户要黑白
IDX_HDR = re.compile(r"指数|中证|沪深|北证")


def _index_cols(ws, maxc):
    """返回"指数列"列号集合(表头含 指数/中证/沪深/北证, 从F列第6起)。"""
    cols = set()
    for c in range(6, maxc + 1):
        for r in range(1, 4):
            if IDX_HDR.search(str(ws.cell(r, c).value or "")):
                cols.add(c)
                break
    return cols


def _font_kind(cell):
    """该格字体大类：'red'/'green'/None（按红绿通道判断，忽略黑色/默认）。"""
    f = cell.font
    if not f or not f.color or f.color.type != "rgb":
        return None
    rgb = f.color.rgb
    if not rgb or rgb in ("00000000", "FF000000"):
        return None
    try:
        r, g = int(rgb[2:4], 16), int(rgb[4:6], 16)
    except (ValueError, TypeError):
        return None
    return "red" if (r >= 0x80 and g < 0x80) else ("green" if (g >= 0x80 and r < 0x80) else None)


def _set_font_color(cell, color, bold=True):
    """给单元格设字体色(默认加粗, 涨跌色要醒目)。color 可为 rgb 字符串或 Color 对象(主题色)。
    指数列用 BLACK + bold=False 还原黑白。"""
    f = cell.font
    cell.font = Font(name=f.name, size=f.size, bold=bold, italic=f.italic,
                     vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=color)


def _cellnum(ws, ref):
    m = re.match(r"\$?([A-Z]+)\$?(\d+)$", ref)
    if not m:
        return None
    v = ws.cell(int(m.group(2)), column_index_from_string(m.group(1))).value
    return v if isinstance(v, (int, float)) else None


def _eval_accum(ws, val):
    """算累计收益格的值(不依赖 Excel 缓存): 支持 A/B-1、PRODUCT(1+col s:e)-1、F-H。"""
    text = val.text if isinstance(val, ArrayFormula) else str(val)
    t = (text[1:] if text.startswith("=") else text).replace(" ", "")
    m = re.fullmatch(r"([A-Z]+\d+)/([A-Z]+\d+)-1", t)
    if m:
        a, b = _cellnum(ws, m.group(1)), _cellnum(ws, m.group(2))
        return (a / b - 1) if (a is not None and b not in (None, 0)) else None
    m = re.fullmatch(r"PRODUCT\(1\+([A-Z]+)(\d+):([A-Z]+)(\d+)\)-1", t)
    if m:
        col, prod, any_ = column_index_from_string(m.group(1)), 1.0, False
        for r in range(int(m.group(2)), int(m.group(4)) + 1):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                prod *= (1 + v); any_ = True
        return (prod - 1) if any_ else None
    m = re.fullmatch(r"([A-Z]+\d+)-([A-Z]+\d+)", t)
    if m:
        def side(ref):
            mm = re.match(r"([A-Z]+)(\d+)", ref)
            cv = ws.cell(int(mm.group(2)), column_index_from_string(mm.group(1))).value
            # 引用格本身可能又是公式(=...)或数组公式(PRODUCT 累计指数)->递归求值, 否则取字面数值
            if isinstance(cv, ArrayFormula) or (isinstance(cv, str) and cv.startswith("=")):
                return _eval_accum(ws, cv)
            return _cellnum(ws, ref)
        a, b = side(m.group(1)), side(m.group(2))
        return (a - b) if (a is not None and b is not None) else None
    return None


def _recolor_accum(ws, accr, maxc):
    """累计行所有"可计算的收益率"格统一按正负上色：盈利红、亏损绿（整表统一）。
    只动累计行；无法计算的格(如跨表引用的指数累计)跳过，保持原样。"""
    idx = _index_cols(ws, maxc)
    for c in range(6, maxc + 1):
        cell = ws.cell(accr, c)
        if cell.value in (None, ""):
            continue
        if c in idx:                                  # 指数列累计: 黑白(保留加粗), 不红绿
            _set_font_color(cell, BLACK, bold=cell.font.bold)
            continue
        v = _eval_accum(ws, cell.value)
        if v is None:
            continue
        _set_font_color(cell, GREEN_COL if v < 0 else RED_COL)


# ---- 类型B指数列(表头形如"500指数/1000指数/2000指数",列里直接放周收益) ----
# 这类列是跨表引用对应"中证XXX"源表(Type A)的 H(指数周收益)列。源表已被 fill_index/
# write 自动填好，这里把空缺单元格按日期补上引用，免去每周手工。
INDEX_RET_SRC = {"中证500": "半鞅", "中证1000": "正合", "中证2000": "千衍"}
TYPEB_RE = re.compile(r"^(500|1000|2000)指数$")


def _typeb_cols(ws, header_row, data_start):
    """返回 {列号: 中证指数名}，扫描表头行~首数据行(指数标签位置不固定)。"""
    out = {}
    top = (data_start or 3)
    for c in range(7, ws.max_column + 1):
        for hr in range(1, top + 1):
            h = ws.cell(hr, c).value
            m = TYPEB_RE.match(str(h).strip()) if h else None
            if m:
                out[c] = "中证" + m.group(1)
                break
    return out


def _date_row_map(ws, data_start):
    m = {}
    for r in range(data_start or 3, ws.max_row + 1):
        e = ws.cell(r, 5).value
        if isinstance(e, (int, float)):
            m[int(e)] = r
    return m


def fill_crossref_index(wb, reg):
    """填类型B指数列的空缺格：按日期引用对应中证源表的 H 列(=源!H行)。
    只填空格、只填源表有该日期的；不动已填的格。返回 (filled, still_blank)。"""
    filled, still_blank = [], []
    srcmap = {}
    for sheet in wb.sheetnames:
        info = reg.get(sheet)
        if not info:
            continue
        ws = wb[sheet]
        cols = _typeb_cols(ws, info.get("header_row") or 1, info.get("data_start") or 3)
        if not cols:
            continue
        accr = next((r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "累计"), ws.max_row + 1)
        ds = info.get("data_start") or 3
        for c, idxname in cols.items():
            src = INDEX_RET_SRC.get(idxname)
            if not src or src not in wb.sheetnames:
                continue
            if src not in srcmap:
                srcmap[src] = _date_row_map(wb[src], (reg.get(src) or {}).get("data_start") or 3)
            for r in range(ds, accr):
                cell = ws.cell(r, c)
                if cell.value not in (None, ""):
                    continue
                e = ws.cell(r, 5).value
                if not isinstance(e, (int, float)):
                    continue
                srow = srcmap[src].get(int(e))
                if srow:
                    cell.value = f"={src}!H{srow}"
                    filled.append(f"{sheet}!{get_column_letter(c)}{r} = {src}!H{srow}")
                else:
                    still_blank.append(f"{sheet}!{get_column_letter(c)}{r} ({idxname} 源表无该日期)")
    return filled, still_blank

def _bump_array(af, col_idx, old_last, new_last, new_row):
    """Rebuild a 累计-row ArrayFormula (e.g. =PRODUCT(1+H4:H39)-1) after rows
    were inserted: keep each column's own start row, extend the range END from
    the old last data row to the new one, and re-anchor ref to the moved cell.
    Only the range-end token (col+old_last, e.g. 'H39') is bumped; start rows
    (H4, G27 …) are preserved."""
    col = get_column_letter(col_idx)
    text = af.text if isinstance(af, ArrayFormula) else str(af)
    text = re.sub(r"(?<![A-Za-z0-9$])\$?%s\$?%d(?![0-9])" % (col, old_last),
                  "%s%d" % (col, new_last), text)
    return ArrayFormula(ref="%s%d" % (col, new_row), text=text)

def detect_index(ws, info):
    g = ws.cell(info["header_row"] or 1, 7).value
    if g and IDX_RE.search(str(g)):
        return str(g).strip()
    return None

def main():
    commit = "--commit" in sys.argv
    # --book PATH: 就地在指定工作簿(预览副本)上读写, 不备份、不碰正式表。
    # 供新链路用(write 只算进预览, 再由 com_sync 用 COM 把新行搬进正式表)。
    book = sys.argv[sys.argv.index("--book") + 1] if "--book" in sys.argv else None
    src_path = book or L.CFG["master_path"]
    R = compute_proposals()
    props = R["proposals"]
    by_sheet = {}
    for r in props:
        by_sheet.setdefault(r["sheet"], []).append(r)
    for v in by_sheet.values():
        v.sort(key=lambda r: r["date"])

    reg = L.load_registry()
    cache_path = os.path.join(L.HERE, "index_cache.json")
    _cache = json.load(open(cache_path, encoding="utf-8")) if os.path.exists(cache_path) else {}
    closes_map = _cache.get("closes", {})
    rejected = set(_cache.get("rejected", []))

    wb = openpyxl.load_workbook(src_path, data_only=False)

    written, deferred, skipped, blank_idx = [], [], [], []
    for sheet, rows in by_sheet.items():
        ws = wb[sheet]
        info = reg[sheet]
        last_dr = info["last_data_row"]
        sum_row = last_dr + 1
        maxc = ws.max_column
        # safety: row right below data must be the 累计 summary; rows beyond must be empty
        if str(ws.cell(sum_row, 1).value).strip() != "累计":
            skipped.append((sheet, "结构异常(末行不是累计行)，跳过"))
            continue
        # 累计行后必须为空才安全追加；空白/纯空格视为空(忽略残留空格,如钇远B10)
        def _blank(v):
            return v is None or (isinstance(v, str) and v.strip() == "")
        tail_dirty = any(not _blank(ws.cell(r, c).value)
                         for r in range(sum_row + 1, ws.max_row + 1) for c in range(1, maxc + 1))
        if tail_dirty:
            skipped.append((sheet, "累计行后有内容，跳过"))
            continue
        base = info["return_base"] or "D"
        first = info["data_start"]
        index_name = detect_index(ws, info)
        # index mode: None=no index col; fill=have validated closes; blank=index
        # cannot be obtained (write NAV, leave G blank+flag); wait=cache missing -> defer
        if not index_name:
            idx_mode = None
        elif index_name in closes_map:
            idx_mode = "fill"
        elif index_name in rejected:
            idx_mode = "blank"
        else:
            idx_mode = "wait"

        # capture the summary row template (values + styles) before overwriting
        tmpl = [(ws.cell(sum_row, c).value, copy(ws.cell(sum_row, c)._style)) for c in range(1, maxc + 1)]
        # style source = last real data row
        src = last_dr

        # decide writable (contiguous; stop at first row needing an unavailable index)
        writable = []
        for r in rows:
            if idx_mode == "wait":
                deferred.append((sheet, r["date"], "待指数(未取到) " + index_name))
                break
            if idx_mode == "fill" and not closes_map[index_name].get(r["date"].isoformat()):
                deferred.append((sheet, r["date"], "待指数收盘 " + index_name))
                break
            writable.append(r)
        for r in rows[len(writable):]:
            if not any(d[0] == sheet and d[1] == r["date"] for d in deferred):
                deferred.append((sheet, r["date"], "待前序行"))

        if not writable:
            continue
        N = len(writable)

        def put(rr, col, value):
            c = ws.cell(rr, col)
            c.value = value
            c._style = copy(ws.cell(src, col)._style)

        def _clean(v):                                   # 去掉源行代码/名称的首尾换行/空白(否则Excel显示乱码方块)
            return v.strip() if isinstance(v, str) else v

        for i, r in enumerate(writable):
            RR = last_dr + 1 + i
            put(RR, 1, _clean(ws.cell(src, 1).value))    # A code
            put(RR, 2, _clean(ws.cell(src, 2).value))    # B name
            put(RR, 3, round(r["unit"], 4))              # C unit
            put(RR, 4, round(r["cum"], 4))               # D cumulative
            put(RR, 5, (r["date"] - EPOCH).days)         # E date serial
            put(RR, 6, f"={base}{RR}/{base}{RR-1}-1")    # F weekly return
            set_cols = {1, 2, 3, 4, 5, 6}
            if idx_mode == "fill":
                close = closes_map[index_name].get(r["date"].isoformat())
                put(RR, 7, close)                        # G index level
                put(RR, 8, f"=G{RR}/G{RR-1}-1")          # H index return
                put(RR, 9, f"=F{RR}-H{RR}")              # I excess
                set_cols |= {7, 8, 9}
            # explicitly clear every other column so nothing leaks from the old 累计 row
            for c in range(1, maxc + 1):
                if c not in set_cols:
                    put(RR, c, None)
            # flag a sheet whose G column is actually used (index / cross-sheet ref) but
            # we couldn't auto-fill (rejected index, or 跨表引用 like 爱凡哲/大麓/和美)
            if idx_mode != "fill" and maxc >= 7 and ws.cell(src, 7).value not in (None, ""):
                if not any(b[0] == sheet for b in blank_idx):
                    blank_idx.append((sheet, index_name or "G列(跨表引用)需手工"))
            written.append(dict(sheet=sheet, date=r["date"], unit=r["unit"],
                                cum=r["cum"], source=r["source"], ret=r.get("ret")))

        # rewrite 累计 summary row at its new position
        new_sum = last_dr + 1 + N
        new_last = new_sum - 1
        for c in range(1, maxc + 1):
            val, style = tmpl[c - 1]
            cell = ws.cell(new_sum, c)
            cell._style = copy(style)
            if c == 6:
                cell.value = f"={base}{new_last}/{base}{first}-1"
            elif c == 8 and idx_mode == "fill":
                cell.value = f"=G{new_last}/G{first}-1"
            elif c == 9 and idx_mode == "fill":
                cell.value = f"=F{new_sum}-H{new_sum}"
            elif isinstance(val, ArrayFormula):
                # manual cross-fund columns (G–L) carry a 累计 array formula like
                # =PRODUCT(1+H4:H39)-1; re-anchor + extend it to the inserted rows
                cell.value = _bump_array(val, c, last_dr, new_last, new_sum)
            else:
                cell.value = val
        # 累计行收益配色随正负维护(盈红亏绿), 防自动写入后颜色陈旧; 仅动已染色的格
        _recolor_accum(ws, new_sum, maxc)

    # 类型B指数列(大麓/和美 500指数, 爱凡哲 2000/1000指数)空缺格 -> 自动按日期引用源表
    idx_filled, idx_blank = fill_crossref_index(wb, reg)
    if idx_filled:
        _filled_sheets = {f.split("!")[0] for f in idx_filled}
        blank_idx = [b for b in blank_idx if b[0] not in _filled_sheets]

    # output
    if book:
        wb.save(book)
        target = book
        print(f"[--book] 已就地写入工作簿(预览副本) -> {book}\n(正式表未改动; 由 com_sync 用 COM 移植新行)")
    elif commit:
        os.makedirs(os.path.join(L.HERE, "backups"), exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = os.path.splitext(os.path.basename(L.CFG["master_path"]))[0]
        bak = os.path.join(L.HERE, "backups", f"{stem}_{ts}.xlsx")
        shutil.copy2(L.CFG["master_path"], bak)
        wb.save(L.CFG["master_path"])
        target = L.CFG["master_path"]
        print(f"已备份 -> {bak}\n已写入正式表 -> {target}")
    else:
        d = os.path.dirname(L.CFG["master_path"])
        stem = os.path.splitext(os.path.basename(L.CFG["master_path"]))[0]
        target = os.path.join(d, stem + "_自动更新预览.xlsx")
        wb.save(target)
        print(f"[预览模式] 已写入副本 -> {target}\n(正式表未改动；确认无误后加 --commit 写正式表)")

    # changelog
    rt = datetime.datetime.now().isoformat(timespec="seconds")
    mode = "commit" if commit else "preview"
    if written:
        clog = os.path.join(L.HERE, "changelog.csv")
        new = not os.path.exists(clog)
        with open(clog, "a", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)
            if new:
                w.writerow(["run_time", "mode", "sheet", "date", "unit_nav", "cum_nav", "source"])
            for x in written:
                w.writerow([rt, mode, x["sheet"], x["date"].isoformat(), x["unit"], x["cum"], x["source"]])

    # structured run result for the notifier (notify.py)
    def _ser(rows):
        out = []
        for r in rows:
            d = dict(r)
            if isinstance(d.get("date"), datetime.date):
                d["date"] = d["date"].isoformat()
            out.append(d)
        return out
    run = dict(
        time=rt, mode=mode,
        written=_ser(written),
        deferred=[(s, d.isoformat(), why) for (s, d, why) in deferred],
        blank_idx=blank_idx,
        skipped=skipped,
        anomalies=_ser(R["anomalies"]),
        preview=_ser(R["preview"]),
        held=_ser(R.get("held", [])),
        dormant=R["dormant"],
        excluded=R["excluded"],
    )
    json.dump(run, open(os.path.join(L.HERE, "last_run.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    print(f"\n写入 {len(written)} 行：")
    for x in written:
        print(f"   {x['sheet']:10s} {x['date']}  单位={x['unit']} 累计={x['cum']}  ({x['source']})")
    if idx_filled:
        print(f"\n指数列自动补全 {len(idx_filled)} 格（按日期引用中证源表）：")
        for f in idx_filled:
            print(f"   {f}")
    if blank_idx:
        print("\n以下产品已写净值，但指数列(G)无法自动获取、留空待手工补：")
        for (sheet, name) in blank_idx:
            print(f"   {sheet:10s} 指数={name}")
    if deferred:
        print(f"\n暂缓 {len(deferred)} 行（待指数收盘或前序行）：")
        for (sheet, d, why) in deferred:
            print(f"   {sheet:10s} {d}  {why}")
    if R.get("held"):
        print(f"\n待本周五净值·暂缓 {len(R['held'])} 行（最新仅到周内非周五，下次自动补，不会写错日期）：")
        for r in R["held"]:
            print(f"   {r['sheet']:10s} {r['date']}  单位={r['unit']}")
    if skipped:
        print("\n跳过：", skipped)

if __name__ == "__main__":
    main()
