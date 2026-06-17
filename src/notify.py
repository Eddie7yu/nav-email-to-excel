# -*- coding: utf-8 -*-
"""Weekly summary email (boss-facing). Reads last_run.json + index.json (+
validation.json if present) and the master workbook, then emails a formal
Chinese report: what was updated this week, performance rankings over four
horizons (本周 / 近一月 / 近三月 / 记录以来), items needing attention, and a
forwarding-staleness alert. Sends via QQ SMTP (same authorization code as IMAP).

    python notify.py          # compose and send
    python notify.py --dry    # print only, do not send
"""
import os, sys, json, datetime
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import openpyxl
import navlib as L

EPOCH = datetime.date(1899, 12, 30)
# Products confirmed by the user as redeemed / stopped reporting — keep them in
# the workbook but do NOT raise them in the weekly email every time.
CONFIRMED_INACTIVE = {"基金12(友）(赎)", "基金28(赎)", "基金08(赎)", "基金09(赎)", "基金18(赎)"}
FRESH_DAYS = 10   # a product counts as "active" if its latest record is this recent

def _load(name):
    p = os.path.join(L.HERE, name)
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None

def pct(x):
    return "—" if x is None else f"{x*100:+.2f}%"

def _s2d(v):
    if isinstance(v, (int, float)):
        return EPOCH + datetime.timedelta(days=int(v))
    if isinstance(v, datetime.datetime):
        return v.date()
    return v

def workbook_stats():
    """Per active product: weekly / 1-month / 3-month / since-record cumulative
    return, computed from raw 单位净值(C) and 累计单位净值(D) values (these are
    literals, not formulas, so openpyxl reads them reliably)."""
    reg = L.load_registry()
    try:
        wb = openpyxl.load_workbook(L.CFG["master_path"], data_only=False)
    except Exception:
        return {}
    today = datetime.date.today()
    stats = {}
    for sheet in L.all_scope():
        if sheet in CONFIRMED_INACTIVE or sheet not in wb.sheetnames:
            continue
        info = reg.get(sheet)
        if not info:
            continue
        ws = wb[sheet]
        rows = []
        for r in range(info["data_start"], ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a in (None, "", "累计"):
                continue
            u = L.num(ws.cell(r, 3).value)
            cu = L.num(ws.cell(r, 4).value)
            d = _s2d(ws.cell(r, 5).value)
            if u is None or not isinstance(d, datetime.date):
                continue
            rows.append((d, u, cu if cu is not None else u))
        if len(rows) < 2:
            continue
        rows.sort(key=lambda x: x[0])
        last_d, last_u, last_cu = rows[-1]
        if (today - last_d).days > FRESH_DAYS:   # dormant -> skip from rankings
            continue

        def ret_since(days):
            cutoff = last_d - datetime.timedelta(days=days)
            base = None
            for d, u, cu in rows[:-1]:
                if d <= cutoff:
                    base = cu
            return (last_cu / base - 1) if base else None

        stats[sheet] = dict(
            weekly=(last_cu / rows[-2][2] - 1) if rows[-2][2] else None,
            m1=ret_since(30), m3=ret_since(90),
            since=(last_cu / rows[0][2] - 1) if rows[0][2] else None,
            latest=last_d.isoformat(),
        )
    return stats

def rank_block(stats, key, title, top=3):
    vals = [(s, v[key]) for s, v in stats.items() if v.get(key) is not None]
    if not vals:
        return []
    vals.sort(key=lambda x: x[1], reverse=True)
    out = [f"【{title}】"]
    out.append("　居前：" + "，".join(f"{s} {pct(r)}" for s, r in vals[:top]))
    out.append("　靠后：" + "，".join(f"{s} {pct(r)}" for s, r in vals[-top:][::-1]))
    return out

def staleness(skip_set, stale_days):
    """Active products whose newest email is older than stale_days -> possible
    forwarding break / stopped sending."""
    idx = _load("index.json")
    if not idx:
        return []
    newest = {}
    for r in idx["index"]:
        if r["date"]:
            d = datetime.date.fromisoformat(r["date"])
            if r["sheet"] not in newest or d > newest[r["sheet"]]:
                newest[r["sheet"]] = d
    today = datetime.date.today()
    out = []
    for sheet in L.all_scope():
        if sheet in skip_set:
            continue
        nd = newest.get(sheet)
        if nd is None or (today - nd).days > stale_days:
            out.append((sheet, nd.isoformat() if nd else "无邮件"))
    return out

def compose(run):
    written = run.get("written", [])
    held = run.get("held", [])
    preview = run.get("preview", [])
    dormant = run.get("dormant", [])
    excluded = run.get("excluded", [])
    deferred = run.get("deferred", [])
    blank_idx = run.get("blank_idx", [])
    anomalies = run.get("anomalies", [])
    cfg = L.CFG.get("notify", {})
    stats = workbook_stats()
    # forwarding-staleness: skip confirmed-inactive, this-run dormant, excluded
    skip = CONFIRMED_INACTIVE | {d[0] for d in dormant} | set(excluded)
    stale = staleness(skip, cfg.get("stale_days", 12))
    new_dormant = [(s, nd) for (s, nd) in dormant if s not in CONFIRMED_INACTIVE]

    # latest data day + how many products are updated to that latest week —
    # computed from the workbook so the digest reflects the whole period, not
    # just this single run's writes (we may run several times a week).
    latest_days = [v["latest"] for v in stats.values() if v.get("latest")]
    data_day = max(latest_days) if latest_days else None
    n_upd = 0
    if data_day:
        wk = datetime.date.fromisoformat(data_day).isocalendar()[:2]
        n_upd = sum(1 for v in stats.values()
                    if v.get("latest") and datetime.date.fromisoformat(v["latest"]).isocalendar()[:2] == wk)

    L_ = []
    a = L_.append
    a("本期私募净值已自动更新，整体情况汇报如下。")
    a("")

    # 一、本期更新
    a("【本期净值更新】")
    if data_day:
        a(f"　最新一期净值已更新，数据日 {data_day}，本期共 {n_upd} 只产品净值到位。")
    else:
        a("　暂无最新净值。")
    if held:
        a(f"　另有 {len(held)} 只因托管方净值尚未送达，已暂缓，将于下次运行自动补录，不影响已更新部分。")
    a("")

    # 表现排名（四档）
    if stats:
        wk = [(s, v["weekly"]) for s, v in stats.items() if v.get("weekly") is not None]
        ups = sum(1 for _, r in wk if r > 0)
        downs = sum(1 for _, r in wk if r < 0)
        a(f"【近一周表现】纳入统计 {len(wk)} 只：上涨 {ups} 只，下跌 {downs} 只")
        wk.sort(key=lambda x: x[1], reverse=True)
        a("　居前：" + "，".join(f"{s} {pct(r)}" for s, r in wk[:3]))
        a("　靠后：" + "，".join(f"{s} {pct(r)}" for s, r in wk[-3:][::-1]))
        a("")
        for key, title in [("m1", "近一月表现"), ("m3", "近三月表现"), ("since", "记录以来累计表现")]:
            blk = rank_block(stats, key, title)
            if blk:
                a("\n".join(blk)); a("")

    # 二、需关注（人工跟进）
    attn = []
    if anomalies:
        attn.append(f"　· 数据异常待核对 {len(anomalies)} 条：" +
                    "，".join(f"{x.get('sheet')}({x.get('date')} {x.get('flag')})" for x in anomalies[:8]))
    if stale:
        attn.append("　· ⚠️ 转发可能掉线/停发（活跃产品超时未收到邮件，请检查 QQ 自动转发）：" +
                    "，".join(f"{s}(最新{nd})" for s, nd in stale))
    if new_dormant:
        attn.append("　· 疑似已赎回/停更（请确认）：" + "，".join(f"{s}(最新邮件{nd})" for s, nd in new_dormant))
    if excluded:
        attn.append("　· 来源口径待确认、暂未自动更新：" + "，".join(excluded))
    if "validation" in run:
        v = run["validation"]
        if v.get("diffs"):
            attn.append(f"　· 回归校验 {v.get('matched')}/{v.get('total')} 一致，其中 {len(v['diffs'])} 处需关注")
    a("【需关注】")
    if attn:
        L_.extend(attn)
    else:
        a("　本周运行正常，无异常事项。")
    a("")

    # 三、系统备注（操作项，低优先）
    notes = []
    if held:
        notes.append("　· 待本周五净值（下次自动补）：" + "，".join(f"{x['sheet']}({x['date']})" for x in held[:12]))
    if blank_idx:
        notes.append("　· 已写净值、指数/对比列(G)需手工补：" + "，".join(s for s, _ in blank_idx))
    if deferred:
        notes.append(f"　· 待指数收盘暂缓 {len(deferred)} 行（下次自动补）")
    if "validation" in run and not run["validation"].get("diffs"):
        v = run["validation"]
        notes.append(f"　· 回归校验：{v.get('matched')}/{v.get('total')} 与历史记录一致，全部通过 ✅")
    if notes:
        a("【系统备注】")
        L_.extend(notes)
        a("")

    a(f"——本周报由系统自动生成于 {run.get('time','')}")

    n_attn = len(attn)
    subject = f"【私募净值周报】{datetime.date.today()} 在册{n_upd}只" + (f" 待办{n_attn}" if n_attn else " 运行正常")
    return subject, "\n".join(L_)

def send(subject, body):
    cfg = L.CFG.get("notify", {})
    user = L.CFG["imap"]["user"]
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = Header(subject, "utf-8")
    msg["From"] = user
    msg["To"] = ",".join(cfg.get("to", [user]))
    s = smtplib.SMTP_SSL(cfg.get("smtp_host", "smtp.qq.com"), cfg.get("smtp_port", 465), timeout=30)
    s.login(user, L.get_password())
    s.sendmail(user, cfg.get("to", [user]), msg.as_string())
    s.quit()

def main():
    run = _load("last_run.json")
    if not run:
        print("无 last_run.json，跳过通知"); return
    v = _load("validation.json")
    if v:
        try:
            age = (datetime.datetime.now() - datetime.datetime.fromisoformat(v["time"])).days
            if age <= 2:
                run["validation"] = v
        except Exception:
            pass
    subject, body = compose(run)
    if "--dry" in sys.argv or not L.CFG.get("notify", {}).get("enabled", True):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
        print("SUBJECT:", subject); print("-" * 60); print(body)
        return
    try:
        send(subject, body)
        print("已发送汇总邮件:", subject)
    except Exception as e:
        print("发送失败:", type(e).__name__, str(e)[:200])

if __name__ == "__main__":
    main()
