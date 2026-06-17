# -*- coding: utf-8 -*-
"""Fetch benchmark index daily closes into index_cache.json, using the free
Eastmoney API (index_source.py; only needs `requests`). Self-validating: before
caching an index it checks the closes reproduce the values Linda already recorded
in the sheet's G column; if they don't match, that index is NOT cached (so NAV
rows just defer instead of getting bad data). No futu/OpenD required.
"""
import os, json, re, datetime
import openpyxl
import navlib as L
import index_source as IS

TOL = 0.003  # relative tolerance vs recorded index values
IDX_RE = re.compile(r"中证\s*(500|1000|2000)|北证[^\s]*")

def s2d(v):
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
    if isinstance(v, datetime.datetime):
        return v.date()
    return v

def main():
    reg = L.load_registry()
    wb = openpyxl.load_workbook(L.CFG["master_path"], data_only=True)

    sheet_index, recorded = {}, {}
    for sheet in L.CFG["scope_sheets"]:
        if sheet not in wb.sheetnames:
            continue
        info = reg[sheet]; ws = wb[sheet]
        if ws.max_column < 7:
            continue
        g = ws.cell(info["header_row"] or 1, 7).value
        if not (g and IDX_RE.search(str(g))):
            continue
        name = str(g).strip()
        sheet_index[sheet] = name
        for r in range(info["data_start"], ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a in (None, "", "累计"):
                continue
            d = s2d(ws.cell(r, 5).value); gv = ws.cell(r, 7).value
            if d and isinstance(gv, (int, float)):
                recorded.setdefault(name, []).append((d, gv))

    needed = sorted(set(sheet_index.values()))
    print("需要的指数:", needed)
    closes_map, rejected = {}, []
    for name in needed:
        closes, secid, _ = IS.fetch_closes(name, start="20250101")
        if not closes:
            print(f"  [{name}] 拉取失败（网络？），本次不缓存 -> 相关行将暂缓重试")
            continue
        samples = recorded.get(name, [])[-8:]
        ok = bad = 0
        for d, gv in samples:
            f = closes.get(d.isoformat())
            if f is None:
                continue
            if abs(f - gv) <= max(abs(gv) * TOL, 0.5):
                ok += 1
            else:
                bad += 1
                if bad <= 2:
                    print(f"     校验不符 {name} {d}: 记录={gv} 接口={f}")
        if ok >= 1 and bad == 0:
            closes_map[name] = closes
            print(f"  [{name}] {secid} 校验通过 ({ok} 样本) -> 缓存 {len(closes)} 天")
        else:
            rejected.append(name)
            print(f"  [{name}] {secid} 校验失败 (ok={ok} bad={bad}) -> 标记为无法获取（该列留空+标注）")

    cache = {"built": datetime.datetime.now().isoformat(timespec="seconds"),
             "closes": closes_map, "rejected": rejected}
    json.dump(cache, open(os.path.join(L.HERE, "index_cache.json"), "w", encoding="utf-8"),
              ensure_ascii=False)
    print("已写 index_cache.json | 已校验:", list(closes_map.keys()), "| 无法获取:", rejected)

if __name__ == "__main__":
    main()
