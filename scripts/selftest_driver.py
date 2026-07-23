#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import gc
import io
import json
import shutil
import subprocess
import sys
import zipfile
from copy import deepcopy
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.formula import ArrayFormula
from pypdf import PdfWriter


def check(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def create_book(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Demo Fund"
    sheet.append(
        [
            "NAV Date",
            "Fund Return",
            "Product Name",
            "Unit NAV",
            "Product Code",
            "Cumulative NAV",
            "Audit Formula",
            "Benchmark Return",
            "Excess",
            "Benchmark Level",
        ]
    )
    sheet.append(
        [
            dt.date(2026, 1, 2),
            None,
            "Example Fund",
            1.0,
            "DEMO01",
            1.0,
            "=D2*2",
            None,
            None,
            100.0,
        ]
    )
    sheet.append(
        [
            dt.date(2026, 1, 9),
            "=F3/F2-1",
            "Example Fund",
            1.01,
            "DEMO01",
            1.01,
            "=D3*2",
            0.005,
            "=B3-H3",
            101.0,
        ]
    )
    sheet.append(
        [
            "TOTAL",
            "=F3/F2-1",
            None,
            None,
            None,
            None,
            None,
            "=PRODUCT(1+H2:H3)-1",
            "=B4-H4",
            None,
        ]
    )
    for row in (2, 3, 4):
        for column in (2, 8, 9):
            sheet.cell(row, column).font = Font(
                name="Arial", size=10, bold=True, color="00C00000"
            )
            sheet.cell(row, column).number_format = "0.00%"
    benchmark = workbook.create_sheet("Demo Benchmark")
    benchmark.append(["Date", "Benchmark Return"])
    for date, value in (
        (dt.date(2026, 1, 2), 0.0),
        (dt.date(2026, 1, 9), 0.005),
        (dt.date(2026, 1, 16), -0.002),
        (dt.date(2026, 1, 23), 0.003),
    ):
        benchmark.append([date, value])
    workbook.save(path)


def config_for(runtime: Path, book: Path) -> dict:
    return {
        "schema_version": 1,
        "runtime_id": "00000000-0000-4000-8000-000000000001",
        "workbook_path": str(book.resolve()),
        "imap": {
            "host": "imap.example.invalid",
            "port": 993,
            "user": "user@example.invalid",
            "mailbox": "INBOX",
            "lookback_days": 180,
        },
        "routes": [
            {
                "sender": "sender@example.invalid",
                "sheet": "Demo Fund",
                "code": "DEMO01",
                "parser": "auto",
                "allow_sender_only": False,
                "cumulative_policy": "require",
                "cumulative_offset": None,
                "return_basis": "cumulative",
                "return_frequency": "weekly",
                "data_frequency": "auto",
                "series_start": "2026-01-02",
                "benchmark": {
                    "source_sheet": "Demo Benchmark",
                    "source_type": "aligned_return",
                    "source_date": "A",
                    "source_value": "B",
                },
            }
        ],
        "column_overrides": {},
        "style": {"mode": "infer", "zero_threshold": 0.00005},
        "schedule": [],
        "validation": {"minimum_history_dates": 2, "tolerance": 0.000001},
    }


def parser_tests(runtime: Path) -> None:
    sys.path.insert(0, str(runtime))
    import nav_mail
    import nav_parse
    import nav_products
    import navctl
    from nav_parse import (
        ParseError,
        choose_route_rows,
        parse_number,
        rows_from_html,
        rows_from_message,
        rows_from_text,
    )
    from nav_mail import (
        MailError,
        exact_from_matches,
        fetch_authorized_messages,
        fetch_candidate_headers,
        fetch_candidate_messages,
        imap_date,
        needs_imap_id,
        non_nav_subject_category,
        single_from_address,
    )
    from runtime_secret import (
        MASKED_INPUT_PROMPT,
        SecretInputCancelled,
        _read_interactive_secret,
        _read_masked,
    )

    utf8_args_path = runtime / "utf8-args.txt"
    utf8_args_path.write_text(
        "\ufeff# UTF-8 参数文件\n"
        "propose\n"
        "--sender\n"
        "sender@example.invalid\n"
        "--subject-contains\n"
        "示例产品净值通知\n",
        encoding="utf-8",
    )
    expanded_args, used_argfile = navctl.expand_utf8_argfiles(
        [f"@{utf8_args_path}"]
    )
    check(
        used_argfile
        and expanded_args
        == [
            "propose",
            "--sender",
            "sender@example.invalid",
            "--subject-contains",
            "示例产品净值通知",
        ],
        "UTF-8 argument file did not preserve Chinese filter text",
    )
    navctl.validate_filter_encoding("示例产品净值通知", "--subject-contains")
    mojibake_filter = "净值通知".encode("utf-8").decode("latin1")
    try:
        navctl.validate_filter_encoding(
            mojibake_filter, "--subject-contains"
        )
    except ValueError as exc:
        check(
            "疑似被终端错误解码" in str(exc)
            and "UTF-8 参数文件" in str(exc),
            "mojibake filter returned the wrong diagnostic",
        )
    else:
        raise AssertionError("mojibake filter was accepted")
    invalid_args_path = runtime / "invalid-args.txt"
    invalid_args_path.write_bytes(b"\xff")
    try:
        navctl.expand_utf8_argfiles([f"@{invalid_args_path}"])
    except ValueError as exc:
        check(
            "必须使用 UTF-8" in str(exc),
            "non-UTF-8 argument file returned the wrong diagnostic",
        )
    else:
        raise AssertionError("non-UTF-8 argument file was accepted")

    coverage_config = {
        "routes": [
            {"sheet": "活动产品"},
            {
                "sheet": "基准待审产品",
                "benchmark_review_only": True,
            },
            {"sheet": "暂停产品", "paused": True},
        ],
        "sheet_reviews": {
            "归档产品": {"status": "excluded", "reason": "已完成本地核实"},
            "解析待补": {
                "status": "local_parser_required",
                "reason": "正式通知需要专用解析器",
            },
            "业务待核": {
                "status": "business_review",
                "reason": "持续来源尚未确认",
            },
            "无邮件证据": {
                "status": "no_mail_evidence",
                "reason": "有界回看期无代码或名称命中",
            },
        },
    }
    coverage = nav_products._coverage_matrix(
        coverage_config,
        [
            "活动产品",
            "基准待审产品",
            "暂停产品",
            "指数参考",
            "归档产品",
            "解析待补",
            "业务待核",
            "无邮件证据",
            "未分类页",
        ],
        {"活动产品", "基准待审产品", "暂停产品"},
        {"指数参考"},
    )
    check(
        coverage["total_workbook_sheets"] == 9
        and coverage["classified_sheets"] == 8
        and not coverage["all_sheets_classified"]
        and coverage["action_required_sheets"] == 4
        and coverage["counts"]["active"] == 1
        and coverage["counts"]["active_review_required"] == 1
        and coverage["counts"]["reference"] == 1
        and coverage["counts"]["excluded"] == 1
        and coverage["counts"]["no_mail_evidence"] == 1
        and coverage["counts"]["local_parser_required"] == 1
        and coverage["counts"]["business_review"] == 1
        and coverage["counts"]["unclassified"] == 1,
        "workbook coverage matrix did not classify every sheet outcome",
    )

    try:
        _read_interactive_secret(
            lambda: (_ for _ in ()).throw(
                AssertionError("non-terminal input reader was called")
            ),
            io.StringIO(),
            io.StringIO(),
        )
    except RuntimeError as exc:
        check(
            str(exc) == "此命令需要用户在真实终端中运行",
            "non-terminal secret input returned the wrong error",
        )
    else:
        raise AssertionError("non-terminal secret input was allowed to wait for input")

    class FakeTerminal:
        def isatty(self):
            return True

    cancelled_keys = iter(["a", "\x03"])
    cancelled_mask = io.StringIO()
    try:
        _read_interactive_secret(
            lambda: next(cancelled_keys), FakeTerminal(), cancelled_mask
        )
    except SecretInputCancelled:
        pass
    else:
        raise AssertionError("Ctrl+C did not cancel interactive secret input")
    check(
        cancelled_mask.getvalue().endswith("*\n"),
        "Ctrl+C did not finish the masked input line cleanly",
    )

    keys = iter(["a", "b", "\b", "C", "\x00", "K", "\r"])
    masked_output = io.StringIO()
    masked_value = _read_masked(lambda: next(keys), masked_output)
    check(
        masked_value == "aC"
        and masked_output.getvalue().startswith(f"{MASKED_INPUT_PROMPT}\n")
        and masked_output.getvalue().endswith("**\b \b*\n"),
        "Windows masked secret input did not echo stars or handle backspace",
    )
    original_set_password = navctl.set_password
    navctl.set_password = lambda _runtime_id: runtime / "secret.json"
    secret_stdout = io.StringIO()
    secret_stderr = io.StringIO()
    original_stdout, original_stderr = sys.stdout, sys.stderr
    try:
        sys.stdout, sys.stderr = secret_stdout, secret_stderr
        result = navctl.command_secret(
            {"runtime_id": "00000000-0000-4000-8000-000000000001"},
            argparse.Namespace(secret_action="set"),
        )
    finally:
        sys.stdout, sys.stderr = original_stdout, original_stderr
        navctl.set_password = original_set_password
    check(
        result == 0
        and secret_stderr.getvalue() == "已加密保存\n"
        and json.loads(secret_stdout.getvalue())["stored"],
        "secret set did not print the encrypted-save confirmation",
    )
    original_launch_secret_prompt = navctl.launch_secret_prompt
    navctl.launch_secret_prompt = lambda: 4321
    launch_stdout = io.StringIO()
    try:
        sys.stdout = launch_stdout
        launched = navctl.command_secret(
            {"runtime_id": "00000000-0000-4000-8000-000000000001"},
            argparse.Namespace(secret_action="launch"),
        )
    finally:
        sys.stdout = original_stdout
        navctl.launch_secret_prompt = original_launch_secret_prompt
    launch_report = json.loads(launch_stdout.getvalue())
    check(
        launched == 0
        and launch_report["launched"]
        and launch_report["process_id"] == 4321
        and "secret status" in launch_report["next"],
        "secret launch did not open a dedicated prompt and require verification",
    )
    navctl.set_password = lambda _runtime_id: (_ for _ in ()).throw(
        SecretInputCancelled
    )
    cancelled_stdout = io.StringIO()
    cancelled_stderr = io.StringIO()
    original_stdout, original_stderr = sys.stdout, sys.stderr
    try:
        sys.stdout, sys.stderr = cancelled_stdout, cancelled_stderr
        cancelled = navctl.command_secret(
            {"runtime_id": "00000000-0000-4000-8000-000000000001"},
            argparse.Namespace(secret_action="set"),
        )
    finally:
        sys.stdout, sys.stderr = original_stdout, original_stderr
        navctl.set_password = original_set_password
    check(
        cancelled == 130
        and cancelled_stderr.getvalue() == ""
        and json.loads(cancelled_stdout.getvalue())
        == {"passed": False, "error": "已取消"},
        "Ctrl+C during secret input did not exit cleanly",
    )

    check(
        imap_date(dt.date(2026, 7, 9)) == "09-Jul-2026",
        "IMAP dates must use fixed English month names",
    )
    check(
        needs_imap_id("IMAP.163.COM.")
        and needs_imap_id("imap.126.com")
        and needs_imap_id("imap.yeah.net")
        and not needs_imap_id("imap.qq.com"),
        "NetEase IMAP host detection is incorrect",
    )
    check(
        non_nav_subject_category("示例产品季报") == "periodic-report"
        and non_nav_subject_category("示例产品季报净值通知") is None
        and non_nav_subject_category("Example quarterly report NAV") is None,
        "built-in report exclusions swallowed a subject with explicit NAV semantics",
    )

    original_ssl = nav_mail.imaplib.IMAP4_SSL
    original_read_password = nav_mail.read_password
    events: list[tuple[Any, ...]] = []

    class FakeIMAP:
        id_status = "OK"

        def __init__(self, host, port, **_kwargs):
            events.append(("connect", host, port))

        def login(self, user, _password):
            events.append(("login", user))
            return "OK", [b""]

        def xatom(self, name, payload):
            events.append((name, payload))
            return self.id_status, [b""]

        def select(self, mailbox, readonly=False):
            events.append(("select", mailbox, readonly))
            return "OK", [b""]

        def logout(self):
            events.append(("logout",))
            return "BYE", [b""]

    mail_config = {
        "runtime_id": "00000000-0000-4000-8000-000000000001",
        "imap": {
            "host": "imap.163.com",
            "port": 993,
            "user": "user@example.invalid",
            "mailbox": "INBOX",
        },
    }
    try:
        nav_mail.imaplib.IMAP4_SSL = FakeIMAP
        nav_mail.read_password = lambda _runtime_id: "fixture-secret"
        nav_mail.connect(mail_config)
        check(
            [event[0] for event in events] == ["connect", "login", "ID", "select"],
            "NetEase IMAP ID must run after login and before mailbox selection",
        )
        check(
            "user@example.invalid" not in str(events[2])
            and "fixture-secret" not in str(events[2]),
            "IMAP ID leaked account information",
        )

        events.clear()
        mail_config["imap"]["host"] = "imap.qq.com"
        nav_mail.connect(mail_config)
        check(
            [event[0] for event in events] == ["connect", "login", "select"],
            "Non-NetEase IMAP received an unnecessary ID command",
        )

        events.clear()
        mail_config["imap"]["host"] = "imap.126.com"
        FakeIMAP.id_status = "BAD"
        try:
            nav_mail.connect(mail_config)
        except MailError as exc:
            check(
                "ID handshake" in str(exc),
                "Rejected NetEase IMAP ID did not return a specific error",
            )
        else:
            raise AssertionError("Rejected NetEase IMAP ID was accepted")
        check(events[-1][0] == "logout", "Failed IMAP connection was not closed")
    finally:
        FakeIMAP.id_status = "OK"
        nav_mail.imaplib.IMAP4_SSL = original_ssl
        nav_mail.read_password = original_read_password

    sender_message = EmailMessage()
    sender_message["From"] = "Example Sender <sender@example.invalid>"
    check(
        exact_from_matches(sender_message, "sender@example.invalid"),
        "exact From matching rejected a display name",
    )
    spoofed_message = EmailMessage()
    spoofed_message["From"] = "other@example.invalid"
    check(
        not exact_from_matches(spoofed_message, "sender@example.invalid"),
        "substring From matching was accepted",
    )
    candidate_message = EmailMessage()
    candidate_message["From"] = "NAV Desk <sender@example.invalid>"
    candidate_message["Subject"] = "NAV update"
    candidate_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-01-09 | 1.01 | 1.01"
    )
    candidate_payload = candidate_message.as_bytes()
    candidate_header_payload = (
        b"From: NAV Desk <sender@example.invalid>\r\nSubject: NAV update\r\n\r\n"
    )
    candidate_calls: list[tuple[str, tuple[Any, ...]]] = []

    class CandidateIMAP:
        def uid(self, command, *_args):
            candidate_calls.append((command, _args))
            if command == "search":
                return "OK", [b"1 2"]
            query = str(_args[-1])
            if "HEADER.FIELDS" in query:
                requested = _args[0].split(b",")
                return "OK", [
                    (
                        uid + b" (UID " + uid + b" BODY[HEADER] {1})",
                        candidate_header_payload,
                    )
                    for uid in requested
                ]
            if "RFC822.SIZE" in query:
                return "OK", [
                    f"1 (UID 1 RFC822.SIZE {len(candidate_payload)})".encode(),
                    f"2 (UID 2 RFC822.SIZE {len(candidate_payload)})".encode(),
                ]
            uid = _args[0]
            return "OK", [(uid + b" (BODY[])", candidate_payload), b")"]

        def close(self):
            return "OK", [b""]

        def logout(self):
            return "BYE", [b""]

    original_connect = nav_mail.connect
    nav_mail.connect = lambda _config: CandidateIMAP()
    try:
        candidates, scan = fetch_candidate_messages(mail_config)
    finally:
        nav_mail.connect = original_connect
    check(
        len(candidates) == 2
        and single_from_address(candidates[0]) == "sender@example.invalid"
        and scan["messages_fetched"] == 2
        and len(
            [
                call
                for call in candidate_calls
                if call[0] == "fetch" and "RFC822.SIZE" in str(call[1][-1])
            ]
        )
        == 1,
        "mailbox-wide sender discovery did not return the local NAV candidate",
    )
    header_call_index = next(
        index
        for index, call in enumerate(candidate_calls)
        if call[0] == "fetch" and "HEADER.FIELDS" in str(call[1][-1])
    )
    size_call_index = next(
        index
        for index, call in enumerate(candidate_calls)
        if call[0] == "fetch" and "RFC822.SIZE" in str(call[1][-1])
    )
    check(
        header_call_index < size_call_index,
        "candidate discovery downloaded size/body data before minimum headers",
    )
    candidate_calls.clear()
    nav_mail.connect = lambda _config: CandidateIMAP()
    try:
        headers, header_scan = fetch_candidate_headers(mail_config, limit=1)
    finally:
        nav_mail.connect = original_connect
    check(
        len(headers) == 1
        and header_scan["headers_fetched"] == 1
        and header_scan["messages_fetched"] == 0
        and not any(
            call[0] == "fetch"
            and (
                "RFC822.SIZE" in str(call[1][-1]) or str(call[1][-1]) == "(BODY.PEEK[])"
            )
            for call in candidate_calls
        ),
        "header-only proposal scan downloaded full message data",
    )

    routed_payloads: dict[bytes, bytes] = {}
    routed_headers: dict[bytes, bytes] = {}
    for uid, subject in (
        (b"1", "TARGET weekly NAV"),
        (b"2", "quarterly report"),
        (b"3", "TARGET valuation"),
        (b"4", "another unrelated notice"),
    ):
        routed = EmailMessage()
        routed["From"] = "NAV Desk <sender@example.invalid>"
        routed["Subject"] = subject
        routed.set_content(
            "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
            f"DEMO01 | 2026-01-{int(uid) + 10:02d} | 1.0{int(uid)} | 1.0{int(uid)}"
        )
        routed_payloads[uid] = routed.as_bytes()
        routed_headers[uid] = (
            f"From: NAV Desk <sender@example.invalid>\r\nSubject: {subject}\r\n\r\n"
        ).encode("ascii")
    routed_calls: list[tuple[str, tuple[Any, ...]]] = []

    class RoutedIMAP:
        def uid(self, command, *_args):
            routed_calls.append((command, _args))
            if command == "search":
                query = str(_args[-1]).upper()
                if "FROM" in query and "SUBJECT" in query:
                    return "OK", [b"1 3"]
                if "FROM" in query:
                    return "OK", [b"1 2 3 4"]
                return "OK", [b" ".join(str(uid).encode("ascii") for uid in range(1, 3005))]
            request = str(_args[-1])
            requested = _args[0].split(b",")
            if "HEADER.FIELDS" in request:
                return "OK", [
                    (uid + b" (UID " + uid + b" BODY[HEADER] {1})", routed_headers[uid])
                    for uid in requested
                ]
            if "RFC822.SIZE" in request:
                return "OK", [
                    uid
                    + b" (UID "
                    + uid
                    + b" RFC822.SIZE "
                    + str(len(routed_payloads[uid])).encode("ascii")
                    + b")"
                    for uid in requested
                ]
            uid = _args[0]
            return "OK", [(uid + b" (BODY[])", routed_payloads[uid]), b")"]

        def close(self):
            return "OK", [b""]

        def logout(self):
            return "BYE", [b""]

    routed_config = {
        "runtime_id": "00000000-0000-4000-8000-000000000001",
        "imap": {
            "host": "imap.example.invalid",
            "port": 993,
            "user": "user@example.invalid",
            "mailbox": "INBOX",
            "lookback_days": 30,
            "max_messages": 2,
            "max_message_bytes": 1024 * 1024,
            "max_total_bytes": 2 * 1024 * 1024,
        },
        "routes": [
            {
                "sender": "sender@example.invalid",
                "subject_contains": "TARGET",
                "sheet": "Demo Fund",
                "code": "DEMO01",
            }
        ],
    }
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        routed_messages = fetch_authorized_messages(routed_config)
    finally:
        nav_mail.connect = original_connect
    full_fetch_uids = [
        call[1][0]
        for call in routed_calls
        if call[0] == "fetch" and str(call[1][-1]) == "(BODY.PEEK[])"
    ]
    check(
        len(routed_messages["sender@example.invalid"]) == 2
        and full_fetch_uids == [b"1", b"3"],
        "authorized sender mail was not header-filtered before message limits and full download",
    )
    check(
        [
            getattr(message, "_nav_source_uid", None)
            for message in routed_messages["sender@example.invalid"]
        ]
        == [1, 3],
        "authorized messages did not retain mailbox order for correction auditing",
    )
    transient_disconnects = {"remaining": 1, "connections": 0}

    class ReconnectingRoutedIMAP(RoutedIMAP):
        def __init__(self):
            transient_disconnects["connections"] += 1

        def uid(self, command, *_args):
            if (
                command == "fetch"
                and str(_args[-1]) == "(BODY.PEEK[])"
                and transient_disconnects["remaining"]
            ):
                transient_disconnects["remaining"] -= 1
                raise nav_mail.imaplib.IMAP4.abort(
                    "fixture transient disconnect"
                )
            return super().uid(command, *_args)

    routed_calls.clear()
    nav_mail.connect = lambda _config: ReconnectingRoutedIMAP()
    try:
        reconnected_messages = fetch_authorized_messages(routed_config)
    finally:
        nav_mail.connect = original_connect
    check(
        len(reconnected_messages["sender@example.invalid"]) == 2
        and reconnected_messages.reconnect_count == 1
        and transient_disconnects["connections"] == 2
        and [
            getattr(message, "_nav_source_uid", None)
            for message in reconnected_messages["sender@example.invalid"]
        ]
        == [1, 3],
        "authorized discovery did not reconnect and resume at the current UID",
    )

    changed_uidvalidity_connections = {"count": 0}

    class ChangedUIDValidityIMAP(RoutedIMAP):
        def __init__(self):
            changed_uidvalidity_connections["count"] += 1
            self.connection_number = changed_uidvalidity_connections["count"]

        def response(self, _name):
            return "UIDVALIDITY", [
                str(100 + self.connection_number).encode("ascii")
            ]

        def uid(self, command, *_args):
            if (
                self.connection_number == 1
                and command == "fetch"
                and str(_args[-1]) == "(BODY.PEEK[])"
            ):
                raise nav_mail.imaplib.IMAP4.abort(
                    "fixture mailbox identity changed"
                )
            return super().uid(command, *_args)

    nav_mail.connect = lambda _config: ChangedUIDValidityIMAP()
    try:
        fetch_authorized_messages(routed_config)
    except MailError as exc:
        check(
            "UIDVALIDITY" in str(exc)
            and "fixture mailbox identity changed" not in str(exc),
            "changed mailbox identity returned the wrong reconnect error",
        )
    else:
        raise AssertionError(
            "authorized discovery resumed after UIDVALIDITY changed"
        )
    finally:
        nav_mail.connect = original_connect
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        routed_scope = fetch_authorized_messages(
            routed_config, load_bodies=False
        )
    finally:
        nav_mail.connect = original_connect
    check(
        routed_scope.scope_fingerprint
        == routed_messages.scope_fingerprint
        and routed_scope.messages_selected == 2
        and not any(
            call[0] == "fetch"
            and str(call[1][-1]) == "(BODY.PEEK[])"
            for call in routed_calls
        ),
        "lightweight mailbox scope verification downloaded full message bodies",
    )
    excluded_config = json.loads(json.dumps(routed_config))
    excluded_config["routes"][0]["subject_excludes"] = ["valuation"]
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        excluded_messages = fetch_authorized_messages(excluded_config)
    finally:
        nav_mail.connect = original_connect
    excluded_full_fetch_uids = [
        call[1][0]
        for call in routed_calls
        if call[0] == "fetch" and str(call[1][-1]) == "(BODY.PEEK[])"
    ]
    check(
        len(excluded_messages["sender@example.invalid"]) == 1
        and excluded_full_fetch_uids == [b"1"]
        and excluded_messages.excluded_non_nav_messages == 1
        and excluded_messages.excluded_non_nav_reasons
        == {"configured-subject-exclude": 1},
        "confirmed subject exclusions were not applied before full message download",
    )
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        selected_candidates, selected_scan = fetch_candidate_messages(
            routed_config,
            sender="sender@example.invalid",
            subject_contains="TARGET",
        )
    finally:
        nav_mail.connect = original_connect
    selected_full_fetch_uids = [
        call[1][0]
        for call in routed_calls
        if call[0] == "fetch" and str(call[1][-1]) == "(BODY.PEEK[])"
    ]
    check(
        len(selected_candidates) == 2
        and selected_scan["headers_fetched"] == 2
        and selected_scan["server_since_matches"] == 3004
        and selected_scan["server_sender_matches"] == 2
        and selected_scan["server_subject_filter_applied"]
        and selected_scan["messages_selected"] == 2
        and selected_full_fetch_uids == [b"3", b"1"],
        "selected proposal scan did not narrow at the server before exact local verification",
    )
    import nav_service

    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        first_chunk = nav_service.propose_routes(
            routed_config,
            sender="sender@example.invalid",
            subject_contains="TARGET",
            lookback_days=30,
            batch_messages=1,
            time_budget_seconds=120,
        )
        from nav_products import ProductError, sync as sync_products

        try:
            sync_products(routed_config, refresh=False)
        except ProductError as exc:
            check(
                "尚未完成" in str(exc),
                "partial proposal blocked products with the wrong error",
            )
        else:
            raise AssertionError("products accepted a stale proposal during resume")
        second_chunk = nav_service.propose_routes(
            routed_config,
            batch_messages=1,
            time_budget_seconds=120,
            resume=True,
        )
    finally:
        nav_mail.connect = original_connect
    progress_report = json.loads(
        (runtime / "route-proposal-progress.json").read_text(encoding="utf-8")
    )
    check(
        not first_chunk["passed"]
        and first_chunk["partial"]
        and first_chunk["resume_available"]
        and second_chunk["passed"]
        and not second_chunk["partial"]
        and second_chunk["scan"]["chunks_completed"] == 2
        and len(second_chunk["candidates"][0]["observations"]) == 2
        and progress_report["status"] == "complete"
        and not (runtime / "route-proposals.partial.json").exists(),
        "bounded proposal scan did not publish progress or resume from its UID cursor",
    )
    unfiltered_config = json.loads(json.dumps(routed_config))
    unfiltered_config["routes"][0].pop("subject_contains")
    unfiltered_config["imap"]["max_messages"] = 4
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        unfiltered_messages = fetch_authorized_messages(unfiltered_config)
    finally:
        nav_mail.connect = original_connect
    check(
        len(unfiltered_messages["sender@example.invalid"]) == 3
        and unfiltered_messages.excluded_non_nav_messages == 1
        and unfiltered_messages.excluded_non_nav_reasons == {"periodic-report": 1},
        "built-in non-NAV exclusions did not report a controlled reason before download",
    )
    snapshot_config = json.loads(json.dumps(routed_config))
    snapshot_config["routes"][0]["max_staleness_days"] = 366
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        snapshot_rows, snapshot_report = nav_service.collect_route_rows(
            snapshot_config
        )
    finally:
        nav_mail.connect = original_connect
    reused_config = json.loads(json.dumps(snapshot_config))
    reused_config["routes"][0]["benchmark_review_only"] = True
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        reused_rows, reused_report = nav_service.reuse_discovery_snapshot(
            reused_config
        )
    finally:
        nav_mail.connect = original_connect
    check(
        snapshot_report["passed"]
        and len(snapshot_rows["Demo Fund"]) == 2
        and len(reused_rows["Demo Fund"]) == 2
        and reused_report["discovery_reused"]
        and reused_report["mail_scope_reverified"]
        and not any(
            call[0] == "fetch"
            and str(call[1][-1]) == "(BODY.PEEK[])"
            for call in routed_calls
        ),
        "verified discovery snapshot was not reused without full body downloads",
    )
    original_scope_header = routed_headers[b"3"]
    routed_headers[b"3"] = (
        b"From: NAV Desk <sender@example.invalid>\r\n"
        b"Subject: TARGET valuation revised\r\n\r\n"
    )
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        try:
            nav_service.reuse_discovery_snapshot(reused_config)
        except RuntimeError as exc:
            check(
                "邮箱范围" in str(exc),
                "changed mailbox scope rejected the snapshot for the wrong reason",
            )
        else:
            raise AssertionError(
                "changed mailbox scope reused a stale discovery snapshot"
            )
    finally:
        nav_mail.connect = original_connect
        routed_headers[b"3"] = original_scope_header
    changed_snapshot_config = json.loads(json.dumps(snapshot_config))
    changed_snapshot_config["routes"][0]["subject_contains"] = "CHANGED"
    try:
        nav_service.reuse_discovery_snapshot(changed_snapshot_config)
    except RuntimeError as exc:
        check(
            "路由" in str(exc),
            "changed routing rules rejected the snapshot for the wrong reason",
        )
    else:
        raise AssertionError("changed routing rules reused a stale discovery snapshot")
    bounded_config = json.loads(json.dumps(routed_config))
    bounded_config["routes"][0].pop("subject_contains")
    bounded_config["imap"]["max_header_messages"] = 3
    bounded_config["imap"]["max_messages"] = 2
    routed_calls.clear()
    nav_mail.connect = lambda _config: RoutedIMAP()
    try:
        fetch_authorized_messages(bounded_config)
    except MailError as exc:
        check(
            "邮件头扫描边界" in str(exc)
            and not any(
                call[0] == "fetch" and "HEADER.FIELDS" in str(call[1][-1])
                for call in routed_calls
            ),
            "oversized header scan did not stop before downloading message headers",
        )
    else:
        raise AssertionError("mail header scan limit was not enforced")
    finally:
        nav_mail.connect = original_connect

    class DisconnectingIMAP(CandidateIMAP):
        def uid(self, command, *_args):
            if command == "fetch" and "BODY.PEEK" in str(_args[-1]):
                raise nav_mail.imaplib.IMAP4.abort("fixture disconnect")
            return super().uid(command, *_args)

    nav_mail.connect = lambda _config: DisconnectingIMAP()
    try:
        fetch_candidate_messages(mail_config)
    except MailError as exc:
        check(
            "意外断开" in str(exc) and "fixture disconnect" not in str(exc),
            "IMAP disconnect did not return a readable, sanitized stage error",
        )
    else:
        raise AssertionError("IMAP disconnect was not converted to MailError")
    finally:
        nav_mail.connect = original_connect
    check(
        parse_number("1.02%") is None, "percentage text must not be accepted as a NAV"
    )

    text = "Product Code | NAV Date | Unit NAV | Cumulative NAV\nDEMO01 | 2026-01-09 | 1.01 | 1.01"
    rows = rows_from_text(text, "body")
    check(
        len(rows) == 1 and rows[0].code == "DEMO01" and rows[0].unit == 1.01,
        "labelled table parsing failed",
    )

    attachment = openpyxl.Workbook()
    cover = attachment.active
    cover.title = "Cover"
    cover.append(["Cover page"])
    data = attachment.create_sheet("Data")
    for _ in range(10):
        data.append(["Note"])
    data.append(["Product Code", "估值基准日", "Cumulative NAV", "Unit NAV"])
    data.append(["DEMO01", "2026-01-16", 1.02, 1.02])
    buffer = io.BytesIO()
    attachment.save(buffer)
    message = EmailMessage()
    message["Subject"] = "NAV notice"
    message.set_content("See attachment")
    message.add_attachment(
        buffer.getvalue(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="report.xlsx",
    )
    parsed = rows_from_message(message)
    check(
        any(row.date == dt.date(2026, 1, 16) for row in parsed),
        "multi-sheet attachment parsing or 估值基准日 alias recognition failed",
    )
    warning_buffer = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as source_archive:
        with zipfile.ZipFile(warning_buffer, "w") as target_archive:
            for item in source_archive.infolist():
                payload = source_archive.read(item.filename)
                if item.filename == "xl/styles.xml":
                    payload = (
                        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        b'<styleSheet xmlns="http://schemas.openxmlformats.org/'
                        b'spreadsheetml/2006/main"/>'
                    )
                target_archive.writestr(item, payload)
    warning_message = EmailMessage()
    warning_message["From"] = "Warning Desk <warning@example.invalid>"
    warning_message["Subject"] = "WARNING NAV fixture"
    warning_message.set_content("See attachment")
    warning_message.add_attachment(
        warning_buffer.getvalue(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="warning-fixture.xlsx",
    )
    warning_rows = rows_from_message(warning_message)
    library_warnings = nav_parse.consume_parse_library_warnings()
    check(
        warning_rows
        and library_warnings
        and library_warnings[0] == {
            "code": "parser-library-warning",
            "library": "openpyxl",
            "source_type": "xlsx-attachment",
            "category": "UserWarning",
        }
        and "Workbook contains no stylesheet" not in json.dumps(library_warnings),
        "openpyxl warnings were not recorded as sanitized structured diagnostics",
    )
    original_candidate_fetch = nav_service.fetch_candidate_messages
    nav_service.fetch_candidate_messages = lambda *_args, **_kwargs: (
        [warning_message],
        {
            "messages_found": 3004,
            "server_since_matches": 3004,
            "server_sender_matches": 1,
            "headers_fetched": 1,
            "messages_selected": 1,
            "messages_fetched": 1,
            "bytes_fetched": len(warning_buffer.getvalue()),
            "skipped_oversize": 0,
            "truncated": False,
            "selection_applied": True,
            "selection": {
                "mode": "sender-subject",
                "sender": "warning@example.invalid",
                "subject_contains": "WARNING",
            },
        },
    )
    try:
        warning_report = nav_service.propose_routes(
            mail_config,
            sender="warning@example.invalid",
            subject_contains="WARNING",
        )
    finally:
        nav_service.fetch_candidate_messages = original_candidate_fetch
    serialized_warning_report = json.dumps(warning_report, ensure_ascii=False)
    check(
        warning_report["passed"]
        and warning_report["selection"]["subject_contains"] == "WARNING"
        and warning_report["candidates"][0]["selection"]
        == warning_report["selection"]
        and warning_report["parser_library_warnings"][0]["count"] == 1
        and "Workbook contains no stylesheet" not in serialized_warning_report
        and "warning-fixture.xlsx" not in serialized_warning_report,
        "selected proposal did not preserve scope or sanitize real library warnings",
    )

    binding_message = EmailMessage()
    binding_message["From"] = "Binding Desk <binding@example.invalid>"
    binding_message["Subject"] = "BIND01 weekly NAV"
    binding_message.set_content(
        "NAV Date | Unit NAV | Cumulative NAV\n"
        "2026-01-16 | 1.03 | 1.05"
    )
    nav_service.fetch_candidate_messages = lambda *_args, **_kwargs: (
        [binding_message],
        {
            "messages_found": 1,
            "server_since_matches": 1,
            "server_sender_matches": 1,
            "headers_fetched": 1,
            "messages_selected": 1,
            "matching_messages_in_range": 1,
            "messages_fetched": 1,
            "bytes_fetched": 100,
            "skipped_oversize": 0,
            "truncated": False,
            "timed_out": False,
            "range_complete": True,
            "resume_before_uid": None,
            "selection_applied": True,
            "selection": {
                "mode": "sender-subject",
                "sender": "binding@example.invalid",
                "subject_contains": "BIND01",
            },
        },
    )
    try:
        binding_report = nav_service.propose_routes(
            mail_config,
            sender="binding@example.invalid",
            subject_contains="BIND01",
            subject_product_code="BIND01",
        )
    finally:
        nav_service.fetch_candidate_messages = original_candidate_fetch
    check(
        binding_report["passed"]
        and binding_report["candidates"][0]["detected_codes"] == ["BIND01"]
        and binding_report["subject_code_binding"]["messages"] == 1
        and binding_report["selection"]["subject_product_code"] == "BIND01"
        and binding_report["candidates"][0]["observations"][0]["source"].endswith(
            ":subject-product-code"
        ),
        "explicit subject product code did not bind a unique code-less NAV series",
    )

    gap_message = EmailMessage()
    gap_message["From"] = "Gap Desk <gap@example.invalid>"
    gap_message["Subject"] = "GAP01 report"
    gap_message.set_content("Please review the proprietary attachment")
    gap_message.add_attachment(
        b"not a supported NAV format",
        maintype="application",
        subtype="octet-stream",
        filename="private-name.bin",
    )
    nav_service.fetch_candidate_messages = lambda *_args, **_kwargs: (
        [gap_message],
        {
            "messages_found": 1,
            "headers_fetched": 1,
            "messages_selected": 1,
            "matching_messages_in_range": 1,
            "messages_fetched": 1,
            "bytes_fetched": 100,
            "skipped_oversize": 0,
            "truncated": False,
            "timed_out": False,
            "range_complete": True,
            "resume_before_uid": None,
            "selection": {
                "mode": "sender-subject",
                "sender": "gap@example.invalid",
                "subject_contains": "GAP01",
            },
        },
    )
    try:
        gap_report = nav_service.propose_routes(
            mail_config,
            sender="gap@example.invalid",
            subject_contains="GAP01",
        )
    finally:
        nav_service.fetch_candidate_messages = original_candidate_fetch
    gap_serialized = json.dumps(gap_report, ensure_ascii=False)
    check(
        not gap_report["passed"]
        and gap_report["parse_gap_summary"]["headers_matched"] == 1
        and gap_report["parse_gap_summary"]["parsed_records"] == 0
        and gap_report["parse_gap_summary"]["local_parser_recommended"]
        and gap_report["parse_gap_summary"]["attachment_types"]
        == {"application/octet-stream": 1}
        and "private-name.bin" not in gap_serialized,
        "matched-but-unparsed mail did not produce a sanitized parser-gap summary",
    )

    html_body = """
    <html><body><table>
      <tr>
        <th rowspan="2"><span>Product</span> Code</th>
        <th rowspan="2">NAV Date</th>
        <th colspan="2">NAV</th>
      </tr>
      <tr><th>Unit NAV</th><th>Cumulative NAV</th></tr>
      <tr><td><span>DEMO</span>&nbsp;01</td><td>2026-01-09</td>
          <td><strong>1.0100</strong></td><td>1.0100</td></tr>
      <tr><td>OTHER01</td><td>2026-01-09</td><td>2.0200</td><td>2.0200</td></tr>
    </table></body></html>
    """
    html_message = EmailMessage()
    html_message["Subject"] = "HTML NAV fixture"
    html_message.set_content(html_body, subtype="html")
    html_rows = rows_from_message(html_message)
    check(
        len(html_rows) == 2
        and {row.code for row in html_rows} == {"DEMO01", "OTHER01"}
        and html_rows[0].date == dt.date(2026, 1, 9),
        "HTML-only table, nested markup, entities, or merged headers were not parsed",
    )
    alternative = EmailMessage()
    alternative["Subject"] = "Multipart NAV fixture"
    alternative.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-01-09 | 1.0100 | 1.0100"
    )
    alternative.add_alternative(html_body, subtype="html")
    alternative_rows = rows_from_message(alternative)
    check(
        len(alternative_rows) == 2,
        "plain-text and HTML alternatives were not parsed and deduplicated together",
    )
    try:
        rows_from_html(
            '<table><tr><td rowspan="0">invalid</td></tr></table>',
            "invalid-html-fixture",
        )
    except ParseError:
        pass
    else:
        raise AssertionError("invalid HTML spans did not fail closed")
    try:
        rows_from_html(
            "x" * (nav_parse.MAX_TEXT_CHARS + 1),
            "oversize-html-fixture",
        )
    except ParseError:
        pass
    else:
        raise AssertionError("oversized HTML body did not fail closed")

    parser_dir = runtime / "parsers"
    parser_dir.mkdir(exist_ok=True)
    strict_example_name = "fixed_label_xlsx_example"
    shutil.copy2(
        runtime / "parser-examples" / "fixed_label_xlsx.py",
        parser_dir / f"{strict_example_name}.py",
    )

    def fixed_label_message(
        *,
        nav_date: dt.date = dt.date(2026, 1, 30),
        subject_date: dt.date | None = None,
        unit: float = 1.03,
        cumulative: float | None = 1.03,
        product_name: str = "虚构示例产品",
        manager: str = "虚构示例管理人",
        duplicate_xlsx: bool = False,
        include_pdf: bool = True,
    ) -> EmailMessage:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "净值通知"
        values = (
            ("产品代码", "SAMPLE01"),
            ("产品名称", product_name),
            ("净值日期", nav_date),
            ("单位净值", unit),
            ("累计净值", cumulative),
            ("管理人", manager),
            ("托管人", "虚构示例托管人"),
        )
        for row, (label, value) in enumerate(values, 1):
            sheet.cell(row, 1, label)
            sheet.cell(row, 2, value)
        xlsx_buffer = io.BytesIO()
        workbook.save(xlsx_buffer)
        workbook.close()
        actual_subject_date = subject_date or nav_date
        message = EmailMessage()
        message["Subject"] = (
            f"SAMPLE01 净值通知 {actual_subject_date.isoformat()}"
        )
        message.set_content("请见附件。")
        message.add_attachment(
            xlsx_buffer.getvalue(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="sample.xlsx",
        )
        if duplicate_xlsx:
            message.add_attachment(
                xlsx_buffer.getvalue(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename="duplicate.xlsx",
            )
        if include_pdf:
            pdf_buffer = io.BytesIO()
            writer = PdfWriter()
            writer.add_blank_page(width=200, height=200)
            writer.write(pdf_buffer)
            message.add_attachment(
                pdf_buffer.getvalue(),
                maintype="application",
                subtype="pdf",
                filename="sample.pdf",
            )
        return message

    strict_rows = rows_from_message(
        fixed_label_message(), f"local:{strict_example_name}"
    )
    check(
        len(strict_rows) == 1
        and strict_rows[0].code == "SAMPLE01"
        and strict_rows[0].date == dt.date(2026, 1, 30)
        and strict_rows[0].unit == 1.03
        and strict_rows[0].cumulative == 1.03,
        "strict fixed-label XLSX local parser example rejected a valid fixture",
    )
    invalid_strict_messages = (
        fixed_label_message(product_name="错误身份"),
        fixed_label_message(manager="错误管理人"),
        fixed_label_message(subject_date=dt.date(2026, 1, 31)),
        fixed_label_message(cumulative=None),
        fixed_label_message(duplicate_xlsx=True),
        fixed_label_message(include_pdf=False),
    )
    for invalid_message in invalid_strict_messages:
        try:
            rows_from_message(
                invalid_message, f"local:{strict_example_name}"
            )
        except ParseError:
            pass
        else:
            raise AssertionError(
                "strict fixed-label XLSX parser accepted invalid identity, "
                "date, cumulative NAV, or attachment structure"
            )
    strict_conflict_rows = [
        *strict_rows,
        *rows_from_message(
            fixed_label_message(unit=1.04, cumulative=1.04),
            f"local:{strict_example_name}",
        ),
    ]
    try:
        choose_route_rows(
            strict_conflict_rows,
            {"code": "SAMPLE01", "sheet": "示例产品", "allow_sender_only": False},
            True,
        )
    except ParseError:
        pass
    else:
        raise AssertionError(
            "strict fixed-label XLSX parser regression did not expose a same-day conflict"
        )

    (parser_dir / "fixture.py").write_text(
        "from datetime import date\n"
        "from nav_parse import NavRow\n\n"
        "def parse_message(message):\n"
        "    return [NavRow(date(2026, 1, 30), 1.03, 1.03, 'DEMO01', 'local')]\n",
        encoding="utf-8",
    )
    sensitive_path = "\\".join(
        ("C:", "Users", "Example User", "Synced Folder - Team", "客户甲", "估值表.xlsx")
    )
    sensitive_exception = f"{sensitive_path} user@example.invalid 产品甲净值 1.2345"
    (parser_dir / "sensitive_failure.py").write_text(
        "class 客户甲净值1234(Exception):\n"
        "    pass\n\n"
        "def parse_message(message):\n"
        f"    raise 客户甲净值1234({sensitive_exception!r})\n",
        encoding="utf-8",
    )
    local_rows = rows_from_message(EmailMessage(), "local:fixture")
    check(
        len(local_rows) == 1 and local_rows[0].date == dt.date(2026, 1, 30),
        "trusted local parser extension failed",
    )

    route = {"code": "DEMO01", "allow_sender_only": False, "sheet": "Demo Fund"}
    conflict = rows_from_text(
        "Product Code | NAV Date | Unit NAV\nDEMO01 | 2026-01-09 | 1.01\nDEMO01 | 2026-01-09 | 1.02",
        "body",
    )
    try:
        choose_route_rows(conflict, route, True)
    except ParseError:
        pass
    else:
        raise AssertionError("same-day value conflicts must fail closed")
    try:
        rows_from_text(
            "NAV Date: 2026-01-09\nNAV Date: 2026-01-16\nUnit NAV: 1.01", "body"
        )
    except ParseError:
        pass
    else:
        raise AssertionError("repeated labelled fields must fail closed")
    try:
        rows_from_text(
            "估值日期 | 估值基准日 | 单位净值\n2026-01-09 | 2026-01-10 | 1.01",
            "ambiguous-date-fixture",
        )
    except ParseError as exc:
        check(
            "Ambiguous date columns" in str(exc),
            "multiple semantic date columns failed for the wrong reason",
        )
    else:
        raise AssertionError("multiple semantic date columns were accepted silently")
    explanatory_rows = rows_from_text(
        "估值日期说明 | 业务日期口径\n"
        "本页为说明 | 请以下表为准\n"
        "估值基准日 | 单位净值 | 产品代码\n"
        "2026-01-09 | 1.01 | DEMO01",
        "cover-before-data-fixture",
    )
    check(
        len(explanatory_rows) == 1 and explanatory_rows[0].date == dt.date(2026, 1, 9),
        "a non-header explanation row with two date terms blocked the real data table",
    )


def route_state_tests(runtime: Path) -> None:
    sys.path.insert(0, str(runtime))
    import nav_service
    import navctl

    book = runtime / "route-state-placeholder.xlsx"
    create_book(book)
    config = config_for(runtime, book)
    empty = json.loads(json.dumps(config))
    empty["routes"] = []
    proposal_message = EmailMessage()
    proposal_message["From"] = "NAV Desk <sender@example.invalid>"
    proposal_message["Subject"] = "Weekly NAV"
    proposal_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-01-09 | 1.01 | 1.01"
    )
    original_candidates = nav_service.fetch_candidate_messages
    nav_service.fetch_candidate_messages = lambda _config: (
        [proposal_message],
        {
            "messages_found": 1,
            "messages_fetched": 1,
            "bytes_fetched": 100,
            "skipped_oversize": 0,
            "truncated": False,
        },
    )
    try:
        proposal = nav_service.propose_routes(empty)
    finally:
        nav_service.fetch_candidate_messages = original_candidates
    check(
        proposal["passed"]
        and proposal["candidates"][0]["sender"] == "sender@example.invalid"
        and proposal["candidates"][0]["detected_codes"] == ["DEMO01"],
        "AI route proposal did not discover the sender and product code",
    )
    _, report = nav_service.collect_route_rows(empty)
    check(
        not report["passed"] and "No active routes" in report["errors"][0],
        "empty route configuration reported discovery success",
    )
    paused = json.loads(json.dumps(config))
    paused["routes"][0]["paused"] = True
    paused["routes"][0]["pause_reason"] = "fixture pause"
    _, report = nav_service.collect_route_rows(paused)
    check(
        not report["passed"] and report["warnings"],
        "paused route was not excluded with a visible warning",
    )
    original_fetch = nav_service.fetch_authorized_messages
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": []
    }
    try:
        _, report = nav_service.collect_route_rows(config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        not report["passed"]
        and any("no routed NAV rows" in item for item in report["errors"]),
        "active route with no messages reported discovery success",
    )
    binding_config = json.loads(json.dumps(config))
    binding_route = binding_config["routes"][0]
    binding_route["code"] = "BIND01"
    binding_route["subject_contains"] = "BIND01"
    binding_route["subject_product_code"] = "BIND01"
    binding_route["max_staleness_days"] = 366
    binding_message = EmailMessage()
    binding_message["From"] = "NAV Desk <sender@example.invalid>"
    binding_message["Subject"] = "BIND01 weekly NAV"
    binding_message.set_content(
        "NAV Date | Unit NAV | Cumulative NAV\n"
        "2026-07-17 | 1.03 | 1.05"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [binding_message]
    }
    try:
        binding_rows, binding_route_report = nav_service.collect_route_rows(
            binding_config
        )
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        binding_route_report["passed"]
        and binding_rows["Demo Fund"][0].code == "BIND01"
        and binding_route_report["routes"][0]["subject_code_binding_messages"] == 1
        and any("主题产品代码证据" in item for item in binding_route_report["warnings"]),
        "configured subject product code binding was not visible in routed preview evidence",
    )
    sensitive_subject = "机密产品甲净值通知"
    sensitive_attachment = "客户私密估值表.xlsx"
    failed_message = EmailMessage()
    failed_message["From"] = "NAV Desk <sender@example.invalid>"
    failed_message["Subject"] = sensitive_subject
    failed_message.set_content("See attachment")
    failed_message.add_attachment(
        b"not-an-xlsx",
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=sensitive_attachment,
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [failed_message]
    }
    try:
        _, failed_report = nav_service.collect_route_rows(config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_failure = json.dumps(failed_report, ensure_ascii=False)
    check(
        not failed_report["passed"]
        and failed_report["diagnostics"]
        and failed_report["diagnostics"][0]["stage"] == "message_parse"
        and failed_report["diagnostics"][0]["subject_id"]
        and failed_report["diagnostics"][0]["attachment_types"][0]["suffix"] == ".xlsx"
        and sensitive_subject not in serialized_failure
        and sensitive_attachment not in serialized_failure,
        "parse diagnostics did not identify the failure without leaking sensitive names",
    )
    non_nav_message = EmailMessage()
    non_nav_message["From"] = "NAV Desk <sender@example.invalid>"
    non_nav_subject = "DEMO01 account service notice"
    non_nav_message["Subject"] = non_nav_subject
    non_nav_message.set_content("This fixture contains no structured values.")
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [non_nav_message]
    }
    try:
        _, non_nav_report = nav_service.collect_route_rows(config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_non_nav = json.dumps(non_nav_report, ensure_ascii=False)
    check(
        not non_nav_report["passed"]
        and non_nav_report["diagnostics"][0]["notice_classification"]
        == "likely-non-nav-notice"
        and non_nav_report["diagnostics"][0]["suggested_filter"]["field"]
        == "subject_excludes"
        and non_nav_subject not in serialized_non_nav,
        "non-NAV messages did not produce a redacted exclusion diagnostic",
    )

    pdf_buffer = io.BytesIO()
    pdf_writer = PdfWriter()
    pdf_writer.add_blank_page(width=72, height=72)
    pdf_writer.write(pdf_buffer)
    pdf_message = EmailMessage()
    pdf_message["From"] = "NAV Desk <sender@example.invalid>"
    pdf_message["Subject"] = "DEMO01 NAV PDF fixture"
    pdf_message.set_content("See attached reports.")
    pdf_names = ["private-report-one.pdf", "private-report-two.pdf"]
    for pdf_name in pdf_names:
        pdf_message.add_attachment(
            pdf_buffer.getvalue(),
            maintype="application",
            subtype="pdf",
            filename=pdf_name,
        )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [pdf_message]
    }
    try:
        _, pdf_report = nav_service.collect_route_rows(config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_pdf = json.dumps(pdf_report, ensure_ascii=False)
    pdf_diagnostics = pdf_report["diagnostics"][0]["attachment_diagnostics"]
    check(
        not pdf_report["passed"]
        and len(pdf_diagnostics) == 2
        and len({item["attachment_id"] for item in pdf_diagnostics}) == 2
        and all(item["suffix"] == ".pdf" for item in pdf_diagnostics)
        and all(item["status"] == "no-nav-records" for item in pdf_diagnostics)
        and all(name not in serialized_pdf for name in pdf_names),
        "PDF parser gaps were not reported per attachment with redacted fingerprints",
    )
    local_failure_config = json.loads(json.dumps(config))
    local_failure_config["routes"][0]["parser"] = "local:sensitive_failure"
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [failed_message]
    }
    try:
        _, local_failure_report = nav_service.collect_route_rows(local_failure_config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_local_failure = json.dumps(local_failure_report, ensure_ascii=False)
    check(
        local_failure_report["diagnostics"]
        and local_failure_report["diagnostics"][0]["error_type"] == "LocalParserError"
        and local_failure_report["diagnostics"][0]["root_error_type"]
        == "ControlledParseFailure"
        and "Synced Folder" not in serialized_local_failure
        and "客户甲" not in serialized_local_failure
        and "客户甲净值1234" not in serialized_local_failure
        and "1.2345" not in serialized_local_failure
        and "user@example.invalid" not in serialized_local_failure,
        "local parser exception text leaked through structured diagnostics",
    )
    mixed = json.loads(json.dumps(config))
    mixed["routes"][0].update({"parser": "local:fixture", "max_staleness_days": 366})
    second_route = json.loads(json.dumps(mixed["routes"][0]))
    second_route.update({"sheet": "Second Fund", "code": "DEMO02", "parser": "auto"})
    mixed["routes"].append(second_route)
    mixed_message = EmailMessage()
    mixed_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO02 | 2026-02-06 | 1.04 | 1.04"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [mixed_message]
    }
    try:
        mixed_rows, report = nav_service.collect_route_rows(mixed)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        report["passed"]
        and len(mixed_rows["Demo Fund"]) == 1
        and len(mixed_rows["Second Fund"]) == 1,
        "multiple trusted parsers for one sender were not merged and routed safely",
    )
    check(
        report["route_overlaps"]
        and report["route_overlaps"][0]["reason"]
        == "overlapping-message-scope"
        and len(report["route_overlaps"][0]["routes"]) == 2
        and report["route_overlaps"][0]["suggested_filter"]["rerun_sheets"],
        "overlapping route scopes were not reported as a minimal route pair",
    )

    conflict_config = json.loads(json.dumps(config))
    conflict_config["routes"][0]["max_staleness_days"] = 3660
    conflict_message = EmailMessage()
    conflict_message["Subject"] = "Weekly NAV conflict fixture"
    conflict_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-07-17 | 1.10 | 1.10\n"
        "DEMO01 | 2026-07-17 | 1.11 | 1.11"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [conflict_message]
    }
    try:
        _, conflict_report = nav_service.collect_route_rows(conflict_config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_conflict = json.dumps(conflict_report, ensure_ascii=False)
    check(
        not conflict_report["passed"]
        and conflict_report["date_conflicts"]
        and conflict_report["date_conflicts"][0]["date"] == "2026-07-17"
        and len(conflict_report["date_conflicts"][0]["candidate_sources"]) == 2
        and conflict_report["date_conflicts"][0]["suggested_filter"]["rerun_sheet"]
        == "Demo Fund"
        and "1.10" not in serialized_conflict
        and "1.11" not in serialized_conflict,
        "same-date conflicts did not expose actionable redacted fingerprints",
    )

    bounded_conflict_config = json.loads(json.dumps(conflict_config))
    bounded_conflict_config["routes"][0]["series_start"] = "2026-07-18"
    bounded_conflict_message = EmailMessage()
    bounded_conflict_message["Subject"] = "Weekly NAV bounded conflict fixture"
    bounded_conflict_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-07-17 | 1.10 | 1.10\n"
        "DEMO01 | 2026-07-17 | 1.11 | 1.11\n"
        "DEMO01 | 2026-07-18 | 1.12 | 1.12"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [bounded_conflict_message]
    }
    try:
        bounded_rows, bounded_conflict_report = nav_service.collect_route_rows(
            bounded_conflict_config
        )
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        bounded_conflict_report["passed"]
        and not bounded_conflict_report["date_conflicts"]
        and bounded_conflict_report["pre_managed_rows_ignored"] == 2
        and [row.date for row in bounded_rows["Demo Fund"]]
        == [dt.date(2026, 7, 18)],
        "pre-managed same-date conflicts blocked or entered the managed route",
    )
    overlap_config = json.loads(json.dumps(conflict_config))
    overlap_config["routes"][0]["series_start"] = "2026-07-18"
    overlap_config["routes"][0]["baseline_overlap"] = "last_existing_point"
    overlap_message = EmailMessage()
    overlap_message["Subject"] = "Weekly NAV overlap fixture"
    overlap_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-07-17 | 1.10 | 1.10"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [overlap_message]
    }
    try:
        overlap_rows, overlap_report = nav_service.collect_route_rows(
            overlap_config
        )
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        overlap_report["passed"]
        and overlap_report["routes"][0]["verification_anchor_date"]
        == "2026-07-17"
        and [row.date for row in overlap_rows["Demo Fund"]]
        == [dt.date(2026, 7, 17)],
        "last-existing-point overlap did not survive routing as a read-only anchor",
    )

    def correction_fixture(
        subject: str,
        uid: int,
        value_rows: list[tuple[float, float]],
        message_id: str,
    ) -> EmailMessage:
        message = EmailMessage()
        message["From"] = "NAV Desk <sender@example.invalid>"
        message["Subject"] = subject
        message["Date"] = (
            f"Fri, 17 Jul 2026 {8 + uid % 10:02d}:00:00 +0800"
        )
        message["Message-ID"] = f"<{message_id}@example.invalid>"
        lines = ["Product Code | NAV Date | Unit NAV | Cumulative NAV"]
        lines.extend(
            f"DEMO01 | 2026-07-17 | {unit:.4f} | {cumulative:.4f}"
            for unit, cumulative in value_rows
        )
        message.set_content("\n".join(lines))
        setattr(message, "_nav_source_uid", uid)
        return message

    original_notice = correction_fixture(
        "Weekly NAV", 100, [(1.10, 1.10)], "original-notice"
    )
    correction_subject = "Weekly NAV 更正通知（以此为准）"
    correction_notice = correction_fixture(
        correction_subject,
        101,
        [(1.11, 1.11)],
        "correction-notice",
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [original_notice, correction_notice]
    }
    try:
        corrected_rows, corrected_report = nav_service.collect_route_rows(
            conflict_config
        )
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_corrected = json.dumps(corrected_report, ensure_ascii=False)
    check(
        corrected_report["passed"]
        and len(corrected_rows["Demo Fund"]) == 1
        and corrected_rows["Demo Fund"][0].unit == 1.11
        and corrected_report["corrections_applied"]
        and corrected_report["corrections_applied"][0]["order_basis"]
        == "mailbox-uid"
        and len(corrected_report["corrections_applied"][0]["replaced"]) == 1
        and correction_subject not in serialized_corrected
        and "1.1000" not in serialized_corrected
        and "1.1100" not in serialized_corrected,
        "a later explicit correction did not replace the original with redacted audit evidence",
    )

    ordinary_late_notice = correction_fixture(
        "Weekly NAV resend", 102, [(1.12, 1.12)], "ordinary-late-notice"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [original_notice, ordinary_late_notice]
    }
    try:
        _, ordinary_late_report = nav_service.collect_route_rows(conflict_config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        not ordinary_late_report["passed"]
        and ordinary_late_report["date_conflicts"]
        and not ordinary_late_report["corrections_applied"],
        "a normal later message incorrectly overwrote an earlier same-date value",
    )

    early_correction = correction_fixture(
        "Weekly NAV corrected", 99, [(1.11, 1.11)], "early-correction"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [early_correction, original_notice]
    }
    try:
        _, early_correction_report = nav_service.collect_route_rows(conflict_config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        not early_correction_report["passed"]
        and early_correction_report["date_conflicts"]
        and not early_correction_report["corrections_applied"],
        "an earlier correction-labeled message incorrectly overwrote a later original",
    )

    internally_conflicting_correction = correction_fixture(
        "Weekly NAV 修正通知",
        101,
        [(1.11, 1.11), (1.12, 1.12)],
        "internally-conflicting-correction",
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [original_notice, internally_conflicting_correction]
    }
    try:
        _, internal_conflict_report = nav_service.collect_route_rows(conflict_config)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        not internal_conflict_report["passed"]
        and internal_conflict_report["date_conflicts"]
        and not internal_conflict_report["corrections_applied"],
        "multiple values inside one correction message did not fail closed",
    )

    scoped_config = json.loads(json.dumps(config))
    scoped_config["routes"][0]["max_staleness_days"] = 3660
    scoped_second = json.loads(json.dumps(scoped_config["routes"][0]))
    scoped_second.update({"sheet": "Second Fund", "code": "DEMO02"})
    scoped_config["routes"].append(scoped_second)
    scoped_message = EmailMessage()
    scoped_message["Subject"] = "Weekly NAV scoped fixture"
    scoped_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-07-17 | 1.10 | 1.10"
    )
    captured_scope: list[list[str]] = []

    def scoped_fetch(test_config: dict[str, Any]) -> dict[str, list[EmailMessage]]:
        captured_scope.append(
            [str(route["sheet"]) for route in test_config.get("routes") or []]
        )
        return {"sender@example.invalid": [scoped_message]}

    nav_service.fetch_authorized_messages = scoped_fetch
    try:
        scoped_rows, scoped_report = nav_service.discover(
            scoped_config, ["Demo Fund"]
        )
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    parsed_args = navctl.parser().parse_args(
        ["discover", "--sheet", "Demo Fund", "--sheet", "Second Fund"]
    )
    check(
        scoped_report["passed"]
        and captured_scope == [["Demo Fund"]]
        and list(scoped_rows) == ["Demo Fund"]
        and scoped_report["scoped"]
        and scoped_report["final_full_preview_required"]
        and parsed_args.sheets == ["Demo Fund", "Second Fund"],
        "read-only per-sheet discovery did not isolate the requested route",
    )
    scoped_preview_config = json.loads(json.dumps(scoped_config))
    scoped_preview_config["routes"][0].pop("benchmark", None)
    scoped_preview_config["routes"][1].pop("benchmark", None)
    scoped_preview_message = EmailMessage()
    scoped_preview_message["Subject"] = "Weekly NAV scoped preview fixture"
    scoped_preview_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "DEMO01 | 2026-01-09 | 1.01 | 1.01"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [scoped_preview_message]
    }
    try:
        scoped_plan = nav_service.preview(
            scoped_preview_config, sheets=["Demo Fund"]
        )
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    persisted_scoped_plan = json.loads(
        (runtime / "plan.json").read_text(encoding="utf-8")
    )
    check(
        scoped_plan["scoped"]
        and scoped_plan["scope_sheets"] == ["Demo Fund"]
        and scoped_plan["diagnostic_only"]
        and not scoped_plan["committable"]
        and scoped_plan["final_full_preview_required"]
        and persisted_scoped_plan["diagnostic_only"]
        and any(
            item["issue"] == "scoped-diagnostic-preview"
            for item in scoped_plan["blocking_reviews"]
        ),
        "per-sheet preview was not marked diagnostic-only and non-committable",
    )
    from nav_commit import CommitError as ScopedCommitError
    from nav_commit import commit as commit_scoped_preview

    try:
        commit_scoped_preview(scoped_preview_config)
    except ScopedCommitError as exc:
        check(
            "review-only preview" in str(exc),
            "scoped preview was rejected for the wrong reason",
        )
    else:
        raise AssertionError("a scoped diagnostic preview was accepted for commit")
    parsed_preview_args = navctl.parser().parse_args(
        ["preview", "--sheet", "Demo Fund", "--sheet", "Second Fund"]
    )
    check(
        parsed_preview_args.sheets == ["Demo Fund", "Second Fund"],
        "preview CLI did not accept repeated per-sheet diagnostic scopes",
    )

    selective = json.loads(json.dumps(config))
    selective["routes"][0]["max_staleness_days"] = 3660
    paused_route = json.loads(json.dumps(selective["routes"][0]))
    paused_route.update(
        {
            "sheet": "Paused Fund",
            "code": "PAUSED01",
            "paused": True,
            "pause_reason": "fixture pause",
        }
    )
    selective["routes"].append(paused_route)
    selective_message = EmailMessage()
    selective_message["Subject"] = "Bulk HTML NAV fixture"
    selective_message.set_content(
        """
        <table>
          <tr><th>Product Code</th><th>NAV Date</th><th>Unit NAV</th><th>Cumulative NAV</th></tr>
          <tr><td>DEMO01</td><td>2026-07-17</td><td>1.10</td><td>1.10</td></tr>
          <tr><td>PAUSED01</td><td>2026-07-17</td><td>1.20</td><td>1.20</td></tr>
          <tr><td>OTHER01</td><td>2026-07-17</td><td>1.30</td><td>1.30</td></tr>
        </table>
        """,
        subtype="html",
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [selective_message]
    }
    try:
        selective_rows, selective_report = nav_service.collect_route_rows(selective)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    serialized_selective = json.dumps(selective_report, ensure_ascii=False)
    check(
        selective_report["passed"]
        and len(selective_rows["Demo Fund"]) == 1
        and "Paused Fund" not in selective_rows
        and selective_report["ignored_unconfigured_rows"] == 2
        and selective_report["messages_with_ignored_unconfigured_rows"] == 1
        and "PAUSED01" not in serialized_selective
        and "OTHER01" not in serialized_selective,
        "explicit unconfigured or paused product rows were not ignored safely",
    )

    unrelated_message = EmailMessage()
    unrelated_message.set_content(
        "Product Code | NAV Date | Unit NAV | Cumulative NAV\n"
        "OTHER01 | 2026-07-17 | 1.30 | 1.30"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [unrelated_message]
    }
    try:
        unrelated_rows, unrelated_report = nav_service.collect_route_rows(selective)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        not unrelated_report["passed"]
        and not unrelated_rows["Demo Fund"]
        and unrelated_report["ignored_unconfigured_rows"] == 1
        and any("no routed NAV rows" in item for item in unrelated_report["errors"]),
        "an all-unconfigured lookback incorrectly reported managed data success",
    )

    ambiguous_message = EmailMessage()
    ambiguous_message.set_content(
        "NAV Date | Unit NAV | Cumulative NAV\n2026-07-17 | 1.10 | 1.10"
    )
    nav_service.fetch_authorized_messages = lambda _config: {
        "sender@example.invalid": [ambiguous_message]
    }
    try:
        _, ambiguous_report = nav_service.collect_route_rows(selective)
    finally:
        nav_service.fetch_authorized_messages = original_fetch
    check(
        not ambiguous_report["passed"]
        and ambiguous_report["ignored_unconfigured_rows"] == 0
        and any(
            "not routed to exactly one" in item for item in ambiguous_report["errors"]
        ),
        "a code-less ambiguous NAV row was ignored instead of failing closed",
    )


def lock_tests(runtime: Path) -> None:
    sys.path.insert(0, str(runtime))
    from nav_schedule import (
        _local_timezone,
        _task_time,
        _task_time_detail,
        record_scheduled_run,
        status,
    )

    com_wall_clock = dt.datetime(2026, 7, 23, 12, 0, 0, tzinfo=dt.timezone.utc)
    local_timezone = _local_timezone()
    local_detail = _task_time_detail(com_wall_clock, local_timezone)
    check(
        _task_time(com_wall_clock) == "2026-07-23T12:00:00"
        and local_detail is not None
        and local_detail["datetime"] == "2026-07-23T12:00:00"
        and local_detail["timezone_id"] == local_timezone["id"]
        and "+00:00" not in local_detail["datetime"],
        "Task Scheduler wall-clock time was mislabeled as UTC",
    )

    holder_code = (
        "import time\n"
        "from navctl import run_lock\n"
        "with run_lock():\n"
        "    print('ready', flush=True)\n"
        "    time.sleep(60)\n"
    )
    contender_code = (
        "import sys\n"
        "from navctl import run_lock\n"
        "try:\n"
        "    with run_lock():\n"
        "        print('acquired')\n"
        "except RuntimeError:\n"
        "    print('blocked')\n"
        "    raise SystemExit(3)\n"
    )
    holder = subprocess.Popen(
        [sys.executable, "-X", "utf8", "-c", holder_code],
        cwd=runtime,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
    )
    try:
        ready = holder.stdout.readline().strip() if holder.stdout else ""
        check(ready == "ready", "runtime lock holder did not start")
        blocked = subprocess.run(
            [sys.executable, "-X", "utf8", "-c", contender_code],
            cwd=runtime,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        check(
            blocked.returncode == 3 and "blocked" in blocked.stdout,
            "runtime lock allowed concurrent access",
        )
        check(holder.poll() is None, "runtime lock probe interrupted the lock holder")
    finally:
        holder.terminate()
        holder.wait(timeout=10)
    recovered = subprocess.run(
        [sys.executable, "-X", "utf8", "-c", contender_code],
        cwd=runtime,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    check(
        recovered.returncode == 0 and "acquired" in recovered.stdout,
        "runtime lock did not recover after a crashed holder",
    )
    state = json.loads((runtime / "run.lock").read_text(encoding="utf-8"))
    check(state.get("status") == "idle", "runtime lock did not record idle state")
    scheduled = {
        "started": "2026-01-01T09:30:00",
        "finished": "2026-01-01T09:30:05",
        "passed": False,
        "exit_code": 2,
        "error": "fixture failure",
    }
    record_scheduled_run(scheduled)
    check(
        status().get("last_run") == scheduled,
        "schedule status did not expose the latest run result",
    )
    check(
        "scheduled-update" in (runtime / "run-update.cmd").read_text(encoding="utf-8"),
        "scheduled wrapper does not run automatic updates",
    )


def workbook_tests(runtime: Path, use_com: bool) -> str | None:
    sys.path.insert(0, str(runtime))
    from nav_automation import (
        AutomationError,
        approve,
        automatic_update,
        revoke,
        status as automation_status,
    )
    from nav_commit import CommitError, _process_id, commit, ensure_process_exit
    from nav_config import ConfigError, load_config, validate_config
    from nav_parse import NavRow
    from nav_service import preview as service_preview
    from nav_workbook import (
        WorkbookError,
        build_preview,
        file_sha256,
        make_file_writable,
        review_file_is_read_only,
        validate_history,
    )

    book = runtime / "脱敏 示例.xlsx"
    create_book(book)
    config = config_for(runtime, book)
    array_workbook = openpyxl.load_workbook(book)
    try:
        array_sheet = array_workbook.create_sheet("Array Audit")
        array_sheet["A1"] = ArrayFormula(ref="A1:A1", text="=SUM(1,1)")
        array_workbook.save(book)
    finally:
        array_workbook.close()
    revoke()
    check(
        not automation_status(config)["approved"],
        "automatic updates started approved before the first reviewed commit",
    )
    approve(config)
    check(
        automation_status(config)["approved"],
        "reviewed configuration could not approve future automatic updates",
    )
    changed_config = json.loads(json.dumps(config))
    changed_config["routes"][0]["return_frequency"] = "daily"
    check(
        not automation_status(changed_config)["approved"],
        "automatic approval survived a write-rule configuration change",
    )
    revoke()
    invalid = json.loads(json.dumps(config))
    invalid["routes"][0]["series_start"] = "2026-99-99"
    invalid["routes"][0]["baseline_overlap"] = "unsafe"
    invalid["routes"][0]["parser"] = "local:../unsafe"
    invalid["routes"][0]["data_frequency"] = "hourly"
    invalid["unexpected"] = True
    try:
        validate_config(invalid)
    except ConfigError:
        pass
    else:
        raise AssertionError("invalid dates and unknown config fields must be rejected")
    local_parser_config = json.loads(json.dumps(config))
    local_parser_config["routes"][0]["parser"] = "local:fixture"
    local_parser_config["routes"][0]["subject_excludes"] = ["fee notice"]
    validate_config(local_parser_config)
    invalid_excludes = json.loads(json.dumps(config))
    invalid_excludes["routes"][0]["subject_contains"] = "NAV"
    invalid_excludes["routes"][0]["subject_excludes"] = ["nav", "NAV"]
    try:
        validate_config(invalid_excludes)
    except ConfigError:
        pass
    else:
        raise AssertionError("duplicate or self-cancelling subject exclusions were accepted")
    mixed_parser_config = json.loads(json.dumps(config))
    second_route = json.loads(json.dumps(mixed_parser_config["routes"][0]))
    second_route.update(
        {"sheet": "Second Fund", "code": "DEMO02", "parser": "local:fixture"}
    )
    mixed_parser_config["routes"].append(second_route)
    validate_config(mixed_parser_config)
    paused_config = json.loads(json.dumps(config))
    paused_config["routes"][0]["paused"] = True
    paused_config["routes"][0]["pause_reason"] = "fixture pause"
    validate_config(paused_config)
    reviewed_sheet_config = json.loads(json.dumps(config))
    reviewed_sheet_config["sheet_reviews"] = {
        "Archived Example": {
            "status": "excluded",
            "reason": "local evidence says this page is no longer updated",
        },
        "Parser Example": {
            "status": "local_parser_required",
            "reason": "official notices require a vendor-specific parser",
        },
    }
    validate_config(reviewed_sheet_config)
    conflicting_review_config = json.loads(json.dumps(config))
    conflicting_review_config["sheet_reviews"] = {
        "Demo Fund": {
            "status": "business_review",
            "reason": "must not duplicate a managed route",
        }
    }
    try:
        validate_config(conflicting_review_config)
    except ConfigError:
        pass
    else:
        raise AssertionError("sheet review duplicated a managed route")
    invalid_review_config = json.loads(json.dumps(config))
    invalid_review_config["sheet_reviews"] = {
        "Unknown Example": {"status": "guess", "reason": ""}
    }
    try:
        validate_config(invalid_review_config)
    except ConfigError:
        pass
    else:
        raise AssertionError("invalid workbook sheet review was accepted")
    invalid_header_limit = json.loads(json.dumps(config))
    invalid_header_limit["imap"]["max_messages"] = 100
    invalid_header_limit["imap"]["max_header_messages"] = 99
    try:
        validate_config(invalid_header_limit)
    except ConfigError:
        pass
    else:
        raise AssertionError(
            "max_header_messages smaller than max_messages was accepted"
        )
    invalid_append = json.loads(json.dumps(config))
    invalid_append["routes"][0].update(
        {"sheet_mode": "append", "code": None, "product_name": None}
    )
    try:
        validate_config(invalid_append)
    except ConfigError:
        pass
    else:
        raise AssertionError("append mode accepted a route without product identity")
    invalid_formula_name = json.loads(json.dumps(config))
    invalid_formula_name["routes"][0]["product_name"] = (
        '=HYPERLINK("https://example.invalid")'
    )
    try:
        validate_config(invalid_formula_name)
    except ConfigError:
        pass
    else:
        raise AssertionError("route accepted a formula-like product name")
    integrity_book = runtime / "history-integrity-fixture.xlsx"
    integrity_workbook = openpyxl.Workbook()
    integrity_sheet = integrity_workbook.active
    integrity_sheet.title = "Integrity Fund"
    integrity_sheet.append(
        ["NAV Date", "Product Code", "Unit NAV", "Cumulative NAV"]
    )
    for row in (
        (dt.date(2026, 1, 2), "DEMO01", 1.00, 1.50),
        (dt.date(2026, 1, 9), "DEMO02", 1.10, 1.10),
        (dt.date(2026, 1, 16), "DEMO03", 1.20, 1.70),
    ):
        integrity_sheet.append(row)
    integrity_workbook.save(integrity_book)
    integrity_workbook.close()
    integrity_config = config_for(runtime, integrity_book)
    integrity_config["routes"][0].update(
        {
            "sheet": "Integrity Fund",
            "sheet_mode": "append",
            "benchmark": None,
            "max_staleness_days": 366,
        }
    )
    integrity_rows = {
        "Integrity Fund": [
            NavRow(dt.date(2026, 1, 2), 1.00, 1.50, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 9), 1.10, 1.60, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 16), 1.20, 1.70, "DEMO01", "fixture"),
        ]
    }
    integrity_report = validate_history(integrity_config, integrity_rows)
    serialized_integrity = json.dumps(integrity_report, ensure_ascii=False)
    integrity_checks = integrity_report["routes"][0]["history_integrity"]
    check(
        not integrity_report["passed"]
        and integrity_report["history_repairs_required"] == 1
        and not integrity_checks["code_column"]["passed"]
        and len(integrity_checks["code_column"]["unexpected_rows"]) == 2
        and not integrity_checks["cumulative_sequence"]["passed"]
        and integrity_checks["cumulative_sequence"]["anomalies"][0]["issue"]
        == "isolated-unit-cumulative-spread-break"
        and "DEMO02" not in serialized_integrity
        and "DEMO03" not in serialized_integrity
        and "1.6000" not in serialized_integrity,
        "historical code drift and isolated cumulative discontinuity were not diagnosed safely",
    )
    boundary_book = runtime / "series-start-boundary-fixture.xlsx"
    boundary_workbook = openpyxl.Workbook()
    boundary_sheet = boundary_workbook.active
    boundary_sheet.title = "Boundary Fund"
    boundary_sheet.append(
        ["NAV Date", "Product Code", "Unit NAV", "Cumulative NAV"]
    )
    boundary_sheet.append([dt.date(2026, 1, 2), "LEGACY01", 0.90, None])
    boundary_sheet.append([dt.date(2026, 1, 9), "DEMO01", 1.00, 1.50])
    boundary_sheet.append([dt.date(2026, 1, 16), "DEMO01", 1.10, 1.60])
    boundary_sheet.append(["TOTAL", None, None, None])
    boundary_workbook.save(boundary_book)
    boundary_workbook.close()
    boundary_config = config_for(runtime, boundary_book)
    boundary_config["routes"][0].update(
        {
            "sheet": "Boundary Fund",
            "product_name": None,
            "series_start": "2026-01-09",
            "benchmark": None,
            "max_staleness_days": 366,
        }
    )
    boundary_rows = {
        "Boundary Fund": [
            NavRow(dt.date(2026, 1, 9), 1.00, 1.50, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 16), 1.10, 1.60, "DEMO01", "fixture"),
        ]
    }
    boundary_report = validate_history(boundary_config, boundary_rows)
    boundary_integrity = boundary_report["routes"][0]["history_integrity"]
    check(
        boundary_report["passed"]
        and boundary_report["history_repairs_required"] == 0
        and boundary_integrity["passed"]
        and boundary_integrity["scope"]["pre_managed_rows"] == 1
        and boundary_integrity["scope"]["managed_rows"] == 2
        and boundary_integrity["pre_managed_diagnostics"]["repair_required"]
        and boundary_report["warnings"],
        "pre-managed integrity diagnostics blocked the managed series tail",
    )
    boundary_master = boundary_book.read_bytes()
    boundary_plan = service_preview(boundary_config, boundary_rows)
    check(
        boundary_plan["approval_kind"] == "validated-no-change"
        and not boundary_plan["sheets"]
        and boundary_book.read_bytes() == boundary_master,
        "series_start boundary preview changed preserved history",
    )
    managed_break_book = runtime / "series-start-managed-break.xlsx"
    shutil.copy2(boundary_book, managed_break_book)
    managed_break_workbook = openpyxl.load_workbook(managed_break_book)
    try:
        managed_break_workbook["Boundary Fund"]["D3"] = None
        managed_break_workbook.save(managed_break_book)
    finally:
        managed_break_workbook.close()
    managed_break_config = json.loads(json.dumps(boundary_config))
    managed_break_config["workbook_path"] = str(managed_break_book.resolve())
    managed_break_report = validate_history(managed_break_config, boundary_rows)
    check(
        not managed_break_report["passed"]
        and managed_break_report["history_repairs_required"] == 1
        and not managed_break_report["routes"][0]["history_integrity"][
            "cumulative_sequence"
        ]["passed"],
        "a cumulative gap at the first managed point was ignored",
    )
    anchor_config = config_for(runtime, book)
    anchor_config["routes"][0]["series_start"] = "2026-01-10"
    anchor_config["routes"][0]["baseline_overlap"] = "last_existing_point"
    anchor_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 9), 1.01, 1.01, "DEMO01", "fixture")
        ]
    }
    anchor_validation = validate_history(anchor_config, anchor_rows)
    anchor_plan = service_preview(anchor_config, anchor_rows)
    check(
        anchor_validation["passed"]
        and anchor_validation["routes"][0]["boundary_anchor_verified"]
        and anchor_validation["routes"][0]["boundary_anchor_date"]
        == "2026-01-09"
        and anchor_plan["approval_kind"] == "validated-no-change"
        and not anchor_plan["sheets"],
        "a verified pre-managed tail anchor did not unlock a zero-add baseline",
    )
    bad_anchor_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 9), 1.02, 1.02, "DEMO01", "fixture")
        ]
    }
    bad_anchor_validation = validate_history(anchor_config, bad_anchor_rows)
    check(
        not bad_anchor_validation["passed"]
        and not bad_anchor_validation["routes"][0]["boundary_anchor_verified"],
        "a mismatched last-existing-point overlap was accepted",
    )
    continuing_anchor_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 9), 1.01, 1.01, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, "DEMO01", "fixture"),
        ]
    }
    continuing_anchor_validation = validate_history(
        anchor_config, continuing_anchor_rows
    )
    continuing_anchor_plan = build_preview(
        anchor_config,
        continuing_anchor_rows,
        continuing_anchor_validation["warnings"],
    )
    continuing_anchor_preview = openpyxl.load_workbook(
        continuing_anchor_plan["preview_path"], data_only=False
    )
    try:
        continuing_anchor_sheet = continuing_anchor_preview["Demo Fund"]
        check(
            continuing_anchor_sheet["B4"].value == "=F4/F3-1"
            and continuing_anchor_sheet["B5"].value == "=F4/F2-1",
            "adoption boundary truncated the first new return or full-history summary",
        )
    finally:
        continuing_anchor_preview.close()
    (runtime / "config.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    config = load_config(runtime / "config.json")
    rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 2), 1.0, 1.0, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 9), 1.01, 1.01, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 23), 1.015, 1.015, "DEMO01", "fixture"),
        ]
    }
    unsafe_book = runtime / "unsafe-summary.xlsx"
    create_book(unsafe_book)
    unsafe_workbook = openpyxl.load_workbook(unsafe_book)
    try:
        unsafe_workbook["Demo Fund"]["G4"] = "=SUM(D2:D3)"
        unsafe_workbook.save(unsafe_book)
    finally:
        unsafe_workbook.close()
    unsafe_config = config_for(runtime, unsafe_book)
    try:
        build_preview(unsafe_config, rows)
    except WorkbookError as exc:
        check(
            "summary formula" in str(exc),
            "unsafe summary formula failed for the wrong reason",
        )
    else:
        raise AssertionError("unmanaged summary formula must fail closed")

    managed_array_book = runtime / "managed-array-formula.xlsx"
    create_book(managed_array_book)
    managed_array_workbook = openpyxl.load_workbook(managed_array_book)
    try:
        managed_array_sheet = managed_array_workbook["Demo Fund"]
        managed_array_sheet["G4"] = None
        managed_array_sheet["G3"] = ArrayFormula(ref="G3:G4", text="=D3:D4*2")
        managed_array_workbook.save(managed_array_book)
    finally:
        managed_array_workbook.close()
    managed_array_config = config_for(runtime, managed_array_book)
    try:
        build_preview(managed_array_config, rows)
    except WorkbookError as exc:
        check(
            "array formula" in str(exc).lower(),
            "managed array formula failed closed for the wrong reason",
        )
    else:
        raise AssertionError(
            "managed array formula was copied or moved during automatic insertion"
        )

    review_multi_array_book = runtime / "review-multi-cell-array.xlsx"
    create_book(review_multi_array_book)
    review_multi_array_workbook = openpyxl.load_workbook(review_multi_array_book)
    try:
        review_multi_array_workbook["Demo Fund"]["H4"] = ArrayFormula(
            ref="H4:H5", text="=PRODUCT(1+H2:H3)-1"
        )
        review_multi_array_workbook.save(review_multi_array_book)
    finally:
        review_multi_array_workbook.close()
    review_multi_array_config = config_for(runtime, review_multi_array_book)
    review_multi_array_config["routes"][0]["benchmark"] = None
    review_multi_array_config["routes"][0]["benchmark_review_only"] = True
    try:
        build_preview(review_multi_array_config, rows)
    except WorkbookError as exc:
        check(
            "array formula" in str(exc).lower(),
            "multi-cell review array failed for the wrong reason",
        )
    else:
        raise AssertionError(
            "review-only exception accepted a multi-cell summary array"
        )

    validation = validate_history(config, rows)
    check(validation["passed"], f"historical validation failed: {validation['errors']}")

    reviewed_config = json.loads(json.dumps(config))
    reviewed_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, "DEMO01", "fixture"),
        ]
    }
    reviewed_validation = validate_history(reviewed_config, reviewed_rows)
    check(
        reviewed_validation["passed"]
        and reviewed_validation["routes"][0]["cold_start_kind"]
        == "summary-reviewed-preview"
        and reviewed_validation["routes"][0]["matched_history_dates"] == 0
        and reviewed_validation["warnings"],
        "a unique product-code match with one new date did not reach first preview",
    )
    reviewed_plan = build_preview(reviewed_config, reviewed_rows)
    check(
        reviewed_plan["sheets"][0]["new_dates"] == ["2026-01-16"],
        "review-gated onboarding preview did not append its new date",
    )

    name_only_config = json.loads(json.dumps(config))
    name_only_config["routes"][0].update(
        {"code": None, "product_name": "Example Fund", "allow_sender_only": True}
    )
    name_only_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, None, "fixture"),
        ]
    }
    name_only_validation = validate_history(name_only_config, name_only_rows)
    check(
        name_only_validation["passed"]
        and name_only_validation["routes"][0]["cold_start_kind"]
        == "summary-reviewed-preview",
        "a unique product-name match did not reach first preview without a code",
    )
    name_only_plan = build_preview(name_only_config, name_only_rows)
    name_only_preview = openpyxl.load_workbook(
        name_only_plan["preview_path"], data_only=False
    )
    try:
        check(
            name_only_preview["Demo Fund"]["C4"].value == "Example Fund",
            "name-only onboarding did not preserve the product identity in preview",
        )
    finally:
        name_only_preview.close()

    mixed_name_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, "CODEA", "fixture"),
            NavRow(dt.date(2026, 1, 23), 1.03, 1.03, "CODEB", "fixture"),
        ]
    }
    mixed_name_validation = validate_history(name_only_config, mixed_name_rows)
    check(
        not mixed_name_validation["passed"],
        "name-only onboarding accepted multiple product codes in one routed range",
    )

    one_match_config = json.loads(json.dumps(config))
    one_match_rows = {
        "Demo Fund": [
            NavRow(dt.date(2026, 1, 9), 1.01, 1.01, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, "DEMO01", "fixture"),
        ]
    }
    one_match_validation = validate_history(one_match_config, one_match_rows)
    check(
        one_match_validation["passed"]
        and one_match_validation["routes"][0]["matched_history_dates"] == 1,
        "one matching historical date still blocked a review-gated preview",
    )

    wrong_name_config = json.loads(json.dumps(name_only_config))
    wrong_name_config["routes"][0]["product_name"] = "Different Example Fund"
    wrong_name_validation = validate_history(wrong_name_config, name_only_rows)
    check(
        not wrong_name_validation["passed"],
        "review-gated onboarding accepted a conflicting product name",
    )
    wrong_code_config = json.loads(json.dumps(reviewed_config))
    wrong_code_config["routes"][0]["code"] = "OTHER01"
    wrong_code_validation = validate_history(wrong_code_config, reviewed_rows)
    check(
        not wrong_code_validation["passed"],
        "review-gated onboarding accepted a conflicting product code",
    )
    plan = build_preview(config, rows)
    check(
        len(plan["sheets"]) == 1
        and plan["sheets"][0]["new_dates"] == ["2026-01-16", "2026-01-23"],
        "catch-up plan is incomplete",
    )
    preview_path = Path(plan["preview_path"])
    check(
        "preview-只读审查-" in preview_path.name
        and plan["preview_read_only"]
        and review_file_is_read_only(preview_path)
        and isinstance(plan.get("master_manifest"), dict),
        "review workbook is not clearly named, read-only, or bound to a manifest",
    )
    preview = openpyxl.load_workbook(plan["preview_path"], data_only=False)
    try:
        sheet = preview["Demo Fund"]
        array_formula = preview["Array Audit"]["A1"].value
        check(
            isinstance(array_formula, ArrayFormula)
            and array_formula.ref == "A1:A1"
            and array_formula.text == "=SUM(1,1)",
            "unmanaged array formula was not preserved semantically",
        )
        check(
            sheet["A4"].value == dt.datetime(2026, 1, 16, 0, 0)
            or sheet["A4"].value == dt.date(2026, 1, 16),
            "first catch-up date is wrong",
        )
        check(
            sheet["A5"].value == dt.datetime(2026, 1, 23, 0, 0)
            or sheet["A5"].value == dt.date(2026, 1, 23),
            "second catch-up date is wrong",
        )
        check(
            isinstance(sheet["B4"].value, str) and sheet["B4"].value.startswith("="),
            "weekly return formula is missing",
        )
        check(
            sheet["H4"].value == "='Demo Benchmark'!B4",
            "benchmark date mapping is wrong",
        )
        check(sheet["I4"].value == "=B4-H4", "excess formula is wrong")
        check(sheet["G4"].value == "=D4*2", "unmanaged formula translation failed")
        check(
            sheet["H2"].value is None and sheet["I2"].value is None,
            "existing historical benchmark cells must remain authoritative",
        )
        check(
            sheet["B6"].value == "=F5/F2-1",
            "weekly product summary does not use completed-period anchors",
        )
        check(
            sheet["B3"].value == "=F3/F2-1" and sheet["H3"].value == 0.005,
            "historical formulas were rewritten",
        )
    finally:
        preview.close()

    master_before = book.read_bytes()
    make_file_writable(preview_path)
    tampered = openpyxl.load_workbook(plan["preview_path"])
    tampered["Demo Fund"]["D4"] = 9.99
    tampered.save(plan["preview_path"])
    tampered.close()
    try:
        commit(config)
    except CommitError:
        pass
    else:
        raise AssertionError("commit accepted a preview changed after review")
    check(
        book.read_bytes() == master_before,
        "preview tampering changed the master workbook",
    )

    concurrency_book = runtime / "concurrency-fixture.xlsx"
    shutil.copyfile(book, concurrency_book)
    concurrency_config = deepcopy(config)
    concurrency_config["workbook_path"] = str(concurrency_book)
    concurrency_plan = build_preview(concurrency_config, rows)
    concurrency_workbook = openpyxl.load_workbook(concurrency_book)
    try:
        concurrency_sheet = concurrency_workbook["Demo Fund"]
        concurrency_sheet.insert_rows(4, 1)
        concurrency_sheet["A4"] = dt.date(2026, 1, 16)
        concurrency_sheet["C4"] = "Example Fund"
        concurrency_sheet["D4"] = 1.019
        concurrency_sheet["E4"] = "DEMO01"
        concurrency_sheet["F4"] = 1.019
        concurrency_workbook.save(concurrency_book)
    finally:
        concurrency_workbook.close()
    external_sha256 = file_sha256(concurrency_book)
    try:
        commit(concurrency_config)
    except CommitError:
        pass
    else:
        raise AssertionError("commit accepted an externally changed master workbook")
    check(
        file_sha256(concurrency_book) == external_sha256,
        "concurrency block rolled back or overwrote the external workbook change",
    )
    concurrency_report_path = runtime / "concurrency-report.json"
    concurrency_report = json.loads(
        concurrency_report_path.read_text(encoding="utf-8")
    )
    report_json = json.dumps(
        concurrency_report, ensure_ascii=False, sort_keys=True
    )
    changed_demo = next(
        (
            sheet
            for sheet in concurrency_report["sheets"]
            if sheet.get("sheet") == "Demo Fund"
        ),
        None,
    )
    check(
        concurrency_report["blocked"]
        and concurrency_report["phase"] == "commit-preflight"
        and concurrency_report["plan_id"] == concurrency_plan["plan_id"]
        and concurrency_report["binary_hash_changed"]
        and concurrency_report["changed_sheet_count"] >= 1
        and changed_demo is not None
        and changed_demo["row_delta"] == 1
        and changed_demo["sampled_changed_cells"],
        "external workbook change did not produce a useful structured concurrency report",
    )
    check(
        "1.019" not in report_json
        and "DEMO01" not in report_json
        and "Example Fund" not in report_json,
        "concurrency report leaked workbook cell values",
    )

    dense_rows = {
        "Demo Fund": [
            NavRow(date, value, value, "DEMO01", "fixture")
            for date, value in (
                (dt.date(2026, 1, 2), 1.0),
                (dt.date(2026, 1, 5), 1.002),
                (dt.date(2026, 1, 6), 1.004),
                (dt.date(2026, 1, 7), 1.006),
                (dt.date(2026, 1, 8), 1.008),
                (dt.date(2026, 1, 9), 1.01),
                (dt.date(2026, 1, 12), 1.012),
                (dt.date(2026, 1, 13), 1.014),
                (dt.date(2026, 1, 14), 1.016),
                (dt.date(2026, 1, 15), 1.018),
                (dt.date(2026, 1, 16), 1.02),
                (dt.date(2026, 1, 19), 1.019),
                (dt.date(2026, 1, 20), 1.018),
                (dt.date(2026, 1, 21), 1.017),
                (dt.date(2026, 1, 22), 1.016),
                (dt.date(2026, 1, 23), 1.015),
            )
        ]
    }
    dense_validation = validate_history(config, dense_rows)
    check(
        dense_validation["passed"]
        and dense_validation["routes"][0]["data_frequency"] == "weekly"
        and dense_validation["routes"][0]["data_frequency_source"]
        == "workbook-history",
        "weekly template frequency was not inferred from existing history",
    )
    dense_plan = build_preview(config, dense_rows)
    check(
        dense_plan["sheets"][0]["new_dates"] == ["2026-01-16", "2026-01-23"]
        and dense_plan["sheets"][0]["data_frequency"] == "weekly",
        "daily email history was not reduced to the weekly template dates",
    )
    dense_preview = openpyxl.load_workbook(dense_plan["preview_path"], data_only=False)
    try:
        dense_sheet = dense_preview["Demo Fund"]
        check(
            dense_sheet.max_row == 6
            and dense_sheet["A6"].value == "TOTAL"
            and dense_sheet["B4"].value == "=F4/F3-1",
            "weekly preview did not preserve the summary row and formulas",
        )
    finally:
        dense_preview.close()

    current_week_date = dt.date.today()
    current_week_book = runtime / "current-week-baseline.xlsx"
    create_book(current_week_book)
    current_week_workbook = openpyxl.load_workbook(current_week_book)
    try:
        current_week_sheet = current_week_workbook["Demo Fund"]
        current_week_sheet["A2"] = current_week_date - dt.timedelta(days=14)
        current_week_sheet["A3"] = current_week_date - dt.timedelta(days=7)
        current_week_workbook.save(current_week_book)
    finally:
        current_week_workbook.close()
    current_week_config = config_for(runtime, current_week_book)
    current_week_config["routes"][0]["series_start"] = (
        current_week_date - dt.timedelta(days=14)
    ).isoformat()
    current_week_config["routes"][0]["benchmark"] = None
    current_week_rows = {
        "Demo Fund": [
            NavRow(
                current_week_date,
                1.011,
                1.011,
                "DEMO01",
                "fixture-current-week",
            )
        ]
    }
    current_week_master = current_week_book.read_bytes()
    current_week_validation = validate_history(
        current_week_config, current_week_rows
    )
    check(
        current_week_validation["passed"]
        and current_week_validation["routes"][0][
            "pending_current_week_baseline"
        ]
        and current_week_validation["routes"][0][
            "withheld_current_week_dates"
        ]
        == [current_week_date.isoformat()],
        "a verified unfinished natural week did not reach a zero-add baseline",
    )
    current_week_plan = service_preview(current_week_config, current_week_rows)
    current_week_review = Path(current_week_plan["review_path"])
    check(
        not current_week_plan["sheets"]
        and current_week_plan["approval_kind"] == "validated-no-change"
        and current_week_plan["committable"]
        and current_week_plan["preview_path"] is None
        and "未完成自然周暂缓日期数：1"
        in current_week_review.read_text(encoding="utf-8")
        and current_week_book.read_bytes() == current_week_master,
        "unfinished-week baseline wrote data or did not produce a reviewable report",
    )
    wrong_current_week_rows = {
        "Demo Fund": [
            NavRow(
                current_week_date,
                1.011,
                1.011,
                "OTHER01",
                "fixture-current-week",
            )
        ]
    }
    wrong_current_week_validation = validate_history(
        current_week_config, wrong_current_week_rows
    )
    check(
        not wrong_current_week_validation["passed"],
        "unfinished-week baseline accepted a conflicting product code",
    )
    stale_current_week_validation = validate_history(config, current_week_rows)
    check(
        not stale_current_week_validation["passed"]
        and not stale_current_week_validation["routes"][0][
            "pending_current_week_baseline"
        ],
        "unfinished-week baseline accepted an unverified stale workbook tail",
    )

    forced_daily = json.loads(json.dumps(config))
    forced_daily["routes"][0]["data_frequency"] = "daily"
    try:
        build_preview(forced_daily, dense_rows)
    except WorkbookError as exc:
        check("conflicts" in str(exc), "frequency mismatch failed unclearly")
    else:
        raise AssertionError("an explicit daily override replaced a weekly template")

    gap_book = runtime / "daily-gap.xlsx"
    gap_workbook = openpyxl.Workbook()
    gap_sheet = gap_workbook.active
    gap_sheet.title = "Daily Fund"
    gap_sheet.append(["NAV Date", "Product Code", "Unit NAV", "Cumulative NAV"])
    gap_sheet.append([dt.date(2026, 1, 2), "DEMO01", 1.0, 1.0])
    gap_sheet.append([dt.date(2026, 1, 4), "DEMO01", 1.02, 1.02])
    gap_sheet.append(["TOTAL", None, None, None])
    gap_workbook.save(gap_book)
    gap_workbook.close()
    gap_config = config_for(runtime, gap_book)
    gap_config["routes"][0].update(
        {
            "sheet": "Daily Fund",
            "product_name": None,
            "benchmark": None,
            "return_frequency": "daily",
        }
    )
    gap_rows = {
        "Daily Fund": [
            NavRow(dt.date(2026, 1, 2), 1.0, 1.0, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 3), 1.01, 1.01, "DEMO01", "fixture"),
            NavRow(dt.date(2026, 1, 4), 1.02, 1.02, "DEMO01", "fixture"),
        ]
    }
    try:
        build_preview(gap_config, gap_rows)
    except WorkbookError:
        check(
            not (runtime / "plan.json").exists(),
            "failed preview left a committable plan",
        )
    else:
        raise AssertionError("internal historical gaps must require supervised repair")

    plan = build_preview(config, rows)

    application = None
    if use_com:
        from win32com.client import DispatchEx

        master_before_locked_commit = book.read_bytes()
        backups_before = set((runtime / "backups").glob("*"))
        blocking_app = DispatchEx("Excel.Application")
        blocking_app.Visible = False
        blocking_app.DisplayAlerts = False
        blocking_pid = _process_id(blocking_app)
        blocking_book = blocking_app.Workbooks.Open(str(book.resolve()))
        try:
            locked = subprocess.run(
                [
                    sys.executable,
                    "-X",
                    "utf8",
                    "navctl.py",
                    "commit",
                    "--yes-reviewed-preview",
                ],
                cwd=runtime,
                capture_output=True,
                text=True,
                encoding="utf-8",
            )
            locked_report = json.loads(locked.stdout)
            check(
                locked.returncode == 2
                and not locked_report["passed"]
                and "关闭" in locked_report["error"]
                and "Traceback" not in locked.stderr,
                "open-workbook conflict did not return a readable JSON error",
            )
            check(
                book.read_bytes() == master_before_locked_commit,
                "open-workbook conflict changed the master",
            )
            check(
                set((runtime / "backups").glob("*")) == backups_before,
                "failed open-workbook commit retained a backup",
            )
        finally:
            blocking_book.Close(False)
            blocking_app.Quit()
            blocking_book = None
            blocking_app = None
            gc.collect()
            ensure_process_exit(blocking_pid)
        result = commit(config)
        application = str(result["application"])
        check(
            result["changed"] and result["rows"] == 2,
            "COM commit did not apply both rows",
        )
        calculated = openpyxl.load_workbook(book, data_only=True)
        try:
            check(
                abs(calculated["Demo Fund"]["B4"].value - (1.02 / 1.01 - 1)) < 1e-10,
                "COM product return value is wrong",
            )
            check(
                abs(calculated["Demo Fund"]["H4"].value - (-0.002)) < 1e-12,
                "COM benchmark value is wrong",
            )
            check(
                abs(calculated["Demo Fund"]["I4"].value - ((1.02 / 1.01 - 1) + 0.002))
                < 1e-10,
                "COM excess value is wrong",
            )
        finally:
            calculated.close()
    else:
        shutil.copyfile(plan["preview_path"], book)

    second_validation = validate_history(config, rows)
    check(second_validation["passed"], "post-commit history validation failed")
    second = build_preview(config, rows)
    review_path = Path(second["review_path"])
    check(
        not second["sheets"]
        and second["preview_path"] is None
        and second["approval_kind"] == "validated-no-change"
        and review_path.is_file()
        and "待写入日期数：0" in review_path.read_text(encoding="utf-8"),
        "second run must produce a reviewable no-change baseline without a workbook copy",
    )
    check(
        (runtime / "plan.json").is_file(),
        "a validated no-change preview must leave an approvable plan",
    )
    check(
        second["preview_read_only"] and review_file_is_read_only(review_path),
        "validated no-change review is not system read-only",
    )
    no_change_master = book.read_bytes()
    no_change_backups = set((runtime / "backups").glob("*"))
    make_file_writable(review_path)
    review_path.write_text(
        review_path.read_text(encoding="utf-8") + "tampered\n", encoding="utf-8"
    )
    try:
        commit(config)
    except CommitError:
        pass
    else:
        raise AssertionError("commit accepted a changed no-change baseline report")
    check(
        book.read_bytes() == no_change_master,
        "tampered no-change approval changed the master workbook",
    )
    second = build_preview(config, rows)
    reviewed = subprocess.run(
        [
            sys.executable,
            "-X",
            "utf8",
            "navctl.py",
            "commit",
            "--yes-reviewed-preview",
        ],
        cwd=runtime,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    reviewed_report = json.loads(reviewed.stdout)
    check(
        reviewed.returncode == 0
        and not reviewed_report["changed"]
        and reviewed_report["approved_baseline"]
        and reviewed_report["automatic_updates"]["approved"],
        "reviewed no-change baseline did not approve automatic updates",
    )
    check(
        book.read_bytes() == no_change_master
        and set((runtime / "backups").glob("*")) == no_change_backups,
        "no-change baseline approval wrote the workbook or created a backup",
    )
    automatic_noop = automatic_update(config, rows)
    check(
        not automatic_noop["changed"]
        and automatic_noop["rows"] == 0
        and not (runtime / "plan.json").exists(),
        "approved scheduled no-op did not finish successfully and clean its staging plan",
    )
    revoke()

    level_book = runtime / "level-benchmark.xlsx"
    create_book(level_book)
    level_workbook = openpyxl.load_workbook(level_book)
    try:
        for row, value in enumerate((100.0, 101.0, 102.0, 103.0), 2):
            level_workbook["Demo Benchmark"].cell(row, 2).value = value
        level_workbook.save(level_book)
    finally:
        level_workbook.close()
    level_config = config_for(runtime, level_book)
    level_config["routes"][0]["benchmark"]["source_type"] = "level"
    level_plan = build_preview(level_config, rows)
    level_preview = openpyxl.load_workbook(level_plan["preview_path"], data_only=False)
    try:
        level_sheet = level_preview["Demo Fund"]
        check(
            level_sheet["J4"].value == "='Demo Benchmark'!B4",
            "benchmark level mapping is wrong",
        )
        check(
            level_sheet["H4"].value == "=J4/J3-1",
            "benchmark level anchor return is wrong",
        )
        check(
            level_sheet["I4"].value == "=B4-H4", "level-based excess formula is wrong"
        )
        check(
            level_sheet["H6"].value == "='Demo Benchmark'!B5/'Demo Benchmark'!B2-1",
            "benchmark summary anchors are wrong",
        )
    finally:
        level_preview.close()

    plain_book = runtime / "no-return-column.xlsx"
    plain_workbook = openpyxl.Workbook()
    plain_sheet = plain_workbook.active
    plain_sheet.title = "Plain Fund"
    plain_sheet.append(["NAV Date", "Unit NAV", "Cumulative NAV"])
    plain_sheet.append([dt.date(2026, 1, 2), 1.0, 1.0])
    plain_sheet.append([dt.date(2026, 1, 9), 1.01, 1.01])
    plain_sheet.append(["TOTAL", None, None])
    plain_workbook.save(plain_book)
    plain_workbook.close()
    plain_config = config_for(runtime, plain_book)
    plain_config["routes"][0].update(
        {
            "sheet": "Plain Fund",
            "code": None,
            "product_name": "Plain Fund",
            "benchmark": None,
        }
    )
    plain_rows = {
        "Plain Fund": [
            NavRow(dt.date(2026, 1, 2), 1.0, 1.0, None, "fixture"),
            NavRow(dt.date(2026, 1, 9), 1.01, 1.01, None, "fixture"),
            NavRow(dt.date(2026, 1, 16), 1.02, 1.02, None, "fixture"),
        ]
    }
    plain_plan = build_preview(plain_config, plain_rows)
    check(
        plain_plan["sheets"][0]["new_dates"] == ["2026-01-16"],
        "workbook without a return column could not produce a preview",
    )
    plain_sparse_rows = {
        "Plain Fund": [NavRow(dt.date(2026, 1, 16), 1.02, 1.02, None, "fixture")]
    }
    plain_sparse_validation = validate_history(plain_config, plain_sparse_rows)
    check(
        plain_sparse_validation["passed"]
        and plain_sparse_validation["routes"][0]["cold_start_kind"]
        == "summary-reviewed-preview",
        "a worksheet-name product identity did not unlock a sparse first preview",
    )

    reserved_book = runtime / "summary-reserved-cold-start.xlsx"
    reserved_workbook = openpyxl.Workbook()
    reserved_sheet = reserved_workbook.active
    reserved_sheet.title = "Reserved Fund"
    reserved_sheet.append(
        ["Product Code", "NAV Date", "Unit NAV", "Cumulative NAV", "Weekly Return"]
    )
    reserved_sheet.append(["DEMO43", dt.date(2026, 1, 1), None, None, None])
    reserved_sheet.append(["累计", None, None, None, None])
    reserved_workbook.save(reserved_book)
    reserved_workbook.close()
    reserved_config = config_for(runtime, reserved_book)
    reserved_config["routes"][0].update(
        {
            "sheet": "Reserved Fund",
            "code": "DEMO43",
            "benchmark": None,
            "return_basis": "cumulative",
        }
    )
    single_reserved_rows = {
        "Reserved Fund": [NavRow(dt.date(2026, 1, 2), 1.0, 1.0, "DEMO43", "fixture")]
    }
    single_reserved_validation = validate_history(reserved_config, single_reserved_rows)
    check(
        single_reserved_validation["passed"]
        and single_reserved_validation["routes"][0]["cold_start_kind"]
        == "summary-reserved-row",
        "one real email date did not unlock a reserved-row first preview",
    )
    single_reserved_plan = build_preview(
        reserved_config,
        single_reserved_rows,
        single_reserved_validation["warnings"],
    )
    check(
        single_reserved_plan["sheets"][0]["filled_existing_rows"] == [2]
        and single_reserved_plan["sheets"][0]["new_dates"] == ["2026-01-02"],
        "one-date reserved-row preview did not populate the placeholder",
    )
    reserved_rows = {
        "Reserved Fund": [
            NavRow(date, value, value, "DEMO43", "fixture")
            for date, value in (
                (dt.date(2026, 1, 2), 1.0),
                (dt.date(2026, 1, 5), 1.002),
                (dt.date(2026, 1, 6), 1.004),
                (dt.date(2026, 1, 7), 1.006),
                (dt.date(2026, 1, 8), 1.008),
                (dt.date(2026, 1, 9), 1.01),
                (dt.date(2026, 1, 12), 1.012),
                (dt.date(2026, 1, 13), 1.014),
                (dt.date(2026, 1, 14), 1.016),
                (dt.date(2026, 1, 15), 1.018),
                (dt.date(2026, 1, 16), 1.02),
            )
        ]
    }
    reserved_validation = validate_history(reserved_config, reserved_rows)
    check(
        reserved_validation["passed"]
        and reserved_validation["warnings"]
        and reserved_validation["routes"][0]["cold_start_kind"]
        == "summary-reserved-row"
        and reserved_validation["routes"][0]["data_frequency"] == "weekly",
        "reserved summary row was not recognized as a safe cold start",
    )
    reserved_plan = build_preview(
        reserved_config, reserved_rows, reserved_validation["warnings"]
    )
    reserved_sheet_plan = reserved_plan["sheets"][0]
    check(
        reserved_sheet_plan["insert_count"] == 2
        and reserved_sheet_plan["populated_count"] == 3
        and reserved_sheet_plan["filled_existing_rows"] == [2]
        and reserved_sheet_plan["new_dates"]
        == ["2026-01-02", "2026-01-09", "2026-01-16"],
        "reserved-row plan did not distinguish the filled row from inserted rows",
    )
    reserved_preview = openpyxl.load_workbook(
        reserved_plan["preview_path"], data_only=False
    )
    try:
        preview_sheet = reserved_preview["Reserved Fund"]
        check(
            (
                preview_sheet["B2"].value == dt.datetime(2026, 1, 2)
                or preview_sheet["B2"].value == dt.date(2026, 1, 2)
            )
            and preview_sheet["C2"].value == 1.0
            and preview_sheet["D2"].value == 1.0
            and preview_sheet["C4"].value == 1.02,
            "reserved first row or later email history was not populated",
        )
        check(
            preview_sheet["A5"].value == "累计"
            and preview_sheet["E4"].value == "=D4/D3-1"
            and preview_sheet["E5"].value == "=D4/D2-1",
            "summary row or managed formulas were not preserved after cold start",
        )
    finally:
        reserved_preview.close()
    if use_com:
        reserved_result = commit(reserved_config)
        application = str(reserved_result["application"])
        check(
            reserved_result["changed"] and reserved_result["rows"] == 3,
            "COM reserved-row cold start did not report all populated dates",
        )
    else:
        shutil.copyfile(reserved_plan["preview_path"], reserved_book)
    reserved_second_validation = validate_history(reserved_config, reserved_rows)
    check(
        reserved_second_validation["passed"]
        and not reserved_second_validation["warnings"],
        "reserved-row cold start did not become ordinary verified history",
    )
    reserved_second = build_preview(reserved_config, reserved_rows)
    check(
        not reserved_second["sheets"] and reserved_second["preview_path"] is None,
        "reserved-row cold start was not idempotent after the first write",
    )

    unsafe_reserved_book = runtime / "unsafe-reserved-row.xlsx"
    unsafe_reserved_workbook = openpyxl.Workbook()
    unsafe_reserved_sheet = unsafe_reserved_workbook.active
    unsafe_reserved_sheet.title = "Reserved Fund"
    unsafe_reserved_sheet.append(
        ["Product Code", "NAV Date", "Unit NAV", "Cumulative NAV", "Note"]
    )
    unsafe_reserved_sheet.append(
        ["DEMO43", dt.date(2026, 1, 1), None, None, "keep this content"]
    )
    unsafe_reserved_sheet.append(["累计", None, None, None, None])
    unsafe_reserved_workbook.save(unsafe_reserved_book)
    unsafe_reserved_workbook.close()
    unsafe_reserved_config = config_for(runtime, unsafe_reserved_book)
    unsafe_reserved_config["routes"][0].update(
        {"sheet": "Reserved Fund", "code": "DEMO43", "benchmark": None}
    )
    unsafe_reserved_validation = validate_history(unsafe_reserved_config, reserved_rows)
    check(
        not unsafe_reserved_validation["passed"]
        and not unsafe_reserved_validation["routes"][0]["cold_start"],
        "a row containing extra business content was treated as a placeholder",
    )

    header_book = runtime / "append-header-only.xlsx"
    header_workbook = openpyxl.Workbook()
    header_sheet = header_workbook.active
    header_sheet.title = "Header Fund"
    header_sheet.append(["净值日期", "产品名称", "单位净值"])
    header_workbook.save(header_book)
    header_workbook.close()
    header_config = config_for(runtime, header_book)
    header_config["routes"][0].update(
        {
            "sheet": "Header Fund",
            "sheet_mode": "append",
            "code": None,
            "product_name": "Example Name Only Fund",
            "benchmark": None,
            "return_basis": "unit",
        }
    )
    header_config["column_overrides"] = {}
    header_rows = {
        "Header Fund": [NavRow(dt.date(2026, 1, 2), 1.0, None, None, "fixture")]
    }
    header_validation = validate_history(header_config, header_rows)
    check(
        header_validation["passed"] and header_validation["warnings"],
        "name-only header sheet could not enter append cold start",
    )
    header_plan = build_preview(
        header_config, header_rows, header_validation["warnings"]
    )
    header_preview = openpyxl.load_workbook(
        header_plan["preview_path"], data_only=False
    )
    try:
        check(
            header_preview["Header Fund"]["B2"].value == "Example Name Only Fund"
            and header_preview["Header Fund"]["C2"].value == 1.0,
            "name-only header sheet did not receive the minimum record",
        )
    finally:
        header_preview.close()

    append_book = runtime / "append-cold-start.xlsx"
    append_workbook = openpyxl.Workbook()
    append_sheet = append_workbook.active
    append_sheet.title = "New Fund"
    analysis_sheet = append_workbook.create_sheet("Analysis")
    analysis_sheet["A1"] = "Example metric"
    analysis_sheet["B1"] = "=MAX('New Fund'!D:D)"
    append_workbook.save(append_book)
    append_workbook.close()
    append_config = config_for(runtime, append_book)
    append_config["routes"][0].update(
        {
            "sheet": "New Fund",
            "sheet_mode": "append",
            "code": "NEW01",
            "product_name": "Example New Fund",
            "benchmark": None,
            "cumulative_policy": "unit",
            "return_basis": "unit",
        }
    )
    append_config["column_overrides"] = {}
    validate_config(append_config)
    append_rows = {
        "New Fund": [
            NavRow(dt.date(2026, 1, 2), 1.0, None, "NEW01", "fixture"),
            NavRow(dt.date(2026, 1, 9), 1.01, None, "NEW01", "fixture"),
        ]
    }
    append_validation = validate_history(append_config, append_rows)
    check(
        append_validation["passed"]
        and append_validation["warnings"]
        and append_validation["routes"][0]["cold_start"],
        "append cold start did not downgrade missing history to a visible warning",
    )
    append_plan = service_preview(append_config, append_rows)
    check(
        append_plan["sheets"][0]["sheet_mode"] == "append"
        and append_plan["sheets"][0]["new_dates"] == ["2026-01-02", "2026-01-09"],
        "append cold start plan is incomplete",
    )
    check(
        append_plan["warnings"],
        "append cold start warning was not preserved in the preview plan",
    )
    append_preview = openpyxl.load_workbook(
        append_plan["preview_path"], data_only=False
    )
    try:
        new_sheet = append_preview["New Fund"]
        check(
            [new_sheet.cell(1, column).value for column in range(1, 6)]
            == ["净值日期", "产品代码", "产品名称", "单位净值", "累计单位净值"],
            "blank append sheet did not receive canonical headers",
        )
        check(
            new_sheet["B2"].value == "NEW01"
            and new_sheet["C2"].value == "Example New Fund"
            and new_sheet["D3"].value == 1.01
            and new_sheet["E3"].value == 1.01,
            "append preview did not write product identity and NAV values",
        )
        check(
            new_sheet["A4"].value is None
            and append_preview["Analysis"]["B1"].value == "=MAX('New Fund'!D:D)",
            "append preview added a summary row or changed an analysis sheet",
        )
    finally:
        append_preview.close()
    if use_com:
        append_result = commit(append_config)
        application = str(append_result["application"])
        check(
            append_result["changed"] and append_result["rows"] == 2,
            "COM append commit did not apply both cold-start rows",
        )
    else:
        shutil.copyfile(append_plan["preview_path"], append_book)
    approve(append_config)
    check(
        automation_status(append_config)["approved"],
        "first reviewed append commit did not enable automatic updates",
    )
    committed_append = openpyxl.load_workbook(append_book, data_only=False)
    try:
        check(
            committed_append["New Fund"]["A4"].value is None
            and committed_append["Analysis"]["B1"].value == "=MAX('New Fund'!D:D)",
            "append commit added a summary row or changed the analysis sheet",
        )
    finally:
        committed_append.close()
    append_second_validation = validate_history(append_config, append_rows)
    check(
        append_second_validation["passed"] and not append_second_validation["warnings"],
        "append mode did not leave a verifiable history after first commit",
    )
    append_second = build_preview(append_config, append_rows)
    check(
        not append_second["sheets"] and append_second["preview_path"] is None,
        "append mode is not idempotent after first commit",
    )
    append_rows["New Fund"].append(
        NavRow(dt.date(2026, 1, 16), 1.02, None, "NEW01", "fixture")
    )
    append_next_validation = validate_history(append_config, append_rows)
    check(
        append_next_validation["passed"] and not append_next_validation["warnings"],
        "existing append table did not validate before a later update",
    )
    if use_com:
        altered = json.loads(json.dumps(append_config))
        altered["routes"][0]["product_name"] = "Changed Name"
        try:
            automatic_update(altered, append_rows)
        except AutomationError:
            pass
        else:
            raise AssertionError(
                "automatic update accepted a configuration changed after approval"
            )
        append_next_result = automatic_update(append_config, append_rows)
        application = str(append_next_result["application"])
        check(
            append_next_result["changed"]
            and append_next_result["rows"] == 1
            and not (runtime / "plan.json").exists(),
            "automatic update did not write one row and remove its staging plan",
        )
    else:
        append_next = build_preview(append_config, append_rows)
        check(
            append_next["sheets"][0]["new_dates"] == ["2026-01-16"]
            and append_next["sheets"][0]["copy_template_rows"],
            "existing append table did not plan exactly one later date",
        )
        shutil.copyfile(append_next["preview_path"], append_book)
    append_updated = openpyxl.load_workbook(append_book, data_only=False)
    try:
        check(
            append_updated["New Fund"]["D4"].value == 1.02
            and append_updated["New Fund"]["A5"].value is None
            and append_updated["Analysis"]["B1"].value == "=MAX('New Fund'!D:D)",
            "later append update added a footer or changed the analysis sheet",
        )
    finally:
        append_updated.close()
    return application


def template_tests(runtime: Path, use_com: bool) -> str | None:
    from nav_automation import approve, automatic_update
    from nav_commit import commit, ensure_process_exit, spreadsheet_app
    from nav_config import validate_config
    from nav_parse import NavRow
    from nav_template import TemplateError, init_template
    from nav_workbook import build_preview, validate_history

    target = runtime / "全新脱敏模板.xlsx"
    route_specs = [
        ("周度无指数", "WEEK00", "weekly", None),
        ("周度有指数", "WEEK01", "weekly", "示例指数"),
        ("日度无指数", "DAY000", "daily", None),
        ("日度有指数", "DAY001", "daily", "示例指数"),
    ]
    routes = []
    for index, (sheet, code, frequency, benchmark_name) in enumerate(route_specs, 1):
        benchmark = None
        if benchmark_name:
            benchmark = {
                "source_sheet": "指数数据",
                "source_type": "level",
                "source_date": "A",
                "source_value": "B",
                "display_name": benchmark_name,
            }
        routes.append(
            {
                "sender": f"sender{index}@example.invalid",
                "subject_contains": code,
                "sheet": sheet,
                "sheet_mode": "template",
                "code": code,
                "product_name": f"示例产品{index}",
                "parser": "auto",
                "allow_sender_only": False,
                "cumulative_policy": "require",
                "return_basis": "cumulative",
                "return_frequency": frequency,
                "data_frequency": frequency,
                "series_start": "2026-01-02",
                "benchmark": benchmark,
            }
        )
    config = {
        "schema_version": 1,
        "runtime_id": "00000000-0000-4000-8000-000000000170",
        "workbook_path": str(target.resolve()),
        "workbook_mode": "bundled-template",
        "imap": {
            "host": "imap.example.invalid",
            "port": 993,
            "user": "user@example.invalid",
            "mailbox": "INBOX",
            "lookback_days": 180,
        },
        "routes": routes,
        "column_overrides": {},
        "style": {"mode": "cn-red-up-green-down", "zero_threshold": 0.00005},
        "schedule": [{"days": ["MON", "WED", "FRI"], "time": "09:00"}],
        "validation": {"minimum_history_dates": 2, "tolerance": 0.000001},
    }
    validate_config(config)
    initialized = init_template(config)
    check(
        initialized["product_sheets"]
        == ["周度无指数", "周度有指数", "日度无指数", "日度有指数"]
        and initialized["benchmark_source_sheets"] == ["指数数据"]
        and target.is_file(),
        "template initializer did not create four product sheets and one shared source sheet",
    )
    try:
        init_template(config)
    except TemplateError as exc:
        check("refusing to overwrite" in str(exc), "overwrite refusal is unclear")
    else:
        raise AssertionError("template initializer overwrote an existing workbook")

    workbook = openpyxl.load_workbook(target, data_only=False)
    dates_by_frequency = {
        "weekly": [
            dt.date(2026, 1, 2),
            dt.date(2026, 1, 9),
            dt.date(2026, 1, 16),
        ],
        "daily": [
            dt.date(2026, 1, 2),
            dt.date(2026, 1, 5),
            dt.date(2026, 1, 6),
        ],
    }
    try:
        check(
            workbook.sheetnames
            == ["周度无指数", "周度有指数", "日度无指数", "日度有指数", "指数数据"],
            "generated workbook sheet order is wrong",
        )
        expected_headers = {
            "周度无指数": [
                "产品代码",
                "产品名称",
                "单位净值",
                "累计单位净值",
                "净值日期",
                "周收益",
            ],
            "周度有指数": [
                "产品代码",
                "产品名称",
                "单位净值",
                "累计单位净值",
                "净值日期",
                "周收益",
                "示例指数",
                "指数收益(周度)",
                "超额(周度)",
            ],
            "日度无指数": [
                "产品代码",
                "产品名称",
                "单位净值",
                "累计单位净值",
                "净值日期",
                "日收益",
                "周收益",
            ],
            "日度有指数": [
                "产品代码",
                "产品名称",
                "单位净值",
                "累计单位净值",
                "净值日期",
                "日收益",
                "周收益",
                "示例指数收益(日度)",
                "超额(日度)",
            ],
        }
        for name, headers in expected_headers.items():
            sheet = workbook[name]
            check(
                [sheet.cell(2, column).value for column in range(1, len(headers) + 1)]
                == headers,
                f"{name}: generated headers are wrong",
            )
            check(
                sheet["A1"].value is None
                and list(sheet.merged_cells.ranges)
                and sheet["B2"].font.name == "等线"
                and sheet["B2"].font.sz == 10
                and sheet["F3"].fill.fgColor.rgb.endswith("FFF2CC")
                and sum(len(item.rules) for item in sheet.conditional_formatting) >= 2,
                f"{name}: template style or dynamic return coloring is incomplete",
            )
        check(
            workbook["日度有指数"]["I3"].fill.fgColor.rgb.endswith("FFFF00"),
            "daily excess column is not bright yellow",
        )
        source = workbook["指数数据"]
        check(
            [source.cell(1, column).value for column in range(1, 4)]
            == ["日期", "示例指数点位", "来源"],
            "shared index source headers are wrong",
        )
        source_points = {
            dt.date(2026, 1, 2): 100.0,
            dt.date(2026, 1, 5): 100.2,
            dt.date(2026, 1, 6): 100.3,
            dt.date(2026, 1, 9): 101.0,
            dt.date(2026, 1, 16): 100.5,
        }
        for row, (date, level) in enumerate(sorted(source_points.items()), 2):
            source.cell(row, 1).value = date
            source.cell(row, 2).value = level
            source.cell(row, 3).value = "https://example.invalid/index"
        workbook.save(target)
    finally:
        workbook.close()

    values = {
        "周度无指数": [1.0, 1.02, 1.03],
        "周度有指数": [1.0, 1.01, 1.02],
        "日度无指数": [1.0, 0.99, 1.01],
        "日度有指数": [1.0, 0.98, 0.97],
    }

    def rows(count: int) -> dict[str, list[NavRow]]:
        return {
            sheet: [
                NavRow(date, value, value, code, "fixture")
                for date, value in zip(
                    dates_by_frequency[frequency][:count], values[sheet][:count]
                )
            ]
            for sheet, code, frequency, _benchmark in route_specs
        }

    first_rows = rows(1)
    first_validation = validate_history(config, first_rows)
    check(
        first_validation["passed"]
        and len(first_validation["warnings"]) == 4
        and all(
            item["cold_start_kind"] == "bundled-template"
            for item in first_validation["routes"]
        ),
        "one-date bundled template cold start was not accepted with visible warnings",
    )
    first_plan = build_preview(config, first_rows, first_validation["warnings"])
    check(
        len(first_plan["sheets"]) == 4
        and all(item["new_dates"] == ["2026-01-02"] for item in first_plan["sheets"]),
        "one-date template preview is incomplete",
    )
    application = None
    if use_com:
        application = str(commit(config)["application"])
    else:
        shutil.copyfile(first_plan["preview_path"], target)
    one_date_validation = validate_history(config, first_rows)
    check(
        one_date_validation["passed"] and one_date_validation["warnings"],
        "template did not keep the cold-start warning until two dates were verified",
    )

    second_rows = rows(2)
    second_validation = validate_history(config, second_rows)
    check(
        second_validation["passed"] and second_validation["warnings"],
        "second date was not allowed to enter the reviewed template preview",
    )
    second_plan = build_preview(config, second_rows, second_validation["warnings"])
    preview = openpyxl.load_workbook(second_plan["preview_path"], data_only=False)
    try:
        weekly = preview["周度有指数"]
        daily = preview["日度有指数"]
        check(
            weekly["F4"].value == "=D4/D3-1"
            and weekly["G3"].value == "='指数数据'!B2"
            and weekly["G4"].value == "='指数数据'!B5"
            and weekly["H4"].value == "=G4/G3-1"
            and weekly["I4"].value == "=F4-H4",
            "weekly benchmark return or excess formulas are wrong: "
            f"{[weekly[cell].value for cell in ('F4', 'G3', 'G4', 'H4', 'I4')]}",
        )
        check(
            daily["F4"].value == "=D4/D3-1"
            and daily["G4"].value == "=D4/D3-1"
            and daily["H4"].value == "='指数数据'!B3/'指数数据'!B2-1"
            and daily["I4"].value == "=F4-H4"
            and daily["F5"].value == "=D4/D3-1"
            and daily["G5"].value is None,
            "daily/weekly product returns or daily excess formulas are wrong",
        )
        check(
            set(second_plan["sheets"][3]["return_columns"]) == {6, 7, 8, 9},
            "daily benchmark style plan does not include all return columns",
        )
    finally:
        preview.close()
    if use_com:
        application = str(commit(config)["application"])
        app, _progid, process_id = spreadsheet_app()
        book = None
        try:
            book = app.Workbooks.Open(
                str(target.resolve()), ReadOnly=True, UpdateLinks=0
            )
            app.CalculateFull()
            red = int(book.Worksheets("周度无指数").Range("F4").Font.Color)
            green = int(book.Worksheets("日度有指数").Range("F4").Font.Color)
            check(
                red == 255 and green == 176 * 256 + 80 * 65536,
                "Excel/WPS did not display red-up/green-down fixed colors",
            )
        finally:
            if book is not None:
                book.Close(SaveChanges=False)
            app.Quit()
            app = None
            gc.collect()
            ensure_process_exit(process_id)
    else:
        shutil.copyfile(second_plan["preview_path"], target)
    strict = validate_history(config, second_rows)
    check(
        strict["passed"] and not strict["warnings"],
        "two verified template dates did not restore strict validation",
    )
    no_op = build_preview(config, second_rows)
    check(
        not no_op["sheets"] and no_op["preview_path"] is None,
        "template update is not idempotent after two verified dates",
    )

    third_rows = rows(3)
    third_validation = validate_history(config, third_rows)
    check(
        third_validation["passed"] and not third_validation["warnings"],
        "later template increment did not pass strict history validation",
    )
    if use_com:
        approve(config)
        result = automatic_update(config, third_rows)
        application = str(result["application"])
        check(
            result["changed"] and result["rows"] == 4,
            "scheduled automatic update did not write one date to every product sheet",
        )
    else:
        third_plan = build_preview(config, third_rows)
        expected_increment = {
            "周度无指数": ["2026-01-16"],
            "周度有指数": ["2026-01-16"],
            "日度无指数": ["2026-01-06"],
            "日度有指数": ["2026-01-06"],
        }
        check(
            all(
                item["new_dates"] == expected_increment[item["sheet"]]
                for item in third_plan["sheets"]
            ),
            "later template preview did not contain exactly one incremental date",
        )
        shutil.copyfile(third_plan["preview_path"], target)
    final = validate_history(config, third_rows)
    check(
        final["passed"]
        and not final["warnings"]
        and not build_preview(config, third_rows)["sheets"],
        "template workflow is not strict and idempotent after the later increment",
    )
    return application


def product_lifecycle_tests(runtime: Path, use_com: bool) -> str | None:
    from nav_automation import approve, status as approval_status
    from nav_commit import CommitError, commit
    from nav_config import ConfigError, load_config, validate_config
    from nav_parse import NavRow, choose_route_rows
    from nav_product_workbook import _same_cell, prepare_clone_spec
    from nav_products import (
        ProductError,
        add,
        add_code_alias,
        adopt,
        clone,
        pause,
        resume,
        status,
        sync,
    )
    from nav_template import init_template
    from nav_workbook import WorkbookError, build_preview, validate_history

    target = runtime / "产品生命周期模板.xlsx"
    config_path = runtime / "product-lifecycle-config.json"
    initial_route = {
        "sender": "initial@example.invalid",
        "subject_contains": "BASE001",
        "sheet": "初始产品",
        "sheet_mode": "template",
        "code": "BASE001",
        "product_name": "初始示例产品",
        "parser": "auto",
        "paused": False,
        "allow_sender_only": False,
        "cumulative_policy": "require",
        "return_basis": "cumulative",
        "return_frequency": "weekly",
        "data_frequency": "weekly",
        "max_staleness_days": 14,
        "benchmark": {
            "source_sheet": "指数数据",
            "source_type": "level",
            "source_date": "A",
            "source_value": "B",
            "display_name": "示例指数",
        },
    }
    config = {
        "schema_version": 1,
        "runtime_id": "00000000-0000-4000-8000-000000000171",
        "workbook_path": str(target.resolve()),
        "workbook_mode": "bundled-template",
        "imap": {
            "host": "imap.example.invalid",
            "port": 993,
            "user": "user@example.invalid",
            "mailbox": "INBOX",
            "lookback_days": 180,
        },
        "routes": [initial_route],
        "sheet_reviews": {
            "新增产品": {
                "status": "business_review",
                "reason": "等待邮箱候选与工作簿身份核实",
            }
        },
        "column_overrides": {},
        "style": {"mode": "cn-red-up-green-down", "zero_threshold": 0.00005},
        "schedule": [{"days": ["MON", "WED", "FRI"], "time": "09:00"}],
        "validation": {"minimum_history_dates": 2, "tolerance": 0.000001},
    }
    validate_config(config)
    init_template(config)
    config_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    proposal = {
        "passed": True,
        "scan": {},
        "selection": {
            "mode": "sender-subject",
            "sender": "new@example.invalid",
            "subject_contains": "NEW002",
            "subject_product_code": "NEW002",
        },
        "candidates": [
            {
                "sender": "new@example.invalid",
                "message_count": 3,
                "recent_message_times": [],
                "subject_examples": ["NEW002 周净值"],
                "detected_codes": ["NEW002"],
                "first_date": "2026-01-02",
                "latest_date": "2026-01-16",
                "selection": {
                    "mode": "sender-subject",
                    "sender": "new@example.invalid",
                    "subject_contains": "NEW002",
                    "subject_product_code": "NEW002",
                },
                "observations": [
                    {
                        "date": "2026-01-02",
                        "code": "NEW002",
                        "unit": 1.01,
                        "cumulative": 1.01,
                        "source": "body",
                    },
                    {
                        "date": "2026-01-09",
                        "code": "NEW002",
                        "unit": 1.02,
                        "cumulative": 1.02,
                        "source": "body",
                    },
                ],
            }
        ],
        "warnings": [],
        "errors": [],
    }
    (runtime / "route-proposals.json").write_text(
        json.dumps(proposal, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    before = sync(config, refresh=False)
    check(
        before["passed"]
        and len(before["new_candidates"]) == 1
        and before["new_candidates"][0]["code"] == "NEW002",
        "products sync did not identify the new candidate",
    )
    result = add(
        config,
        config_path,
        proposal_index=1,
        sheet="新增产品",
        frequency="weekly",
        product_name="新增示例产品",
        benchmark_source_sheet="新增指数数据",
        benchmark_display_name="新增示例指数",
    )
    check(
        result["changed"]
        and result["requires_preview_approval"]
        and result["route"]["subject_contains"] == "NEW002"
        and result["route"]["subject_product_code"] == "NEW002"
        and result["template"]["product_sheet"] == "新增产品"
        and Path(result["template"]["backup"]).is_file(),
        "products add did not create a backed-up template page",
    )
    updated = load_config(config_path)
    workbook = openpyxl.load_workbook(target, data_only=False)
    try:
        check(
            workbook.sheetnames == ["初始产品", "新增产品", "指数数据", "新增指数数据"]
            and workbook["新增产品"]["F3"].fill.fgColor.rgb.endswith("FFF2CC")
            and workbook["新增产品"]["G2"].value == "新增示例指数"
            and workbook["新增指数数据"]["B1"].value == "新增示例指数点位"
            and sum(
                len(item.rules) for item in workbook["新增产品"].conditional_formatting
            )
            >= 2,
            "new template product page has wrong order or style",
        )
    finally:
        workbook.close()
    check(
        len(updated["routes"]) == 2
        and "新增产品" not in updated.get("sheet_reviews", {})
        and not approval_status(updated)["approved"]
        and not sync(updated, refresh=False)["new_candidates"],
        "new product route or approval state is wrong",
    )
    alias_proposal = deepcopy(proposal)
    alias_candidate = alias_proposal["candidates"][0]
    alias_candidate["detected_codes"] = ["NEW002A"]
    alias_candidate["observations"] = [
        {
            "date": "2026-01-16",
            "code": "NEW002A",
            "unit": 1.03,
            "cumulative": 1.03,
            "source": "body",
        }
    ]
    (runtime / "route-proposals.json").write_text(
        json.dumps(alias_proposal, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    alias_analysis = sync(updated, refresh=False)
    check(
        len(alias_analysis["alias_pending"]) == 1
        and alias_analysis["alias_pending"][0]["possible_sheet"] == "新增产品"
        and alias_analysis["alias_pending"][0]["requires_confirmation"],
        "share-class suffix difference was not reported as a confirmation-only alias",
    )
    alias_result = add_code_alias(
        updated, config_path, sheet="新增产品", code="NEW002A"
    )
    updated = load_config(config_path)
    alias_rows = choose_route_rows(
        [
            NavRow(
                dt.date(2026, 1, 16),
                1.03,
                1.03,
                "NEW002A",
                "body",
            )
        ],
        updated["routes"][1],
        False,
    )
    check(
        alias_result["changed"]
        and updated["routes"][1]["code_aliases"] == ["NEW002A"]
        and alias_rows[0].code == "NEW002",
        "confirmed exact code alias was not persisted and routed to the primary code",
    )

    approve(updated)
    paused_result = pause(
        updated, config_path, sheet="新增产品", reason="示例产品暂时停更"
    )
    paused = load_config(config_path)
    check(
        paused_result["approval_preserved"]
        and approval_status(paused)["approved"]
        and paused["routes"][1]["paused"]
        and target.is_file(),
        "pausing one product did not preserve history and narrowed approval",
    )
    resumed_result = resume(paused, config_path, sheet="新增产品")
    resumed = load_config(config_path)
    check(
        resumed_result["requires_preview_approval"]
        and not approval_status(resumed)["approved"]
        and not resumed["routes"][1]["paused"],
        "resuming a product did not require renewed preview approval",
    )
    report = status(resumed)
    check(
        report["active"] == 2
        and report["paused"] == 0
        and not report["workbook_missing_sheets"],
        "products status returned the wrong lifecycle state",
    )
    cli_status = subprocess.run(
        [
            sys.executable,
            "-X",
            "utf8",
            "navctl.py",
            "--config",
            str(config_path),
            "products",
            "status",
        ],
        cwd=runtime,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    check(
        cli_status.returncode == 0 and json.loads(cli_status.stdout)["active"] == 2,
        "navctl products status CLI is not wired correctly",
    )

    workbook_before = target.read_bytes()
    config_before = config_path.read_bytes()
    try:
        add(
            resumed,
            config_path,
            proposal_index=1,
            sheet="新增产品",
            frequency="weekly",
        )
    except (RuntimeError, ValueError):
        pass
    else:
        raise AssertionError("products add accepted a duplicate managed worksheet")
    check(
        target.read_bytes() == workbook_before
        and config_path.read_bytes() == config_before,
        "duplicate product rejection changed the workbook or configuration",
    )
    try:
        add(
            resumed,
            config_path,
            proposal_index=1,
            sheet="新增产品",
            frequency="weekly",
            subject_contains="BROADER",
        )
    except RuntimeError as exc:
        check(
            "主题范围" in str(exc),
            "selected proposal scope mismatch did not return a controlled error",
        )
    else:
        raise AssertionError("products add allowed selected mailbox scope replacement")
    check(
        target.read_bytes() == workbook_before
        and config_path.read_bytes() == config_before,
        "selected proposal scope rejection changed the workbook or configuration",
    )

    existing_target = runtime / "产品生命周期已有表.xlsx"
    existing_book = openpyxl.Workbook()
    source_sheet = existing_book.active
    source_sheet.title = "参考产品"
    source_sheet.append(["参考产品费率说明", None, None, None, None, None])
    source_sheet.append(
        ["产品代码", "产品名称", "单位净值", "累计单位净值", "净值日期", "收益（周度）"]
    )
    source_sheet.append(
        ["BASE001", "参考示例产品", 1.0, 1.0, dt.date(2025, 12, 26), "/"]
    )
    source_sheet.append(
        ["BASE001", "参考示例产品", 1.01, 1.01, dt.date(2026, 1, 2), "=D4/D3-1"]
    )
    source_sheet.append(["累计", None, None, None, None, "=D4/D3-1"])
    thin = Side(style="thin", color="000000")
    for row in range(1, 6):
        for column in range(1, 7):
            cell = source_sheet.cell(row, column)
            cell.font = Font(name="等线", size=10, bold=row == 2)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row == 2:
                cell.fill = PatternFill("solid", fgColor="9BC2E6")
            if column == 6 and row >= 3:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.number_format = "0.00%"
    source_sheet.column_dimensions["B"].width = 28
    prepared = existing_book.create_sheet("用户新建页")
    prepared.append(["用户自行填写的产品说明", None, None, None, None, None])
    prepared.append(
        [
            "产品代码",
            "产品名称",
            "单位净值",
            "累计单位净值",
            "净值日期",
            "收益（周度）",
            "示例指数A",
            "示例指数A（日收益）",
            "超额收益",
        ]
    )
    prepared.append(
        ["NEW002", "新增示例产品", 1.01, None, dt.date(2026, 1, 2), None]
    )
    prepared.append(["累计", None, None, None, None, None])
    review_sheet = existing_book.create_sheet("待审基准页")
    review_sheet.append(["待审基准示例", None, None, None, None, None, None, None, None])
    review_sheet.append(
        [
            "产品代码",
            "产品名称",
            "单位净值",
            "累计单位净值",
            "净值日期",
            "收益（周度）",
            "示例指数A",
            "示例指数A（日收益）",
            "超额收益",
        ]
    )
    review_sheet.append(
        [
            "REVIEW03",
            "待审示例产品",
            1.0,
            1.0,
            dt.date(2025, 12, 26),
            "/",
            100.0,
            "/",
            "/",
        ]
    )
    review_sheet.append(
        [
            "REVIEW03",
            "待审示例产品",
            1.01,
            1.01,
            dt.date(2026, 1, 2),
            "=D4/D3-1",
            101.0,
            "=G4/G3-1",
            "=F4-H4",
        ]
    )
    review_sheet.append(
        ["累计", None, None, None, None, "=D4/D3-1", None, "=G4/G3-1", "=F5-H5"]
    )
    review_sheet["H5"] = ArrayFormula(
        ref="H5:H5", text="=PRODUCT(1+H3:H4)-1"
    )
    benchmark_source = existing_book.create_sheet("待审指数源")
    benchmark_source.append(["日期", "指数点位"])
    benchmark_source.append([dt.date(2025, 12, 26), 100.0])
    benchmark_source.append([dt.date(2026, 1, 2), 101.0])
    benchmark_source.append([dt.date(2026, 1, 9), 102.0])
    existing_book.create_sheet("分析页")["A1"] = "用户分析"
    existing_book.save(existing_target)
    existing_book.close()
    existing_config_path = runtime / "existing-product-config.json"
    existing_config = deepcopy(config)
    existing_config["runtime_id"] = "00000000-0000-4000-8000-000000000172"
    existing_config["workbook_path"] = str(existing_target.resolve())
    existing_config["workbook_mode"] = "existing"
    existing_config["style"] = {"mode": "infer", "zero_threshold": 0.00005}
    existing_config["routes"] = [
        {
            **initial_route,
            "sheet": "参考产品",
            "sheet_mode": "summary",
            "data_frequency": "auto",
            "benchmark": None,
        }
    ]
    validate_config(existing_config)
    existing_config_path.write_text(
        json.dumps(existing_config, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    adopt_proposal = deepcopy(proposal)
    adopt_proposal["candidates"][0]["detected_codes"] = ["NEW002", "OTHER999"]
    adopt_proposal["candidates"][0]["observations"].append(
        {
            "date": "2026-01-02",
            "code": "OTHER999",
            "unit": 2.01,
            "cumulative": 2.01,
            "source": "body",
        }
    )
    adopt_proposal["candidates"].append(
        {
            "sender": "review@example.invalid",
            "message_count": 2,
            "recent_message_times": [],
            "subject_examples": ["REVIEW03 周净值"],
            "detected_codes": ["REVIEW03"],
            "first_date": "2025-12-26",
            "latest_date": "2026-01-09",
            "selection": {
                "mode": "sender-subject",
                "sender": "review@example.invalid",
                "subject_contains": "REVIEW03",
            },
            "observations": [
                {
                    "date": "2025-12-26",
                    "code": "REVIEW03",
                    "unit": 1.0,
                    "cumulative": 1.0,
                    "source": "body",
                },
                {
                    "date": "2026-01-02",
                    "code": "REVIEW03",
                    "unit": 1.01,
                    "cumulative": 1.01,
                    "source": "body",
                },
                {
                    "date": "2026-01-09",
                    "code": "REVIEW03",
                    "unit": 1.02,
                    "cumulative": 1.02,
                    "source": "body",
                },
            ],
        }
    )
    (runtime / "route-proposals.json").write_text(
        json.dumps(adopt_proposal, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    existing_before = existing_target.read_bytes()
    existing_result = adopt(
        existing_config,
        existing_config_path,
        proposal_index=1,
        sheet="用户新建页",
    )
    existing_updated = load_config(existing_config_path)
    existing_status = status(existing_updated)
    check(
        existing_result["changed"]
        and existing_result["action"] == "adopted-existing-sheet",
        "products adopt did not report a successful adoption",
    )
    check(
        existing_result["inference"]["sheet_mode"] == "summary"
        and existing_result["inference"]["frequency"] == "weekly",
        f"products adopt inferred the wrong layout: {existing_result['inference']}",
    )
    check(
        not existing_result["inference"]["benchmark_review_only"],
        "empty benchmark/excess columns incorrectly blocked a normal adoption",
    )
    check(
        existing_result["inference"]["code"] == "NEW002",
        "products adopt did not use the worksheet to resolve a multi-code sender",
    )
    check(
        existing_result["route"]["subject_contains"] == "NEW002",
        "products adopt did not inherit the selected proposal subject scope",
    )
    check(
        existing_target.read_bytes() == existing_before,
        "products adopt changed the user workbook",
    )
    check(
        len(existing_updated["routes"]) == 2
        and existing_status["unmanaged_workbook_sheets"]
        == ["待审基准页", "待审指数源", "分析页"],
        f"products adopt saved the wrong route status: {existing_status}",
    )
    adopted_validation = validate_history(
        existing_updated,
        {
            "参考产品": [
                NavRow(dt.date(2025, 12, 26), 1.0, 1.0, "BASE001", "fixture"),
                NavRow(dt.date(2026, 1, 2), 1.01, 1.01, "BASE001", "fixture"),
            ],
            "用户新建页": [
                NavRow(dt.date(2026, 1, 2), 1.01, 1.01, "NEW002", "fixture"),
                NavRow(dt.date(2026, 1, 9), 1.02, 1.02, "NEW002", "fixture"),
            ],
        },
    )
    adopted_report = next(
        item for item in adopted_validation["routes"] if item["sheet"] == "用户新建页"
    )
    check(
        adopted_validation["passed"]
        and adopted_report["cold_start_kind"] == "summary-reserved-row",
        f"partial user-prepared onboarding row was not accepted safely: {adopted_validation}",
    )
    review_before = existing_target.read_bytes()
    review_config_before = existing_config_path.read_bytes()
    history_inspection = adopt(
        existing_updated,
        existing_config_path,
        proposal_index=2,
        sheet="待审基准页",
        history_scope="mail-history",
        inspect_only=True,
    )
    check(
        history_inspection["inference"]["history_scope"] == "mail-history"
        and history_inspection["inference"]["history_scope_default"] == "tail"
        and history_inspection["inference"]["series_start"] is None
        and history_inspection["inference"]["baseline_overlap"] is None
        and existing_config_path.read_bytes() == review_config_before
        and existing_target.read_bytes() == review_before,
        "mail-history adoption scope was not available as a read-only alternative",
    )
    try:
        adopt(
            existing_updated,
            existing_config_path,
            proposal_index=2,
            sheet="待审基准页",
            history_scope="unsupported",
            inspect_only=True,
        )
    except ProductError:
        pass
    else:
        raise AssertionError("products adopt accepted an unknown history scope")
    review_inspection = adopt(
        existing_updated,
        existing_config_path,
        proposal_index=2,
        sheet="待审基准页",
        inspect_only=True,
    )
    check(
        review_inspection["review_required"]
        and not review_inspection["ready_for_direct_adoption"]
        and review_inspection["inference"]["series_start"] == "2026-01-03"
        and review_inspection["inference"]["series_start_reason"]
        == "existing-sheet-tail-plus-one-day"
        and review_inspection["inference"]["baseline_overlap"]
        == "last_existing_point"
        and review_inspection["inference"]["column_roles"]["benchmark_level"]
        == {"column": "G", "header": "示例指数A"}
        and review_inspection["inference"]["column_roles"]["benchmark_return"]
        == {"column": "H", "header": "示例指数A（日收益）"}
        and review_inspection["master_unchanged"]
        and existing_config_path.read_bytes() == review_config_before
        and existing_target.read_bytes() == review_before,
        "read-only adoption inspection did not isolate a benchmark-review "
        f"product: {review_inspection}",
    )
    review_result = adopt(
        existing_updated,
        existing_config_path,
        proposal_index=2,
        sheet="待审基准页",
    )
    review_config = load_config(existing_config_path)
    check(
        review_result["inference"]["benchmark_review_only"]
        and review_result["route"]["series_start"] == "2026-01-03"
        and review_result["route"]["baseline_overlap"]
        == "last_existing_point"
        and review_config["routes"][-1]["benchmark_review_only"] is True
        and status(review_config)["review_required"] == 1
        and existing_target.read_bytes() == review_before,
        "active unresolved benchmark columns did not enter review-only adoption safely",
    )
    review_rows = {
        "参考产品": [
            NavRow(dt.date(2025, 12, 26), 1.0, 1.0, "BASE001", "fixture"),
            NavRow(dt.date(2026, 1, 2), 1.01, 1.01, "BASE001", "fixture"),
        ],
        "用户新建页": [
            NavRow(dt.date(2026, 1, 2), 1.01, 1.01, "NEW002", "fixture"),
            NavRow(dt.date(2026, 1, 9), 1.02, 1.02, "NEW002", "fixture"),
        ],
        "待审基准页": [
            NavRow(dt.date(2025, 12, 26), 1.0, 1.0, "REVIEW03", "fixture"),
            NavRow(dt.date(2026, 1, 2), 1.01, 1.01, "REVIEW03", "fixture"),
            NavRow(dt.date(2026, 1, 9), 1.02, 1.02, "REVIEW03", "fixture"),
        ],
    }
    check(
        _same_cell(
            ArrayFormula(ref="H5:H5", text="=SUM(H3:H4)"),
            ArrayFormula(ref="H5", text="=SUM(H3:H4)"),
        )
        and not _same_cell(
            ArrayFormula(ref="H5:H6", text="=SUM(H3:H4)"),
            ArrayFormula(ref="H5", text="=SUM(H3:H4)"),
        ),
        "single-cell array range normalization weakened array formula comparison",
    )
    review_validation = validate_history(review_config, review_rows)
    check(
        review_validation["passed"]
        and any("只生成审查预览" in item for item in review_validation["warnings"]),
        f"review-only benchmark route did not pass into supervised preview: {review_validation}",
    )
    baseline_config = deepcopy(review_config)
    baseline_config["routes"] = [deepcopy(review_config["routes"][-1])]
    baseline_rows = {
        "待审基准页": review_rows["待审基准页"][:2],
    }
    baseline_validation = validate_history(baseline_config, baseline_rows)
    baseline_plan = build_preview(
        baseline_config, baseline_rows, baseline_validation["warnings"]
    )
    check(
        baseline_validation["passed"]
        and not baseline_plan["sheets"]
        and baseline_plan["committable"] is False
        and Path(baseline_plan["review_path"]).is_file(),
        "zero-change unresolved benchmark state did not produce a review-only baseline",
    )
    try:
        commit(baseline_config)
    except CommitError:
        pass
    else:
        raise AssertionError("review-only zero-change baseline was accepted for approval")
    review_plan = build_preview(
        review_config, review_rows, review_validation["warnings"]
    )
    check(
        review_plan["committable"] is False
        and review_plan["blocking_reviews"]
        == [{"sheet": "待审基准页", "issue": "benchmark-source-unresolved"}],
        f"unresolved benchmark preview was not marked non-committable: {review_plan}",
    )
    review_preview = openpyxl.load_workbook(review_plan["preview_path"], data_only=False)
    try:
        check(
            review_preview["待审基准页"]["E5"].value.date()
            == dt.date(2026, 1, 9)
            and review_preview["待审基准页"]["F5"].value == "=D5/D4-1"
            and review_preview["待审基准页"]["F6"].value == "=D5/D3-1"
            and review_preview["待审基准页"]["G5"].value is None
            and review_preview["待审基准页"]["H5"].value is None
            and review_preview["待审基准页"]["I5"].value is None
            and review_preview["待审基准页"]["H6"].value is None
            and review_preview["待审基准页"]["I6"].value is None,
            "review-only preview did not preserve NAV while blanking unresolved "
            f"benchmark cells: F5={review_preview['待审基准页']['F5'].value!r}, "
            f"F6={review_preview['待审基准页']['F6'].value!r}",
        )
        check(
            review_plan["sheets"][-1]["review_array_formulas_cleared"]
            == ["H6"],
            "review-only preview did not audit the cleared single-cell summary array",
        )
    finally:
        review_preview.close()
    try:
        commit(review_config)
    except CommitError as exc:
        check(
            "review-only" in str(exc),
            "review-only commit rejection did not explain the blocking state",
        )
    else:
        raise AssertionError("review-only preview was accepted for formal commit")
    check(
        existing_target.read_bytes() == review_before,
        "review-only preview or rejected commit changed the formal workbook",
    )

    license_review_config = deepcopy(review_config)
    license_review_route = license_review_config["routes"][-1]
    license_review_route.pop("benchmark_review_only")
    license_review_route["benchmark"] = {
        "source_sheet": "待审指数源",
        "source_type": "level",
        "source_date": "A",
        "source_value": "B",
        "display_name": "待审示例指数",
        "technical_source_verified": True,
        "license_source": "公开条款技术审查记录",
        "review_only": True,
    }
    validate_config(license_review_config)
    license_status = status(license_review_config)
    license_validation = validate_history(license_review_config, review_rows)
    license_plan = build_preview(
        license_review_config,
        review_rows,
        license_validation["warnings"],
    )
    check(
        license_status["review_required"] == 1
        and license_status["routes"][-1]["review_issue"]
        == "benchmark-license-unresolved"
        and license_validation["passed"]
        and any("使用许可尚未确认" in item for item in license_validation["warnings"])
        and license_plan["committable"] is False
        and license_plan["blocking_reviews"]
        == [{"sheet": "待审基准页", "issue": "benchmark-license-unresolved"}],
        "technically verified benchmark did not remain blocked for license review",
    )
    license_outputs = json.dumps(
        [license_status, license_validation, license_plan],
        ensure_ascii=False,
        sort_keys=True,
    )
    check(
        "公开条款技术审查记录" not in license_outputs
        and "license_approved_by" not in license_outputs,
        "benchmark review reports exposed license evidence or approval identity",
    )
    license_preview = openpyxl.load_workbook(
        license_plan["preview_path"], data_only=False
    )
    try:
        check(
            license_preview["待审基准页"]["G5"].value is None
            and license_preview["待审基准页"]["H5"].value is None
            and license_preview["待审基准页"]["I5"].value is None,
            "license-review preview populated unapproved benchmark values",
        )
    finally:
        license_preview.close()
    try:
        commit(license_review_config)
    except CommitError:
        pass
    else:
        raise AssertionError("license-review benchmark preview was committable")
    invalid_license_review = deepcopy(license_review_config)
    invalid_license_review["routes"][-1]["benchmark"][
        "license_approved_by"
    ] = "示例审批人"
    try:
        validate_config(invalid_license_review)
    except ConfigError:
        pass
    else:
        raise AssertionError("review-only benchmark accepted an approval identity")
    invalid_technical_review = deepcopy(license_review_config)
    invalid_technical_review["routes"][-1]["benchmark"][
        "technical_source_verified"
    ] = False
    try:
        validate_config(invalid_technical_review)
    except ConfigError:
        pass
    else:
        raise AssertionError("license review accepted an unverified technical source")

    resolved_config = deepcopy(license_review_config)
    resolved_route = resolved_config["routes"][-1]
    resolved_route["benchmark"].update(
        {
            "license_source": "企业授权数据源审批记录",
            "license_approved_by": "示例数据负责人",
            "review_only": False,
        }
    )
    missing_license_source = deepcopy(resolved_config)
    missing_license_source["routes"][-1]["benchmark"].pop("license_source")
    try:
        validate_config(missing_license_source)
    except ConfigError:
        pass
    else:
        raise AssertionError("approved benchmark omitted its license source")
    validate_config(resolved_config)
    resolved_validation = validate_history(resolved_config, review_rows)
    try:
        build_preview(
            resolved_config, review_rows, resolved_validation["warnings"]
        )
    except WorkbookError as exc:
        check(
            "array formula" in str(exc).lower(),
            "resolved benchmark array failed for the wrong reason",
        )
    else:
        raise AssertionError(
            "review-only array exception leaked into a committable preview"
        )
    resolved_formula_book = runtime / "resolved-benchmark-formula.xlsx"
    shutil.copy2(existing_target, resolved_formula_book)
    resolved_formula_workbook = openpyxl.load_workbook(resolved_formula_book)
    try:
        resolved_formula_workbook["待审基准页"]["H5"] = (
            "=PRODUCT(1+H3:H4)-1"
        )
        resolved_formula_workbook.save(resolved_formula_book)
    finally:
        resolved_formula_workbook.close()
    resolved_config["workbook_path"] = str(resolved_formula_book.resolve())
    resolved_validation = validate_history(resolved_config, review_rows)
    resolved_plan = build_preview(
        resolved_config, review_rows, resolved_validation["warnings"]
    )
    check(
        resolved_validation["passed"]
        and resolved_plan["committable"] is True
        and not resolved_plan["blocking_reviews"],
        "verified benchmark source did not restore a normal committable preview",
    )
    spec = prepare_clone_spec(
        existing_updated, existing_updated["routes"][0], "照参考新增"
    )
    check(
        spec.header_row == 2 and spec.data_row == 3 and spec.target_summary_row == 4,
        "clone specification did not preserve the reference sheet structure",
    )

    application = None
    if use_com:
        clone_proposal = deepcopy(proposal)
        clone_proposal["candidates"][0] = {
            "sender": "clone@example.invalid",
            "message_count": 2,
            "recent_message_times": [],
            "subject_examples": ["CLONE003 周净值"],
            "detected_codes": ["CLONE003"],
            "first_date": "2026-02-06",
            "latest_date": "2026-02-13",
            "selection": {
                "mode": "sender-subject",
                "sender": "clone@example.invalid",
                "subject_contains": "CLONE003",
            },
            "observations": [
                {
                    "date": "2026-02-06",
                    "code": "CLONE003",
                    "unit": 1.1,
                    "cumulative": 1.1,
                    "source": "body",
                },
                {
                    "date": "2026-02-13",
                    "code": "CLONE003",
                    "unit": 1.2,
                    "cumulative": 1.2,
                    "source": "body",
                },
            ],
        }
        (runtime / "route-proposals.json").write_text(
            json.dumps(clone_proposal, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        clone_result = clone(
            review_config,
            existing_config_path,
            proposal_index=1,
            sheet="照参考新增",
            copy_from="参考产品",
            product_name="复制新增示例产品",
        )
        application = clone_result["workbook"]["application"]
        cloned = openpyxl.load_workbook(existing_target, data_only=False)
        try:
            check(
                cloned.sheetnames
                == [
                    "参考产品",
                    "照参考新增",
                    "用户新建页",
                    "待审基准页",
                    "待审指数源",
                    "分析页",
                ]
                and cloned["照参考新增"]["A1"].value is None
                and cloned["照参考新增"]["A2"].value == "产品代码"
                and cloned["照参考新增"]["A3"].value == "CLONE003"
                and cloned["照参考新增"]["B3"].value == "复制新增示例产品"
                and cloned["照参考新增"]["C3"].value is None
                and cloned["照参考新增"]["D3"].value is None
                and cloned["照参考新增"]["E3"].value.date() == dt.date(2026, 2, 6)
                and cloned["照参考新增"]["F3"].value is None
                and cloned["照参考新增"]["A4"].value == "累计"
                and cloned["照参考新增"]["F4"].value is None
                and cloned["参考产品"]["A1"].value == "参考产品费率说明",
                "products clone did not preserve the format or clear source business data",
            )
        finally:
            cloned.close()
        cloned_config = load_config(existing_config_path)
        check(
            len(cloned_config["routes"]) == 4
            and cloned_config["routes"][-1]["sheet"] == "照参考新增"
            and cloned_config["routes"][-1]["subject_contains"] == "CLONE003"
            and Path(clone_result["workbook"]["backup"]).is_file(),
            "products clone did not save its route or backup",
        )
    return application


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--runtime", required=True)
    parser.add_argument("--com", action="store_true")
    args = parser.parse_args()
    runtime = Path(args.runtime).resolve()
    print("[1/6] 检查自动解析、本地解析器、精确发件人和冲突拦截")
    parser_tests(runtime)
    print("      PASS")
    print("[2/6] 检查空路由、暂停路由和无邮件失败关闭")
    route_state_tests(runtime)
    print("      PASS")
    print("[3/6] 检查崩溃恢复和并发运行锁")
    lock_tests(runtime)
    print("      PASS")
    print("[4/6] 检查严格表、空表冷启动、分析页保留、补录、基准和幂等性")
    application = workbook_tests(runtime, args.com)
    print("      PASS")
    print("[5/6] 检查四类脱敏模板、共享指数、冷启动、增量和拒绝覆盖")
    template_application = template_tests(runtime, args.com)
    application = template_application or application
    print("      PASS")
    print("[6/6] 检查产品发现、接管已建 Sheet、照参考新建、暂停、恢复和备份")
    product_application = product_lifecycle_tests(runtime, args.com)
    application = product_application or application
    print("      PASS")
    if args.com:
        print(f"      Excel/WPS COM、文件占用提示及缓存数值：PASS（{application}）")
    else:
        print("      未启用 COM；如需验证正式写入，请添加 --com")
    print("selftest_driver: PASS（全程仅使用虚构数据）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
